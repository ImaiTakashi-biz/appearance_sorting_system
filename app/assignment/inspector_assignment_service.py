"""
検査員割当てロジック
検査員の割当て、スキルマッチング、新製品チーム対応などの機能を提供
"""

from typing import Optional, List, Dict, Any, Tuple, Callable, Set, Union
from collections import defaultdict
from datetime import date, timedelta
from time import perf_counter
import pandas as pd
import numpy as np
import logging
import copy
import os
import openpyxl
from pathlib import Path
import re
from loguru import logger as loguru_logger

from app.utils.perf import perf_timer

logger = logging.getLogger(__name__)

# 定数定義
# 4時間上限ルールの2段階化
PRODUCT_LIMIT_DRAFT_THRESHOLD = 4.5  # ドラフトフェーズでの許容上限（4.5h未満まで許容）
PRODUCT_LIMIT_HARD_THRESHOLD = 4.0   # 最適化フェーズでの厳格上限（4.0h）
PRODUCT_LIMIT_FINAL_TOLERANCE = 4.2  # 最終検証での許容上限（4.2h未満まで許容、代替検査員が見つからない場合）
# 同一品番の割当回数制約
# 環境変数で設定可能（デフォルト値を使用）
try:
    MAX_ASSIGNMENTS_PER_PRODUCT = int(os.getenv("MAX_ASSIGNMENTS_PER_PRODUCT", "1").strip() or "1")
except Exception:
    MAX_ASSIGNMENTS_PER_PRODUCT = 1  # 同一品番の通常割当は1回まで（デフォルト）
MAX_ASSIGNMENTS_PER_PRODUCT = max(1, min(MAX_ASSIGNMENTS_PER_PRODUCT, 5))  # 1回以上5回以下に制限

try:
    MAX_ASSIGNMENTS_PER_PRODUCT_RELAXED = int(os.getenv("MAX_ASSIGNMENTS_PER_PRODUCT_RELAXED", "2").strip() or "2")
except Exception:
    MAX_ASSIGNMENTS_PER_PRODUCT_RELAXED = 2  # 緩和時のみ最大2回まで許容（デフォルト）
MAX_ASSIGNMENTS_PER_PRODUCT_RELAXED = max(1, min(MAX_ASSIGNMENTS_PER_PRODUCT_RELAXED, 5))  # 1回以上5回以下に制限

# 緩和条件の明確化
# 以下の条件のいずれかが満たされる場合、MAX_ASSIGNMENTS_PER_PRODUCT_RELAXEDが適用される：
# 1. relax_work_hours=True が指定された場合（勤務時間制約の緩和時）
# 2. 当日洗浄上がり品で、必要人数に達しない場合
# 3. 未割当ロットの再処理時（フェーズ3）

# 勤務時間チェックの余裕時間
WORK_HOURS_BUFFER = 0.1  # 0.1h（6分）の余裕を確保（累積誤差を考慮して0.05hから増加）
WORK_HOURS_BUFFER_BASE = 0.05  # 基本バッファ（0.05h、3分）
WORK_HOURS_BUFFER_DYNAMIC_FACTOR = 1.0  # 動的バッファ係数（累積誤差に応じて調整可能）
# 追加：勤務時間の10%超過を許容
WORK_HOURS_OVERRUN_RATE = 0.1
SAME_DAY_WORK_HOURS_OVERRUN_RATE = 0.2
HIGH_PRIORITY_RESERVED_CAPACITY_HOURS = 3.5  # 優先ロットの間、各検査員が確保すべき時間（h）
HIGH_PRIORITY_SHIPPING_THRESHOLD = 2        # この番付以下（当日・当日洗浄）を優先ロットとする
SAME_DAY_FORCE_SECOND_INSPECTOR_HOURS = 3.0  # 当日洗浄上がり品で最低2人以上とする検査時間の境界（h）
SAME_DAY_RELAXATION_THRESHOLD_HOURS = 3.0    # 当日洗浄上がり品の制約緩和を始める検査時間の境界（h）

# 品番切替ペナルティ係数
PENALTY_LOT_COUNT_ALPHA = 2.0 / 60.0  # 割当ロット数に対するペナルティ（2分相当を時間に変換）
PENALTY_PRODUCT_VARIETY_BETA = 5.0 / 60.0  # 担当品番種類数に対するペナルティ（5分相当を時間に変換）

# フェーズ間スラッシング防止
TABU_LIST_MAX_ITERATIONS = 5  # 再配置直後のロットを何回のイテレーションで除外するか（3から5に増加）
# 当日洗浄品の同一品名制約緩和を試みる最大回数
MAX_SAME_DAY_SAME_NAME_RELAXATIONS = 2

# 当日洗浄上がり品の制約緩和条件の明確化
# 以下の条件のいずれかが満たされる場合、制約緩和が適用される：
# 1. 検査時間が SAME_DAY_RELAXATION_THRESHOLD_HOURS (3.0h) 以上の場合
# 2. 未割当ロットの再処理時（フェーズ3）で、必要人数に達しない場合
# 3. _should_force_assign_same_day() が True を返す場合（当日洗浄上がり品の場合）
# 緩和される制約：
# - 品番/品名単位の重複禁止制約（同一検査員が複数の品番/品名を担当可能）
# - 同一品番4.0h上限制約（ignore_product_limit=True の場合）
# - 勤務時間制約（relax_work_hours=True の場合）

# 新規品の保護条件
# 環境変数で設定可能（デフォルト値を使用）
try:
    NEW_PRODUCT_PROTECTION_ENABLED = os.getenv("NEW_PRODUCT_PROTECTION_ENABLED", "true").strip().lower() == "true"
except Exception:
    NEW_PRODUCT_PROTECTION_ENABLED = True  # デフォルトは有効

try:
    NEW_PRODUCT_PROTECTION_DAYS = int(os.getenv("NEW_PRODUCT_PROTECTION_DAYS", "14").strip() or "14")
except Exception:
    NEW_PRODUCT_PROTECTION_DAYS = 14  # デフォルトは14日（2週間）
NEW_PRODUCT_PROTECTION_DAYS = max(1, min(NEW_PRODUCT_PROTECTION_DAYS, 90))  # 1日以上90日以下に制限

# 新規品の保護条件の明確化
# 以下の条件のすべてが満たされる場合、新規品は保護される：
# 1. NEW_PRODUCT_PROTECTION_ENABLED=True の場合
# 2. 品番がスキルマスタに存在しない（新規品）
# 3. 出荷予定日が現在日から NEW_PRODUCT_PROTECTION_DAYS 日以内
# 保護される違反：
# - 同一品番4時間超過（relaxed_product_limit_assignmentsに追加）
# - 勤務時間超過（割当を維持）
# 保護される処理：
# - フェーズ1での再割当て（移動対象外）
# - フェーズ2.5での未割当への変更（割当を維持）

# 固定検査員の保護条件
# 固定検査員の保護条件の明確化
# 以下の条件のすべてが満たされる場合、固定検査員は保護される：
# 1. 品番と工程名（現在工程名）に対して固定検査員が設定されている
# 2. その固定検査員がロットに割り当てられている
# 保護される処理：
# - フェーズ1での違反検出・是正対象からの除外（固定検査員が割当済みのロットは違反検出・是正対象外）
# - フェーズ1での再割当て（移動対象外、_is_locked_fixed_preinspection_lotがTrueを返す場合）
# - 偏り是正フェーズでの再割当て（固定検査員が割り当てられている場合は再割当てをスキップ）
# 保護される違反：
# - 勤務時間超過（固定検査員が割当済みのロットは違反検出・是正対象外）
# - 同一品番4時間超過（固定検査員が割当済みのロットは違反検出・是正対象外）


class InspectorAssignmentManager:
    """検査員割当て管理クラス"""
    
    def __init__(
        self,
        log_callback: Optional[Callable[[str], None]] = None,
        debug_mode: bool = False,
        product_limit_hard_threshold: Optional[float] = None,
        required_inspectors_threshold: Optional[float] = None
    ) -> None:
        """
        初期化
        
        Args:
            log_callback: ログ出力用のコールバック関数
            debug_mode: デバッグモード（Trueの場合、詳細なデバッグログを出力）
            product_limit_hard_threshold: 同一品番の4時間上限（Noneの場合はデフォルト値4.0を使用）
            required_inspectors_threshold: 必要人数計算の3時間基準（Noneの場合はデフォルト値3.0を使用）
        """
        self.log_callback = log_callback
        self.debug_mode = debug_mode
        
        # 設定値の適用（Noneの場合はデフォルト値を使用）
        self.product_limit_hard_threshold = (
            product_limit_hard_threshold 
            if product_limit_hard_threshold is not None 
            else PRODUCT_LIMIT_HARD_THRESHOLD
        )
        self.required_inspectors_threshold = (
            required_inspectors_threshold 
            if required_inspectors_threshold is not None 
            else 3.5
        )
        self.same_day_relaxation_threshold = SAME_DAY_RELAXATION_THRESHOLD_HOURS
        self.same_day_force_second_inspector_threshold = SAME_DAY_FORCE_SECOND_INSPECTOR_HOURS
        # 検査員の割り当て履歴を追跡（公平な割り当てのため）
        self.inspector_assignment_count = {}
        self.inspector_last_assignment = {}
        # 検査員の勤務時間を追跡（勤務時間超過を防ぐため）
        self.inspector_work_hours = {}
        self.inspector_daily_assignments = {}
        # 品番ごとの累計作業時間を検査員別に追跡（同一品番の4時間上限判定に使用）
        # 形式: { inspector_code: { product_number: hours } }
        self.inspector_product_hours = {}
        # 初期割当フェーズで4時間上限を緩和した検査員を追跡（後続最適化で優先的に是正する）
        self.relaxed_product_limit_assignments = set()
        # フェーズ間スラッシング防止用のタブーリスト
        # 形式: { lot_index: iteration_count } - 再配置されたロットを一定回数除外
        self.tabu_list = {}
        # 検査員ごとの担当品番種類数を追跡（品番切替ペナルティ用）
        # 形式: { inspector_code: set(product_numbers) }
        self.inspector_product_variety = {}
        # 警告の重複出力を防ぐためのセット
        # 形式: {(警告タイプ, キー)} - 同じ警告を1回だけ出力
        self.logged_warnings = set()
        # 新製品チーム列の確認ログを1回だけ出力するためのフラグ
        self.new_product_team_logged = False
        # 新製品チームメンバー数のログを1回だけ出力するためのフラグ
        self._new_product_team_count_logged = False
        # スキルマスタの列構造ログを1回に抑制（ログ肥大化防止）
        self._skill_master_schema_logged = False
        # 【重要】当日洗浄上がり品および先行検査品の検査員を追跡（品番単位で管理）
        # 形式: {品番: set(検査員コード)} - 各品番ごとに割り当てられた検査員のセット
        # 
        # 制約ルール（このロジックの核心）:
        # - 別品番であれば同一検査員を割り当ててOK
        # - 同一品番の複数ロットには同一検査員を割り当てない（必須制約）
        # 
        # このデータ構造により、理想的な割当て（各ロットに異なる検査員を割り当て）を実現
        # 変更時は慎重に検討すること（再現性の高い割当てロジックの基盤）
        self.same_day_cleaning_inspectors = {}
        # 【追加】当日洗浄上がり品の検査員を追跡（品名単位で管理）
        # 品名が同じで品番が異なる場合、同じ検査員を割り当てない制約用
        # 形式: {品名: set(検査員コード)} - 各品名ごとに割り当てられた検査員のセット
        # 例: "3D025-G4960"と"3D025-M006A"は別品番だが品名が"ｷﾞﾔB"で同じ場合、同じ検査員を割り当てない
        self.same_day_cleaning_inspectors_by_product_name = {}
        # 当日洗浄品の同一品名制約を緩和した回数を追跡（製品/品名単位）
        self.same_day_same_name_relaxation_attempts = {}
        # 品番ごとの割当回数を追跡
        # 形式: {検査員コード: {品番: 回数}}
        self.inspector_product_assignment_counts = {}
        # 【追加】休暇情報を保持
        self.vacation_data = {}  # {検査員名: 休暇情報辞書}
        self.vacation_date = None  # 休暇情報の対象日付
        self.inspector_name_to_vacation = {}  # {検査員名: 休暇情報辞書} - 名前マッピング用
        self.logged_vacation_messages = set()  # (inspector_name, code, interpretation, date)
        # 【追加】swap実施率追跡用
        self.swap_count = 0  # swapが実行された回数
        self.violation_count = 0  # 総違反件数（swap対象となった違反の数）
        # 【追加】勤務時間バッファの効果測定メトリクス
        self.buffer_usage_metrics = {
            'total_checks': 0,  # バッファチェックの総回数
            'buffer_exceeded_count': 0,  # バッファを超えた回数
            'buffer_exceeded_by': [],  # バッファ超過量のリスト
            'dynamic_buffer_adjustments': 0,  # 動的バッファ調整の回数
        }
        # 【追加】タブーリストの効果測定メトリクス
        self.tabu_list_metrics = {
            'total_additions': 0,  # タブーリストへの追加回数
            'total_skips': 0,  # タブーリストによるスキップ回数
            'lot_reassignment_counts': {},  # ロットごとの再配置回数 {lot_index: count}
            'thrashing_detected': [],  # スラッシングが検出されたロットのリスト
        }
        # 【追加】同一品番の割当回数制約の効果測定メトリクス
        self.product_assignment_metrics = {
            'total_checks': 0,  # 制約チェックの総回数
            'constraint_violations': 0,  # 制約違反の回数（通常制約）
            'relaxed_assignments': 0,  # 緩和条件が適用された割当回数
            'max_assignments_reached': {},  # 最大割当回数に達した検査員・品番の組み合わせ {inspector_code: {product_number: count}}
            'relaxation_reasons': [],  # 緩和理由のリスト（デバッグ用）
        }
        # 【追加】当日洗浄上がり品の制約緩和条件の効果測定メトリクス
        self.same_day_relaxation_metrics = {
            'total_relaxations': 0,  # 制約緩和が適用された総回数
            'relaxation_by_reason': {},  # 緩和理由別の回数 {reason: count}
            'relaxation_history': [],  # 緩和履歴のリスト [{lot_index, product_number, reason, inspection_time, ...}]
            'constraints_relaxed': {
                'product_name_constraint': 0,  # 品名単位の重複禁止制約を緩和した回数
                'product_number_constraint': 0,  # 品番単位の重複禁止制約を緩和した回数
                'product_limit_constraint': 0,  # 同一品番4.0h上限制約を緩和した回数
                'work_hours_constraint': 0,  # 勤務時間制約を緩和した回数
            },
            'relaxation_threshold_checks': 0,  # SAME_DAY_RELAXATION_THRESHOLD_HOURS によるチェック回数
            'relaxation_threshold_applied': 0,  # SAME_DAY_RELAXATION_THRESHOLD_HOURS による緩和適用回数
        }
        # 【追加】新規品の保護条件の効果測定メトリクス
        self.new_product_protection_metrics = {
            'total_protections': 0,  # 保護が適用された総回数
            'protection_by_violation_type': {},  # 違反タイプ別の保護回数 {violation_type: count}
            'protection_by_phase': {},  # フェーズ別の保護回数 {phase: count}
            'protection_history': [],  # 保護履歴のリスト [{lot_index, product_number, violation_type, phase, ...}]
            'protected_lots': set(),  # 保護されたロットのインデックス集合
            'protection_enabled': NEW_PRODUCT_PROTECTION_ENABLED,  # 保護が有効かどうか
            'protection_days': NEW_PRODUCT_PROTECTION_DAYS,  # 保護期間（日数）
        }
        # 【追加】固定検査員の保護条件の効果測定メトリクス
        self.fixed_inspector_protection_metrics = {
            'total_protections': 0,  # 保護が適用された総回数
            'protection_by_phase': {},  # フェーズ別の保護回数 {phase: count}
            'protection_by_reason': {},  # 保護理由別の回数 {reason: count}
            'protection_history': [],  # 保護履歴のリスト [{lot_index, product_number, inspector_name, phase, reason, ...}]
            'protected_lots': set(),  # 保護されたロットのインデックス集合
            'protected_inspectors': set(),  # 保護された検査員名の集合
        }
        # 【追加】固定検査員情報を保持（品番ごとの固定検査員リスト）
        # 形式: {品番: [検査員名1, 検査員名2, ...]}
        # 品番ごとの固定検査員情報:
        # {品番: [ {'process': '工程名', 'inspectors': ['A', 'B']}, ... ]}
        self.fixed_inspectors_by_product: Dict[str, List[Dict[str, Any]]] = {}
        # 同日洗浄品の制約緩和フラグ（品番: ロットインデックス）
        self.same_day_constraint_relaxations = set()
        # 【高速化】検査員マスタのインデックス（O(1)アクセス用）
        # 形式: {氏名: Series行データ} または {コード: Series行データ}
        self.inspector_name_to_row = {}  # 氏名→行データのマッピング
        self.inspector_code_to_row = {}  # コード→行データのマッピング
        self.inspector_id_to_row = {}  # ID→行データのマッピング
        self._inspector_master_df_hash = None  # マスタのハッシュ（変更検知用）
        # 【高速化】工程マスタのキャッシュ
        self._process_master_cache = None
        self._process_master_cache_path = None
        self._process_master_cache_mtime = None
        # 【高速化】ログ出力のバッチ化（オプション）
        self.log_batch_enabled = False  # デフォルトは無効（既存動作を維持）
        self.log_buffer = []  # ログをバッファリング
        self.log_batch_size = 10  # バッチサイズ

        # 通常運用はログ量が多くなりやすいため、UI更新/ファイル書き込み負荷を軽減する目的でバッチ化を有効化（任意でOFF可）
        try:
            enabled = os.getenv("ASSIGN_LOG_BATCH_ENABLED", "1").strip().lower() not in {"0", "false", "off", "no"}
            batch_size = int(os.getenv("ASSIGN_LOG_BATCH_SIZE", "50").strip() or "50")
            if enabled and not self.debug_mode:
                self.enable_log_batching(batch_size=max(1, batch_size))
        except Exception:
            pass

        # ログ削減: 频出メッセージを集計して最終だけ出す（通常運用のログ量を減らす）
        self._suppressed_relax_assign_total = 0
        self._suppressed_relax_assign_by_product = defaultdict(int)
    
    def log_message(
        self,
        message: str,
        debug: bool = False,
        level: str = 'info'
    ) -> None:
        """
        ログメッセージを出力（バッチ化対応・高速化）
        
        Args:
            message: ログメッセージ
            debug: Trueの場合、debug_modeがTrueの時のみ出力
            level: ログレベル ('info', 'warning', 'error')
        """
        if debug and not self.debug_mode:
            return

        # 通常運用ではログが増えすぎるため、特定の頻出メッセージは集計して最後にまとめて出す
        try:
            if not self.debug_mode:
                msg_str = str(message or "")
                if msg_str.startswith("制約を一部緩和して割り当て:"):
                    self._suppressed_relax_assign_total += 1
                    m = re.search(r"品番:\s*([^,)\s]+)", msg_str)
                    if m:
                        self._suppressed_relax_assign_by_product[m.group(1)] += 1
                    return
        except Exception:
            pass

        # 絵文字等が混ざると環境によって文字化け/エンコードエラーになりやすいので除去する
        message = self._sanitize_log_message(message)
        
        # バッチ化が有効な場合はバッファに追加
        if self.log_batch_enabled:
            self.log_buffer.append((message, level))
            # バッチサイズに達したらまとめて出力
            if len(self.log_buffer) >= self.log_batch_size:
                self._flush_log_buffer()
        else:
            # 従来通り即座に出力
            if self.log_callback:
                try:
                    self.log_callback(message, level=level, channel="ASSIGN")
                    return
                except TypeError:
                    self.log_callback(message)
                    return

            loguru_logger_bound = loguru_logger.bind(channel="ASSIGN")
            if level == 'warning':
                loguru_logger_bound.warning(message)
            elif level == 'error':
                loguru_logger_bound.error(message)
            else:
                loguru_logger_bound.info(message)

    @staticmethod
    def _sanitize_log_message(message: str) -> str:
        if message is None:
            return ""
        s = str(message)

        # remove variation selectors / ZWJ
        s = s.replace("\uFE0F", "").replace("\u200D", "")

        # remove emoji/symbol blocks that often cause mojibake
        cleaned = []
        for ch in s:
            o = ord(ch)
            if 0x1F300 <= o <= 0x1FAFF:
                continue
            if 0x2600 <= o <= 0x27BF:
                continue
            cleaned.append(ch)
        s = "".join(cleaned)

        # normalize consecutive spaces (keep newlines)
        s = re.sub(r"[ \t]{2,}", " ", s)
        return s.strip()

    def _flush_log_buffer(self) -> None:
        """ログバッファをフラッシュ（まとめて出力）"""
        if not self.log_buffer:
            return

        def _emit_block(messages: list[str], level: str) -> None:
            combined = "\n".join(messages)
            if self.log_callback:
                try:
                    self.log_callback(combined, level=level, channel="ASSIGN")
                    return
                except TypeError:
                    self.log_callback(combined)
                    return

            loguru_logger_bound = loguru_logger.bind(channel="ASSIGN")
            if level == 'warning':
                loguru_logger_bound.warning(combined)
            elif level == 'error':
                loguru_logger_bound.error(combined)
            else:
                loguru_logger_bound.info(combined)

        # 連続する同一レベルはまとめて出力（UI更新とI/Oを削減しつつ、順序は維持）
        current_level = self.log_buffer[0][1]
        chunk: list[str] = []
        for message, level in self.log_buffer:
            if level != current_level and chunk:
                _emit_block(chunk, current_level)
                chunk = []
                current_level = level
            chunk.append(message)
        if chunk:
            _emit_block(chunk, current_level)

        # バッファをクリア
        self.log_buffer.clear()

    def _normalize_shipping_date_string(self, shipping_date: Any) -> str:
        """出荷予定日を文字列化して空白をトリム"""
        if pd.isna(shipping_date):
            return ''
        return str(shipping_date).strip()

    def _is_same_day_cleaning_label(self, shipping_date: Any) -> bool:
        """出荷予定日が当日洗浄・先行検査系かどうか判定"""
        shipping_date_str = self._normalize_shipping_date_string(shipping_date)
        if not shipping_date_str:
            return False
        if shipping_date_str in {"当日洗浄上がり品", "当日洗浄品", "当日洗浄", "当日先行検査", "先行検査"}:
            return True
        if ("当日洗浄" in shipping_date_str) or ("当日洗流E" in shipping_date_str):
            return True
        return False

    def _is_preinspection_label(self, shipping_date: Any) -> bool:
        """出荷予定日が先行検査系かどうか判定（登録済み品番の先行検査ロット判定に使用）"""
        shipping_date_str = self._normalize_shipping_date_string(shipping_date)
        if not shipping_date_str:
            return False
        return "先行検査" in shipping_date_str

    def _is_locked_fixed_preinspection_lot(
        self,
        result_df: pd.DataFrame,
        lot_index: int,
        restrict_to_preinspection: bool = False
    ) -> bool:
        """
        固定検査員が割り当て済みなら最適化で動かさない（固定維持）。

        restrict_to_preinspection=True の場合のみ、先行検査ラベルのロットに限定する。
        """
        try:
            if lot_index < 0 or lot_index >= len(result_df):
                return False
            shipping_date = result_df.at[lot_index, '出荷予定日'] if '出荷予定日' in result_df.columns else None
            if restrict_to_preinspection and not self._is_preinspection_label(shipping_date):
                return False
            product_number = str(result_df.at[lot_index, '品番']).strip() if '品番' in result_df.columns else ''
            if not product_number:
                return False
            process_name = ''
            if '現在工程名' in result_df.columns:
                process_name = str(result_df.at[lot_index, '現在工程名'] or '').strip()
            fixed_names = self._collect_fixed_inspector_names(product_number, process_name)
            if not fixed_names:
                return False
            for i in range(1, 6):
                col = f'検査員{i}'
                if col not in result_df.columns:
                    continue
                current_name = str(result_df.at[lot_index, col] or '').strip()
                if current_name and current_name in fixed_names:
                    return True
            return False
        except Exception:
            return False
        return "当日洗浄" in shipping_date_str or "当日洗流" in shipping_date_str

    def _should_force_assign_same_day(self, shipping_date: Any) -> bool:
        """
        当日洗浄上がり品（＋先行検査）について、未割当を極力なくすために
        一部の制約（品番/品名単位の重複禁止、同一品番4.0h上限）を無視して割り当てるか。
        """
        return self._is_same_day_cleaning_label(shipping_date)

    def _should_force_keep_assignment(self, shipping_date: Any) -> bool:
        """
        割当を「維持」する最優先対象かどうか。

        - 当日洗浄上がり品（＋先行検査）
        - 出荷予定日が「当日（または過去）」のロット（要日出荷品を含む）
        - 3営業日以内のロット（未割当を避けるため）
        - 2週間以内のロット（未割当を避けるため）

        ※ ここで True の場合、最終検証で違反が検出されても未割当へ落とさず、割当維持を優先する。
        """
        if self._should_force_assign_same_day(shipping_date):
            return True

        try:
            if shipping_date is None or (isinstance(shipping_date, float) and pd.isna(shipping_date)):
                return False
            if isinstance(shipping_date, pd.Timestamp):
                ship_date = shipping_date.date()
            else:
                ship_ts = pd.to_datetime(shipping_date, errors="coerce")
                if pd.isna(ship_ts):
                    return False
                ship_date = ship_ts.date()
            today = pd.Timestamp.now().date()
            # 当日または過去の日付の場合は維持
            if ship_date <= today:
                return True
            
            # 3営業日以内のロットも保護対象に含める
            def add_business_days(start: date, business_days: int) -> date:
                def next_business_day(date_val: date) -> date:
                    weekday = date_val.weekday()
                    if weekday == 4:  # Friday
                        return date_val + timedelta(days=3)
                    if weekday == 5:  # Saturday
                        return date_val + timedelta(days=2)
                    return date_val + timedelta(days=1)
                current = start
                added = 0
                while added < business_days:
                    current = next_business_day(current)
                    added += 1
                return current
            
            three_business_days_ahead = add_business_days(today, 3)
            # 3営業日以内のロットは保護対象
            if ship_date <= three_business_days_ahead:
                return True
            
            # 2週間以内のロットも保護対象に含める（未割当ロット再処理と同様）
            two_weeks_later = today + timedelta(days=14)
            if ship_date <= two_weeks_later:
                return True
            
            return False
        except Exception:
            return False

    def _rebuild_assignment_histories(self, result_df: pd.DataFrame, inspector_master_df: pd.DataFrame) -> None:
        """
        result_df の割当結果から、勤務時間・品番時間などの履歴を再構築する。
        （割当解除/再割当を行った直後に使用）
        """
        current_date = pd.Timestamp.now().date()
        self.inspector_daily_assignments = {}
        self.inspector_work_hours = {}
        self.inspector_product_hours = {}

        # 当日洗浄上がり品の追跡（品番/品名単位）も再構築
        self.same_day_cleaning_inspectors = {}
        self.same_day_cleaning_inspectors_by_product_name = {}

        if result_df is None or result_df.empty:
            return

        cols = {col: idx for idx, col in enumerate(result_df.columns)}
        if '品番' not in cols:
            return

        for row in result_df.itertuples(index=False):
            product_number = row[cols['品番']]
            if product_number is None or (isinstance(product_number, float) and pd.isna(product_number)):
                continue
            product_number_str = str(product_number).strip()
            if not product_number_str:
                continue

            inspector_count = 0
            if '検査員人数' in cols:
                inspector_count_val = row[cols['検査員人数']]
                try:
                    inspector_count = int(inspector_count_val) if pd.notna(inspector_count_val) else 0
                except Exception:
                    inspector_count = 0
            if inspector_count <= 0:
                continue

            inspection_time = 0.0
            if '検査時間' in cols:
                inspection_time_val = row[cols['検査時間']]
                try:
                    inspection_time = float(inspection_time_val) if pd.notna(inspection_time_val) else 0.0
                except Exception:
                    inspection_time = 0.0
            if inspection_time <= 0.0:
                continue

            divided_time = inspection_time / inspector_count

            shipping_date_raw = row[cols.get('出荷予定日', -1)] if '出荷予定日' in cols else None
            is_same_day = self._should_force_assign_same_day(shipping_date_raw)
            product_name_str = ''
            if '品名' in cols:
                product_name_val = row[cols['品名']]
                product_name_str = str(product_name_val).strip() if pd.notna(product_name_val) else ''

            for i in range(1, 6):
                inspector_col = f'検査員{i}'
                if inspector_col not in cols:
                    continue
                inspector_name_raw = row[cols[inspector_col]]
                if pd.isna(inspector_name_raw) or str(inspector_name_raw).strip() == '':
                    continue
                inspector_name = str(inspector_name_raw).strip()
                if '(' in inspector_name:
                    inspector_name = inspector_name.split('(')[0].strip()
                if not inspector_name:
                    continue

                inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                if inspector_info.empty:
                    continue
                inspector_code = inspector_info.iloc[0]['#ID']

                self.inspector_daily_assignments.setdefault(inspector_code, {}).setdefault(current_date, 0.0)
                self.inspector_daily_assignments[inspector_code][current_date] += divided_time
                self.inspector_work_hours[inspector_code] = self.inspector_work_hours.get(inspector_code, 0.0) + divided_time
                self.inspector_product_hours.setdefault(inspector_code, {}).setdefault(product_number_str, 0.0)
                self.inspector_product_hours[inspector_code][product_number_str] += divided_time

                if is_same_day:
                    self.same_day_cleaning_inspectors.setdefault(product_number_str, set()).add(inspector_code)
                    if product_name_str:
                        self.same_day_cleaning_inspectors_by_product_name.setdefault(product_name_str, set()).add(inspector_code)

    def _apply_work_hours_overrun(self, hours: float) -> float:
        """勤務時間上限に許容率を適用（10%超過まで許容）"""
        return hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
    
    def _get_dynamic_work_hours_buffer(self, inspector_code: str, current_date: date, assignment_count: int = 0) -> float:
        """
        動的バッファを計算
        
        Args:
            inspector_code: 検査員コード
            current_date: 現在の日付
            assignment_count: この検査員への割当回数（累積誤差の指標として使用）
        
        Returns:
            動的バッファ値（時間）
        """
        # 基本バッファ
        base_buffer = WORK_HOURS_BUFFER_BASE
        
        # 累積誤差を考慮した動的調整
        # 割当回数が多いほど、累積誤差が大きくなる可能性があるため、バッファを増やす
        dynamic_adjustment = min(0.05, assignment_count * 0.001)  # 最大0.05h（3分）まで増加
        
        # 動的バッファ係数を適用
        dynamic_buffer = base_buffer + dynamic_adjustment * WORK_HOURS_BUFFER_DYNAMIC_FACTOR
        
        # 最大バッファはWORK_HOURS_BUFFER（0.1h）を超えないようにする
        return min(dynamic_buffer, WORK_HOURS_BUFFER)
    
    def _record_buffer_usage(self, daily_hours: float, allowed_max_hours: float, buffer_value: float, exceeded: bool = False, exceeded_by: float = 0.0) -> None:
        """
        バッファ使用状況を記録（効果測定メトリクス用）
        
        Args:
            daily_hours: 現在の勤務時間
            allowed_max_hours: 許容最大勤務時間
            buffer_value: 使用したバッファ値
            exceeded: バッファを超えたかどうか
            exceeded_by: バッファ超過量（時間）
        """
        self.buffer_usage_metrics['total_checks'] += 1
        if exceeded:
            self.buffer_usage_metrics['buffer_exceeded_count'] += 1
            if exceeded_by > 0:
                self.buffer_usage_metrics['buffer_exceeded_by'].append(exceeded_by)
    
    def _analyze_phase1_non_convergence(
        self,
        overworked_assignments: List[Tuple],
        product_limit_violations: List[Tuple],
        result_df: pd.DataFrame,
        inspector_master_df: pd.DataFrame,
        inspector_max_hours: Dict[str, float]
    ) -> None:
        """
        フェーズ1が収束しない原因を分析
        
        Args:
            overworked_assignments: 勤務時間超過の違反リスト
            product_limit_violations: 同一品番時間上限超過の違反リスト
            result_df: 結果データフレーム
            inspector_master_df: 検査員マスタデータ
            inspector_max_hours: 検査員の最大勤務時間辞書
        """
        analysis_results = {
            'overworked_reasons': [],
            'product_limit_reasons': [],
            'common_inspectors': set(),
            'common_products': set(),
        }
        
        current_date = pd.Timestamp.now().date()
        
        # 勤務時間超過の原因分析
        for violation in overworked_assignments:
            index, inspector_code, inspector_name, excess, divided_time, product_number, inspection_time, inspector_col_num = violation
            inspector_info = inspector_master_df[inspector_master_df['#ID'] == inspector_code]
            if not inspector_info.empty:
                max_hours = inspector_max_hours.get(inspector_code, 8.0)
                daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
                allowed_max_hours = max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                
                # 原因を分析
                reason = f"検査員 '{inspector_name}' ({inspector_code}): "
                reason += f"現在{daily_hours:.1f}h, 許容上限{allowed_max_hours:.1f}h, 超過{excess:.1f}h"
                if daily_hours > allowed_max_hours:
                    reason += f" (完全超過)"
                elif daily_hours > allowed_max_hours - WORK_HOURS_BUFFER:
                    reason += f" (バッファ超過)"
                analysis_results['overworked_reasons'].append(reason)
                analysis_results['common_inspectors'].add(inspector_code)
        
        # 同一品番時間上限超過の原因分析
        for violation in product_limit_violations:
            index, inspector_code, inspector_name, excess, divided_time, product_number, inspection_time, inspector_col_num = violation
            product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
            
            reason = f"検査員 '{inspector_name}' ({inspector_code}), 品番 '{product_number}': "
            reason += f"現在{product_hours:.1f}h, 上限{self.product_limit_hard_threshold:.1f}h, 超過{excess:.1f}h"
            analysis_results['product_limit_reasons'].append(reason)
            analysis_results['common_inspectors'].add(inspector_code)
            analysis_results['common_products'].add(product_number)
        
        # 分析結果をログに出力
        self.log_message(
            f"フェーズ1収束しない原因分析: "
            f"勤務時間超過={len(overworked_assignments)}件, "
            f"同一品番時間上限超過={len(product_limit_violations)}件, "
            f"共通検査員数={len(analysis_results['common_inspectors'])}, "
            f"共通品番数={len(analysis_results['common_products'])}",
            level='warning',
        )
        
        # 詳細な原因をdebugログに出力
        if analysis_results['overworked_reasons']:
            self.log_message(
                f"勤務時間超過の詳細: {analysis_results['overworked_reasons'][:5]}",  # 最初の5件のみ
                debug=True,
            )
        if analysis_results['product_limit_reasons']:
            self.log_message(
                f"同一品番時間上限超過の詳細: {analysis_results['product_limit_reasons'][:5]}",  # 最初の5件のみ
                debug=True,
            )
        
        # 共通検査員が複数の違反に関与している場合
        if len(analysis_results['common_inspectors']) < (len(overworked_assignments) + len(product_limit_violations)):
            self.log_message(
                f"複数の違反に関与している検査員が{len(analysis_results['common_inspectors'])}名います。"
                f"これらの検査員への負荷が集中している可能性があります。",
                level='warning',
            )

    def _apply_same_day_work_hours_overrun(self, hours: float) -> float:
        """当日洗浄品用にさらに余裕を広げる（20%超過まで許容）"""
        return hours * (1.0 + SAME_DAY_WORK_HOURS_OVERRUN_RATE)

    def _should_relax_same_day_same_name(self, key: str) -> bool:
        """
        同一品名制約の緩和を許可するか判定し、試行回数を更新する
        """
        attempts = self.same_day_same_name_relaxation_attempts.get(key, 0)
        if attempts >= MAX_SAME_DAY_SAME_NAME_RELAXATIONS:
            return False
        self.same_day_same_name_relaxation_attempts[key] = attempts + 1
        return True

    def _should_relax_hours_for_lot(
        self,
        product_number: str,
        shipping_date_str: str,
        inspection_time: Optional[float] = None
    ) -> bool:
        """
        特定ロットに対して勤務時間ルールや必要人数の緩和を適用すべきか
        """
        keywords = {"当日洗浄上がり品", "当日洗浄品", "当日先行検査", "先行検査"}
        relax_products = {"MHK1017Z-0", "SPD20-0209"}
        if any(product_number.startswith(prefix) for prefix in ("3D025-",)):
            return True
        if product_number in relax_products:
            return True
        if shipping_date_str in keywords or "当日洗浄" in shipping_date_str:
            return True
        if inspection_time is not None and inspection_time >= self.same_day_relaxation_threshold:
            return True
        return False

    def _calc_required_inspectors(self, inspection_time: float, threshold: Optional[float] = None) -> int:
        """検査時間としきい値から必要検査員数を計算"""
        threshold_value = threshold if threshold is not None else self.required_inspectors_threshold
        if inspection_time <= 0 or pd.isna(inspection_time):
            return 1
        if inspection_time <= threshold_value:
            return 1
        calculated = max(2, int(inspection_time / threshold_value) + 1)
        return min(5, calculated)

    def _is_same_day_high_duration(self, inspection_time: float) -> bool:
        """当日洗浄ロットが制約緩和の対象となる長時間ロットか判定"""
        if inspection_time is None or pd.isna(inspection_time):
            return False
        return inspection_time >= self.same_day_relaxation_threshold

    def _should_force_same_day_dual_assignment(self, inspection_time: float) -> bool:
        """当日洗浄ロットに最低2人以上を確保すべきか判定"""
        if inspection_time is None or pd.isna(inspection_time):
            return False
        return inspection_time >= self.same_day_force_second_inspector_threshold
    def _split_process_keywords(self, process_filter: str) -> List[str]:
        """工程名フィルターを / や ／ で分割してキーワードリストを作成"""
        if not process_filter:
            return []
        return [keyword.strip() for keyword in re.split(r"[／/]", process_filter) if keyword.strip()]

    def _process_filter_matches(self, process_filter: str, actual_process_name: Optional[str]) -> bool:
        """登録済み品番の工程フィルタが現在工程名に一致するか判定"""
        if not process_filter:
            return False
        keywords = self._split_process_keywords(process_filter)
        if not keywords:
            return False
        candidate = str(actual_process_name or "").strip()
        if not candidate:
            return False
        for keyword in keywords:
            if keyword in candidate:
                return True
        return False

    def _get_tuple_value(
        self,
        row: Any,
        cols_map: Dict[str, int],
        column_name: str,
        default: str = ''
    ) -> str:
        """itertuplesの結果（namedtuple）から列値を取り出し、文字列化する"""
        idx = cols_map.get(column_name, -1)
        if idx == -1:
            return default
        value = row[idx]
        if pd.isna(value):
            return default
        return str(value).strip()

    def _collect_fixed_inspector_names(
        self,
        product_number: str,
        process_name_context: Optional[str]
    ) -> List[str]:
        """
        品番・工程名にマッチする固定検査員名のリストを取得（重複は除去）
        """
        entries = self.fixed_inspectors_by_product.get(product_number, [])
        if not entries:
            return []
        result: List[str] = []
        seen: Set[str] = set()
        for entry in entries:
            process_filter = entry.get('process', '').strip()
            if process_filter:
                if not self._process_filter_matches(process_filter, process_name_context):
                    continue
            for name in entry.get('inspectors', []):
                if not name:
                    continue
                name_str = str(name).strip()
                if not name_str or name_str in seen:
                    continue
                seen.add(name_str)
                result.append(name_str)
        return result

    def _is_fixed_inspector_for_lot(
        self,
        product_number: Any,
        process_name_context: Optional[str],
        inspector_name: Any
    ) -> bool:
        """
        対象ロットに対して、割当済み検査員が「固定検査員」かどうかを判定する。

        仕様:
        - 工程フィルタが一致している固定検査員のみを固定扱いにする
        - 終日休暇（work_status='休み'）の場合は固定扱いしない
        """
        if not product_number or pd.isna(product_number):
            return False
        if not inspector_name or pd.isna(inspector_name):
            return False

        product_number_str = str(product_number).strip()
        inspector_name_str = str(inspector_name).strip()
        if not product_number_str or not inspector_name_str:
            return False

        # 括弧内の情報を除去
        if '(' in inspector_name_str:
            inspector_name_str = inspector_name_str.split('(')[0].strip()
        if not inspector_name_str:
            return False

        # 休暇は固定扱いしない（固定0hは可だが、休暇は不可）
        if self.is_inspector_on_vacation(inspector_name_str):
            return False

        fixed_names = self._collect_fixed_inspector_names(product_number_str, process_name_context)
        return inspector_name_str in fixed_names
    
    def enable_log_batching(self, batch_size: int = 10) -> None:
        """
        ログ出力のバッチ化を有効化（高速化オプション）
        
        Args:
            batch_size: バッチサイズ（デフォルト: 10）
        """
        # 既存のバッファをフラッシュ
        if self.log_buffer:
            self._flush_log_buffer()
        self.log_batch_enabled = True
        self.log_batch_size = batch_size
    
    def disable_log_batching(self) -> None:
        """ログ出力のバッチ化を無効化（従来の動作に戻す）"""
        # 既存のバッファをフラッシュ
        if self.log_buffer:
            self._flush_log_buffer()
        self.log_batch_enabled = False
    
    def _build_inspector_index(self, inspector_master_df: pd.DataFrame) -> None:
        """
        検査員マスタのインデックスを作成（高速化：O(1)アクセス用）
        
        Args:
            inspector_master_df: 検査員マスタのDataFrame
        """
        import hashlib
        # マスタのハッシュを計算（変更検知用）
        try:
            df_hash = hashlib.md5(pd.util.hash_pandas_object(inspector_master_df).values).hexdigest()
            if df_hash == self._inspector_master_df_hash:
                # 変更がない場合は再構築をスキップ
                return
            self._inspector_master_df_hash = df_hash
        except Exception:
            pass  # ハッシュ計算に失敗した場合は再構築
        
        # インデックスをクリア
        self.inspector_name_to_row = {}
        self.inspector_code_to_row = {}
        self.inspector_id_to_row = {}
        
        # インデックスを構築（itertuples()で高速化）
        # 列インデックスを事前に取得
        name_col_idx = inspector_master_df.columns.get_loc('#氏名') if '#氏名' in inspector_master_df.columns else -1
        code_col_idx = inspector_master_df.columns.get_loc('#コード') if '#コード' in inspector_master_df.columns else -1
        id_col_idx = inspector_master_df.columns.get_loc('#ID') if '#ID' in inspector_master_df.columns else -1
        
        for row_tuple in inspector_master_df.itertuples(index=True):
            idx = row_tuple[0]  # インデックス
            row = inspector_master_df.loc[idx]  # Seriesとして取得（互換性のため）
            
            # 氏名→行データのマッピング
            if name_col_idx >= 0:
                name = row_tuple[name_col_idx + 1]  # itertuplesはインデックスを含むため+1
                if pd.notna(name) and str(name).strip():
                    name_key = str(name).strip()
                    self.inspector_name_to_row[name_key] = row
            
            # コード→行データのマッピング
            if code_col_idx >= 0:
                code = row_tuple[code_col_idx + 1]
                if pd.notna(code) and str(code).strip():
                    code_key = str(code).strip()
                    self.inspector_code_to_row[code_key] = row
            
            # ID→行データのマッピング
            if id_col_idx >= 0:
                inspector_id = row_tuple[id_col_idx + 1]
                if pd.notna(inspector_id) and str(inspector_id).strip():
                    id_key = str(inspector_id).strip()
                    self.inspector_id_to_row[id_key] = row
    
    def _get_inspector_by_name(
        self,
        inspector_name: Any,
        inspector_master_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        検査員名から検査員情報を取得（高速化：O(1)アクセス）
        
        Args:
            inspector_name: 検査員名
            inspector_master_df: 検査員マスタのDataFrame（フォールバック用）
        
        Returns:
            DataFrame: 検査員情報（見つからない場合は空のDataFrame）
        """
        if not inspector_name or pd.isna(inspector_name):
            return pd.DataFrame()
        
        name_key = str(inspector_name).strip()
        # 括弧内の情報を除去
        if '(' in name_key:
            name_key = name_key.split('(')[0].strip()
        
        # インデックスから取得
        inspector_row = self.inspector_name_to_row.get(name_key)
        if inspector_row is not None:
            return pd.DataFrame([inspector_row])
        
        # フォールバック：従来の方法（互換性のため）
        inspector_info = inspector_master_df[inspector_master_df['#氏名'] == name_key]
        if not inspector_info.empty:
            # 見つかった場合はインデックスに追加
            self.inspector_name_to_row[name_key] = inspector_info.iloc[0]
        return inspector_info
    
    def _get_inspector_by_code(
        self,
        inspector_code: Any,
        inspector_master_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        検査員コードから検査員情報を取得（高速化：O(1)アクセス）
        
        Args:
            inspector_code: 検査員コード
            inspector_master_df: 検査員マスタのDataFrame（フォールバック用）
        
        Returns:
            DataFrame: 検査員情報（見つからない場合は空のDataFrame）
        """
        if not inspector_code or pd.isna(inspector_code):
            return pd.DataFrame()
        
        code_key = str(inspector_code).strip()
        
        # インデックスから取得
        inspector_row = self.inspector_code_to_row.get(code_key)
        if inspector_row is not None:
            return pd.DataFrame([inspector_row])
        
        # フォールバック：従来の方法（互換性のため）
        inspector_info = inspector_master_df[inspector_master_df['#コード'] == code_key]
        if not inspector_info.empty:
            # 見つかった場合はインデックスに追加
            self.inspector_code_to_row[code_key] = inspector_info.iloc[0]
        return inspector_info
    
    def _get_inspector_by_id(
        self,
        inspector_id: Any,
        inspector_master_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        検査員IDから検査員情報を取得（高速化：O(1)アクセス）
        
        Args:
            inspector_id: 検査員ID
            inspector_master_df: 検査員マスタのDataFrame（フォールバック用）
        
        Returns:
            DataFrame: 検査員情報（見つからない場合は空のDataFrame）
        """
        if not inspector_id or pd.isna(inspector_id):
            return pd.DataFrame()
        
        id_key = str(inspector_id).strip()
        
        # インデックスから取得
        inspector_row = self.inspector_id_to_row.get(id_key)
        if inspector_row is not None:
            return pd.DataFrame([inspector_row])
        
        # フォールバック：従来の方法（互換性のため）
        inspector_info = inspector_master_df[inspector_master_df['#ID'] == id_key]
        if not inspector_info.empty:
            # 見つかった場合はインデックスに追加
            self.inspector_id_to_row[id_key] = inspector_info.iloc[0]
        return inspector_info

    def _normalize_shipping_date(self, shipping_date: Any) -> pd.Timestamp:
        """
        出荷予定日の文字列表現などを一貫した Timestamp に変換する。
        当日洗浄上がり品や当日先行検査などの優先案件は最優先となるよう最小値へマップする。
        
        Args:
            shipping_date: 出荷予定日（文字列、日付、Timestampなど）
        
        Returns:
            pd.Timestamp: 正規化されたTimestamp
        """
        try:
            if shipping_date is None or (isinstance(shipping_date, float) and pd.isna(shipping_date)):
                return pd.Timestamp.max

            if isinstance(shipping_date, str):
                shipping_date_str = shipping_date.strip()
                if not shipping_date_str:
                    return pd.Timestamp.max

                # 当日洗浄・先行検査など文字列表現の優先案件
                same_day_keywords = [
                    "当日洗浄上がり品",
                    "当日洗浄上がり",
                    "当日洗浄あがり",
                    "当日洗浄品",
                    "当日洗浄",
                    "当日先行検査",
                    "先行検査",
                ]
                if any(keyword in shipping_date_str for keyword in same_day_keywords):
                    return pd.Timestamp.min

            normalized = pd.to_datetime(shipping_date, errors='coerce')
            if pd.isna(normalized):
                return pd.Timestamp.max
            return normalized
        except Exception:
            return pd.Timestamp.max

    def _convert_shipping_date(self, val: Any) -> Union[str, pd.Timestamp, Any]:
        """
        出荷予定日を日付型に変換（当日洗浄品は文字列として保持）
        
        複数のメソッドで使用されるため、共通メソッドとして定義
        
        Args:
            val: 出荷予定日の値（文字列、日付、Timestampなど）
        
        Returns:
            当日洗浄品の場合は文字列、その他の場合はpd.Timestampまたは元の値
        """
        if pd.isna(val):
            return val
        
        val_str = str(val).strip()
        
        # 当日洗浄品の場合は文字列として保持
        if self._is_same_day_cleaning(val_str):
            return val_str
        
        # その他の場合は日付型に変換
        try:
            date_val = pd.to_datetime(val, errors='coerce')
            if pd.notna(date_val):
                # 異常な日付（例: 1677年など）をチェック
                if date_val.year < 1900 or date_val.year > 2100:
                    # 異常な日付の場合はNaTを返す（後続処理で適切に処理される）
                    self.log_message(
                        f"警告: 出荷予定日が異常な値です ({val}, 年: {date_val.year})。無効な日付として処理します。",
                        level='warning'
                    )
                    return pd.NaT
                return date_val
            return date_val
        except Exception:
            return val
    
    def _is_same_day_cleaning(self, val_str: str) -> bool:
        """
        当日洗浄品かどうかを判定
        
        Args:
            val_str: 判定する文字列
        
        Returns:
            当日洗浄品の場合はTrue、それ以外はFalse
        """
        return (
            val_str == "当日洗浄上がり品" or 
            val_str == "当日洗浄品" or
            "当日洗浄" in val_str or
            val_str == "先行検査" or
            val_str == "当日先行検査"
        )

    def _gather_skill_candidates_for_feasibility(
        self,
        product_number: str,
        process_number: Optional[Any],
        skill_master_df: pd.DataFrame,
        inspector_master_df: pd.DataFrame
    ) -> List[Dict[str, Any]]:
        """
        スキルマスタを基に該当品番の候補検査員を抽出する。
        勤務時間や4時間ルールのような動的制約は考慮せず、スキル適合と新製品チームのみで構成。
        
        Args:
            product_number: 品番
            process_number: 工程番号（オプション）
            skill_master_df: スキルマスタのDataFrame
            inspector_master_df: 検査員マスタのDataFrame
        
        Returns:
            候補検査員のリスト（各要素は辞書形式）
        """
        candidates = []
        try:
            skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]

            if skill_rows.empty:
                # 新製品チームのみが候補
                return self.get_new_product_team_inspectors(inspector_master_df)

            filtered_skill_rows = []
            if process_number is None or str(process_number).strip() == '':
                filtered_skill_rows = list(skill_rows.iterrows())
            else:
                # 工程番号列のインデックスを事前に取得（高速化：itertuples()を使用）
                process_col_idx = 1  # iloc[1]に対応
                for row_tuple in skill_rows.itertuples(index=True):
                    row_idx = row_tuple[0]  # インデックス
                    skill_process_number = row_tuple[process_col_idx + 1]  # itertuplesはインデックスを含むため+1
                    if pd.isna(skill_process_number) or str(skill_process_number).strip() == '' or str(skill_process_number) == str(process_number):
                        # Seriesとして扱うために元の行を取得
                        skill_row = skill_rows.loc[row_idx]
                        filtered_skill_rows.append((row_idx, skill_row))

            if not filtered_skill_rows:
                return self.get_new_product_team_inspectors(inspector_master_df)

            skill_columns = skill_master_df.columns[2:]
            seen_codes = set()

            for _, skill_row in filtered_skill_rows:
                for col_name in skill_columns:
                    inspector_code = col_name
                    if pd.isna(inspector_code) or str(inspector_code).strip() == '':
                        continue
                    inspector_code = str(inspector_code).strip()
                    if inspector_code in seen_codes:
                        continue
                    skill_value = skill_row.get(col_name, None)
                    if pd.notna(skill_value) and str(skill_value).strip() in {'1', '2', '3'}:
                        inspector_info = inspector_master_df[inspector_master_df['#ID'] == inspector_code]
                        if inspector_info.empty:
                            continue
                        inspector_name = inspector_info.iloc[0]['#氏名']
                        try:
                            numeric_skill = int(str(skill_value).strip())
                        except ValueError:
                            numeric_skill = 1
                        candidates.append({
                            '氏名': inspector_name,
                            'コード': inspector_code,
                            'スキル': numeric_skill,
                            'is_new_team': False
                        })
                        seen_codes.add(inspector_code)

            if not candidates:
                return self.get_new_product_team_inspectors(inspector_master_df)

            return candidates
        except Exception as exc:  # フェールセーフ: 例外が発生しても空配列を返す
            self.log_message(f"候補抽出中にエラーが発生しました（品番 {product_number}）: {exc}")
            return []

    def _calculate_remaining_capacity(
        self,
        inspector_code: str,
        inspector_master_df: pd.DataFrame
    ) -> float:
        """
        指定検査員の当日残り勤務時間を計算する。0未満の場合は0を返す。
        
        Args:
            inspector_code: 検査員コード
            inspector_master_df: 検査員マスタのDataFrame
        
        Returns:
            残り勤務時間（時間単位）
        """
        current_date = pd.Timestamp.now().date()
        daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
        max_hours = self.get_inspector_max_hours(inspector_code, inspector_master_df)
        allowed_hours = self._apply_work_hours_overrun(max_hours)
        remaining = allowed_hours - daily_hours - WORK_HOURS_BUFFER
        return max(0.0, remaining)

    def _calculate_assignability_status(
        self,
        row: Union[Dict[str, Any], Any],
        base_candidates: List[Dict[str, Any]],
        inspector_master_df: pd.DataFrame
    ) -> Tuple[str, float]:
        """
        事前のアサイン可能性判定を行う。
        base_candidates はスキル的に対応可能な検査員リスト。
        rowは辞書形式のデータを受け取る
        
        改善ポイント: assignability_statusの詳細分類
        - skill_mismatch: スキル該当者が0
        - capacity_shortage: 候補の残時間合計が不足
        - ready: 理論上割当可能
        
        Args:
            row: 行データ（辞書またはオブジェクト）
            base_candidates: スキル的に対応可能な検査員リスト
            inspector_master_df: 検査員マスタのDataFrame
        
        Returns:
            Tuple[status, available_capacity]
            - status: 割当可能性ステータス
            - available_capacity: 利用可能な容量
        """
        # rowが辞書の場合はget、そうでない場合は直接アクセス
        if isinstance(row, dict):
            inspection_time = row.get('検査時間', 0.0) or 0.0
        else:
            inspection_time = getattr(row, '検査時間', 0.0) if hasattr(row, '検査時間') else 0.0
        
        if inspection_time <= 0:
            return "ready", 0.0

        # 改善ポイント: skill_mismatchの識別
        if not base_candidates:
            return "skill_mismatch", 0.0

        total_capacity = 0.0
        for candidate in base_candidates:
            total_capacity += self._calculate_remaining_capacity(candidate['コード'], inspector_master_df)

        # 改善ポイント: capacity_shortageの識別
        if total_capacity + 1e-6 < inspection_time:
            return "capacity_shortage", total_capacity

        return "ready", total_capacity

    def _calculate_feasible_inspector_count(
        self,
        product_number: str,
        process_number: Optional[Any],
        skill_master_df: pd.DataFrame,
        inspector_master_df: pd.DataFrame
    ) -> Tuple[int, List[Dict[str, Any]]]:
        """
        feasible_inspector_count の算出用ヘルパー
        
        Args:
            product_number: 品番
            process_number: 工程番号（オプション）
            skill_master_df: スキルマスタのDataFrame
            inspector_master_df: 検査員マスタのDataFrame
        
        Returns:
            Tuple[候補者数, 候補者リスト]
        """
        candidates = self._gather_skill_candidates_for_feasibility(product_number, process_number, skill_master_df, inspector_master_df)
        return len(candidates), candidates
    
    def _prepare_result_dataframe(self, inspector_df: pd.DataFrame) -> pd.DataFrame:
        """
        結果用のDataFrameを準備
        
        Args:
            inspector_df: 元のDataFrame
        
        Returns:
            新しい列が追加されたDataFrame
        """
        result_df = inspector_df.copy()
        
        # 新しい列を追加
        result_df['検査員人数'] = 0
        result_df['分割検査時間'] = 0.0
        for i in range(1, 6):
            result_df[f'検査員{i}'] = ''
        result_df['チーム情報'] = ''
        # ボトルネック優先度・可視化用カラム
        result_df['feasible_inspector_count'] = 0
        result_df['available_capacity_hours'] = 0.0
        result_df['assignability_status'] = 'pending'
        result_df['remaining_work_hours'] = 0.0
        result_df['over_product_limit_flag'] = False
        
        return result_df
    
    def _add_sorting_columns(
        self,
        result_df: pd.DataFrame,
        skill_master_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        ソート用の補助列を追加
        
        Args:
            result_df: 結果DataFrame
            skill_master_df: スキルマスタのDataFrame
        
        Returns:
            補助列が追加されたDataFrame
        """
        # 【追加】固定検査員が設定されている品番を最優先にするソートキーを追加
        def has_fixed_inspectors(row: Any) -> bool:
            product_number = row['品番']
            process_name = row.get('現在工程名', '')
            fixed_inspector_names = self._collect_fixed_inspector_names(product_number, process_name)
            return len(fixed_inspector_names) > 0
        
        result_df['_has_fixed_inspectors'] = result_df.apply(has_fixed_inspectors, axis=1)
        
        # 新規品かどうかを判定する列を追加（ソート前に）
        # 高速化: ベクトル化（applyの代わりにisinを使用）
        if len(skill_master_df) > 0:
            skill_product_numbers = set(skill_master_df.iloc[:, 0].astype(str).str.strip().unique())
            result_df['_is_new_product'] = ~result_df['品番'].astype(str).str.strip().isin(skill_product_numbers)
        else:
            result_df['_is_new_product'] = True
        
        return result_df
    
    def _calculate_feasibility_and_candidates(
        self,
        result_df: pd.DataFrame,
        skill_master_df: pd.DataFrame,
        inspector_master_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        事前にスキルベースの候補人数と割当可能性を算出
        
        Args:
            result_df: 結果DataFrame
            skill_master_df: スキルマスタのDataFrame
            inspector_master_df: 検査員マスタのDataFrame
        
        Returns:
            候補情報が追加されたDataFrame
        """
        base_candidate_cache: Dict[int, List[Dict[str, Any]]] = {}
        feasible_counts: List[int] = []
        assign_statuses: List[str] = []
        capacity_list: List[float] = []
        
        # 列名のインデックスマップを作成（itertuples用）
        result_cols = {col: idx for idx, col in enumerate(result_df.columns)}

        for row_idx, row in enumerate(result_df.itertuples(index=False)):
            product_number = row[result_cols['品番']]
            process_number = row[result_cols.get('現在工程番号', -1)] if '現在工程番号' in result_cols else ''
            if process_number == -1:
                process_number = ''
            
            # 行データを辞書形式に変換（_calculate_assignability_status用）
            row_dict = {col: row[result_cols[col]] for col in result_df.columns}
            
            feasible_count, base_candidates = self._calculate_feasible_inspector_count(
                product_number,
                process_number,
                skill_master_df,
                inspector_master_df
            )
            base_candidate_cache[row_idx] = base_candidates
            feasible_counts.append(feasible_count)
            status, total_capacity = self._calculate_assignability_status(row_dict, base_candidates, inspector_master_df)
            assign_statuses.append(status)
            capacity_list.append(round(total_capacity, 2))

        result_df['feasible_inspector_count'] = feasible_counts
        result_df['assignability_status'] = assign_statuses
        result_df['available_capacity_hours'] = capacity_list
        # 後続フェーズで利用するためにベース候補を保持
        # 高速化: 参照で保持（必要に応じてassign_inspectors内でコピー）
        result_df['_base_candidates'] = [base_candidate_cache[idx] for idx in result_df.index]
        
        return result_df
    
    def _sort_lots_by_priority(
        self,
        result_df: pd.DataFrame,
        result_cols: Dict[str, int]
    ) -> pd.DataFrame:
        """
        ロットを優先順位でソート
        
        Args:
            result_df: 結果DataFrame
            result_cols: 列名のインデックスマップ
        
        Returns:
            ソートされたDataFrame
        """
        # 出荷予定日の優先順位を設定（厳守ルール）
        # 優先度: 0=先行検査(固定検査員あり/登録済み優先), 1=当日, 2=当日洗浄/先行検査, 3=2営業日以内, 4=固定検査員あり, 5=その他
        today = pd.Timestamp.now().normalize()
        today_date = today.date()

        def add_business_days(start: date, business_days: int) -> date:
            result_date = start
            added = 0
            while added < business_days:
                result_date += timedelta(days=1)
                if result_date.weekday() < 5:
                    added += 1
            return result_date

        three_business_days_ahead = add_business_days(today_date, 3)
        two_business_days_ahead = add_business_days(today_date, 2)
        
        def calculate_priority(row: Any) -> int:
            shipping_date = row['出荷予定日']
            shipping_date_str = str(shipping_date).strip() if pd.notna(shipping_date) else ''
            product_number = str(row.get('品番', row.get('蜩∫分', ''))).strip()
            # 0. 先行検査（固定検査員あり）: 最優先（登録済み品番の固定検査員ロット）
            if row.get('_has_fixed_inspectors', False) and self._is_preinspection_label(shipping_date):
                return 0
            # 1. 3D025-G4960 は当日出荷と同等の優先度に引き上げ
            if product_number == "3D025-G4960":
                return 1
            # 1. 当日出荷
            try:
                parsed = pd.to_datetime(shipping_date, errors='coerce')
            except Exception:
                parsed = pd.NaT

            if pd.notna(parsed) and parsed.date() == today_date:
                return 1
            if shipping_date_str == "当日":
                return 1

            # 2. 当日洗浄品
            if self._is_same_day_cleaning_label(shipping_date):
                if product_number == "3D025-G4960":
                    return 1
                return 2

            # 3. 今日より後〜3営業日以内（週末除外）
            if pd.notna(parsed):
                parsed_date = parsed.date()
                if parsed_date > today_date and parsed_date <= three_business_days_ahead:
                    return 3
            # 4. その他（固定検査員ありの特別扱いは削除）
            return 4
        
        result_df['_shipping_priority'] = result_df.apply(calculate_priority, axis=1)
        result_df['_is_high_priority'] = result_df['_shipping_priority'] <= HIGH_PRIORITY_SHIPPING_THRESHOLD
        
        # 同一品番の当日洗浄上がり品/先行検査品のロット数を事前にカウント（各ロットに均等に検査員を分散させるため）
        same_day_cleaning_product_counts: Dict[str, int] = {}
        for row_idx, row in enumerate(result_df.itertuples(index=False)):
            shipping_date = row[result_cols.get('出荷予定日', -1)] if '出荷予定日' in result_cols else 'N/A'
            shipping_date_str = str(shipping_date).strip() if pd.notna(shipping_date) else ''
            is_same_day_cleaning = self._is_same_day_cleaning_label(shipping_date)
            if is_same_day_cleaning:
                product_number = row[result_cols['品番']]
                same_day_cleaning_product_counts[product_number] = same_day_cleaning_product_counts.get(product_number, 0) + 1
        
        # 各ロットにロット数を記録
        # 高速化: ベクトル化（applyの代わりに条件分岐をベクトル化）
        is_same_day_cleaning_mask = result_df['出荷予定日'].apply(self._is_same_day_cleaning_label)
        product_numbers = result_df['品番'].astype(str).str.strip()
        result_df['_same_day_cleaning_lot_count'] = (
            is_same_day_cleaning_mask * product_numbers.map(same_day_cleaning_product_counts).fillna(0)
        ).astype(int)
        
        result_df['_sort_product_id'] = result_df['品番'].astype(str)
        # 【変更】固定検査員が設定されている品番を最優先にソート
        # 登録済み品番リストの固定検査員が設定されている品番は、出荷予定日よりも優先して割り当てる
        def is_within_two_business_days(shipping_date: Any) -> bool:
            if pd.isna(shipping_date):
                return False
            try:
                parsed = pd.to_datetime(shipping_date, errors='coerce')
                if pd.isna(parsed):
                    return False
                return parsed.date() <= two_business_days_ahead
            except Exception:
                return False

        result_df['_within_two_business_days'] = result_df['出荷予定日'].apply(is_within_two_business_days)

        # 「それ以降の日付(=翌営業日より後)」だけは日付昇順を優先し、その上で候補数の少ないロットを優先する。
        # それ以外は従来どおり候補数優先（=割付しづらいロットを先に処理）を維持する。
        next_business_day = add_business_days(today_date, 1)
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            parsed_shipping_date = pd.to_datetime(result_df['出荷予定日'], errors='coerce')
        later_than_next_business_day_mask = parsed_shipping_date > pd.Timestamp(next_business_day)
        result_df['_later_shipping_date'] = parsed_shipping_date.where(later_than_next_business_day_mask, pd.NaT)

        # 優先度4（その他）は出荷予定日昇順でソート
        # 指示日のソートキーを追加（FIFO用：指示日が古い順）
        if '指示日' in result_df.columns:
            def instruction_date_sort_key_for_priority(val):
                if pd.isna(val):
                    return pd.Timestamp.max  # 最後に
                try:
                    date_val = pd.to_datetime(val, errors='coerce')
                    if pd.notna(date_val):
                        return date_val
                except:
                    pass
                return pd.Timestamp.max
            result_df['_instruction_date_sort_key'] = result_df['指示日'].apply(instruction_date_sort_key_for_priority)
        else:
            result_df['_instruction_date_sort_key'] = pd.Timestamp.max
        
        # 同一品番内で指示日が古い順でソート（FIFO）
        # 重要：FIFOを極力順守するため、指示日を品番の直後に配置（優先度を上げる）
        # 候補数（feasible_inspector_count）は指示日の後に配置し、指示日順を優先する
        result_df = result_df.sort_values(
            ['_shipping_priority', '_within_two_business_days', '_is_new_product',
             '_later_shipping_date', '_sort_product_id', '_instruction_date_sort_key', 'feasible_inspector_count', '出荷予定日'],
            ascending=[True, False, False, True, True, True, True, True],
            na_position='last'
        ).reset_index(drop=True)
        
        # FIFO確認用ログ（デバッグ用：同一品番・同一出荷予定日のロットの指示日順を確認）
        # 同一品番・同一出荷予定日のロットが複数ある場合、指示日順になっているか確認
        if '指示日' in result_df.columns and '品番' in result_df.columns and '出荷予定日' in result_df.columns:
            for product_num in result_df['品番'].unique():
                product_lots = result_df[result_df['品番'] == product_num]
                for shipping_date in product_lots['出荷予定日'].unique():
                    same_product_same_shipping = product_lots[product_lots['出荷予定日'] == shipping_date]
                    if len(same_product_same_shipping) > 1:
                        instruction_dates = same_product_same_shipping['指示日'].dropna()
                        if len(instruction_dates) > 1:
                            # 指示日がソートされているか確認（デバッグ用）
                            sorted_dates = instruction_dates.sort_values()
                            if not instruction_dates.equals(sorted_dates):
                                self.log_message(
                                    f"⚠️ FIFO警告: 品番 {product_num} (出荷予定日: {shipping_date}) で指示日順が崩れています",
                                    level='warning',
                                    debug=True
                                )
        
        # 指示日ソートキー列は削除しない（FIFO維持のため、割当て処理で使用する）
        # result_df = result_df.drop(columns=['_instruction_date_sort_key'], errors='ignore')
        
        # 優先度4の先頭数件を出荷予定日昇順でログ出力（確認用）
        priority_4_lots = result_df[result_df['_shipping_priority'] == 4]
        if not priority_4_lots.empty:
            preview_count = min(10, len(priority_4_lots))
            preview_dates = priority_4_lots['出荷予定日'].head(preview_count).apply(
                lambda x: str(x).strip() if pd.notna(x) else 'N/A'
            ).tolist()
            self.log_message(
                f"優先度4（その他）の先頭{preview_count}件を出荷予定日昇順でプレビュー: {preview_dates}",
                level='info'
            )

        if '_later_shipping_date' in result_df.columns:
            later_date_df = result_df[pd.notna(result_df['_later_shipping_date'])].copy()
            if not later_date_df.empty:
                min_date = later_date_df['_later_shipping_date'].min().date()
                max_date = later_date_df['_later_shipping_date'].max().date()
                # 先頭側の出荷予定日を少量だけ出して、ログで「日付昇順になっているか」を確認できるようにする
                first_dates = (
                    later_date_df['_later_shipping_date']
                    .dt.strftime('%Y-%m-%d')
                    .head(8)
                    .tolist()
                )
                self.log_message(
                    f"それ以降の日付ロットは出荷予定日昇順で処理します: {len(later_date_df)}件、範囲 {min_date}～{max_date}、先頭 {first_dates}"
                )
        
        # 固定検査員が設定されている品番のロット数をログ出力
        fixed_inspector_lots = result_df[result_df['_has_fixed_inspectors'] == True]
        if not fixed_inspector_lots.empty:
            fixed_products = fixed_inspector_lots['品番'].unique()
            self.log_message(f"固定検査員が設定されている品番のロットを最優先で割り当てます: {len(fixed_inspector_lots)}ロット（品番: {list(fixed_products)}）")

        self.log_message("並び順ロジック: 緊急度 → 3営業日以内の出荷 → 新製品の順で並び替えました。")

        return result_df
    
    def create_inspector_assignment_table(
        self,
        assignment_df: pd.DataFrame,
        product_master_df: pd.DataFrame,
        product_master_path: Optional[str] = None,
        process_master_path: Optional[str] = None,
        inspection_target_keywords: Optional[List[str]] = None
    ) -> Optional[pd.DataFrame]:
        """
        検査員割振りテーブルを作成（高速化版：itertuplesとマージ最適化）
        
        Args:
            assignment_df: ロット割当結果のDataFrame
            product_master_df: 製品マスタのDataFrame
            product_master_path: 製品マスタファイルのパス（オプション）
            process_master_path: 工程マスタファイルのパス（オプション）
            inspection_target_keywords: 検査対象キーワードリスト（オプション）
        
        Returns:
            検査員割当てテーブルのDataFrame、失敗時はNone
        """
        try:
            if assignment_df.empty:
                self.log_message("ロット割り当て結果がありません")
                return None
            
            if product_master_df is None or product_master_df.empty:
                self.log_message("製品マスタが読み込まれていません")
                return None
            
            # ロット数量が0の行を事前にフィルタリング
            original_count_before_filter = len(assignment_df)
            assignment_df = assignment_df[assignment_df['ロット数量'].notna() & (assignment_df['ロット数量'] != 0)].copy()
            filtered_count = original_count_before_filter - len(assignment_df)
            
            if filtered_count > 0:
                self.log_message(f"⚠️ ロット数量が0またはNaNのロットを{filtered_count}件除外しました")
            
            if assignment_df.empty:
                self.log_message("ロット数量が0のため、検査員割り当てをスキップします")
                return None
            
            # 列名のインデックスマップを作成（itertuples用）
            assignment_cols = {col: idx for idx, col in enumerate(assignment_df.columns)}
            product_cols = {col: idx for idx, col in enumerate(product_master_df.columns)}
            
            # 製品マスタを品番と工程番号でインデックス化（高速検索用）
            # スキルマスタと同様のロジック: 工程番号が空のものも保存
            product_master_dict = {}
            for row in product_master_df.itertuples(index=False):
                product_num = str(row[product_cols['品番']]).strip() if pd.notna(row[product_cols['品番']]) else None
                if product_num is None:
                    continue
                    
                process_num = row[product_cols.get('工程番号', -1)] if '工程番号' in product_cols else None
                inspection_time = row[product_cols['検査時間']]
                
                if product_num not in product_master_dict:
                    product_master_dict[product_num] = {}
                
                # 工程番号を正規化（数値型と文字列型の両方に対応）
                if process_num is not None and pd.notna(process_num):
                    # 数値型の場合は整数に変換してから文字列化、文字列型の場合はそのまま使用
                    if isinstance(process_num, (int, float)):
                        process_key = str(int(process_num))
                    else:
                        process_key = str(process_num).strip()
                    
                    if process_key != '':
                        product_master_dict[product_num][process_key] = inspection_time
                else:
                    # 工程番号が空の場合は、空文字列キーとして保存（スキルマスタと同様）
                    product_master_dict[product_num][''] = inspection_time
            
            # 工程マスタを読み込む（先行検査品・当日洗浄品用）
            process_master_df = None
            if process_master_path:
                process_master_df = self.load_process_master(process_master_path)
            
            inspector_results = []
            new_products_to_add = []  # 製品マスタに追加する品番のリスト
            
            # itertuples()を使用して高速化
            for row in assignment_df.itertuples(index=False):
                product_number = row[assignment_cols['品番']]
                product_name = row[assignment_cols.get('品名', -1)] if '品名' in assignment_cols else ''
                current_process_number = row[assignment_cols.get('現在工程番号', -1)] if '現在工程番号' in assignment_cols else None
                lot_quantity = row[assignment_cols['ロット数量']]
                shipping_date = row[assignment_cols.get('出荷予定日', -1)] if '出荷予定日' in assignment_cols else None
                
                # 先行検査品・当日洗浄品の場合、工程マスタから工程番号と工程名を取得
                inferred_process_name = None  # 推定された工程名を保持
                is_same_day_cleaning = False
                
                # 出荷予定日を確認して当日洗浄品かどうかを判定
                shipping_date_str = str(shipping_date).strip() if pd.notna(shipping_date) else ''
                is_same_day_cleaning = (
                    shipping_date_str == "当日洗浄上がり品" or
                    shipping_date_str == "当日洗浄品" or
                    "当日洗浄" in shipping_date_str or
                    shipping_date_str == "先行検査" or
                    shipping_date_str == "当日先行検査"
                )
                
                # 当日洗浄品の場合は、工程番号が空でなくても工程マスタから取得（既存の工程番号を上書き）
                needs_inference = (
                    current_process_number is None
                    or pd.isna(current_process_number)
                    or str(current_process_number).strip() == ''
                )
                if (
                    is_same_day_cleaning
                    and process_master_df is not None
                    and inspection_target_keywords
                    and needs_inference
                ):
                    # 工程マスタから工程番号を推定（既存の工程番号を上書き）
                    inferred_process = self.infer_process_number_from_process_master(
                        product_number,
                        process_master_df,
                        inspection_target_keywords
                    )
                    if inferred_process:
                        old_process_number = current_process_number if current_process_number is not None else 'なし'
                        current_process_number = inferred_process
                        self.log_message(
                            f"先行検査品・当日洗浄品: 品番 '{product_number}' の工程番号を "
                            f"'{old_process_number}' → '{inferred_process}' に更新しました（工程マスタから取得）"
                        )
                        
                        # 工程マスタから工程名を取得
                        inferred_process_name = self.get_process_name_from_process_master(
                            product_number,
                            inferred_process,
                            process_master_df,
                            inspection_target_keywords
                        )
                
                # 製品マスタから検査時間を取得（スキルマスタと同様のロジック）
                inspection_time_per_unit = None
                
                # 品番を正規化（文字列型に統一）
                product_number_normalized = str(product_number).strip() if pd.notna(product_number) else None
                
                if product_number_normalized and product_number_normalized in product_master_dict:
                    product_dict = product_master_dict[product_number_normalized]
                    
                    # スキルマスタと同様の検索ロジック
                    # 1. 現在工程番号が空欄の場合は、工程フィルタをスキップして工程番号が空の行を検索
                    # 洗浄指示から取得したロットの場合、工程番号が複数ある場合は数字が若い方から処理
                    if current_process_number is None or (pd.notna(current_process_number) and str(current_process_number).strip() == ''):
                        # 工程番号が空の行を検索
                        inspection_time_per_unit = product_dict.get('')
                        if inspection_time_per_unit is None:
                            # 工程番号が空の行が見つからない場合、工程番号が複数ある場合は数字が若い方から選択
                            process_keys = [k for k in product_dict.keys() if k != '']  # 空文字列を除外
                            if process_keys:
                                # 工程番号を数値としてソート（数字が若い方から）
                                def sort_key(k):
                                    try:
                                        return int(k)
                                    except (ValueError, TypeError):
                                        return float('inf')  # 数値に変換できない場合は最後に
                                
                                sorted_keys = sorted(process_keys, key=sort_key)
                                selected_key = sorted_keys[0]  # 数字が最も若い工程番号を選択
                                inspection_time_per_unit = product_dict.get(selected_key)
                            else:
                                self.log_message(f"⚠️ 製品マスタ検索失敗: 品番={product_number_normalized}, 工程番号=(空) が見つかりません。利用可能な工程番号: {list(product_dict.keys())}")
                    else:
                        # 2. 工程番号を正規化して検索
                        if isinstance(current_process_number, (int, float)):
                            process_key = str(int(current_process_number))
                        else:
                            process_key = str(current_process_number).strip()
                        
                        if process_key != '':
                            # まず工程番号が一致するものを検索
                            inspection_time_per_unit = product_dict.get(process_key)
                            if inspection_time_per_unit is None:
                                # 工程番号が一致しない場合、工程番号が空の行を検索（スキルマスタと同様）
                                inspection_time_per_unit = product_dict.get('')
                            if inspection_time_per_unit is None:
                                self.log_message(f"⚠️ 製品マスタ検索失敗: 品番={product_number_normalized}, 工程番号={process_key} および工程番号=(空) が見つかりません。利用可能な工程番号: {list(product_dict.keys())}")
                    
                    # 検索結果が取得できない場合
                    if inspection_time_per_unit is None or pd.isna(inspection_time_per_unit):
                        available_processes = list(product_dict.keys())
                        self.log_message(f"⚠️ 品番 '{product_number_normalized}' (工程番号: {current_process_number if current_process_number else 'なし'}) に一致する検査時間が見つかりません。利用可能な工程番号: {available_processes}")
                
                # 製品マスタに存在しない場合は、新規品扱いとして製品マスタに追加する情報を記録
                if product_number_normalized not in product_master_dict:
                    # デフォルト値として15秒/個を使用（製品マスタに追加するため）
                    inspection_time_per_unit = 15.0  # 15秒/個
                    
                    # 製品マスタに追加する情報を記録（重複チェック）
                    if not any(p['品番'] == product_number_normalized for p in new_products_to_add):
                        new_products_to_add.append({
                            '品番': product_number_normalized,
                            '品名': product_name if product_name != -1 and pd.notna(product_name) else '',
                            '工程番号': current_process_number if current_process_number is not None and pd.notna(current_process_number) else '',
                            '検査時間': 15.0,
                            '自動追加': True
                        })
                        self.log_message(f"ℹ️ 品番 '{product_number_normalized}' が製品マスタに存在しません。製品マスタに追加予定です (工程番号: {current_process_number if current_process_number else 'なし'}, 出荷予定日: {shipping_date}, ロット数量: {lot_quantity})")
                elif inspection_time_per_unit is None or pd.isna(inspection_time_per_unit):
                    # 製品マスタに存在するが、工程番号が一致する検査時間が取得できない場合
                    inspection_time_per_unit = 15.0  # 15秒/個
                    available_keys = list(product_dict.keys()) if product_number_normalized in product_master_dict else []
                    self.log_message(f"⚠️ 品番 '{product_number_normalized}' (工程番号: {current_process_number if current_process_number else 'なし'}) の検査時間が取得できません。デフォルト検査時間(15秒/個)を使用します。利用可能な工程番号: {available_keys} (出荷予定日: {shipping_date}, ロット数量: {lot_quantity})")
                
                # 検査時間を計算（秒 × ロット数量）
                total_inspection_time_seconds = inspection_time_per_unit * lot_quantity
                
                # 時間表示に変換（○.○H）
                total_inspection_time_hours = total_inspection_time_seconds / 3600
                
                # 秒/個はそのまま使用（既に秒単位）
                seconds_per_unit = inspection_time_per_unit
                
                # 出荷予定日を取得し、「当日洗浄上がり品」または「先行検査」の場合は文字列として保持
                shipping_date_value = row[assignment_cols.get('出荷予定日', -1)] if '出荷予定日' in assignment_cols else None
                if shipping_date_value != -1 and pd.notna(shipping_date_value):
                    shipping_date_str = str(shipping_date_value).strip()
                    # 「当日洗浄上がり品」または「先行検査」の場合は文字列として保持（元の値を保持）
                    if shipping_date_str == "当日洗浄上がり品" or shipping_date_str == "当日洗浄品" or "当日洗浄" in shipping_date_str or shipping_date_str == "先行検査" or shipping_date_str == "当日先行検査":
                        shipping_date_final = shipping_date_str  # 元の値を保持（「先行検査」は「先行検査」のまま）
                    else:
                        shipping_date_final = shipping_date_value  # その他の場合は元の値を保持
                else:
                    shipping_date_final = None
                
                # 現在工程名の設定
                # 先行検査品・当日洗浄品で工程番号を推定した場合：工程マスタから取得した工程名を使用
                # それ以外：Accessデータベースから取得した現在工程名を使用
                if is_same_day_cleaning and inferred_process_name:
                    process_name = inferred_process_name
                else:
                    # 通常ロットの場合はAccessデータベースから取得した現在工程名を使用
                    process_name = row[assignment_cols.get('現在工程名', -1)] if '現在工程名' in assignment_cols else ''
                    if process_name == -1:
                        process_name = ''
                    elif pd.notna(process_name):
                        process_name = str(process_name).strip()
                    else:
                        process_name = ''
                
                inspector_result = {
                    '出荷予定日': shipping_date_final,
                    '品番': product_number,
                    '品名': row[assignment_cols.get('品名', -1)] if '品名' in assignment_cols else '',
                    '客先': row[assignment_cols.get('客先', -1)] if '客先' in assignment_cols else '',
                    '生産ロットID': row[assignment_cols.get('生産ロットID', -1)] if '生産ロットID' in assignment_cols else '',
                    'ロット数量': lot_quantity,
                    '指示日': row[assignment_cols.get('指示日', -1)] if '指示日' in assignment_cols else '',
                    '号機': row[assignment_cols.get('号機', -1)] if '号機' in assignment_cols else '',
                    '現在工程名': process_name,  # 先行検査品・当日洗浄品は工程マスタから、それ以外はAccessデータベースから
                    '現在工程番号': current_process_number if current_process_number is not None else '',  # 推定された工程番号を反映
                    '秒/個': round(seconds_per_unit, 1),
                    '検査時間': round(total_inspection_time_hours, 1)
                }
                
                inspector_results.append(inspector_result)
            
            # 製品マスタに存在しない品番を製品マスタファイルに追加
            if new_products_to_add and product_master_path:
                try:
                    self.add_products_to_master(new_products_to_add, product_master_path)
                    self.log_message(f"✅ 製品マスタに {len(new_products_to_add)}件の品番を追加しました")
                except Exception as e:
                    self.log_message(f"⚠️ 製品マスタへの追加に失敗しました: {str(e)}")
            
            # ロット割当結果と検査員割当てテーブルの件数を比較
            original_count = len(assignment_df)
            created_count = len(inspector_results)
            
            self.log_message(f"ロット割当結果: {original_count}件 → 検査員割当てテーブル: {created_count}件")
            
            if not inspector_results:
                self.log_message("検査員割振りデータが生成されませんでした")
                return None
            
            inspector_df = pd.DataFrame(inspector_results)
            self.log_message(f"検査員割振りテーブルを作成しました: {len(inspector_df)}件")
            
            return inspector_df
            
        except Exception as e:
            error_msg = f"検査員割振りテーブルの作成に失敗しました: {str(e)}"
            self.log_message(error_msg)
            logger.error(error_msg)
            return None
    
    def assign_inspectors(
        self,
        inspector_df: pd.DataFrame,
        inspector_master_df: pd.DataFrame,
        skill_master_df: pd.DataFrame,
        show_skill_values: bool = False,
        process_master_df: Optional[pd.DataFrame] = None,
        inspection_target_keywords: Optional[List[str]] = None
    ) -> pd.DataFrame:
        """
        検査員を割り当てる
        
        Args:
            inspector_df: 検査対象のロットデータ（品番、不足数、検査時間など）
            inspector_master_df: 検査員マスタデータ（氏名、コード、最大勤務時間など）
            skill_master_df: スキルマスタデータ（品番ごとの検査員スキル）
            show_skill_values: スキル値を表示するかどうか
            process_master_df: 工程マスタデータ（オプション）
            inspection_target_keywords: 検査対象キーワードリスト（オプション）
        
        Returns:
            割り当て結果を含むDataFrame（検査員1〜5列が追加される）
        """
        try:
            if inspector_df is None or inspector_df.empty:
                return inspector_df
            
            if inspector_master_df is None or inspector_master_df.empty:
                self.log_message("検査員マスタが読み込まれていません")
                return inspector_df
            
            if skill_master_df is None or skill_master_df.empty:
                self.log_message("スキルマスタが読み込まれていません")
                return inspector_df
            
            self.same_day_same_name_relaxation_attempts.clear()
            self.logged_vacation_messages.clear()
            # 【高速化】検査員マスタのインデックスを構築
            with perf_timer(loguru_logger, "inspector_assignment.manager.build_inspector_index"):
                self._build_inspector_index(inspector_master_df)
            self.same_day_constraint_relaxations.clear()
            
            # 結果用のDataFrameを準備
            with perf_timer(loguru_logger, "inspector_assignment.manager.prepare_result_dataframe"):
                result_df = self._prepare_result_dataframe(inspector_df)
            
            process_name_context: str = ''
            # 出荷予定日でソート（古い順）- 最優先ルール
            # 日付形式を統一してからソート（当日洗浄品は文字列として保持）
            result_df['出荷予定日'] = result_df['出荷予定日'].apply(self._convert_shipping_date)
            
            # ソート用の補助列を追加
            with perf_timer(loguru_logger, "inspector_assignment.manager.add_sorting_columns"):
                result_df = self._add_sorting_columns(result_df, skill_master_df)

            # 事前にスキルベースの候補人数と割当可能性を算出
            with perf_timer(loguru_logger, "inspector_assignment.manager.calculate_candidates"):
                result_df = self._calculate_feasibility_and_candidates(
                    result_df, skill_master_df, inspector_master_df
                )
            
            # 並び順ロジック: 出荷予定日の優先順位ごとに候補数が少ないロットを優先
            # 列名のインデックスマップを作成（itertuples用）
            result_cols = {col: idx for idx, col in enumerate(result_df.columns)}
            with perf_timer(loguru_logger, "inspector_assignment.manager.sort_lots_by_priority"):
                result_df = self._sort_lots_by_priority(result_df, result_cols)
            high_priority_mask = result_df.get('_is_high_priority', pd.Series([False] * len(result_df), index=result_df.index)).astype(bool)
            high_priority_total = int(high_priority_mask.sum())
            high_priority_remaining = high_priority_total
            if high_priority_total > 0:
                self.log_message(
                    f"優先度の高いロット {high_priority_total}件（当日/当日洗浄）を優先するため、各検査員の割当枠から "
                    f"{HIGH_PRIORITY_RESERVED_CAPACITY_HOURS:.1f}h を確保します"
                )

            # 固定検査員が設定されている品番のロット配分を平準化するためのカウンタ（工数ベース）
            # NOTE:
            # - 現在工程名はロットごとに揺れる（例: 外観 / 外観検査）ためキーに使うとカウンタが分断される。
            # - 固定検査員の「集合」をキー化して、同一品番の固定ロットが同じカウンタを共有するようにする。
            # - 平準化は「ロット数」ではなく「工数（検査時間h）」で行う。
            # key: (品番, 固定検査員名集合) / value: {検査員名: 既割当工数(h)}
            fixed_workload_hours: Dict[Tuple[str, Tuple[str, ...]], Dict[str, float]] = {}
            
            # 各ロットに対して検査員を割り当て
            result_cols_after_sort = {col: idx for idx, col in enumerate(result_df.columns)}
            _first_pass_t0 = perf_counter()
            for row_idx, row in enumerate(result_df.itertuples(index=False)):
                index = result_df.index[row_idx]
                inspection_time = row[result_cols_after_sort['検査時間']]
                product_number = row[result_cols_after_sort['品番']]
                process_number = row[result_cols_after_sort.get('現在工程番号', -1)] if '現在工程番号' in result_cols_after_sort else ''
                if process_number == -1:
                    process_number = ''
                lot_quantity = row[result_cols_after_sort.get('ロット数量', -1)] if 'ロット数量' in result_cols_after_sort else 0
                if lot_quantity == -1:
                    lot_quantity = 0
                pre_status = result_df.at[index, 'assignability_status']
                
                # 当日洗浄上がり品かどうかを判定（先に判定）
                shipping_date = row[result_cols_after_sort.get('出荷予定日', -1)] if '出荷予定日' in result_cols_after_sort else 'N/A'
                shipping_date_str = str(shipping_date).strip() if pd.notna(shipping_date) else ''
                is_same_day_cleaning = self._is_same_day_cleaning_label(shipping_date)
                high_priority_idx = result_cols_after_sort.get('_is_high_priority', -1)
                is_high_priority = bool(row[high_priority_idx]) if high_priority_idx != -1 else False
                if is_high_priority and high_priority_remaining > 0:
                    high_priority_remaining -= 1
                reserve_for_high_priority = (high_priority_remaining > 0) and not is_high_priority

                two_business_idx = result_cols_after_sort.get('_within_two_business_days', -1)
                is_two_business_day = bool(row[two_business_idx]) if two_business_idx != -1 else False
                # 3営業日以内かどうかを判定
                shipping_priority_idx = result_cols_after_sort.get('_shipping_priority', -1)
                is_within_three_business_days = (shipping_priority_idx != -1 and 
                                                  row[shipping_priority_idx] == 3) if shipping_priority_idx != -1 else False
                is_high_priority_urgent = is_same_day_cleaning or is_two_business_day or is_within_three_business_days
                is_same_day_high_duration = is_same_day_cleaning and self._is_same_day_high_duration(inspection_time)
                
                # ロット数量が0の場合は検査員を割り当てない
                if lot_quantity == 0 or pd.isna(lot_quantity) or inspection_time == 0 or pd.isna(inspection_time):
                    reason = "ロット数量0" if (lot_quantity == 0 or pd.isna(lot_quantity)) else "検査時間0"
                    self.log_message(f"ロット数量が0または検査時間が0のため、品番 {product_number} の検査員割り当てをスキップします ({reason})")
                    result_df.at[index, '検査員人数'] = 0
                    result_df.at[index, '分割検査時間'] = 0.0
                    for i in range(1, 6):
                        result_df.at[index, f'検査員{i}'] = ''
                    result_df.at[index, 'チーム情報'] = f'未割当({reason})'
                    result_df.at[index, 'assignability_status'] = 'quantity_zero'
                    result_df.at[index, 'remaining_work_hours'] = round(inspection_time or 0.0, 2)
                    continue
                
                # 改善ポイント: 必要人数と検査時間の割り方（非対称＋部分割当）
                # 必要人数を満たせなかった場合でも、確保できた人数分だけ部分的に割当を行う
                # まずベース候補を取得してから、非対称分配を実行
                base_candidates = result_df.at[index, '_base_candidates'] if '_base_candidates' in result_df.columns else []
                if not isinstance(base_candidates, list):
                    base_candidates = []
                # 高速化: 浅いコピーで十分（変更が必要な場合のみ深いコピー）
                available_inspectors = [insp.copy() for insp in base_candidates] if base_candidates else []
                
                # 【追加】固定検査員が設定されている品番の場合、固定検査員を優先的に配置
                # 登録済み品番リストの固定検査員が設定されている品番は、出荷予定日よりも優先して割り当てる
                process_name = self._get_tuple_value(row, result_cols_after_sort, '現在工程名')
                fixed_inspector_names = self._collect_fixed_inspector_names(product_number, process_name)
                # 固定検査員でも「終日休暇」の場合は割当不可のため除外（0hは許容）
                if fixed_inspector_names:
                    before_count = len(fixed_inspector_names)
                    fixed_inspector_names = [n for n in fixed_inspector_names if not self.is_inspector_on_vacation(n)]
                    if before_count != len(fixed_inspector_names):
                        excluded = before_count - len(fixed_inspector_names)
                        self.log_message(
                            f"固定検査員のうち終日休暇の {excluded}名 を除外しました: 品番 '{product_number}'",
                            debug=True,
                            level='warning'
                        )
                fixed_primary_name: Optional[str] = None
                force_fixed_assignment = False
                fixed_candidate_count = 0
                fixed_balance_key: Optional[Tuple[str, Tuple[str, ...]]] = None
                if fixed_inspector_names:
                    # 固定検査員ロットは、固定検査員のロット数ベースで平準化して「主担当」を1名選ぶ
                    fixed_balance_key = (
                        str(product_number).strip(),
                        tuple(sorted(str(n).strip() for n in fixed_inspector_names if str(n).strip())),
                    )
                    workloads = fixed_workload_hours.setdefault(fixed_balance_key, {name: 0.0 for name in fixed_inspector_names})
                    for name in fixed_inspector_names:
                        workloads.setdefault(name, 0.0)

                    def _inspector_code_sort_key_from_name(n: str) -> Tuple[int, Any]:
                        try:
                            row = self.inspector_name_to_row.get(str(n).strip())
                            raw = row.get('#ID') if row is not None else ''
                        except Exception:
                            raw = ''
                        try:
                            code_str = str(raw).strip()
                        except Exception:
                            code_str = ''
                        try:
                            code_int = int(code_str) if code_str.isdigit() else None
                        except Exception:
                            code_int = None
                        return (0, code_int) if code_int is not None else (1, code_str)

                    fixed_primary_name = min(
                        fixed_inspector_names,
                        key=lambda name: (workloads.get(name, 0.0), _inspector_code_sort_key_from_name(name), str(name).strip()),
                    )
                    # 固定検査員は制約を多少破っても必ず入れる
                    force_fixed_assignment = True

                    # 固定検査員とそれ以外に分離
                    fixed_inspectors = []
                    other_inspectors = []
                    available_inspector_names = {insp['氏名'] for insp in available_inspectors}
                    
                    for inspector in available_inspectors:
                        inspector_name = inspector.get('氏名', '')
                        if inspector_name in fixed_inspector_names:
                            fixed_inspectors.append(inspector)
                        else:
                            other_inspectors.append(inspector)
                    
                    # 【追加】base_candidatesに含まれていない固定検査員を追加
                    missing_fixed_inspectors = [name for name in fixed_inspector_names if name not in available_inspector_names]
                    if missing_fixed_inspectors:
                        # get_available_inspectorsを再度呼び出して、固定検査員を含める
                        process_number = row[result_cols_after_sort.get('現在工程番号', -1)] if '現在工程番号' in result_cols_after_sort else ''
                        if process_number == -1:
                            process_number = ''
                        shipping_date = row[result_cols_after_sort.get('出荷予定日', -1)] if '出荷予定日' in result_cols_after_sort else None
                        
                        # 固定検査員を含む完全な候補リストを取得
                        complete_candidates = self.get_available_inspectors(
                            product_number, process_number, skill_master_df, inspector_master_df,
                            shipping_date=shipping_date, allow_new_team_fallback=False,
                            process_master_df=process_master_df, inspection_target_keywords=inspection_target_keywords,
                            process_name_context=process_name
                        )
                        
                        # 追加された固定検査員を確認
                        complete_inspector_names = {insp['氏名'] for insp in complete_candidates}
                        for missing_name in missing_fixed_inspectors:
                            if missing_name in complete_inspector_names:
                                # complete_candidatesから該当する検査員を取得
                                for insp in complete_candidates:
                                    if insp.get('氏名', '') == missing_name:
                                        fixed_inspectors.append(insp)
                                        available_inspector_names.add(missing_name)
                                        warning_key = ("fixed_inspector_added_from_base", str(product_number).strip(), str(missing_name).strip())
                                        if warning_key not in self.logged_warnings:
                                            self.logged_warnings.add(warning_key)
                                            self.log_message(f"固定検査員 '{missing_name}' をbase_candidatesから追加しました（登録済み品番の特別処置）")
                                        break

                        # スキルマスタに存在しない等で complete_candidates に載らない固定検査員も、
                        # 検査員マスタから強制的に候補へ追加する（固定検査員は制約を多少破っても必ず入れる）。
                        for missing_name in missing_fixed_inspectors:
                            if missing_name in available_inspector_names:
                                continue
                            inspector_info = self._get_inspector_by_name(missing_name, inspector_master_df)
                            if inspector_info.empty:
                                self.log_message(
                                    f"⚠️ 警告: 品番 '{product_number}' の固定検査員 '{missing_name}' が検査員マスタに存在しないため追加できません"
                                )
                                continue
                            inspector_row = inspector_info.iloc[0]
                            fixed_inspectors.append(
                                {
                                    '氏名': str(missing_name).strip(),
                                    'スキル': 99,
                                    '就業時間': inspector_row.get('開始時刻', ''),
                                    'コード': inspector_row.get('#ID', ''),
                                }
                            )
                            available_inspector_names.add(missing_name)
                            warning_key = ("fixed_inspector_added_from_master", str(product_number).strip(), str(missing_name).strip())
                            if warning_key not in self.logged_warnings:
                                self.logged_warnings.add(warning_key)
                                self.log_message(f"固定検査員 '{missing_name}' を検査員マスタから追加しました（強制固定）")
                    
                    # 固定検査員を優先的にリストの先頭に配置
                    if fixed_inspectors:
                        # 平準化で選んだ主担当が候補に存在しない場合は、存在する固定検査員から選び直す
                        if fixed_primary_name:
                            fixed_names_available = [str(insp.get('氏名', '')).strip() for insp in fixed_inspectors if str(insp.get('氏名', '')).strip()]
                            if fixed_primary_name not in fixed_names_available:
                                workloads = fixed_workload_hours.get(fixed_balance_key or ("", tuple()), {})
                                fixed_primary_name = min(
                                    fixed_names_available,
                                    key=lambda name: (workloads.get(name, 0.0), _inspector_code_sort_key_from_name(name), str(name).strip()),
                                ) if fixed_names_available else None

                        # 1ロットに固定検査員を複数人入れない：主担当固定検査員のみ候補として残す
                        if fixed_primary_name:
                            fixed_inspectors = [
                                insp for insp in fixed_inspectors
                                if str(insp.get('氏名', '')).strip() == fixed_primary_name
                            ]

                        # 平準化対象の主担当固定検査員が決まっている場合は、先頭に来るように並び替える
                        if fixed_primary_name:
                            fixed_inspectors.sort(
                                key=lambda insp: (insp.get('氏名', '') != fixed_primary_name)
                            )
                        available_inspectors = fixed_inspectors + other_inspectors
                        self.log_message(f"固定検査員を優先配置（初期割当）: 品番 '{product_number}' の固定検査員 {len(fixed_inspectors)}名を先頭に配置（設定: {len(fixed_inspector_names)}名）")
                    fixed_candidate_count = len(fixed_inspectors)
                else:
                    fixed_candidate_count = 0
                
                # 必要な検査員人数を計算（先に計算してから、当日洗浄上がり品の制約を適用）
                required_inspectors = self._calc_required_inspectors(inspection_time)
                if is_same_day_cleaning and self._should_force_same_day_dual_assignment(inspection_time):
                    required_inspectors = max(required_inspectors, 2)
                
                force_full_pool_for_product = (str(product_number).strip() == "3D025-G4960")
                balance_across_max_inspectors = False
                if force_full_pool_for_product:
                    required_inspectors = 2
                    balance_across_max_inspectors = True
                
                # 固定検査員は「ロットを固定検査員間で平準化」して割り当てるため、
                # 1ロットに固定検査員を複数人入れない（固定ロットは主担当1名に固定する）。
                if force_fixed_assignment and fixed_primary_name:
                    required_inspectors = 1
                    balance_across_max_inspectors = False

                # 【重要】当日洗浄上がり品および先行検査品の品番単位制約ロジック
                # このロジックは理想的な割当てを実現するため、変更時は慎重に検討すること
                # 
                # 制約ルール:
                # 1. 別品番であれば同一検査員を割り当ててOK
                # 2. 同一品番の複数ロットには同一検査員を割り当てない（必須制約）
                # 3. 同一品番の複数ロットがある場合、各ロットに均等に検査員を分散させる
                # 4. 【追加】品名が同じで品番が異なる場合、同じ検査員を割り当てない（必須制約）
                # 
                # 実装のポイント:
                # - self.same_day_cleaning_inspectors は品番ごとに管理（{品番: set(検査員コード)}）
                # - self.same_day_cleaning_inspectors_by_product_name は品名ごとに管理（{品名: set(検査員コード)}）
                # - 各ロットで既に割り当てられた検査員を除外して候補を絞り込む
                # - 複数ロットがある場合は均等分散ロジックで必要人数を調整
                if is_same_day_cleaning:
                    shipping_date_val = row[result_cols_after_sort.get('出荷予定日', -1)] if '出荷予定日' in result_cols_after_sort else None
                    force_assign_same_day = self._should_force_assign_same_day(shipping_date_val)
                    # この品番に既に割り当てられた検査員を取得（品番単位の制約）
                    already_assigned_to_this_product = set() if force_assign_same_day else self.same_day_cleaning_inspectors.get(product_number, set())
                    
                    # 【追加】品名が同じで品番が異なる場合の制約
                    # 品名を取得
                    product_name = row[result_cols_after_sort.get('品名', -1)] if '品名' in result_cols_after_sort else ''
                    product_name_str = str(product_name).strip() if pd.notna(product_name) and product_name != -1 else ''
                    
                    # 品名が同じ他の品番に既に割り当てられた検査員を取得（品名単位の制約）
                    # ただし、4時間超えのロットについては、最初から品名単位の制約を緩和（候補不足を防ぐため）
                    already_assigned_to_same_product_name = set()
                    is_over_4h_initial = inspection_time > 4.0
                    if not force_assign_same_day and product_name_str and not is_over_4h_initial:
                        # 品名単位の追跡辞書から、同じ品名に既に割り当てられた検査員を取得
                        already_assigned_to_same_product_name = self.same_day_cleaning_inspectors_by_product_name.get(product_name_str, set())
                        if already_assigned_to_same_product_name:
                            self.log_message(
                                f"当日洗浄上がり品/先行検査品 {product_number} (品名: {product_name_str}): "
                                f"同じ品名の他の品番に既に割り当てられた検査員 {len(already_assigned_to_same_product_name)}人を除外します（品名単位の制約）",
                                debug=True,
                            )
                    elif is_over_4h_initial and product_name_str:
                        # 4時間超えのロットについては、最初から品名単位の制約を緩和
                        self.log_message(
                            f"当日洗浄上がり品/先行検査品 {product_number} (品名: {product_name_str}): "
                            f"4時間超え（{inspection_time:.1f}h）のため、最初から品名単位の制約を緩和します（品番単位の制約は維持）"
                        )
                    
                    # 品番単位と品名単位の両方の制約を統合
                    # 4時間超えのロットについては、品番単位の制約のみ適用
                    excluded_codes = already_assigned_to_this_product | already_assigned_to_same_product_name
                    
                    # 同一品番のロット数を取得
                    lot_count = result_df.at[index, '_same_day_cleaning_lot_count'] if '_same_day_cleaning_lot_count' in result_df.columns else 1
                    lot_count = max(1, int(lot_count)) if pd.notna(lot_count) else 1
                    
                    # この品番の既に割り当てられたロット数をカウント（現在のロットを含む）
                    # ソート後の順序で、現在のロットより前のロットをカウント
                    current_assigned_lot_count = 0
                    for prev_idx in range(row_idx):
                        prev_index = result_df.index[prev_idx]
                        if (result_df.at[prev_index, '品番'] == product_number and 
                            result_df.at[prev_index, '_same_day_cleaning_lot_count'] > 0):
                            current_assigned_lot_count += 1
                    # 現在のロットを含める
                    current_assigned_lot_count += 1
                    
                    original_count = len(available_inspectors)
                    # 【改善】元の候補リストを保存（制約緩和時に使用）
                    original_available_inspectors = available_inspectors.copy()
                    
                    # 既にこの品番または同じ品名の他の品番に割り当てられた検査員を除外
                    filtered_inspectors = [insp for insp in available_inspectors if insp['コード'] not in excluded_codes]
                    
                    # 品名単位の制約が適用された場合のログ
                    if already_assigned_to_same_product_name:
                        self.log_message(
                            f"当日洗浄上がり品/先行検査品 {product_number} (品名: {product_name_str}): "
                            f"同じ品名の他の品番に既に割り当てられた検査員 {len(already_assigned_to_same_product_name)}人を除外しました（品名単位の制約）",
                            debug=True,
                        )
                    
                    # 同一品番の複数ロットがある場合、各ロットに均等に検査員を分散させる
                    # 利用可能な検査員数をロット数で割って、各ロットに割り当て可能な検査員数を計算
                    if lot_count > 1 and len(filtered_inspectors) > 0:
                        remaining_lots = lot_count - current_assigned_lot_count + 1
                        if remaining_lots <= 0:
                            remaining_lots = 1
                        unique_pool_size = len(original_available_inspectors)
                        used_up = current_assigned_lot_count - 1
                        remaining_unique = max(1, unique_pool_size - used_up)
                        if remaining_unique < required_inspectors:
                            new_required = max(1, remaining_unique)
                            self.log_message(
                                f"当日洗浄上がり品/先行検査品 {product_number}: 同一品番の複数ロット（{lot_count}ロット）に対し、残り {remaining_unique}人しか候補がいないため必要人数を {required_inspectors}人から {new_required}人に調整します（均等分散）"
                            )
                            required_inspectors = new_required
                    
                    # 【改善】品番単位・品名単位の制約を適用（候補が不足している場合は緩和）
                    same_name_relaxation_used = False
                    if is_same_day_cleaning and len(filtered_inspectors) < required_inspectors and already_assigned_to_same_product_name:
                        relaxed_same_name_candidates = [
                            insp for insp in original_available_inspectors
                            if insp['コード'] not in already_assigned_to_this_product
                        ]
                        if relaxed_same_name_candidates:
                            self.log_message(
                                f"?? 警告: 当日洗浄上がり品/先行検査品 {product_number} は同じ品名の別品番との制約で候補が不足しているため、品名単位の制約を緩和して再利用可能にします",
                                level='warning'
                            )
                            filtered_inspectors = relaxed_same_name_candidates
                            same_name_relaxation_used = True

                    if is_same_day_cleaning and not same_name_relaxation_used and len(filtered_inspectors) == 0 and already_assigned_to_same_product_name:
                        overridden_candidates = [
                            insp for insp in original_available_inspectors
                            if insp['コード'] not in already_assigned_to_this_product
                        ]
                        if overridden_candidates:
                            self.log_message(
                                f"?? 警告: 当日洗浄上がり品/先行検査品 {product_number} は同一品名制約で完全に候補が枯渇したため、再利用を許可して再試行します",
                                level='warning'
                            )
                            filtered_inspectors = overridden_candidates

                    same_day_relaxation_key = f"{product_number}:{row_idx}"
                    if (is_same_day_cleaning and len(filtered_inspectors) < required_inspectors and
                        same_day_relaxation_key not in self.same_day_constraint_relaxations and
                        len(original_available_inspectors) > len(filtered_inspectors)):
                        self.same_day_constraint_relaxations.add(same_day_relaxation_key)
                        self.log_message(
                            f"?? 警告: 当日洗浄上がり品 {product_number} は候補数不足のため制約を緩和し、全員候補リストを再利用します",
                            level='warning'
                        )
                        filtered_inspectors = original_available_inspectors.copy()

                    if is_same_day_high_duration and len(filtered_inspectors) < required_inspectors:
                        self.log_message(
                            f"🔶 当日洗浄上がり品 {product_number} ({inspection_time:.1f}h) は長時間ロットのため制約緩和フラグを適用し、候補リストを維持します（必要: {required_inspectors}人）",
                            level='info'
                        )
                        filtered_inspectors = original_available_inspectors.copy()

                    available_inspectors = filtered_inspectors
                    excluded_count = original_count - len(available_inspectors)
                    if excluded_count > 0:
                        constraint_types = []
                        if len(already_assigned_to_this_product) > 0:
                            constraint_types.append("品番単位")
                        if len(already_assigned_to_same_product_name) > 0:
                            constraint_types.append("品名単位")
                        constraint_msg = "・".join(constraint_types) if constraint_types else "制約"
                        self.log_message(
                            f"当日洗浄上がり品/先行検査品 {product_number}: 既に割り当てられた検査員 {excluded_count}人を除外しました（{constraint_msg}の制約、ロット数: {lot_count}）",
                            debug=True,
                        )
                    
                    # 【改善】当日洗浄上がり品は最優先のため、候補が不足している場合は制約を緩和
                    if len(available_inspectors) < required_inspectors:
                        constraint_types = []
                        if len(already_assigned_to_this_product) > 0:
                            constraint_types.append("品番単位")
                        if len(already_assigned_to_same_product_name) > 0:
                            constraint_types.append("品名単位")
                        constraint_msg = "・".join(constraint_types) if constraint_types else "制約"
                        adjusted_required = max(1, len(available_inspectors))
                        if adjusted_required != required_inspectors:
                            self.log_message(
                                f"当日洗浄上がり品 {product_number}: 候補が {len(available_inspectors)}名しかいないため、必要人数を {required_inspectors}人から {adjusted_required}人に調整します",
                                level='warning'
                            )
                            required_inspectors = adjusted_required
                        
                        # 【改善】候補が0人の場合、制約を完全に緩和
                        if len(available_inspectors) == 0:
                            self.log_message(
                                f"⚠️ 警告: 当日洗浄上がり品 {product_number} の候補が0人ですが、最優先のため{constraint_msg}の制約を緩和して再試行します",
                                level='warning'
                            )
                            # 制約を緩和して元の候補を使用（品番単位・品名単位の制約を無視）
                            available_inspectors = original_available_inspectors
                        # 【改善】候補が不足している場合（0人より多いが必要人数に満たない場合）
                        elif len(available_inspectors) < required_inspectors:
                            self.log_message(
                                f"⚠️ 警告: 当日洗浄上がり品 {product_number} の候補が不足しています（必要: {required_inspectors}人、利用可能: {len(available_inspectors)}人、ロット数: {lot_count}）。可能な限り割り当てます。",
                                level='warning'
                            )
                            # 候補が不足しているが、0人より多い場合はそのまま使用（可能な限り割り当て）
                
                # 新規品かどうかを判定（スキルマスタに登録がない場合）
                is_new_product = False
                
                # get_available_inspectorsは既に新製品チームを返す場合があるが、明示的に確認
                skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                if skill_rows.empty:
                    is_new_product = True
                    self.log_message(f"品番 {product_number} は新規品です（スキルマスタ未登録）")
                    
                    # available_inspectorsが空の場合は、新製品チームを取得
                    if not available_inspectors:
                        self.log_message(f"新製品チームのメンバーを取得します")
                        available_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                        if not available_inspectors:
                            self.log_message(f"新製品チームのメンバーも見つからないため、スキップします")
                            result_df.at[index, 'assignability_status'] = 'capacity_shortage'
                            result_df.at[index, 'remaining_work_hours'] = round(inspection_time, 2)
                            continue
                        self.log_message(f"新製品チームメンバー: {len(available_inspectors)}人取得しました")
                    else:
                        self.log_message(f"新規品 {product_number}: get_available_inspectorsから {len(available_inspectors)}人の検査員が返されました（新製品チームの可能性あり）")
                if not available_inspectors and is_same_day_high_duration:
                    fallback_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                    if fallback_inspectors:
                        available_inspectors = fallback_inspectors
                        self.log_message(
                            f"⚡ 長時間の当日洗浄上がり品 {product_number} ({inspection_time:.1f}h) に対し、新製品チームを投入して候補を確保しました"
                        )
                        continue
                elif not available_inspectors and not (force_fixed_assignment and fixed_primary_name):
                    # 詳細な原因を特定
                    skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                    if skill_rows.empty:
                        reason = "スキルマスタ未登録"
                    else:
                        # 工程番号による絞り込み結果を確認
                        filtered_rows = []
                        if process_number is None or str(process_number).strip() == '':
                            filtered_rows = list(skill_rows.iterrows())
                        else:
                            # 工程番号列のインデックスを事前に取得（高速化：itertuples()を使用）
                            process_col_idx = 1  # iloc[1]に対応
                            for row_tuple in skill_rows.itertuples(index=True):
                                row_idx = row_tuple[0]  # インデックス
                                skill_process = row_tuple[process_col_idx + 1]  # itertuplesはインデックスを含むため+1
                                if pd.isna(skill_process) or skill_process == '' or str(skill_process) == str(process_number):
                                    # Seriesとして扱うために元の行を取得
                                    skill_row = skill_rows.loc[row_idx]
                                    filtered_rows.append(skill_row)
                        
                        if not filtered_rows:
                            reason = f"工程番号'{process_number}'に一致するスキル情報なし"
                        else:
                            reason = "条件に合う検査員がいない"
                    
                    shipping_date = row[result_cols_after_sort.get('出荷予定日', -1)] if '出荷予定日' in result_cols_after_sort else 'N/A'
                    self.log_message(f"⚠️ 品番 {product_number} (出荷予定日: {shipping_date}) の検査員が見つかりません: {reason}")
                    self.log_message(f"   詳細: 工程番号={process_number}, 検査時間={inspection_time:.1f}h, ロット数量={lot_quantity}")
                    result_df.at[index, '検査員人数'] = 0
                    result_df.at[index, '分割検査時間'] = 0.0
                    for i in range(1, 6):
                        result_df.at[index, f'検査員{i}'] = ''
                    result_df.at[index, 'チーム情報'] = f'未割当({reason})'
                    self.log_message(f"   チーム情報を設定: '{result_df.at[index, 'チーム情報']}'")
                    result_df.at[index, 'assignability_status'] = 'capacity_shortage'
                    result_df.at[index, 'remaining_work_hours'] = round(inspection_time, 2)
                    continue
                
                # 改善ポイント: 非対称分配＋部分割当の実装
                # 検査員の残勤務時間に応じた非対称分配（貪欲法）を実行
                # 特例: 一ロットで検査員が5名以上必要になる場合、5名に制限
                
                process_name_context = self._get_tuple_value(row, result_cols_after_sort, '現在工程名')
                assigned_inspectors, remaining_time, assigned_time_sum = self.assign_inspectors_asymmetric(
                    available_inspectors, inspection_time, inspector_master_df, product_number, is_new_product,
                    max_inspectors=required_inspectors, allow_same_day_overrun=is_same_day_cleaning,
                    reserve_capacity_for_high_priority=reserve_for_high_priority,
                    process_name_context=process_name_context,
                    fixed_primary_inspector_name=fixed_primary_name,
                    force_fixed_assignment=force_fixed_assignment,
                    ignore_product_limit=(force_full_pool_for_product or self._should_force_assign_same_day(shipping_date)),
                    balance_across_max_inspectors=balance_across_max_inspectors
                )

                # 3D025-G4960（当日洗浄）だけ: 未割当(残り時間あり)になった場合は勤務時間制限を緩和して再試行
                if (
                    force_full_pool_for_product
                    and remaining_time > 0.01
                    and (not force_fixed_assignment)
                ):
                    assigned_inspectors_retry, remaining_time_retry, assigned_time_sum_retry = self.assign_inspectors_asymmetric(
                        available_inspectors, inspection_time, inspector_master_df, product_number, is_new_product,
                        max_inspectors=required_inspectors, allow_same_day_overrun=is_same_day_cleaning,
                        reserve_capacity_for_high_priority=reserve_for_high_priority,
                        process_name_context=process_name_context,
                        fixed_primary_inspector_name=fixed_primary_name,
                        force_fixed_assignment=force_fixed_assignment,
                        ignore_product_limit=(force_full_pool_for_product or self._should_force_assign_same_day(shipping_date)),
                        relax_work_hours=True,
                        balance_across_max_inspectors=balance_across_max_inspectors
                    )
                    if assigned_inspectors_retry:
                        assigned_inspectors, remaining_time, assigned_time_sum = (
                            assigned_inspectors_retry,
                            remaining_time_retry,
                            assigned_time_sum_retry,
                        )

                # 先行検査×固定検査員のロット数をカウント（平準化用）
                if fixed_primary_name and assigned_inspectors:
                    if any(str(insp.get('氏名', '')).strip() == fixed_primary_name for insp in assigned_inspectors):
                        if fixed_balance_key is not None:
                            fixed_workload_hours.setdefault(fixed_balance_key, {}).setdefault(fixed_primary_name, 0.0)
                            fixed_workload_hours[fixed_balance_key][fixed_primary_name] += float(assigned_time_sum or 0.0)
                
                # デバッグログ出力
                self.log_message(f"品番 {product_number}: 必要時間 {inspection_time:.1f}h → 割当時間 {assigned_time_sum:.1f}h, 残り {remaining_time:.1f}h, 割当人数 {len(assigned_inspectors)}人")
                
                # 【追加】当日洗浄上がり品の場合は、通常の割り当て処理後にも検査員を追跡（再分配処理前に追跡）
                if is_same_day_cleaning and len(assigned_inspectors) > 0:
                    product_name = row[result_cols_after_sort.get('品名', -1)] if '品名' in result_cols_after_sort else ''
                    product_name_str = str(product_name).strip() if pd.notna(product_name) and product_name != -1 else ''
                    for inspector in assigned_inspectors:
                        if isinstance(inspector, dict) and 'コード' in inspector:
                            code = inspector['コード']
                            self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(code)
                            # 品名単位でも検査員を記録
                            if product_name_str:
                                self.same_day_cleaning_inspectors_by_product_name.setdefault(product_name_str, set()).add(code)
                
                # 当日洗浄上がり品で必要人数に達しない場合、検査時間を再分配する
                if is_same_day_cleaning and inspection_time > self.required_inspectors_threshold and len(assigned_inspectors) < required_inspectors:
                    self.log_message(f"当日洗浄上がり品 {product_number}: 必要人数 {required_inspectors}人に対して {len(assigned_inspectors)}人しか割り当てられていないため、検査時間を再分配します")
                    
                    current_date = pd.Timestamp.now().date()
                    
                    # 既に割り当てられた検査員のコードを取得
                    assigned_codes = {insp['コード'] for insp in assigned_inspectors}
                    
                    # 既にこの品番に割り当てられた検査員を取得（品番単位の制約）
                    # 再分配処理の各ステップで最新の状態を取得する
                    already_assigned_to_this_product = self.same_day_cleaning_inspectors.get(product_number, set())
                    
                    # 【修正】品名単位の制約も取得
                    product_name = row[result_cols_after_sort.get('品名', -1)] if '品名' in result_cols_after_sort else ''
                    product_name_str = str(product_name).strip() if pd.notna(product_name) and product_name != -1 else ''
                    already_assigned_to_same_product_name = set()
                    if product_name_str:
                        already_assigned_to_same_product_name = self.same_day_cleaning_inspectors_by_product_name.get(product_name_str, set())
                    
                    # 品番単位と品名単位の両方の制約を統合
                    excluded_codes_for_reassignment = already_assigned_to_this_product | already_assigned_to_same_product_name
                    
                    # 追加の検査員候補を取得（より広い範囲から）
                    process_number = row[result_cols_after_sort.get('現在工程番号', -1)] if '現在工程番号' in result_cols_after_sort else ''
                    if process_number == -1:
                        process_number = ''
                    
                    # 当日洗浄上がり品の場合は、制約を緩和して候補を取得
                    # 新製品チームも含めて、より広い範囲から候補を取得
                    additional_candidates = self.get_available_inspectors(
                        product_number, process_number, skill_master_df, inspector_master_df,
                        shipping_date=shipping_date, allow_new_team_fallback=True,
                        process_master_df=process_master_df, inspection_target_keywords=inspection_target_keywords,
                        process_name_context=process_name
                    )
                    
                    # 既にこの品番または同じ品名の他の品番に割り当てられた検査員を除外（品番単位・品名単位の制約）
                    # 当日洗浄上がり品の場合は、既にこのロットに割り当てられた検査員でも、この品番に既に割り当てられている場合は含めない
                    filtered_candidates = []
                    for candidate in additional_candidates:
                        code = candidate['コード']
                        # 当日洗浄上がり品の場合は、品番単位・品名単位の制約を優先
                        # 既にこの品番または同じ品名の他の品番に割り当てられている検査員は除外
                        if is_same_day_cleaning and code in excluded_codes_for_reassignment:
                            # この品番または同じ品名の他の品番に既に割り当てられている場合は除外
                            continue
                        # まだこの品番に割り当てられていない検査員は含める
                        filtered_candidates.append(candidate)
                        # 品番単位・品名単位の制約を厳格に適用（既にこの品番または同じ品名の他の品番に割り当てられている検査員は除外）
                        # 注: 必要人数に達しない場合でも、品番単位・品名単位の制約は維持する
                    
                    # 既に割り当てられた検査員も含めて、全候補で再割り当て
                    # 既に割り当てられた検査員の情報を保持
                    # 高速化: 浅いコピーで十分（変更が必要な場合のみ）
                    existing_assigned = [insp.copy() for insp in assigned_inspectors]
                    
                    # 最新の状態を再取得（既にこの品番に割り当てられた検査員を取得）
                    already_assigned_to_this_product = self.same_day_cleaning_inspectors.get(product_number, set())
                    
                    # 全候補を統合
                    # 当日洗浄上がり品の場合は、既に割り当てられた検査員を追加しない（品番単位の制約を維持）
                    # 高速化: 浅いコピーで十分
                    all_candidates = [c.copy() for c in filtered_candidates]
                    # 当日洗浄上がり品でない場合のみ、既に割り当てられた検査員を追加
                    if not is_same_day_cleaning:
                        for existing in existing_assigned:
                            code = existing['コード']
                            if code not in {c['コード'] for c in all_candidates}:
                                all_candidates.append(existing)
                    
                    # 【修正】当日洗浄上がり品の場合は、品番単位・品名単位の制約を厳格に適用してから再分配
                    # 既にこの品番または同じ品名の他の品番に割り当てられていない検査員のみを使用
                    if is_same_day_cleaning:
                        # 最新の状態を再取得
                        shipping_date_val = row[result_cols_after_sort.get('出荷予定日', -1)] if '出荷予定日' in result_cols_after_sort else None
                        force_assign_same_day = self._should_force_assign_same_day(shipping_date_val)
                        already_assigned_to_this_product = set() if force_assign_same_day else self.same_day_cleaning_inspectors.get(product_number, set())
                        if (not force_assign_same_day) and product_name_str:
                            already_assigned_to_same_product_name = self.same_day_cleaning_inspectors_by_product_name.get(product_name_str, set())
                        else:
                            already_assigned_to_same_product_name = set()
                        excluded_codes_for_reassignment = already_assigned_to_this_product | already_assigned_to_same_product_name
                        # この品番または同じ品名の他の品番に割り当てられていない検査員のみを使用
                        all_candidates_filtered = [c for c in all_candidates if c['コード'] not in excluded_codes_for_reassignment]
                    else:
                        all_candidates_filtered = all_candidates
                    
                    # 候補を分散優先でソート（既にこの品番に割り当てられていない検査員を優先、総検査時間が少ない検査員を優先）
                    def sort_key_for_distribution(candidate):
                        code = candidate['コード']
                        daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                        # 既にこの品番に割り当てられていない検査員を優先、総検査時間が少ない検査員を優先
                        if code not in already_assigned_to_this_product:
                            return (0, daily_hours, candidate.get('_fairness_score', 0))
                        else:
                            return (1, daily_hours, candidate.get('_fairness_score', 0))
                    
                    all_candidates_filtered.sort(key=sort_key_for_distribution)
                    
                    # 【修正】assign_inspectors_asymmetric呼び出し前に、最新の追跡情報を再取得して再度フィルタリング
                    # これにより、前のロットの処理結果が確実に反映される
                    if is_same_day_cleaning:
                        # 最新の状態を再取得（直前のロットの処理結果を含む）
                        shipping_date_val = row[result_cols_after_sort.get('出荷予定日', -1)] if '出荷予定日' in result_cols_after_sort else None
                        force_assign_same_day = self._should_force_assign_same_day(shipping_date_val)
                        already_assigned_to_this_product = set() if force_assign_same_day else self.same_day_cleaning_inspectors.get(product_number, set())
                        if (not force_assign_same_day) and product_name_str:
                            already_assigned_to_same_product_name = self.same_day_cleaning_inspectors_by_product_name.get(product_name_str, set())
                        else:
                            already_assigned_to_same_product_name = set()
                        excluded_codes_for_reassignment_final = already_assigned_to_this_product | already_assigned_to_same_product_name
                        # この品番または同じ品名の他の品番に割り当てられていない検査員のみを使用
                        all_candidates_filtered = [c for c in all_candidates_filtered if c['コード'] not in excluded_codes_for_reassignment_final]
                    
                    # 検査時間全体を再分配
                    # 必要人数に達するまで、検査時間を強制的に分割する
                    reassigned_inspectors, reassigned_remaining, reassigned_time_sum = self.assign_inspectors_asymmetric(
                        all_candidates_filtered, inspection_time, inspector_master_df, product_number, is_new_product,
                        max_inspectors=required_inspectors, allow_same_day_overrun=is_same_day_cleaning,
                        process_name_context=process_name_context
                    )
                    
                    # 必要人数に達しない場合、検査時間を強制的に分割する
                    # 候補が不足している場合でも、可能な限り割り当てる
                    if len(reassigned_inspectors) < required_inspectors:
                        self.log_message(f"当日洗浄上がり品 {product_number}: 必要人数 {required_inspectors}人に達しないため、検査時間を強制的に分割します（現在の割当人数: {len(reassigned_inspectors)}人、候補数: {len(all_candidates_filtered)}人）")
                        # 検査時間を必要人数で分割
                        divided_time_per_inspector = inspection_time / required_inspectors
                        # 候補を必要人数分選択（候補が不足している場合は可能な限り）
                        # ただし、最低2人は確保する（設定時間基準を満たすため）
                        min_required = max(2, required_inspectors) if inspection_time > self.required_inspectors_threshold else required_inspectors
                        # 【修正】フィルタリング後の候補数を使用（制約を満たす候補のみをカウント）
                        max_available = min(min_required, len(all_candidates_filtered))
                        # 候補が不足している場合は、可能な限り多くの候補を使用
                        if max_available < min_required:
                            self.log_message(f"当日洗浄上がり品 {product_number}: 候補が不足しています（必要: {min_required}人、利用可能: {max_available}人）。可能な限り割り当てます。")
                        # 当日洗浄上がり品の場合は、品番単位・品名単位の制約を厳格に適用
                        # 既にこの品番または同じ品名の他の品番に割り当てられていない検査員のみ選択
                        if is_same_day_cleaning:
                            # 最新の状態を再取得
                            already_assigned_to_this_product = self.same_day_cleaning_inspectors.get(product_number, set())
                            if product_name_str:
                                already_assigned_to_same_product_name = self.same_day_cleaning_inspectors_by_product_name.get(product_name_str, set())
                            else:
                                already_assigned_to_same_product_name = set()
                            excluded_codes_for_selection = already_assigned_to_this_product | already_assigned_to_same_product_name
                            # この品番または同じ品名の他の品番に割り当てられていない検査員のみ選択
                            selected_candidates = [c for c in all_candidates_filtered if c['コード'] not in excluded_codes_for_selection]
                            # 【追加】総検査時間が少ない検査員を優先するソート
                            selected_candidates.sort(key=lambda c: (
                                self.inspector_daily_assignments.get(c['コード'], {}).get(current_date, 0.0),  # 総検査時間が少ない順
                                c.get('_fairness_score', 0)  # 公平性スコア
                            ))
                            selected_candidates = selected_candidates[:max_available]
                        else:
                            # 当日洗浄上がり品でない場合は、既存のロジックを使用
                            unassigned_candidates = [c for c in all_candidates]
                            selected_candidates = unassigned_candidates[:max_available]
                        reassigned_inspectors = []
                        reassigned_time_sum = 0.0
                        for candidate in selected_candidates:
                            # 各検査員に割り当て可能な時間を計算
                            # 当日洗浄上がり品は優先度が高いため、制約を大幅に緩和
                            code = candidate['コード']
                            daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                            max_hours = self.get_inspector_max_hours(code, inspector_master_df)
                            allowed_max_hours = self._apply_work_hours_overrun(max_hours)
                            # 当日洗浄上がり品は制約を大幅に緩和（WORK_HOURS_BUFFERを小さく、上限も緩和）
                            work_hours_buffer = WORK_HOURS_BUFFER * 0.1  # 通常の10%に緩和（0.005h）
                            # 上限を緩和（最大勤務時間の95%まで許容）
                            remaining_capacity = max(0.0, allowed_max_hours * 0.95 - daily_hours - work_hours_buffer)
                            product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                            # 当日洗浄上がり品は4時間上限も大幅に緩和（PRODUCT_LIMIT_DRAFT_THRESHOLDを使用、さらに緩和）
                            product_room_to_4h = max(0.0, PRODUCT_LIMIT_DRAFT_THRESHOLD * 1.1 - product_hours)  # 4.5h * 1.1 = 4.95hまで許容
                            # 割り当て可能な時間は、分割時間、残り容量、4時間上限の最小値
                            assignable_time = min(divided_time_per_inspector, remaining_capacity, product_room_to_4h)
                            # 当日洗浄上がり品は、少しでも割り当て可能な場合は含める（0.05h以上）
                            if assignable_time >= 0.05:
                                assignment = candidate.copy()
                                assignment['割当時間'] = assignable_time
                                reassigned_inspectors.append(assignment)
                                reassigned_time_sum += assignable_time
                        reassigned_remaining = inspection_time - reassigned_time_sum
                        self.log_message(f"当日洗浄上がり品 {product_number}: 強制分割により {len(reassigned_inspectors)}人を割り当てました（必要人数: {required_inspectors}人、割当時間合計: {reassigned_time_sum:.1f}h、残り時間: {reassigned_remaining:.1f}h）")
                    
                    if len(reassigned_inspectors) >= required_inspectors:
                        # 必要人数に達した場合、再割り当て結果を使用
                        assigned_inspectors = reassigned_inspectors
                        remaining_time = reassigned_remaining
                        assigned_time_sum = reassigned_time_sum
                        self.log_message(f"当日洗浄上がり品 {product_number}: 再分配により {len(assigned_inspectors)}人を割り当てました（必要人数: {required_inspectors}人）")
                        # 再分配処理が完了した時点で、当日洗浄上がり品の検査員を即座に追跡（次のロットの処理前に反映、品番単位・品名単位）
                        for inspector in assigned_inspectors:
                            if isinstance(inspector, dict) and 'コード' in inspector:
                                code = inspector['コード']
                                self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(code)
                                # 【追加】品名単位でも検査員を記録
                                if product_name_str:
                                    self.same_day_cleaning_inspectors_by_product_name.setdefault(product_name_str, set()).add(code)
                    else:
                        # まだ必要人数に達しない場合、新製品チームも含めて再試行
                        assigned_codes = {insp['コード'] for insp in reassigned_inspectors}
                        new_product_team = self.get_new_product_team_inspectors(inspector_master_df)
                        if new_product_team:
                            # 既に割り当てられた検査員を除外
                            new_product_candidates = [insp for insp in new_product_team if insp['コード'] not in assigned_codes]
                            if new_product_candidates:
                                # 【修正】当日洗浄上がり品の場合は、品番単位・品名単位の制約を厳格に適用してから再分配
                                if is_same_day_cleaning:
                                    # 最新の状態を再取得
                                    already_assigned_to_this_product = self.same_day_cleaning_inspectors.get(product_number, set())
                                    if product_name_str:
                                        already_assigned_to_same_product_name = self.same_day_cleaning_inspectors_by_product_name.get(product_name_str, set())
                                    else:
                                        already_assigned_to_same_product_name = set()
                                    excluded_codes_for_new_team = already_assigned_to_this_product | already_assigned_to_same_product_name
                                    # この品番または同じ品名の他の品番に割り当てられていない検査員のみを使用
                                    # 【修正】all_candidates_filteredも再度フィルタリング（最新の状態を反映）
                                    all_candidates_with_new_team = [c for c in all_candidates_filtered if c['コード'] not in excluded_codes_for_new_team]
                                    new_product_candidates_filtered = [insp for insp in new_product_candidates if insp['コード'] not in excluded_codes_for_new_team]
                                    all_candidates_with_new_team.extend(new_product_candidates_filtered)
                                else:
                                    # 全候補を統合（元の候補情報を使用）
                                    # 高速化: 浅いコピーで十分
                                    all_candidates_with_new_team = [c.copy() for c in all_candidates]
                                    all_candidates_with_new_team.extend(new_product_candidates)
                                
                                # 【追加】総検査時間が少ない検査員を優先するソート
                                all_candidates_with_new_team.sort(key=lambda c: (
                                    self.inspector_daily_assignments.get(c['コード'], {}).get(current_date, 0.0),  # 総検査時間が少ない順
                                    c.get('_fairness_score', 0)  # 公平性スコア
                                ))
                                
                                # 検査時間全体を再分配
                                final_assigned, final_remaining, final_time_sum = self.assign_inspectors_asymmetric(
                                    all_candidates_with_new_team, inspection_time, inspector_master_df, product_number, is_new_product=True,
                                    max_inspectors=required_inspectors, allow_same_day_overrun=is_same_day_cleaning,
                                    process_name_context=process_name_context
                                )
                                
                                # 必要人数に達しない場合、検査時間を強制的に分割する
                                # 候補が不足している場合でも、可能な限り割り当てる
                                if len(final_assigned) < required_inspectors:
                                    self.log_message(f"当日洗浄上がり品 {product_number}: 新製品チームを含めても必要人数 {required_inspectors}人に達しないため、検査時間を強制的に分割します（現在の割当人数: {len(final_assigned)}人、候補数: {len(all_candidates_with_new_team)}人）")
                                    # 検査時間を必要人数で分割
                                    divided_time_per_inspector = inspection_time / required_inspectors
                                    # 候補を必要人数分選択（候補が不足している場合は可能な限り）
                                    # ただし、最低2人は確保する（設定時間基準を満たすため）
                                    min_required = max(2, required_inspectors) if inspection_time > self.required_inspectors_threshold else required_inspectors
                                    max_available = min(min_required, len(all_candidates_with_new_team))
                                    # 候補が不足している場合は、可能な限り多くの候補を使用
                                    if max_available < min_required:
                                        self.log_message(f"当日洗浄上がり品 {product_number}: 候補が不足しています（必要: {min_required}人、利用可能: {max_available}人）。可能な限り割り当てます。")
                                    # 当日洗浄上がり品の場合は、品番単位・品名単位の制約を厳格に適用
                                    # ただし、4時間超えのロットで候補が不足している場合は、品名単位の制約を緩和
                                    # 最新の状態を再取得
                                    already_assigned_to_this_product = self.same_day_cleaning_inspectors.get(product_number, set())
                                    if product_name_str:
                                        already_assigned_to_same_product_name = self.same_day_cleaning_inspectors_by_product_name.get(product_name_str, set())
                                    else:
                                        already_assigned_to_same_product_name = set()
                                    
                                    # 4時間超えのロットで候補が不足している場合は、品名単位の制約を緩和
                                    # 品番単位の制約は維持（同じ品番には再割当しない）
                                    # 品名単位の制約は緩和（同じ品名の他の品番に割り当てられた検査員も候補に含める）
                                    is_over_4h = inspection_time > 4.0
                                    if is_over_4h and max_available < min_required:
                                        # 4時間超えで候補が不足している場合、品名単位の制約を緩和
                                        excluded_codes_for_final_selection = already_assigned_to_this_product  # 品番単位のみ
                                        self.log_message(f"当日洗浄上がり品 {product_number}: 4時間超え（{inspection_time:.1f}h）で候補が不足しているため、品名単位の制約を緩和します（品番単位の制約は維持）")
                                    else:
                                        # 通常は品番単位・品名単位の両方の制約を適用
                                        excluded_codes_for_final_selection = already_assigned_to_this_product | already_assigned_to_same_product_name
                                    
                                    # この品番（または同じ品名の他の品番）に割り当てられていない検査員のみ選択
                                    selected_candidates = [c for c in all_candidates_with_new_team if c['コード'] not in excluded_codes_for_final_selection]
                                    # 【追加】総検査時間が少ない検査員を優先するソート
                                    selected_candidates.sort(key=lambda c: (
                                        self.inspector_daily_assignments.get(c['コード'], {}).get(current_date, 0.0),  # 総検査時間が少ない順
                                        c.get('_fairness_score', 0)  # 公平性スコア
                                    ))
                                    selected_candidates = selected_candidates[:max_available]
                                    final_assigned = []
                                    final_time_sum = 0.0
                                    for candidate in selected_candidates:
                                        # 各検査員に割り当て可能な時間を計算
                                        # 当日洗浄上がり品は優先度が高いため、制約を緩和
                                        code = candidate['コード']
                                        daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                                        max_hours = self.get_inspector_max_hours(code, inspector_master_df)
                                        allowed_max_hours = self._apply_work_hours_overrun(max_hours)
                                        # 当日洗浄上がり品は制約を大幅に緩和（WORK_HOURS_BUFFERを小さく、上限も緩和）
                                        work_hours_buffer = WORK_HOURS_BUFFER * 0.1  # 通常の10%に緩和（0.005h）
                                        # 上限を緩和（最大勤務時間の95%まで許容）
                                        remaining_capacity = max(0.0, allowed_max_hours * 0.95 - daily_hours - work_hours_buffer)
                                        product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                                        # 当日洗浄上がり品は4時間上限も大幅に緩和（PRODUCT_LIMIT_DRAFT_THRESHOLDを使用、さらに緩和）
                                        product_room_to_4h = max(0.0, PRODUCT_LIMIT_DRAFT_THRESHOLD * 1.1 - product_hours)  # 4.5h * 1.1 = 4.95hまで許容
                                        # 割り当て可能な時間は、分割時間、残り容量、4時間上限の最小値
                                        assignable_time = min(divided_time_per_inspector, remaining_capacity, product_room_to_4h)
                                        # 当日洗浄上がり品は、少しでも割り当て可能な場合は含める（0.05h以上）
                                        if assignable_time >= 0.05:
                                            assignment = candidate.copy()
                                            assignment['割当時間'] = assignable_time
                                            final_assigned.append(assignment)
                                            final_time_sum += assignable_time
                                    final_remaining = inspection_time - final_time_sum
                                    self.log_message(f"当日洗浄上がり品 {product_number}: 新製品チームを含めた強制分割により {len(final_assigned)}人を割り当てました（必要人数: {required_inspectors}人、割当時間合計: {final_time_sum:.1f}h、残り時間: {final_remaining:.1f}h）")
                                
                                if len(final_assigned) >= required_inspectors:
                                    assigned_inspectors = final_assigned
                                    remaining_time = final_remaining
                                    assigned_time_sum = final_time_sum
                                    self.log_message(f"当日洗浄上がり品 {product_number}: 新製品チームを含めた再分配により {len(assigned_inspectors)}人を割り当てました（必要人数: {required_inspectors}人）")
                                    # 再分配処理が完了した時点で、当日洗浄上がり品の検査員を即座に追跡（次のロットの処理前に反映、品番単位・品名単位）
                                    for inspector in assigned_inspectors:
                                        if isinstance(inspector, dict) and 'コード' in inspector:
                                            code = inspector['コード']
                                            self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(code)
                                            # 【追加】品名単位でも検査員を記録
                                            if product_name_str:
                                                self.same_day_cleaning_inspectors_by_product_name.setdefault(product_name_str, set()).add(code)
                                else:
                                    # 最終的に必要人数に達しない場合、再割り当て結果を使用
                                    assigned_inspectors = final_assigned
                                    remaining_time = final_remaining
                                    assigned_time_sum = final_time_sum
                                    self.log_message(f"当日洗浄上がり品 {product_number}: 新製品チームを含めても {len(assigned_inspectors)}人しか割り当てられませんでした（必要人数: {required_inspectors}人）")
                                    # 再分配処理が完了した時点で、当日洗浄上がり品の検査員を即座に追跡（次のロットの処理前に反映、品番単位・品名単位）
                                    for inspector in assigned_inspectors:
                                        if isinstance(inspector, dict) and 'コード' in inspector:
                                            code = inspector['コード']
                                            self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(code)
                                            # 【追加】品名単位でも検査員を記録
                                            if product_name_str:
                                                self.same_day_cleaning_inspectors_by_product_name.setdefault(product_name_str, set()).add(code)
                            else:
                                # 新製品チームの候補がない場合、再割り当て結果を使用
                                assigned_inspectors = reassigned_inspectors
                                remaining_time = reassigned_remaining
                                assigned_time_sum = reassigned_time_sum
                                # 再分配処理が完了した時点で、当日洗浄上がり品の検査員を即座に追跡（次のロットの処理前に反映、品番単位・品名単位）
                                if len(assigned_inspectors) > 0:
                                    for inspector in assigned_inspectors:
                                        if isinstance(inspector, dict) and 'コード' in inspector:
                                            code = inspector['コード']
                                            self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(code)
                                            # 【追加】品名単位でも検査員を記録
                                            if product_name_str:
                                                self.same_day_cleaning_inspectors_by_product_name.setdefault(product_name_str, set()).add(code)
                        else:
                            # 新製品チームがない場合、再割り当て結果を使用
                            assigned_inspectors = reassigned_inspectors
                            remaining_time = reassigned_remaining
                            assigned_time_sum = reassigned_time_sum
                            # 再分配処理が完了した時点で、当日洗浄上がり品の検査員を即座に追跡（次のロットの処理前に反映、品番単位・品名単位）
                            if len(assigned_inspectors) > 0:
                                for inspector in assigned_inspectors:
                                    if isinstance(inspector, dict) and 'コード' in inspector:
                                        code = inspector['コード']
                                        self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(code)
                                        # 【追加】品名単位でも検査員を記録
                                        if product_name_str:
                                            self.same_day_cleaning_inspectors_by_product_name.setdefault(product_name_str, set()).add(code)
                
                # 設定時間基準の最低人数チェック: 検査時間が設定時間を超える場合は最低2人必要
                # 当日洗浄上がり品の場合は優先順位が高いため、可能な限り割り当てる（未割当にしない）
                if inspection_time > self.required_inspectors_threshold and len(assigned_inspectors) < required_inspectors:
                    if is_same_day_cleaning:
                        # 当日洗浄上がり品の場合は、設定時間基準違反でも可能な限り割り当てる（未割当にしない）
                        if len(assigned_inspectors) > 0:
                            # 1人以上割り当てられている場合は、そのまま割り当てを維持
                            self.log_message(f"⚠️ 警告: 当日洗浄上がり品 {product_number} は{self.required_inspectors_threshold:.1f}時間基準違反ですが、優先順位が高いため {len(assigned_inspectors)}人を割り当てます（検査時間: {inspection_time:.1f}h, 必要人数: {required_inspectors}人, 実際の割当人数: {len(assigned_inspectors)}人）", level='warning')
                        else:
                            # 0人の場合は、制約を大幅に緩和して再試行
                            self.log_message(f"⚠️ 警告: 当日洗浄上がり品 {product_number} は{self.required_inspectors_threshold:.1f}時間基準違反で0人ですが、優先順位が高いため制約を大幅に緩和して再試行します（検査時間: {inspection_time:.1f}h, 必要人数: {required_inspectors}人）", level='warning')
                            # 制約を大幅に緩和して再試行（未割当ロット再処理と同じロジック）
                            # この処理は後続の未割当ロット再処理で行われるため、ここでは未割当のままにする
                            # ただし、assignability_statusは'logic_conflict'ではなく、後続処理で再試行できるようにする
                            result_df.at[index, '検査員人数'] = 0
                            result_df.at[index, '分割検査時間'] = 0.0
                            for i in range(1, 6):
                                result_df.at[index, f'検査員{i}'] = ''
                            result_df.at[index, 'チーム情報'] = f'未割当({self.required_inspectors_threshold:.1f}時間基準違反: 必要{required_inspectors}人に対して{len(assigned_inspectors)}人)'
                            result_df.at[index, 'remaining_work_hours'] = round(inspection_time, 2)
                            result_df.at[index, 'assignability_status'] = 'logic_conflict'
                            continue
                    else:
                        # 3営業日以内のロットの場合は、新製品チームを追加して分割割当を試みる
                        if is_within_three_business_days:
                            self.log_message(f"⚠️ 警告: 品番 {product_number} (出荷予定日: {shipping_date}) は{self.required_inspectors_threshold:.1f}時間基準違反ですが、3営業日以内のため新製品チームを追加して分割割当を試みます（検査時間: {inspection_time:.1f}h, 必要人数: {required_inspectors}人, 実際の割当人数: {len(assigned_inspectors)}人）", level='warning')
                            
                            # 新製品チームを追加して再試行
                            current_date = pd.Timestamp.now().date()
                            assigned_codes = {insp['コード'] for insp in assigned_inspectors}
                            new_product_team = self.get_new_product_team_inspectors(inspector_master_df)
                            if new_product_team:
                                # 既に割り当てられた検査員を除外
                                new_product_candidates = [insp for insp in new_product_team if insp['コード'] not in assigned_codes]
                                if new_product_candidates:
                                    # 全候補を統合
                                    all_candidates_with_new_team = [c.copy() for c in available_inspectors]
                                    all_candidates_with_new_team.extend(new_product_candidates)
                                    
                                    # 総検査時間が少ない検査員を優先するソート
                                    all_candidates_with_new_team.sort(key=lambda c: (
                                        self.inspector_daily_assignments.get(c['コード'], {}).get(current_date, 0.0),
                                        c.get('_fairness_score', 0) if isinstance(c, dict) and '_fairness_score' in c else 0
                                    ))
                                    
                                    # 検査時間を必要人数で分割
                                    divided_time_per_inspector = inspection_time / required_inspectors
                                    max_available = min(required_inspectors, len(all_candidates_with_new_team))
                                    
                                    # 候補を選択
                                    selected_candidates = all_candidates_with_new_team[:max_available]
                                    final_assigned = []
                                    final_time_sum = 0.0
                                    
                                    for candidate in selected_candidates:
                                        code = candidate['コード']
                                        daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                                        max_hours = self.get_inspector_max_hours(code, inspector_master_df)
                                        allowed_max_hours = self._apply_work_hours_overrun(max_hours)
                                        remaining_capacity = max(0.0, allowed_max_hours - daily_hours - WORK_HOURS_BUFFER)
                                        product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                                        product_room_to_4h = max(0.0, PRODUCT_LIMIT_HARD_THRESHOLD - product_hours)
                                        assignable_time = min(divided_time_per_inspector, remaining_capacity, product_room_to_4h)
                                        
                                        if assignable_time >= 0.05:
                                            assignment = candidate.copy()
                                            assignment['割当時間'] = assignable_time
                                            final_assigned.append(assignment)
                                            final_time_sum += assignable_time
                                    
                                    if len(final_assigned) >= required_inspectors or len(final_assigned) > len(assigned_inspectors):
                                        # 必要人数に達したか、改善した場合は割当を更新
                                        assigned_inspectors = final_assigned
                                        remaining_time = inspection_time - final_time_sum
                                        assigned_time_sum = final_time_sum
                                        self.log_message(f"品番 {product_number}: 新製品チームを含めた分割割当により {len(assigned_inspectors)}人を割り当てました（必要人数: {required_inspectors}人、割当時間合計: {final_time_sum:.1f}h）")
                                    else:
                                        # 改善しなかった場合は、元の割当を維持（可能な限り割り当て）
                                        if len(assigned_inspectors) > 0:
                                            self.log_message(f"⚠️ 警告: 品番 {product_number} (出荷予定日: {shipping_date}) は{self.required_inspectors_threshold:.1f}時間基準違反ですが、可能な限り {len(assigned_inspectors)}人を割り当てます（検査時間: {inspection_time:.1f}h, 必要人数: {required_inspectors}人）", level='warning')
                                        else:
                                            # 0人の場合は未割当
                                            self.log_message(f"⚠️ 警告: 品番 {product_number} (出荷予定日: {shipping_date}) は{self.required_inspectors_threshold:.1f}時間基準違反のため未割当とします（検査時間: {inspection_time:.1f}h, 必要人数: {required_inspectors}人）")
                                            result_df.at[index, '検査員人数'] = 0
                                            result_df.at[index, '分割検査時間'] = 0.0
                                            for i in range(1, 6):
                                                result_df.at[index, f'検査員{i}'] = ''
                                            result_df.at[index, 'チーム情報'] = f'未割当({self.required_inspectors_threshold:.1f}時間基準違反: 必要{required_inspectors}人に対して{len(assigned_inspectors)}人)'
                                            result_df.at[index, 'remaining_work_hours'] = round(inspection_time, 2)
                                            result_df.at[index, 'assignability_status'] = 'logic_conflict'
                                            continue
                                else:
                                    # 新製品チームの候補がない場合
                                    if len(assigned_inspectors) > 0:
                                        self.log_message(f"⚠️ 警告: 品番 {product_number} (出荷予定日: {shipping_date}) は{self.required_inspectors_threshold:.1f}時間基準違反ですが、可能な限り {len(assigned_inspectors)}人を割り当てます（検査時間: {inspection_time:.1f}h, 必要人数: {required_inspectors}人）", level='warning')
                                    else:
                                        self.log_message(f"⚠️ 警告: 品番 {product_number} (出荷予定日: {shipping_date}) は{self.required_inspectors_threshold:.1f}時間基準違反のため未割当とします（検査時間: {inspection_time:.1f}h, 必要人数: {required_inspectors}人）")
                                        result_df.at[index, '検査員人数'] = 0
                                        result_df.at[index, '分割検査時間'] = 0.0
                                        for i in range(1, 6):
                                            result_df.at[index, f'検査員{i}'] = ''
                                        result_df.at[index, 'チーム情報'] = f'未割当({self.required_inspectors_threshold:.1f}時間基準違反: 必要{required_inspectors}人に対して{len(assigned_inspectors)}人)'
                                        result_df.at[index, 'remaining_work_hours'] = round(inspection_time, 2)
                                        result_df.at[index, 'assignability_status'] = 'logic_conflict'
                                        continue
                            else:
                                # 新製品チームがない場合
                                if len(assigned_inspectors) > 0:
                                    self.log_message(f"⚠️ 警告: 品番 {product_number} (出荷予定日: {shipping_date}) は{self.required_inspectors_threshold:.1f}時間基準違反ですが、可能な限り {len(assigned_inspectors)}人を割り当てます（検査時間: {inspection_time:.1f}h, 必要人数: {required_inspectors}人）", level='warning')
                                else:
                                    self.log_message(f"⚠️ 警告: 品番 {product_number} (出荷予定日: {shipping_date}) は{self.required_inspectors_threshold:.1f}時間基準違反のため未割当とします（検査時間: {inspection_time:.1f}h, 必要人数: {required_inspectors}人）")
                                    result_df.at[index, '検査員人数'] = 0
                                    result_df.at[index, '分割検査時間'] = 0.0
                                    for i in range(1, 6):
                                        result_df.at[index, f'検査員{i}'] = ''
                                    result_df.at[index, 'チーム情報'] = f'未割当({self.required_inspectors_threshold:.1f}時間基準違反: 必要{required_inspectors}人に対して{len(assigned_inspectors)}人)'
                                    result_df.at[index, 'remaining_work_hours'] = round(inspection_time, 2)
                                    result_df.at[index, 'assignability_status'] = 'logic_conflict'
                                    continue
                        else:
                            # 3営業日以内でない場合は、設定時間基準違反で未割当
                            self.log_message(f"⚠️ 警告: 品番 {product_number} (出荷予定日: {shipping_date}) は{self.required_inspectors_threshold:.1f}時間基準違反のため未割当とします（検査時間: {inspection_time:.1f}h, 必要人数: {required_inspectors}人, 実際の割当人数: {len(assigned_inspectors)}人）")
                            result_df.at[index, '検査員人数'] = 0
                            result_df.at[index, '分割検査時間'] = 0.0
                            for i in range(1, 6):
                                result_df.at[index, f'検査員{i}'] = ''
                            result_df.at[index, 'チーム情報'] = f'未割当({self.required_inspectors_threshold:.1f}時間基準違反: 必要{required_inspectors}人に対して{len(assigned_inspectors)}人)'
                            result_df.at[index, 'remaining_work_hours'] = round(inspection_time, 2)
                            result_df.at[index, 'assignability_status'] = 'logic_conflict'
                            continue
                
                # 検査員が選択されなかった場合（ルール違反を避けるため未割当）
                # 当日洗浄上がり品の場合は優先順位が高いため、未割当ロット再処理で再試行される
                target_process_name = self._get_tuple_value(row, result_cols_after_sort, '現在工程名')
                if len(assigned_inspectors) == 0:
                    # 詳細な原因を記録
                    reason_parts = []
                    if not available_inspectors:
                        reason_parts.append("候補検査員なし")
                    else:
                        # 改善ポイント: 非対称分配ではdivided_timeが存在しないため、inspection_timeを使用
                        # filter_available_inspectorsの結果を確認（簡易的なチェック用にinspection_timeを使用）
                        filtered_count = len(self.filter_available_inspectors(
                            available_inspectors,
                            inspection_time,
                            inspector_master_df,
                            product_number,
                            process_name_context=target_process_name
                        ))
                        if filtered_count == 0:
                            reason_parts.append("勤務時間または4時間上限により全員除外")
                        else:
                            reason_parts.append("ルール違反回避")
                    
                    reason = " / ".join(reason_parts) if reason_parts else "未割当"
                    shipping_date = row[result_cols_after_sort.get('出荷予定日', -1)] if '出荷予定日' in result_cols_after_sort else 'N/A'
                    
                    if is_same_day_cleaning:
                        # 当日洗浄上がり品の場合は、未割当ロット再処理で再試行されるため、警告のみ
                        self.log_message(f"⚠️ 警告: 当日洗浄上がり品 {product_number} は{reason}のため一時的に未割当としますが、優先順位が高いため後続の未割当ロット再処理で再試行します", level='warning')
                        self.log_message(f"   詳細: 候補数={len(available_inspectors)}人, 検査時間={inspection_time:.1f}h, ロット数量={lot_quantity}")
                    else:
                        self.log_message(f"⚠️ 警告: 品番 {product_number} (出荷予定日: {shipping_date}) は{reason}のため未割当とします")
                        self.log_message(f"   詳細: 候補数={len(available_inspectors)}人, 検査時間={inspection_time:.1f}h, ロット数量={lot_quantity}")
                    
                    result_df.at[index, '検査員人数'] = 0
                    result_df.at[index, '分割検査時間'] = 0.0
                    for i in range(1, 6):
                        result_df.at[index, f'検査員{i}'] = ''
                    result_df.at[index, 'チーム情報'] = f'未割当({reason})'
                    self.log_message(f"   チーム情報を設定: '{result_df.at[index, 'チーム情報']}'")
                    result_df.at[index, 'remaining_work_hours'] = round(inspection_time, 2)
                    if pre_status != 'capacity_shortage':
                        result_df.at[index, 'assignability_status'] = 'logic_conflict'
                    continue
                # 改善ポイント: 部分割当の処理
                # remaining_work_hoursを非対称分配の結果から設定
                result_df.at[index, 'remaining_work_hours'] = round(remaining_time, 2)
                
                # assignability_statusの更新
                if len(assigned_inspectors) == 0:
                    # 未割当の場合は既に処理済み（上記のif文で処理）
                    pass
                elif remaining_time > 0.01:  # 0.01h以上の残りがある場合
                    if pre_status == 'capacity_shortage':
                        result_df.at[index, 'assignability_status'] = 'capacity_shortage_partial'
                    elif pre_status == 'skill_mismatch':
                        result_df.at[index, 'assignability_status'] = 'skill_mismatch_partial'
                    else:
                        result_df.at[index, 'assignability_status'] = 'partial_assigned'
                else:
                    result_df.at[index, 'remaining_work_hours'] = 0.0
                    if pre_status == 'capacity_shortage':
                        result_df.at[index, 'assignability_status'] = 'capacity_shortage_resolved'
                    elif pre_status == 'skill_mismatch':
                        result_df.at[index, 'assignability_status'] = 'skill_mismatch_resolved'
                    else:
                        result_df.at[index, 'assignability_status'] = 'fully_assigned'
                
                # 結果を設定
                result_df.at[index, '検査員人数'] = len(assigned_inspectors)
                # 分割検査時間の計算: 実際の割り当て時間の平均（非対称分配の場合は各検査員の割当時間が異なる）
                if len(assigned_inspectors) > 0:
                    # 非対称分配の場合、各検査員の割当時間の平均を計算
                    if all('割当時間' in insp for insp in assigned_inspectors):
                        # 各検査員の割当時間の平均
                        divided_time = sum(insp['割当時間'] for insp in assigned_inspectors) / len(assigned_inspectors)
                    else:
                        # フォールバック: 検査時間 ÷ 実際の分割した検査人数
                        divided_time = inspection_time / len(assigned_inspectors)
                    result_df.at[index, '分割検査時間'] = round(divided_time, 1)
                else:
                    result_df.at[index, '分割検査時間'] = 0.0
                # inspectorが辞書でない場合の対処
                over_limit_present = False
                for insp in assigned_inspectors:
                    if not isinstance(insp, dict):
                        if hasattr(insp, 'to_dict'):
                            insp = insp.to_dict()
                        else:
                            continue
                    if insp.get('over_product_limit', False):
                        over_limit_present = True
                        if 'コード' in insp:
                            self.relaxed_product_limit_assignments.add((insp['コード'], product_number))
                result_df.at[index, 'over_product_limit_flag'] = over_limit_present
                
                # 現在の日付を取得（勤務時間の履歴追跡用）
                current_time = pd.Timestamp.now()
                current_date = current_time.date()
                
                # 【重要】当日洗浄上がり品および先行検査品の検査員を追跡（品番単位・品名単位）
                # この追跡処理により、同一品番の複数ロットに同一検査員が割り当てられることを防ぐ
                # また、品名が同じで品番が異なる場合も同じ検査員を割り当てない
                # 割り当てられた検査員コードをログに出力して、割当ての透明性を確保
                if is_same_day_cleaning:
                    assigned_codes_list = []
                    # 品名を取得
                    product_name = result_df.at[index, '品名'] if '品名' in result_df.columns else ''
                    product_name_str = str(product_name).strip() if pd.notna(product_name) else ''
                    
                    for inspector in assigned_inspectors:
                        if isinstance(inspector, dict) and 'コード' in inspector:
                            code = inspector['コード']
                            assigned_codes_list.append(code)
                            # 品番ごとに割り当てられた検査員コードを記録（品番単位の制約管理）
                            self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(code)
                            
                            # 【追加】品名単位でも検査員を記録（品名が同じで品番が異なる場合の制約用）
                            if product_name_str:
                                self.same_day_cleaning_inspectors_by_product_name.setdefault(product_name_str, set()).add(code)
                    
                    if assigned_codes_list:
                        # 割り当てられた検査員コードをログに出力（割当ての透明性確保）
                        if product_name_str:
                            self.log_message(f"当日洗浄上がり品/先行検査品 {product_number} (品名: {product_name_str}): 割り当てられた検査員コード: {', '.join(assigned_codes_list)}（品番単位・品名単位の制約）")
                        else:
                            self.log_message(f"当日洗浄上がり品/先行検査品 {product_number}: 割り当てられた検査員コード: {', '.join(assigned_codes_list)}（品番単位の制約）")
                
                # 検査員名を設定
                team_members = []
                for i, inspector in enumerate(assigned_inspectors):
                    if i < 5:  # 最大5人まで
                        # inspectorが辞書でない場合（pandas Series等）の対処
                        if not isinstance(inspector, dict):
                            if hasattr(inspector, 'to_dict'):
                                inspector = inspector.to_dict()
                            else:
                                self.log_message(f"警告: inspectorが辞書形式ではありません: {type(inspector)}")
                                continue
                        
                        if show_skill_values:
                            # 新規品チームの場合はスキル値を表示せず(新)のみ
                            if inspector.get('is_new_team', False):
                                inspector_name = f"{inspector['氏名']}(新)"
                            else:
                                inspector_name = f"{inspector['氏名']}({inspector['スキル']})"
                        else:
                            # スキル非表示でも新規品チームの場合は(新)を表示
                            if inspector.get('is_new_team', False):
                                inspector_name = f"{inspector['氏名']}(新)"
                            else:
                                inspector_name = inspector['氏名']
                        
                        result_df.at[index, f'検査員{i+1}'] = inspector_name
                        team_members.append(inspector['氏名'])
                        
                        # 改善ポイント: 非対称分配のため、各検査員の割当時間を個別に取得
                        code = inspector['コード']
                        assigned_time = inspector.get('割当時間', 0.0)  # 非対称分配で設定された個別の割当時間
                        
                        # 日次勤務時間の履歴を更新
                        if code not in self.inspector_daily_assignments:
                            self.inspector_daily_assignments[code] = {}
                        if current_date not in self.inspector_daily_assignments[code]:
                            self.inspector_daily_assignments[code][current_date] = 0.0
                        self.inspector_daily_assignments[code][current_date] += assigned_time
                        
                        # 総勤務時間の履歴を更新
                        if code not in self.inspector_work_hours:
                            self.inspector_work_hours[code] = 0.0
                        self.inspector_work_hours[code] += assigned_time
                        
                        # 同一品番の累計時間を更新（4時間上限のためのトラッキング）
                        if code not in self.inspector_product_hours:
                            self.inspector_product_hours[code] = {}
                        new_product_hours = (
                            self.inspector_product_hours[code].get(product_number, 0.0) + assigned_time
                        )
                        self.inspector_product_hours[code][product_number] = new_product_hours
                        
                        # 【改善】同一品番4時間超過を割り当て後にチェックして警告
                        if new_product_hours > self.product_limit_hard_threshold:
                            # 4.2h未満の場合は警告のみ（許容範囲内）
                            if new_product_hours <= PRODUCT_LIMIT_FINAL_TOLERANCE:
                                self.log_message(
                                    f"⚠️ 警告: 検査員 '{inspector['氏名']}' (コード: {code}) の同一品番 {product_number} の累計時間が"
                                    f" {new_product_hours:.1f}h となり、4時間上限をわずかに超過しています（許容範囲内: ≤{PRODUCT_LIMIT_FINAL_TOLERANCE}h）",
                                    level='warning'
                                )
                                self.relaxed_product_limit_assignments.add((code, product_number))
                            else:
                                # 4.2h超過の場合は重大な警告
                                self.log_message(
                                    f"⚠️ 重大警告: 検査員 '{inspector['氏名']}' (コード: {code}) の同一品番 {product_number} の累計時間が"
                                    f" {new_product_hours:.1f}h となり、4時間上限を超過しています（上限: {self.product_limit_hard_threshold:.1f}h）。"
                                    f"最適化フェーズで是正されます。",
                                    level='warning'
                                )
                        
                        # 【改善】当日洗浄上がり品の制約違反を割り当て後にチェック
                        if is_same_day_cleaning:
                            # 品番単位の制約チェック
                            already_assigned_to_this_product = self.same_day_cleaning_inspectors.get(product_number, set())
                            if code in already_assigned_to_this_product and len(already_assigned_to_this_product) > 1:
                                # 既にこの品番に割り当てられている検査員が複数いる場合、重複をチェック
                                # ただし、現在のロットで追加されたものは除外
                                other_lots_with_same_inspector = [
                                    idx for idx, row in result_df.iterrows()
                                    if idx != index
                                    and row.get('品番') == product_number
                                    and any(
                                        str(row.get(f'検査員{j}', '')).split('(')[0].strip() == inspector['氏名']
                                        for j in range(1, 6)
                                    )
                                ]
                                if other_lots_with_same_inspector:
                                    self.log_message(
                                        f"⚠️ 警告: 当日洗浄上がり品 {product_number} のロット {index} に検査員 '{inspector['氏名']}' を割り当てましたが、"
                                        f"同一品番の他のロット {other_lots_with_same_inspector} にも既に割り当てられています。"
                                        f"最適化フェーズで是正されます。",
                                        level='warning'
                                    )
                            
                            # 品名単位の制約チェック
                            if product_name_str:
                                already_assigned_to_same_product_name = self.same_day_cleaning_inspectors_by_product_name.get(product_name_str, set())
                                if code in already_assigned_to_same_product_name and len(already_assigned_to_same_product_name) > 1:
                                    # 同じ品名の他の品番に既に割り当てられているかチェック
                                    other_products_with_same_name = [
                                        (idx, row.get('品番'))
                                        for idx, row in result_df.iterrows()
                                        if idx != index
                                        and str(row.get('品名', '')).strip() == product_name_str
                                        and row.get('品番') != product_number
                                        and any(
                                            str(row.get(f'検査員{j}', '')).split('(')[0].strip() == inspector['氏名']
                                            for j in range(1, 6)
                                        )
                                    ]
                                    if other_products_with_same_name:
                                        self.log_message(
                                            f"⚠️ 警告: 当日洗浄上がり品 {product_number} (品名: {product_name_str}) のロット {index} に検査員 '{inspector['氏名']}' を割り当てましたが、"
                                            f"同じ品名の他の品番 {[p for _, p in other_products_with_same_name]} にも既に割り当てられています。"
                                            f"最適化フェーズで是正されます。",
                                            level='warning'
                                        )
                        
                        # 改善ポイント: 品番切替ペナルティ用の品番種類数を追跡
                        if code not in self.inspector_product_variety:
                            self.inspector_product_variety[code] = set()
                        self.inspector_product_variety[code].add(product_number)
                
                # チーム情報を設定
                if len(assigned_inspectors) > 1:
                    team_info = f"チーム: {', '.join(team_members)}"
                else:
                    team_info = f"個人: {team_members[0] if team_members else ''}"
                
                result_df.at[index, 'チーム情報'] = team_info
                self.log_message(
                    f"[割当結果] 品番 {product_number}: 割当人数 {len(assigned_inspectors)} / "
                    f"残時間 {result_df.at[index, 'remaining_work_hours']:.2f}h / status={result_df.at[index, 'assignability_status']}",
                    debug=True,
                )
            
            loguru_logger.bind(channel="PERF").debug(
                "PERF {}: {:.1f} ms",
                "inspector_assignment.manager.first_pass_assign",
                (perf_counter() - _first_pass_t0) * 1000.0,
            )
            self.log_message(f"第1次割り当てが完了しました: {len(result_df)}件")
            
            # 割り当て統計（第1次）は通常ログでは省略（ログ量削減）
            if self.debug_mode:
                self.log_message("=== 第1次割り当て統計 ===")
                with perf_timer(loguru_logger, "inspector_assignment.manager.print_stats.first_pass"):
                    self.print_assignment_statistics(inspector_master_df)
            
            # 全体最適化を実行（勤務時間超過の調整と偏りの是正）
            self.log_message("=== 全体最適化を開始 ===")
            with perf_timer(loguru_logger, "inspector_assignment.manager.optimize_assignments"):
                result_df = self.optimize_assignments(result_df, inspector_master_df, skill_master_df, show_skill_values, process_master_df, inspection_target_keywords)
            self.log_message("=== 全体最適化が完了 ===")
            
            # 最終割り当て統計を表示
            self.log_message("=== 最終割り当て統計 ===")
            with perf_timer(loguru_logger, "inspector_assignment.manager.print_stats.final"):
                self.print_assignment_statistics(inspector_master_df)
            
            # 最終KPI（通常は簡易、デバッグ時は詳細）
            with perf_timer(loguru_logger, "inspector_assignment.manager.print_detailed_kpi"):
                self.print_detailed_kpi_statistics(result_df, inspector_master_df, skill_master_df)
            
            if '_base_candidates' in result_df.columns:
                result_df = result_df.drop(columns=['_base_candidates'])
            if '_sort_product_id' in result_df.columns:
                result_df = result_df.drop(columns=['_sort_product_id'])
            if '_is_new_product' in result_df.columns:
                result_df = result_df.drop(columns=['_is_new_product'])
            if '_has_fixed_inspectors' in result_df.columns:
                result_df = result_df.drop(columns=['_has_fixed_inspectors'])
            if '_is_high_priority' in result_df.columns:
                result_df = result_df.drop(columns=['_is_high_priority'])
            if '_shipping_priority' in result_df.columns:
                result_df = result_df.drop(columns=['_shipping_priority'])
            if '_later_shipping_date' in result_df.columns:
                result_df = result_df.drop(columns=['_later_shipping_date'])
            
            # 最終的な表示用ソート: 出荷予定日、品番、指示日の順
            # 出荷予定日のソートキー関数
            current_date = pd.Timestamp.now().date()
            
            # 翌営業日の計算（金曜日の場合は翌週の月曜日）
            def get_next_business_day(date_val):
                """翌営業日を取得（金曜日の場合は翌週の月曜日）"""
                weekday = date_val.weekday()  # 0=月曜日, 4=金曜日
                if weekday == 4:  # 金曜日
                    return date_val + timedelta(days=3)  # 翌週の月曜日
                else:
                    return date_val + timedelta(days=1)  # 翌日
            
            next_business_day = get_next_business_day(current_date)
            
            def shipping_date_sort_key(val):
                if pd.isna(val):
                    return (5, None)  # 最後に
                val_str = str(val).strip()
                
                # 1. 当日の日付（優先度0）
                try:
                    date_val = pd.to_datetime(val, errors='coerce')
                    if pd.notna(date_val):
                        date_date = date_val.date()
                        if date_date == current_date:
                            return (0, date_val)
                except:
                    pass
                
                # 2. 当日洗浄上がり品（優先度1）
                if (val_str == "当日洗浄上がり品" or 
                    val_str == "当日洗浄品" or
                    "当日洗浄" in val_str):
                    return (1, val_str)
                
                # 3. 先行検査品（優先度2）
                if (val_str == "先行検査" or
                    val_str == "当日先行検査"):
                    return (2, val_str)
                
                # 4. 翌日または翌営業日（優先度3）
                try:
                    date_val = pd.to_datetime(val, errors='coerce')
                    if pd.notna(date_val):
                        date_date = date_val.date()
                        if date_date == next_business_day:
                            return (3, date_val)
                except:
                    pass
                
                # 5. それ以降の日付（優先度4）
                try:
                    date_val = pd.to_datetime(val, errors='coerce')
                    if pd.notna(date_val):
                        return (4, date_val)
                except:
                    pass
                
                return (5, val_str)  # その他文字列
            
            # 指示日のソートキー関数
            def instruction_date_sort_key(val):
                if pd.isna(val):
                    return None
                try:
                    date_val = pd.to_datetime(val, errors='coerce')
                    if pd.notna(date_val):
                        return date_val
                except:
                    pass
                return val
            
            # ソートキーを追加
            result_df['_shipping_sort_key'] = result_df['出荷予定日'].apply(shipping_date_sort_key)
            if '指示日' in result_df.columns:
                result_df['_instruction_sort_key'] = result_df['指示日'].apply(instruction_date_sort_key)
            else:
                result_df['_instruction_sort_key'] = None
            
            # ソート実行: 出荷予定日、品番、指示日の順
            sort_columns = ['_shipping_sort_key', '品番']
            if '指示日' in result_df.columns:
                sort_columns.append('_instruction_sort_key')
            
            result_df = result_df.sort_values(
                sort_columns,
                ascending=[True, True, True] if '指示日' in result_df.columns else [True, True],
                na_position='last'
            ).reset_index(drop=True)
            
            # ソートキー列を削除
            result_df = result_df.drop(columns=['_shipping_sort_key', '_instruction_sort_key'], errors='ignore')
            
            # 【高速化】ログバッファをフラッシュ
            if self.log_batch_enabled:
                self._flush_log_buffer()
                
            return result_df
            
        except Exception as e:
            # 【高速化】ログバッファをフラッシュ（エラー時も）
            if self.log_batch_enabled:
                self._flush_log_buffer()
            
            error_msg = f"検査員割り当て中にエラーが発生しました: {str(e)}"
            self.log_message(error_msg)
            logger.error(error_msg, exc_info=True)  # スタックトレースも出力
            
            # エラーが発生した場合でも、最低限の列を追加して返す
            if inspector_df is not None and not inspector_df.empty:
                result_df = inspector_df.copy()
                # 必要な列が存在しない場合は追加
                if '検査員人数' not in result_df.columns:
                    result_df['検査員人数'] = 0
                if '分割検査時間' not in result_df.columns:
                    result_df['分割検査時間'] = 0.0
                for i in range(1, 6):
                    col_name = f'検査員{i}'
                    if col_name not in result_df.columns:
                        result_df[col_name] = ''
                if 'チーム情報' not in result_df.columns:
                    result_df['チーム情報'] = 'エラーにより割り当てできませんでした'
                return result_df
            
            return inspector_df
    
    def get_available_inspectors(
        self,
        product_number: str,
        process_number: Optional[Any],
        skill_master_df: pd.DataFrame,
        inspector_master_df: pd.DataFrame,
        shipping_date: Optional[Any] = None,
        allow_new_team_fallback: bool = False,
        process_master_df: Optional[pd.DataFrame] = None,
        inspection_target_keywords: Optional[List[str]] = None,
        process_name_context: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        利用可能な検査員を取得
        
        Args:
            product_number: 品番
            process_number: 工程番号
            skill_master_df: スキルマスタ
            inspector_master_df: 検査員マスタ
            shipping_date: 出荷予定日（新規品対応チームのフォールバック判定に使用）
            allow_new_team_fallback: 新規品対応チームをフォールバックとして使用するか
                True: 出荷予定日が間近で他に割当てられない場合のみ新規品対応チームを使用
                False: 新規品対応チームを使用しない（通常の品番の場合）
            process_master_df: 工程マスタのDataFrame（先行検査品・当日洗浄品用）
            inspection_target_keywords: 検査対象CSVのキーワードリスト（先行検査品・当日洗浄品用）
            process_name_context: 現在工程名（固定検査員の工程フィルタに使用）
        
        Returns:
            利用可能な検査員のリスト
        """
        try:
            available_inspectors = []
            target_process_name = str(process_name_context or '').strip()
            
            # 先行検査品・当日洗浄品で工程番号が空の場合、工程マスタから推定
            if (process_number is None or str(process_number).strip() == ''):
                shipping_date_str = str(shipping_date).strip() if pd.notna(shipping_date) else ''
                is_same_day_cleaning = (
                    shipping_date_str == "当日洗浄上がり品" or
                    shipping_date_str == "当日洗浄品" or
                    "当日洗浄" in shipping_date_str or
                    shipping_date_str == "先行検査" or
                    shipping_date_str == "当日先行検査"
                )
                
                if is_same_day_cleaning and process_master_df is not None and inspection_target_keywords:
                    inferred_process = self.infer_process_number_from_process_master(
                        product_number,
                        process_master_df,
                        inspection_target_keywords
                    )
                    if inferred_process:
                        process_number = inferred_process
                        self.log_message(f"先行検査品・当日洗浄品: 品番 '{product_number}' の工程番号を '{inferred_process}' に設定しました（スキルマスタ検索用）", debug=True)
            
            # 新規品対応チームメンバーのコードリストを取得（優先度調整用）
            new_team_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
            new_team_codes = {insp['コード'] for insp in new_team_inspectors}
            
            # デバッグ情報を出力
            self.log_message(f"品番 '{product_number}' の検査員を検索中...", debug=True)
            self.log_message(f"工程番号: '{process_number}'", debug=True)
            if new_team_codes:
                self.log_message(f"新規品対応チームメンバー（優先度調整対象）: {sorted(new_team_codes)}", debug=True)
            
            # スキルマスタから該当する品番の行を取得（完全一致のみ）
            skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
            
            if skill_rows.empty:
                self.log_message(f"品番 '{product_number}' がスキルマスタに見つかりません", debug=True)
                # 新規品の場合は新製品チームのメンバーを取得
                if allow_new_team_fallback:
                    self.log_message("新規品のため、新製品チームのメンバーを取得します", debug=True)
                    return self.get_new_product_team_inspectors(inspector_master_df)
                else:
                    # 出荷予定日が間近の場合は新規品対応チームを使用
                    if shipping_date is not None:
                        shipping_date = pd.to_datetime(shipping_date, errors='coerce')
                        if pd.notna(shipping_date):
                            shipping_date_date = shipping_date.date()
                            current_date = pd.Timestamp.now().date()
                            two_weeks_later = current_date + timedelta(days=14)
                            if shipping_date_date <= two_weeks_later:
                                self.log_message(f"出荷予定日が間近（{shipping_date_date}）で他に割当てられないため、新規品対応チームを使用します")
                                return self.get_new_product_team_inspectors(inspector_master_df)
                    self.log_message("利用可能な検査員が見つかりません")
                    return []
            # 工程番号による絞り込み処理
            filtered_skill_rows = []
            # 追加仕様: 現在工程番号が空欄の場合は工程による絞り込みを行わず、品番一致行をすべて対象
            # 洗浄指示から取得したロットの場合、工程番号が複数ある場合は数字が若い方から処理
            if process_number is None or str(process_number).strip() == '':
                self.log_message("現在工程番号が空欄のため、工程フィルタをスキップして品番一致行を処理", debug=True)
                
                # 工程番号が空の行を優先的に取得
                empty_process_rows = []
                numeric_process_rows = []
                other_process_rows = []
                
                # 工程番号列のインデックスを事前に取得（高速化：itertuples()を使用）
                process_col_idx = 1  # iloc[1]に対応
                for row_tuple in skill_rows.itertuples(index=True):
                    row_idx = row_tuple[0]  # インデックス
                    skill_process_number = row_tuple[process_col_idx + 1]  # itertuplesはインデックスを含むため+1
                    skill_row = skill_rows.loc[row_idx]  # Seriesとして扱うために元の行を取得
                    
                    # 工程番号が空の行を優先
                    if pd.isna(skill_process_number) or skill_process_number == '':
                        empty_process_rows.append(skill_row)
                    else:
                        # 工程番号を数値として判定
                        try:
                            process_num = int(skill_process_number)
                            numeric_process_rows.append((process_num, skill_row))
                        except (ValueError, TypeError):
                            # 数値に変換できない場合は別のリストに
                            other_process_rows.append(skill_row)
                
                # 工程番号が空の行がある場合はそれを使用
                if empty_process_rows:
                    filtered_skill_rows = empty_process_rows
                    self.log_message(f"工程番号が空の行を優先採用: {len(empty_process_rows)}件", debug=True)
                elif numeric_process_rows:
                    # 数字が若い方からソート
                    numeric_process_rows.sort(key=lambda x: x[0])
                    filtered_skill_rows = [row for _, row in numeric_process_rows]
                    selected_process = numeric_process_rows[0][0]
                    self.log_message(f"工程番号が空の行が見つからず、数字が若い工程番号={selected_process}を選択: {len(filtered_skill_rows)}件", debug=True)
                else:
                    # その他の行も含める
                    filtered_skill_rows = other_process_rows
                    self.log_message(f"工程番号が空の行も数値の行も見つからず、その他の行を採用: {len(filtered_skill_rows)}件", debug=True)
            else:
                # 工程番号列のインデックスを事前に取得（高速化：itertuples()を使用）
                process_col_idx = 1  # iloc[1]に対応
                for row_tuple in skill_rows.itertuples(index=True):
                    row_idx = row_tuple[0]  # インデックス
                    skill_process_number = row_tuple[process_col_idx + 1]  # itertuplesはインデックスを含むため+1
                    skill_row = skill_rows.loc[row_idx]  # Seriesとして扱うために元の行を取得
                    
                    # スキルマスタの工程番号が空欄の場合は、工程番号が一致しなくてもOK
                    if pd.isna(skill_process_number) or skill_process_number == '':
                        filtered_skill_rows.append(skill_row)
                        self.log_message("工程番号が空欄のため、工程番号条件を無視して追加", debug=True)
                    elif str(skill_process_number) == str(process_number):
                        filtered_skill_rows.append(skill_row)
                        self.log_message(f"工程番号 '{process_number}' でマッチしました", debug=True)
                    else:
                        self.log_message(f"工程番号 '{skill_process_number}' は条件に一致しません", debug=True)
            
            if not filtered_skill_rows:
                warning_key = ("skill_process_not_found", str(product_number).strip(), str(process_number).strip())
                if warning_key not in self.logged_warnings:
                    self.logged_warnings.add(warning_key)
                    self.log_message(
                        f"工程番号 '{process_number}' に一致するスキル情報が見つかりません",
                        level="warning",
                    )
                # 出荷予定日が間近で他に割当てられない場合のみ新規品対応チームを使用
                if allow_new_team_fallback and shipping_date is not None:
                    shipping_date = pd.to_datetime(shipping_date, errors='coerce')
                    if pd.notna(shipping_date):
                        shipping_date_date = shipping_date.date()
                        current_date = pd.Timestamp.now().date()
                        two_weeks_later = current_date + timedelta(days=14)
                        if shipping_date_date <= two_weeks_later:
                            self.log_message(f"出荷予定日が間近（{shipping_date_date}）で他に割当てられないため、新規品対応チームを使用します")
                            return self.get_new_product_team_inspectors(inspector_master_df)
                self.log_message("利用可能な検査員が見つかりません")
                return []
            
            # スキル情報から検査員を取得
            # NOTE: 列名の全量出力はログ肥大化の原因になるため、debug かつ1回だけ出力する
            if not self._skill_master_schema_logged:
                self.log_message(f"スキルマスタの列数: {len(skill_master_df.columns)}", debug=True)
                # 検査員列は V002 以降をすべて対象（従来の Z040 までの固定上限を撤廃）
                self.log_message(
                    f"スキルマスタの検査員列数: {max(0, len(skill_master_df.columns) - 2)}（例: {list(skill_master_df.columns[2:12])}）",
                    debug=True
                )

                # 処理対象の検査員コードを事前に確認（ログ用）
                valid_inspector_codes = []
                for i in range(2, len(skill_master_df.columns)):
                    col_name = skill_master_df.columns[i]
                    if pd.notna(col_name) and str(col_name).strip() != '':
                        valid_inspector_codes.append(col_name)
                self.log_message(f"処理対象検査員コード数: {len(valid_inspector_codes)}", debug=True)
                self._skill_master_schema_logged = True
            
            for skill_row in filtered_skill_rows:
                # マッチした行の品番と工程番号をログに出力
                matched_product = skill_row.iloc[0] if len(skill_row) > 0 else 'N/A'
                matched_process = skill_row.iloc[1] if len(skill_row) > 1 else 'N/A'
                self.log_message(
                    f"🔍 スキルマスタ行を処理中: 品番='{matched_product}', 工程番号='{matched_process}'",
                    debug=True
                )
                
                # スキルマスタの列構造: 品番, 工程, V002, V004, V005, ...（右端は可変）
                # 列2以降の全列を検査員コード列として扱う
                skill_values_found = []  # デバッグ用: 見つかったスキル値のリスト
                skill_values_excluded = []  # デバッグ用: 除外されたスキル値のリスト
                
                for i in range(2, len(skill_master_df.columns)):
                    col_name = skill_master_df.columns[i]
                    inspector_code = col_name
                    skill_value = skill_row.iloc[i]  # ilocを使用してインデックスでアクセス
                    
                    # 列名が空またはNaNの場合はスキップ
                    if pd.isna(col_name) or str(col_name).strip() == '':
                        continue
                    
                    # デバッグ: スキル値の詳細情報をログに出力
                    skill_value_str = str(skill_value) if pd.notna(skill_value) else 'NaN'
                    skill_value_type = type(skill_value).__name__
                    skill_value_stripped = str(skill_value).strip() if pd.notna(skill_value) else ''
                    
                    # スキル値が1, 2, 3のいずれかで、かつ空でない場合
                    if pd.notna(skill_value) and str(skill_value).strip() != '' and str(skill_value).strip() in ['1', '2', '3']:
                        skill_values_found.append(f"{inspector_code}={skill_value_str}")
                        # 【変更】新規品対応チームメンバーも通常の品番に割り当て可能にする
                        # ただし、新規品対応チームメンバーは新規品を優先的に割り当てるため、優先度を下げる
                        # （完全に除外するのではなく、通常の検査員より優先度を低くする）
                        if inspector_code in new_team_codes:
                            # 新規品対応チームメンバーも含めるが、優先度を下げる（後でソート時に考慮）
                            self.log_message(
                                f"ℹ️ 検査員コード {inspector_code} は新規品対応チームメンバーですが、通常の品番にも割り当て可能です（新規品を優先）",
                                debug=True
                            )
                        
                        self.log_message(
                            f"✅ スキル値 {skill_value} の検査員コード {inspector_code} を処理中",
                            debug=True
                        )
                    else:
                        # スキル値が1,2,3でない場合のデバッグ情報
                        if pd.isna(skill_value):
                            skill_values_excluded.append(f"{inspector_code}=NaN")
                        elif str(skill_value).strip() == '':
                            skill_values_excluded.append(f"{inspector_code}=(空文字)")
                        else:
                            skill_values_excluded.append(f"{inspector_code}={skill_value_str}(型:{skill_value_type}, 条件外)")
                        # 最初の10件のみ詳細ログを出力（ログが多くなりすぎないように）
                        if len(skill_values_excluded) <= 10:
                            self.log_message(
                                f"⚠️ 検査員コード {inspector_code} のスキル値 '{skill_value_str}' "
                                f"(型: {skill_value_type}, トリム後: '{skill_value_stripped}') は条件に一致しません",
                                debug=True
                            )
                        continue  # スキル値が1,2,3でない場合は次の検査員へ
                    
                    # 検査員マスタから該当する検査員の情報を取得
                    # 検査員コード（V002, V004等）で検索
                    inspector_info = inspector_master_df[inspector_master_df['#ID'] == inspector_code]
                    if not inspector_info.empty:
                        inspector_data = inspector_info.iloc[0]
                        
                        # 勤務時間を事前チェック（0時間の検査員を除外）
                        start_time = inspector_data['開始時刻']
                        end_time = inspector_data['終了時刻']
                        
                        if pd.notna(start_time) and pd.notna(end_time):
                            try:
                                # 時刻文字列を時間に変換
                                if isinstance(start_time, str):
                                    start_hour = float(start_time.split(':')[0]) + float(start_time.split(':')[1]) / 60.0
                                else:
                                    start_hour = start_time.hour + start_time.minute / 60.0
                                    
                                if isinstance(end_time, str):
                                    end_hour = float(end_time.split(':')[0]) + float(end_time.split(':')[1]) / 60.0
                                else:
                                    end_hour = end_time.hour + end_time.minute / 60.0
                                
                                # 基本勤務時間を計算
                                max_daily_hours = end_hour - start_hour
                                
                                # 休憩時間（12:15～13:00）を含む場合は1時間を差し引く
                                if start_hour <= 12.25 and end_hour >= 13.0:
                                    max_daily_hours -= 1.0
                                
                                # 勤務時間が0以下の場合は候補から除外
                                if max_daily_hours <= 0:
                                    # 重複警告を防ぐ
                                    warning_key = (f"勤務時間0時間_一般", inspector_data['#氏名'])
                                    if warning_key not in self.logged_warnings:
                                        self.log_message(
                                            f"警告: 検査員 '{inspector_data['#氏名']}' の勤務時間が0時間以下です "
                                            f"(開始: {start_time}, 終了: {end_time}) - 候補から除外",
                                            level='warning'
                                        )
                                        self.logged_warnings.add(warning_key)
                                    continue
                                    
                            except Exception as e:
                                # 重複警告を防ぐ
                                warning_key = (f"勤務時間計算失敗_一般", inspector_data['#氏名'])
                                if warning_key not in self.logged_warnings:
                                    self.log_message(
                                        f"警告: 検査員 '{inspector_data['#氏名']}' の勤務時間計算に失敗: {e} - 候補から除外",
                                        level='warning'
                                    )
                                    self.logged_warnings.add(warning_key)
                                continue
                        else:
                            # 重複警告を防ぐ
                            warning_key = (f"時刻情報不正_一般", inspector_data['#氏名'])
                            if warning_key not in self.logged_warnings:
                                self.log_message(
                                    f"警告: 検査員 '{inspector_data['#氏名']}' の時刻情報が不正です - 候補から除外",
                                    level='warning'
                                )
                                self.logged_warnings.add(warning_key)
                            continue
                        
                        inspector_name = inspector_data['#氏名']
                        
                        # 【追加】休暇情報をチェック（終日休みの場合は除外）
                        vacation_info = self.get_vacation_info(inspector_name)
                        if vacation_info:
                            code = vacation_info.get("code", "")
                            work_status = vacation_info.get("work_status", "")
                            
                            # 終日休みの場合は除外
                            if code in ["休", "出", "当"]:
                                interpretation = vacation_info.get("interpretation", "")
                                self.log_message(
                                    f"検査員 '{inspector_name}' は終日休暇のため候補から除外 "
                                    f"(休暇コード: {code}, 解釈: {interpretation})",
                                    debug=True
                                )
                                continue
                        
                        # 新規品対応チームメンバーの場合は優先度を下げる（後でソート時に考慮）
                        is_new_team_member = inspector_code in new_team_codes
                        available_inspectors.append({
                            '氏名': inspector_name,
                            'スキル': int(str(skill_value).strip()),
                            '就業時間': inspector_data['開始時刻'],
                            'コード': inspector_code,
                            'is_new_team': is_new_team_member  # 新規品対応チームメンバーの場合はTrue
                        })
                        if is_new_team_member:
                            self.log_message(
                                f"検査員 '{inspector_name}' (コード: {inspector_code}, スキル: {skill_value}, 新規品対応チーム) を追加",
                                debug=True
                            )
                        else:
                            self.log_message(
                                f"検査員 '{inspector_name}' (コード: {inspector_code}, スキル: {skill_value}) を追加",
                                debug=True
                            )
                    else:
                        warning_key = ("inspector_code_not_found", str(inspector_code).strip())
                        if warning_key not in self.logged_warnings:
                            self.logged_warnings.add(warning_key)
                            self.log_message(
                                f"警告: 検査員コード '{inspector_code}' が検査員マスタに見つかりません",
                                level="warning",
                            )
                        # 検査員マスタの全コードを表示
                        self.log_message(
                            f"検査員マスタの利用可能なコード: {list(inspector_master_df['#ID'].values)}",
                            debug=True
                        )
                
                # デバッグ: スキル値の要約をログに出力
                if skill_values_found:
                    self.log_message(f"スキル値1,2,3が見つかった検査員: {', '.join(skill_values_found)}", debug=True)
                else:
                    self.log_message("⚠️ スキル値1,2,3が見つかった検査員: 0人", debug=True)
                
                if skill_values_excluded:
                    excluded_summary = ', '.join(skill_values_excluded[:20])  # 最初の20件のみ表示
                    if len(skill_values_excluded) > 20:
                        excluded_summary += f" ... (他{len(skill_values_excluded) - 20}件)"
                    self.log_message(f"除外されたスキル値: {excluded_summary}", debug=True)
                    self.log_message(f"除外された検査員数: {len(skill_values_excluded)}人", debug=True)
            
            # 【追加】固定検査員を優先的に配置
            fixed_inspector_names = self._collect_fixed_inspector_names(product_number, target_process_name)
            if fixed_inspector_names:
                self.log_message(f"品番 '{product_number}' の固定検査員: {fixed_inspector_names}")
                # 固定検査員とそれ以外に分離
                fixed_inspectors = []
                other_inspectors = []
                available_inspector_names = {insp['氏名'] for insp in available_inspectors}
                
                for inspector in available_inspectors:
                    inspector_name = inspector['氏名']
                    if inspector_name in fixed_inspector_names:
                        fixed_inspectors.append(inspector)
                    else:
                        other_inspectors.append(inspector)
                
                # 【特別処置】固定検査員が候補に含まれていない場合、検査員マスタから直接追加
                # これは登録済み品番リストの固定検査員の特別処置です
                missing_fixed_inspectors = [name for name in fixed_inspector_names if name not in available_inspector_names]
                if missing_fixed_inspectors:
                    # 警告の重複を防ぐ
                    warning_key = ('fixed_inspector_missing', product_number, tuple(sorted(missing_fixed_inspectors)))
                    if warning_key not in self.logged_warnings:
                        self.logged_warnings.add(warning_key)
                        self.log_message(
                            f"⚠️ 警告: 品番 '{product_number}' の固定検査員のうち、以下の検査員が候補に含まれていません: {missing_fixed_inspectors}",
                            level='warning'
                        )
                        self.log_message(
                            f"   理由: スキルマスタに該当品番のスキル情報がないか、スキル値が1,2,3以外の可能性があります"
                        )
                        self.log_message(
                            f"   特別処置: 固定検査員として設定されているため、スキルマスタに含まれていなくても候補に追加します"
                        )
                    
                    # 検査員マスタから固定検査員の情報を取得して追加
                    for missing_name in missing_fixed_inspectors:
                        # 休暇情報をチェック（終日休みの場合は除外）
                        vacation_info = self.get_vacation_info(missing_name)
                        if vacation_info:
                            work_status = vacation_info.get("work_status", "")
                            if work_status == "休み":
                                code = vacation_info.get("code", "")
                                interpretation = vacation_info.get("interpretation", "")
                                self.log_message(
                                    f"   固定検査員 '{missing_name}' は終日休暇のため除外します "
                                    f"(休暇コード: {code}, 解釈: {interpretation})",
                                    debug=True,
                                    level='warning'
                                )
                                continue

                        inspector_info = self._get_inspector_by_name(missing_name, inspector_master_df)
                        if not inspector_info.empty:
                            inspector_data = inspector_info.iloc[0]
                            inspector_code = inspector_data['#ID']
                            
                            # 【特別処置】固定検査員として選択されていれば、新規品対応チームメンバーでも振り分ける
                            # （通常のスキルマスタベースの処理では新規品対応チームメンバーは除外されるが、
                            #  固定検査員として設定されている場合は特別処置として含める）
                            is_new_team_member = inspector_code in new_team_codes
                            if is_new_team_member:
                                self.log_message(
                                    f"   固定検査員 '{missing_name}' は新規品対応チームメンバーですが、固定検査員として設定されているため特別処置として含めます"
                                )
                            
                            # 勤務時間をチェック
                            start_time = inspector_data['開始時刻']
                            end_time = inspector_data['終了時刻']
                            
                            if pd.notna(start_time) and pd.notna(end_time):
                                try:
                                    # 時刻文字列を時間に変換
                                    if isinstance(start_time, str):
                                        start_hour = float(start_time.split(':')[0]) + float(start_time.split(':')[1]) / 60.0
                                    else:
                                        start_hour = start_time.hour + start_time.minute / 60.0
                                        
                                    if isinstance(end_time, str):
                                        end_hour = float(end_time.split(':')[0]) + float(end_time.split(':')[1]) / 60.0
                                    else:
                                        end_hour = end_time.hour + end_time.minute / 60.0
                                    
                                    # 基本勤務時間を計算
                                    max_daily_hours = end_hour - start_hour
                                    
                                    # 休憩時間（12:15～13:00）を含む場合は1時間を差し引く
                                    if start_hour <= 12.25 and end_hour >= 13.0:
                                        max_daily_hours -= 1.0
                                    
                                    # 勤務時間が0以下の場合は除外
                                    if max_daily_hours <= 0:
                                        # 固定検査員は休暇でない限り0hでも割当可能（登録済品番に限る）
                                        self.log_message(
                                            f"   固定検査員 '{missing_name}' の勤務時間が0時間以下ですが、固定検査員のため候補に含めます",
                                            level='warning'
                                        )
                                        
                                except Exception as e:
                                    self.log_message(
                                        f"   固定検査員 '{missing_name}' の勤務時間計算に失敗: {e} - 固定検査員のため候補に含めます",
                                        level='warning'
                                    )
                            else:
                                self.log_message(
                                    f"   固定検査員 '{missing_name}' の時刻情報が不正ですが、固定検査員のため候補に含めます",
                                    level='warning'
                                )
                            
                            # 固定検査員を候補に追加（スキル値はデフォルトで1とする）
                            fixed_inspectors.append({
                                '氏名': missing_name,
                                'スキル': 1,  # スキルマスタにない場合はデフォルトで1
                                '就業時間': inspector_data['開始時刻'],
                                'コード': inspector_code,
                                'is_new_team': is_new_team_member,  # 新規品対応チームメンバーの場合はTrue
                                'is_fixed_inspector': True  # 固定検査員フラグ
                            })
                            team_mark = " (新規品対応チーム)" if is_new_team_member else ""
                            self.log_message(
                                f"   固定検査員 '{missing_name}' (コード: {inspector_code}){team_mark} を特別処置として候補に追加しました"
                            )
                        else:
                            self.log_message(
                                f"   固定検査員 '{missing_name}' が検査員マスタに見つかりません",
                                level='warning'
                            )
                
                # 固定検査員を優先的にリストの先頭に配置
                available_inspectors = fixed_inspectors + other_inspectors
                self.log_message(f"固定検査員を優先配置: {len(fixed_inspectors)}名を先頭に配置（設定: {len(fixed_inspector_names)}名）")
            
            self.log_message(f"利用可能な検査員: {len(available_inspectors)}人", debug=True)
            
            # 利用可能な検査員の詳細をログ出力（デバッグモードのみ）
            if available_inspectors:
                self.log_message("=== 利用可能な検査員一覧 ===", debug=True)
                for insp in available_inspectors:
                    is_fixed = insp['氏名'] in fixed_inspector_names if fixed_inspector_names else False
                    fixed_mark = " [固定]" if is_fixed else ""
                    self.log_message(
                        f"  {insp['氏名']}{fixed_mark} (コード: {insp['コード']}, スキル: {insp['スキル']})",
                        debug=True
                    )
                self.log_message("=============================", debug=True)
            else:
                self.log_message("警告: 利用可能な検査員が0人です")
                # 出荷予定日が間近で他に割当てられない場合のみ新規品対応チームを使用
                if allow_new_team_fallback and shipping_date is not None:
                    shipping_date = pd.to_datetime(shipping_date, errors='coerce')
                    if pd.notna(shipping_date):
                        shipping_date_date = shipping_date.date()
                        current_date = pd.Timestamp.now().date()
                        two_weeks_later = current_date + timedelta(days=14)
                        if shipping_date_date <= two_weeks_later:
                            self.log_message(f"出荷予定日が間近（{shipping_date_date}）で他に割当てられないため、新規品対応チームを使用します")
                            return self.get_new_product_team_inspectors(inspector_master_df)
            
            return available_inspectors
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            self.log_message(f"利用可能な検査員取得中にエラーが発生しました: {str(e)}")
            self.log_message(f"エラー詳細: {error_detail}")
            # エラーが発生した場合は新製品チームにフォールバック
            self.log_message("エラーのため新製品チームにフォールバックします")
            return self.get_new_product_team_inspectors(inspector_master_df)
    
    def load_process_master(
        self,
        process_master_path: str
    ) -> Optional[pd.DataFrame]:
        """
        工程マスタ.xlsxを読み込む（キャッシュ対応・高速化）
        
        Args:
            process_master_path: 工程マスタファイルのパス
            
        Returns:
            工程マスタのDataFrame（Noneの場合は読み込み失敗）
        """
        try:
            if not process_master_path or not Path(process_master_path).exists():
                self.log_message(f"工程マスタファイルが見つかりません: {process_master_path}")
                return None
            
            # キャッシュチェック（ファイル更新時刻も確認）
            import os
            try:
                if (self._process_master_cache is not None and 
                    self._process_master_cache_path == process_master_path):
                    current_mtime = os.path.getmtime(process_master_path)
                    if current_mtime == self._process_master_cache_mtime:
                        logger.debug("工程マスタをキャッシュから読み込みました（ファイル未変更）")
                        return self._process_master_cache
            except (OSError, AttributeError):
                pass  # キャッシュチェックでエラーが発生した場合は通常読み込みに進む
            
            # 通常読み込み
            df = pd.read_excel(process_master_path, engine='openpyxl')
            self.log_message(f"工程マスタを読み込みました: {len(df)}件")
            
            # キャッシュに保存
            try:
                self._process_master_cache = df
                self._process_master_cache_path = process_master_path
                self._process_master_cache_mtime = os.path.getmtime(process_master_path)
            except (OSError, AttributeError):
                pass  # キャッシュ保存でエラーが発生しても続行
            
            return df
        except Exception as e:
            self.log_message(f"工程マスタの読み込みに失敗しました: {str(e)}")
            return None
    
    def infer_process_number_from_process_master(
        self, 
        product_number: str,
        process_master_df: pd.DataFrame,
        inspection_target_keywords: Optional[List[str]]
    ) -> Optional[str]:
        """
        工程マスタから工程番号を推定する
        
        Args:
            product_number: 品番
            process_master_df: 工程マスタのDataFrame
            inspection_target_keywords: 検査対象CSVのA列のキーワードリスト（外観, エアー, バリ, 顕微鏡, 棒通し）
            
        Returns:
            推定された工程番号（見つからない場合はNone）
        """
        try:
            if process_master_df is None or process_master_df.empty:
                return None
            
            # A列（品番）で一致する行を検索
            product_col = process_master_df.columns[0]  # A列
            matching_rows = process_master_df[process_master_df[product_col] == product_number]
            
            if matching_rows.empty:
                self.log_message(f"工程マスタに品番 '{product_number}' が見つかりません")
                return None
            
            # 最初の一致行を取得
            row = matching_rows.iloc[0]
            
            # B列以降を順に検索（1行目のカラム名を取得するため、列名を使用）
            # 1行目がカラム名として読み込まれている場合
            for col_idx in range(1, len(process_master_df.columns)):
                col_name = process_master_df.columns[col_idx]
                cell_value = row.iloc[col_idx]
                
                # セル値が空でない場合のみチェック
                if pd.notna(cell_value):
                    cell_str = str(cell_value).strip()
                    
                    # 検査対象CSVのキーワードで部分一致検索
                    for keyword in inspection_target_keywords:
                        if keyword in cell_str:
                            # 一致したら、その列のカラム名（1行目の値）を工程番号として返す
                            inferred_process = str(col_name).strip()
                            # 同一品番で何度もログが出ると冗長になるため、初回のみINFOで出す
                            try:
                                logged = getattr(self, "_process_infer_logged", None)
                                if logged is None:
                                    logged = set()
                                    setattr(self, "_process_infer_logged", logged)
                                key = (str(product_number), str(keyword), str(inferred_process))
                                if key not in logged:
                                    logged.add(key)
                                    self.log_message(
                                        f"工程マスタから工程番号を推定: 品番='{product_number}', "
                                        f"キーワード='{keyword}', 推定工程番号='{inferred_process}'",
                                        debug=True
                                    )
                            except Exception:
                                # ログ抑制で失敗しても推定処理は継続
                                self.log_message(
                                    f"工程マスタから工程番号を推定: 品番='{product_number}', "
                                    f"キーワード='{keyword}', 推定工程番号='{inferred_process}'",
                                    debug=True
                                )
                            return inferred_process
            
            self.log_message(f"工程マスタで品番 '{product_number}' の工程番号を推定できませんでした", debug=True)
            return None
            
        except Exception as e:
            self.log_message(f"工程番号の推定中にエラーが発生しました: {str(e)}")
            return None
    
    def get_process_name_from_process_master(
        self,
        product_number: str,
        process_number: Optional[Any],
        process_master_df: pd.DataFrame,
        inspection_target_keywords: Optional[List[str]]
    ) -> Optional[str]:
        """
        工程マスタから工程名を取得する
        
        Args:
            product_number: 品番
            process_number: 工程番号
            process_master_df: 工程マスタのDataFrame
            inspection_target_keywords: 検査対象CSVのキーワードリスト
            
        Returns:
            工程名（見つからない場合はNone）
        """
        try:
            if process_master_df is None or process_master_df.empty:
                return None
            
            if not process_number or str(process_number).strip() == '':
                return None
            
            # A列（品番）で一致する行を検索
            product_col = process_master_df.columns[0]  # A列
            matching_rows = process_master_df[process_master_df[product_col] == product_number]
            
            if matching_rows.empty:
                return None
            
            # 最初の一致行を取得
            row = matching_rows.iloc[0]
            
            # B列以降を順に検索して、工程番号（列名）が一致する列のセル値を取得
            for col_idx in range(1, len(process_master_df.columns)):
                col_name = process_master_df.columns[col_idx]
                cell_value = row.iloc[col_idx]
                
                # 列名（工程番号）が一致する場合
                if str(col_name).strip() == str(process_number).strip():
                    # セル値（工程名）を返す
                    if pd.notna(cell_value):
                        process_name = str(cell_value).strip()
                        self.log_message(
                            f"工程マスタから工程名を取得: 品番='{product_number}', "
                            f"工程番号='{process_number}', 工程名='{process_name}'"
                        )
                        return process_name
            
            return None
            
        except Exception as e:
            self.log_message(f"工程名の取得中にエラーが発生しました: {str(e)}")
            return None
    
    def get_new_product_team_inspectors(
        self,
        inspector_master_df: pd.DataFrame
    ) -> List[Dict[str, Any]]:
        """
        新製品チームの検査員を取得
        
        Args:
            inspector_master_df: 検査員マスタのDataFrame
        
        Returns:
            新製品チームの検査員リスト
        """
        try:
            new_product_team_inspectors = []
            
            # デバッグ: 新製品チーム列の内容を確認（最初の1回だけ）
            if not self.new_product_team_logged:
                self.new_product_team_logged = True
            
            # 検査員マスタのH列（新製品チーム）が"★"のメンバーを取得
            new_product_team_rows = inspector_master_df[inspector_master_df.iloc[:, 7] == '★']
            
            if new_product_team_rows.empty:
                warning_key = "new_product_team_empty"
                if warning_key not in self.logged_warnings:
                    self.log_message("新製品チームのメンバーが見つかりません")
                    self.logged_warnings.add(warning_key)
                return []
            
            # 列インデックスを事前に取得（高速化：itertuples()を使用）
            start_time_col_idx = new_product_team_rows.columns.get_loc('開始時刻')
            end_time_col_idx = new_product_team_rows.columns.get_loc('終了時刻')
            name_col_idx = new_product_team_rows.columns.get_loc('#氏名')
            id_col_idx = new_product_team_rows.columns.get_loc('#ID')
            
            for row_tuple in new_product_team_rows.itertuples(index=True):
                row_idx = row_tuple[0]  # インデックス
                inspector_row = new_product_team_rows.loc[row_idx]  # Seriesとして扱うために元の行を取得
                
                # 勤務時間を事前チェック（0時間の検査員を除外）
                start_time = row_tuple[start_time_col_idx + 1]  # itertuplesはインデックスを含むため+1
                end_time = row_tuple[end_time_col_idx + 1]
                
                inspector_name = row_tuple[name_col_idx + 1]
                
                # 【追加】休暇情報をチェック（終日休みの場合は除外）
                vacation_info = self.get_vacation_info(inspector_name)
                if vacation_info:
                    code = vacation_info.get("code", "")
                    work_status = vacation_info.get("work_status", "")
                    
                    # 終日休みの場合は除外
                    if code in ["休", "出", "当"]:
                        interpretation = vacation_info.get("interpretation", "")
                        warning_key = (f"終日休暇_新製品チーム", inspector_name)
                        if warning_key not in self.logged_warnings:
                            self.log_message(
                                f"警告: 新製品チームメンバー '{inspector_name}' は終日休暇のため候補から除外 "
                                f"(休暇コード: {code}, 解釈: {interpretation})",
                                level='warning'
                            )
                            self.logged_warnings.add(warning_key)
                        continue
                
                if pd.notna(start_time) and pd.notna(end_time):
                    try:
                        # 実質勤務時間を取得（休暇情報を考慮）
                        max_daily_hours = self.get_inspector_max_hours(row_tuple[id_col_idx + 1], inspector_master_df)
                        
                        # 勤務時間が0以下の場合は候補から除外
                        if max_daily_hours <= 0:
                            # 重複警告を防ぐ
                            warning_key = (f"勤務時間0時間_新製品チーム", inspector_name)
                            if warning_key not in self.logged_warnings:
                                self.log_message(
                                    f"警告: 新製品チームメンバー '{inspector_name}' の調整後勤務時間が0時間以下です "
                                    f"(開始: {start_time}, 終了: {end_time}) - 候補から除外",
                                    level='warning'
                                )
                                self.logged_warnings.add(warning_key)
                            continue
                            
                    except Exception as e:
                        inspector_name = row_tuple[name_col_idx + 1]
                        warning_key = (f"勤務時間計算失敗_新製品チーム", inspector_name)
                        if warning_key not in self.logged_warnings:
                            self.log_message(
                                f"警告: 新製品チームメンバー '{inspector_name}' の勤務時間計算に失敗: {e} - 候補から除外",
                                level='warning'
                            )
                            self.logged_warnings.add(warning_key)
                        continue
                else:
                    inspector_name = row_tuple[name_col_idx + 1]
                    warning_key = (f"時刻情報不正_新製品チーム", inspector_name)
                    if warning_key not in self.logged_warnings:
                        self.log_message(
                            f"警告: 新製品チームメンバー '{inspector_name}' の時刻情報が不正です - 候補から除外",
                            level='warning'
                        )
                        self.logged_warnings.add(warning_key)
                    continue
                
                new_product_team_inspectors.append({
                    '氏名': row_tuple[name_col_idx + 1],
                    'スキル': 2,  # 新製品チームは中スキルとして扱う
                    '就業時間': start_time,
                    'コード': row_tuple[id_col_idx + 1],
                    'is_new_team': True  # 新規品チームフラグ
                })
            # メンバー数は最初の1回だけ出力
            if not self._new_product_team_count_logged:
                self.log_message(f"新製品チームメンバー: {len(new_product_team_inspectors)}人")
                self._new_product_team_count_logged = True
            
            return new_product_team_inspectors
            
        except Exception as e:
            self.log_message(f"新製品チームメンバー取得中にエラーが発生しました: {str(e)}")
            return []
    
    def assign_inspectors_asymmetric(
        self,
        available_inspectors: List[Dict[str, Any]],
        required_hours: float,
        inspector_master_df: pd.DataFrame,
        product_number: str,
        is_new_product: bool = False,
        max_inspectors: Optional[int] = None,
        allow_same_day_overrun: bool = False,
        reserve_capacity_for_high_priority: bool = False,
        process_name_context: Optional[str] = None,
        fixed_primary_inspector_name: Optional[str] = None,
        force_fixed_assignment: bool = False,
        ignore_product_limit: bool = False,
        relax_work_hours: bool = False,
        balance_across_max_inspectors: bool = False
    ) -> Tuple[List[Dict[str, Any]], float, float]:
        """
        改善ポイント: 非対称分配＋部分割当の実装
        
        検査員の残勤務時間に応じた非対称分配（貪欲法）を実行する。
        必要人数を満たせなかった場合でも、確保できた人数分だけ部分的に割当を行う。
        余裕のある検査員（総検査時間が少ない検査員）を優先的に割り当てる。
        
        Args:
            available_inspectors: 候補検査員リスト（各要素は辞書形式）
            required_hours: 必要な検査時間（時間単位）
            inspector_master_df: 検査員マスタ
            product_number: 品番
            is_new_product: 新規品フラグ
            max_inspectors: 最大検査員数（特例: 5名制限など）
            allow_same_day_overrun: 当日洗浄品の制約緩和時に勤務時間の許容値をさらに広げるか
            
        Returns:
            Tuple[
                List[Dict[str, Any]],  # 割り当てられた検査員リスト（各要素に'割当時間'キーが追加される）
                float,                  # 残りの未割当時間
                float                    # 割り当てられた時間の合計
            ]
        """
        try:
            target_process_name = process_name_context or ''
            available_inspectors = available_inspectors or []

            # 固定検査員は制約を多少破っても必ず入れる（主担当1名に固定）
            if force_fixed_assignment and fixed_primary_inspector_name:
                fixed_name = str(fixed_primary_inspector_name).strip()
                # ただし、終日休暇の場合は固定でも割当不可
                if fixed_name and self.is_inspector_on_vacation(fixed_name):
                    return [], required_hours, 0.0
                fixed_candidate = None
                for inspector in available_inspectors:
                    if str(inspector.get('氏名', '')).strip() == fixed_name:
                        fixed_candidate = inspector.copy()
                        break
                if fixed_candidate is None:
                    inspector_info = self._get_inspector_by_name(fixed_name, inspector_master_df)
                    if not inspector_info.empty:
                        inspector_row = inspector_info.iloc[0]
                        fixed_candidate = {
                            '氏名': fixed_name,
                            'スキル': 99,
                            '就業時間': inspector_row.get('開始時刻', ''),
                            'コード': inspector_row.get('#ID', '')
                        }
                if fixed_candidate is None:
                    # 固定検査員が見つからない場合は通常処理へフォールバック
                    pass
                else:
                    fixed_candidate['割当時間'] = float(required_hours)
                    return [fixed_candidate], 0.0, float(required_hours)

            if not available_inspectors:
                return [], required_hours, 0.0
            
            current_date = pd.Timestamp.now().date()
            remaining = required_hours
            assignments = []
            
            # 各検査員の利用可能容量を計算
            # 【追加】固定検査員情報を取得（登録済み品番の特別処置）
            fixed_inspector_names = self._collect_fixed_inspector_names(product_number, target_process_name)
            effective_fixed_names: Set[str]
            if fixed_primary_inspector_name:
                effective_fixed_names = {str(fixed_primary_inspector_name).strip()}
            else:
                effective_fixed_names = set(fixed_inspector_names)
            
            candidates_with_capacity = []
            for inspector in available_inspectors:
                code = inspector['コード']
                inspector_name = inspector.get('氏名', '')

                # 【追加】固定検査員フラグを設定
                is_fixed_inspector = inspector_name in effective_fixed_names
                
                # 残り勤務時間を計算
                daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                max_hours = self.get_inspector_max_hours(code, inspector_master_df)
                allowed_max_hours = self._apply_work_hours_overrun(max_hours)
                if allow_same_day_overrun:
                    allowed_max_hours = self._apply_same_day_work_hours_overrun(allowed_max_hours)
                remaining_capacity = max(0.0, allowed_max_hours - daily_hours - WORK_HOURS_BUFFER)
                if reserve_capacity_for_high_priority:
                    remaining_capacity = max(0.0, remaining_capacity - HIGH_PRIORITY_RESERVED_CAPACITY_HOURS)
                # relax_work_hoursがTrueでも、10%超過を超えないように制限
                if relax_work_hours:
                    # 10%超過を超えない範囲で緩和
                    max_relaxed_capacity = allowed_max_hours - daily_hours - WORK_HOURS_BUFFER
                    remaining_capacity = max(remaining_capacity, min(float(required_hours), max_relaxed_capacity))
                
                # 品番4時間上限を考慮
                product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                product_room_to_4h = float(required_hours) if ignore_product_limit else max(0.0, PRODUCT_LIMIT_DRAFT_THRESHOLD - product_hours)
                
                # 利用可能容量は両方の制約の小さい方
                cap = min(remaining_capacity, product_room_to_4h)
                
                if cap > 0:
                    inspector_copy = inspector.copy()
                    inspector_copy['_remaining_capacity'] = cap
                    inspector_copy['_product_room'] = product_room_to_4h
                    inspector_copy['_is_fixed_inspector'] = is_fixed_inspector  # 固定検査員フラグ
                    candidates_with_capacity.append(inspector_copy)
            
            if not candidates_with_capacity:
                return [], required_hours, 0.0
            
            # 改善ポイント: 公平性スコアの適用順序変更（同点ブレーカー化）
            # 候補が複数存在するときにのみ公平性スコアを使用
            # 優先順序: (a) スキル適合, (b) 勤務時間内, (c) 品番4h上限 → これらを満たす候補が2名以上いる場合に公平性スコアを適用
            
            # 公平性スコアを計算（品番切替ペナルティ含む）
            # 【改善】分散を考慮した公平性スコアの計算
            # 全検査員の平均時間を計算（分散を考慮するため）
            all_hours = [self.inspector_work_hours.get(code, 0.0) for code in self.inspector_work_hours.keys()]
            avg_hours_all = np.mean(all_hours) if all_hours else 0.0
            
            for candidate in candidates_with_capacity:
                code = candidate['コード']
                assignment_count = self.inspector_assignment_count.get(code, 0)
                total_hours = self.inspector_work_hours.get(code, 0.0)
                product_variety_count = len(self.inspector_product_variety.get(code, set()))
                
                # 改善ポイント: 品番切替ペナルティ導入
                # score = 稼働率 + (割当ロット数 * α) + (担当品番種類数 * β)
                # ただし、ここでは逆に（スコアが小さい方が優先）として使用
                base_score = total_hours + (assignment_count * PENALTY_LOT_COUNT_ALPHA * 60) + (product_variety_count * PENALTY_PRODUCT_VARIETY_BETA * 60)
                
                # 【改善】分散を考慮した公平性スコア（平均からの偏差を考慮）
                # 平均より多い場合はペナルティを大きく、少ない場合はボーナスを小さく
                if avg_hours_all > 0:
                    deviation_penalty = abs(total_hours - avg_hours_all) * 2.0  # 偏差に対するペナルティ（2倍）
                    base_score += deviation_penalty
                
                candidate['_fairness_score'] = base_score
            
            # 容量の大きい順にソート（貪欲法）
            # 改善ポイント: 同一品番の累計時間が少ない人を優先（4時間上限の分散化）
            # ソート安定性確保（検査員IDを追加）
            for candidate in candidates_with_capacity:
                code = candidate['コード']
                product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                candidate['_product_hours'] = product_hours  # 同一品番の累計時間を記録
            
            # 【変更】固定検査員を最優先にソート
            # 登録済み品番リストの固定検査員が設定されている品番は、出荷予定日よりも優先して割り当てる
            # 【追加】余裕のある検査員（総検査時間が少ない検査員）を優先的に割り当てる
            # 【変更】新規品対応チームメンバーは通常の検査員より優先度を下げる（新規品を優先的に割り当てるため）
            for candidate in candidates_with_capacity:
                code = candidate['コード']
                daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                candidate['_daily_hours'] = daily_hours  # 総検査時間を記録（余裕度の指標）
            
            candidates_with_capacity.sort(key=lambda x: (
                not x.get('_is_fixed_inspector', False),  # False=固定検査員を最優先（False < TrueなのでFalseが先）
                x.get('is_new_team', False),  # 新規品対応チームメンバーは後回し（False < TrueなのでFalseが先）
                x['_daily_hours'],  # 総検査時間が少ない順（余裕のある検査員を優先）
                -x['_remaining_capacity'],  # 容量の大きい順
                x['_product_hours'],  # 同一品番の累計時間が少ない順（4時間上限の分散化）
                x['_fairness_score'],  # 公平性スコア（小さい順）
                x['コード']  # 検査員ID（安定性確保）
            ))
            
            # 貪欲に割り当て
            # 特例: 一ロットで検査員が5名以上必要になる場合、5名に制限
            for candidate in candidates_with_capacity:
                if remaining <= 0:
                    break
                
                # 【追加】max_inspectors制約をチェック
                if max_inspectors is not None and len(assignments) >= max_inspectors:
                    break
                
                cap = candidate['_remaining_capacity']
                if cap <= 0:
                    continue
                
                # 割り当て可能な時間を決定
                # 1人の検査員に4時間を超えないようにする制約を追加
                code = candidate['コード']
                product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                max_assignable_per_inspector = float(remaining) if ignore_product_limit else max(0.0, self.product_limit_hard_threshold - product_hours)
                
                # 容量、残り時間、1人あたりの最大割り当て時間の最小値を取る
                take = min(cap, remaining, max_assignable_per_inspector)
                
                # 【追加】割り当て時に10%超過を厳格にチェック
                # 実際の割り当て時点での勤務時間を再計算
                current_daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                candidate_max_hours = self.get_inspector_max_hours(code, inspector_master_df)
                candidate_allowed_max_hours = candidate_max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                if allow_same_day_overrun:
                    candidate_allowed_max_hours = candidate_allowed_max_hours * (1.0 + SAME_DAY_WORK_HOURS_OVERRUN_RATE - WORK_HOURS_OVERRUN_RATE)
                
                # 10%超過を超える場合は、割り当て可能な時間を調整
                if current_daily_hours + take > candidate_allowed_max_hours - WORK_HOURS_BUFFER:
                    take = max(0.0, candidate_allowed_max_hours - WORK_HOURS_BUFFER - current_daily_hours)
                    if take <= 0:
                        continue  # 割り当て不可
                
                # balance_across_max_inspectorsの処理後も、10%超過をチェック
                if balance_across_max_inspectors and max_inspectors is not None and max_inspectors > 1:
                    remaining_slots = max(1, max_inspectors - len(assignments))
                    target_take = remaining / remaining_slots
                    take = min(take, target_take)
                    # 10%超過を超える場合は、再度調整
                    if current_daily_hours + take > candidate_allowed_max_hours - WORK_HOURS_BUFFER:
                        take = max(0.0, candidate_allowed_max_hours - WORK_HOURS_BUFFER - current_daily_hours)
                        if take <= 0:
                            continue  # 割り当て不可
                
                # 割り当てを記録
                assignment = candidate.copy()
                assignment['割当時間'] = take
                assignments.append(assignment)
                
                # 【追加】固定検査員が選択された場合のログ
                if candidate.get('_is_fixed_inspector', False):
                    inspector_name = candidate.get('氏名', '')
                    self.log_message(f"固定検査員 '{inspector_name}' を優先的に割り当てました（登録済み品番の特別処置）")
                
                # 残り時間を更新
                remaining -= take
            
            assigned_time_sum = sum(ass['割当時間'] for ass in assignments)
            
            return assignments, remaining, assigned_time_sum
            
        except Exception as e:
            self.log_message(f"非対称分配中にエラーが発生しました: {str(e)}")
            return [], required_hours, 0.0
    
    def select_inspectors(
        self,
        available_inspectors: List[Dict[str, Any]],
        required_count: int,
        divided_time: float,
        inspector_master_df: pd.DataFrame,
        product_number: str,
        is_new_product: bool = False,
        relax_work_hours: bool = False,
        process_name_context: Optional[str] = None,
        ignore_product_limit: bool = False
    ) -> List[Dict[str, Any]]:
        """
        検査員を選択する（スキル組み合わせ考慮・勤務時間考慮・公平な割り当て方式）
        
        Args:
            available_inspectors: 利用可能な検査員リスト
            required_count: 必要な検査員数
            divided_time: 分割検査時間
            inspector_master_df: 検査員マスタ
            product_number: 品番
            is_new_product: 新規品フラグ
            relax_work_hours: 勤務時間チェックを緩和するか
        
        Returns:
            選択された検査員リスト
        """
        try:
            # 特例: 一ロットで検査員が5名以上必要になる場合、5名に制限
            if required_count > 5:
                required_count = 5
            
            if not available_inspectors:
                if is_new_product:
                    self.log_message(f"新規品 {product_number}: 新製品チームメンバーが利用可能な検査員がいません")
                return []
            
            # 各検査員の割り当て回数と最終割り当て時刻を更新
            current_time = pd.Timestamp.now()
            current_date = current_time.date()
            target_process_name = process_name_context or ''
            
            # 利用可能な検査員に割り当て履歴を追加
            for inspector in available_inspectors:
                inspector_code = inspector['コード']
                if inspector_code not in self.inspector_assignment_count:
                    self.inspector_assignment_count[inspector_code] = 0
                if inspector_code not in self.inspector_last_assignment:
                    self.inspector_last_assignment[inspector_code] = pd.Timestamp.min
                if inspector_code not in self.inspector_work_hours:
                    self.inspector_work_hours[inspector_code] = 0.0
                if inspector_code not in self.inspector_daily_assignments:
                    self.inspector_daily_assignments[inspector_code] = {}
                if current_date not in self.inspector_daily_assignments[inspector_code]:
                    self.inspector_daily_assignments[inspector_code][current_date] = 0.0
            
            # 勤務時間を考慮して利用可能な検査員をフィルタリング
            input_count = len(available_inspectors)
            if is_new_product:
                self.log_message(f"新規品 {product_number}: 新製品チームメンバー {input_count}人をフィルタリング中")
            available_inspectors = self.filter_available_inspectors(
                available_inspectors,
                divided_time,
                inspector_master_df,
                product_number,
                relax_work_hours=relax_work_hours,
                process_name_context=target_process_name,
                ignore_product_limit=ignore_product_limit  # 4時間上限チェックをスキップするか
            )
            filtered_count = len(available_inspectors)
            if input_count > 0 and filtered_count == 0:
                # 緩和候補が全て除外された場合、詳細ログを出力
                mode_str = "緩和モード（20%超過許容）" if relax_work_hours else "通常モード（10%超過許容）"
                self.log_message(f"警告: 未割当ロット再処理: 品番 {product_number} の候補 {input_count}名がfilter_available_inspectorsで全て除外されました（{mode_str}, ignore_product_limit={ignore_product_limit}）", level='warning')
            elif input_count > 0 and filtered_count < input_count and relax_work_hours:
                # 緩和候補作成後のfilter_available_inspectors呼び出し時に、一部が除外された場合のログ
                excluded_count = input_count - filtered_count
                self.log_message(f"未割当ロット再処理: 品番 {product_number} の緩和候補 {input_count}名のうち {filtered_count}名がfilter_available_inspectorsで利用可能（{excluded_count}名が緩和モードでも除外、relax_work_hours=True）")
            elif input_count > 0 and filtered_count == input_count and relax_work_hours:
                # 緩和候補作成後のfilter_available_inspectors呼び出し時に、全てが通過した場合のログ
                self.log_message(f"未割当ロット再処理: 品番 {product_number} の緩和候補 {input_count}名が全てfilter_available_inspectorsで利用可能（緩和モード、relax_work_hours=True）", debug=True)
            if is_new_product:
                self.log_message(f"新規品 {product_number}: 勤務時間チェック後 {filtered_count}人が利用可能（入力: {input_count}人）")

            # 改善ポイント: 4時間上限ルールの2段階化
            # ドラフトフェーズ：4.5h未満までは許容（4.0h超は over_product_limit=True を設定）
            # 最適化フェーズ：ここで4.0h遵守へ是正。置換不可能な場合のみ未割当へ戻す。
            filtered_by_product = []
            excluded_by_product_limit = []  # 4時間上限で除外された検査員の詳細情報
            for insp in available_inspectors:
                code = insp['コード']
                current = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                projected_hours = current + divided_time
                
                if not ignore_product_limit:
                    # ドラフトフェーズでの許容上限チェック（4.5h未満まで許容）
                    if projected_hours >= PRODUCT_LIMIT_DRAFT_THRESHOLD:
                        excluded_by_product_limit.append({
                            'name': insp['氏名'],
                            'code': code,
                            'current': current,
                            'divided_time': divided_time,
                            'projected': projected_hours,
                            'threshold': PRODUCT_LIMIT_DRAFT_THRESHOLD
                        })
                        self.log_message(f"検査員 '{insp['氏名']}' ({code}) は品番 {product_number} の累計が {current:.1f}h のため除外 (+{divided_time:.1f}hで{projected_hours:.1f}h、閾値{PRODUCT_LIMIT_DRAFT_THRESHOLD}h超過)", debug=True)
                        continue
                
                # 設定時間超過の場合はフラグを設定（ドラフトフェーズでは許容、最適化フェーズで是正）
                insp['over_product_limit'] = (False if ignore_product_limit else (projected_hours > self.product_limit_hard_threshold))
                insp['__projected_product_hours'] = projected_hours
                insp['__current_product_hours'] = current
                if not ignore_product_limit and projected_hours >= PRODUCT_LIMIT_FINAL_TOLERANCE:
                    insp['__near_product_limit'] = True
                else:
                    insp.pop('__near_product_limit', None)
                
                # 同一品番を同日複数回割り当てないよう制限
                product_assignment_count = (
                    self.inspector_product_assignment_counts
                    .get(code, {})
                    .get(product_number, 0)
                )
                
                # 【改善】緩和条件の明確化
                # 緩和条件が適用される場合：
                # 1. relax_work_hours=True が指定された場合（勤務時間制約の緩和時）
                # 2. その他の緩和条件（将来の拡張用）
                should_relax = relax_work_hours
                max_assignments_for_product = (
                    MAX_ASSIGNMENTS_PER_PRODUCT_RELAXED if should_relax else MAX_ASSIGNMENTS_PER_PRODUCT
                )
                
                # 【追加】効果測定メトリクス
                self.product_assignment_metrics['total_checks'] += 1
                
                if product_assignment_count >= max_assignments_for_product:
                    # 【追加】制約違反を記録
                    if not should_relax:
                        self.product_assignment_metrics['constraint_violations'] += 1
                    else:
                        # 緩和条件が適用されている場合でも、最大割当回数に達している場合は記録
                        if code not in self.product_assignment_metrics['max_assignments_reached']:
                            self.product_assignment_metrics['max_assignments_reached'][code] = {}
                        if product_number not in self.product_assignment_metrics['max_assignments_reached'][code]:
                            self.product_assignment_metrics['max_assignments_reached'][code][product_number] = 0
                        self.product_assignment_metrics['max_assignments_reached'][code][product_number] += 1
                    
                    self.log_message(
                        f"検査員 '{insp['氏名']}' は品番 {product_number} を既に {product_assignment_count} 回担当しているため候補外 "
                        f"(最大割当回数: {max_assignments_for_product}, 緩和条件: {'適用' if should_relax else '未適用'})"
                    )
                    continue
                
                # 【追加】緩和条件が適用された場合の記録
                if should_relax and product_assignment_count >= MAX_ASSIGNMENTS_PER_PRODUCT:
                    self.product_assignment_metrics['relaxed_assignments'] += 1
                    reason = f"検査員 '{insp['氏名']}' ({code}), 品番 '{product_number}': 緩和条件適用 (現在{product_assignment_count}回, 通常上限{MAX_ASSIGNMENTS_PER_PRODUCT}回, 緩和上限{MAX_ASSIGNMENTS_PER_PRODUCT_RELAXED}回)"
                    self.product_assignment_metrics['relaxation_reasons'].append(reason)
                    if len(self.product_assignment_metrics['relaxation_reasons']) <= 10:  # 最初の10件のみ保持
                        self.log_message(reason, debug=True)
                
                insp['__product_assignment_count'] = product_assignment_count
                # 3.5h以上4.0h以下の場合は警告フラグを付ける（未割当ロット削減のため柔軟に対応）
                
                filtered_by_product.append(insp)
            
            # 総勤務時間制約は削除（検査員マスタの勤務時間制約のみを使用）
            # 検査員マスタの勤務時間が個別に設定されているため、統一的な総勤務時間制約は適用しない

            if not filtered_by_product:
                if ignore_product_limit:
                    # ignore_product_limit=Trueの場合、4時間上限チェックを完全にスキップして全員を候補に追加
                    filtered_by_product = list(available_inspectors)
                    self.log_message(f"4時間上限チェックをスキップ: 品番 {product_number} の候補 {len(filtered_by_product)}人（ignore_product_limit=True）", debug=True)
                else:
                    # 詳細な除外情報をログに記録
                    if excluded_by_product_limit:
                        excluded_details = []
                        for excl in excluded_by_product_limit[:5]:  # 最初の5件のみ表示
                            excluded_details.append(
                                f"{excl['name']}({excl['code']}): {excl['current']:.1f}h+{excl['divided_time']:.1f}h={excl['projected']:.1f}h"
                            )
                        if len(excluded_by_product_limit) > 5:
                            excluded_details.append(f"...他{len(excluded_by_product_limit)-5}名")
                        self.log_message(
                            f"警告: 品番 {product_number} の4時間上限により全員が除外（{len(excluded_by_product_limit)}名）。"
                            f"除外された検査員: {', '.join(excluded_details)}"
                        )
                    if is_new_product:
                        self.log_message(f"警告: 新規品 {product_number} の4時間上限または勤務時間上限により全員が除外。ルール違反を避けるため、このロットは未割当とします")
                    else:
                        self.log_message(f"警告: 品番 {product_number} の4時間上限により全員が除外。ルール違反を避けるため、このロットは未割当とします")
                    return []
            
            if is_new_product:
                self.log_message(f"新規品 {product_number}: 4時間上限チェック後 {len(filtered_by_product)}人が利用可能")
            
            # スキル組み合わせロジックを適用
            selected_inspectors = self.select_inspectors_with_skill_combination(
                filtered_by_product, required_count, divided_time, current_time, current_date, inspector_master_df, product_number,
                process_name_context=target_process_name,
                relax_work_hours=relax_work_hours,  # 緩和モードを引き継ぐ
                ignore_product_limit=ignore_product_limit  # 4時間上限スキップを引き継ぐ
            )
            
            return selected_inspectors
            
        except Exception as e:
            self.log_message(f"検査員選択中にエラーが発生しました: {str(e)}")
            return []
    
    def select_inspectors_with_skill_combination(
        self,
        available_inspectors: List[Dict[str, Any]],
        required_count: int,
        divided_time: float,
        current_time: pd.Timestamp,
        current_date: date,
        inspector_master_df: pd.DataFrame,
        product_number: Optional[str] = None,
        process_name_context: Optional[str] = None,
        relax_work_hours: bool = False,
        ignore_product_limit: bool = False
    ) -> List[Dict[str, Any]]:
        """
        スキル組み合わせを考慮した検査員選択
        
        Args:
            available_inspectors: 利用可能な検査員リスト
            required_count: 必要な検査員数
            divided_time: 分割検査時間
            current_time: 現在時刻
            current_date: 現在日付
            inspector_master_df: 検査員マスタ
            product_number: 品番（オプション）
        
        Returns:
            選択された検査員リスト
        """
        try:
            target_process_name = process_name_context or ''
            # 特例: 一ロットで検査員が5名以上必要になる場合、5名に制限
            if required_count > 5:
                required_count = 5
            
            # 【追加】固定検査員情報を取得（登録済み品番の特別処置）
            fixed_inspector_names = []
            if product_number:
                fixed_inspector_names = self._collect_fixed_inspector_names(product_number, target_process_name)
            
            # スキルレベル別に検査員を分類
            skill_groups = {
                1: [],
                2: [],
                3: [],
                'new': []  # 新製品チーム
            }
            skill_order_map = {1: 0, 2: 1, 3: 2, 'new': 3}
            
            for inspector in available_inspectors:
                # 固定検査員フラグを設定
                inspector_name = inspector.get('氏名', '')
                inspector['__is_fixed'] = inspector_name in fixed_inspector_names if fixed_inspector_names else False
                
                if inspector.get('is_new_team', False):
                    skill_groups['new'].append(inspector)
                else:
                    skill = inspector.get('スキル', 1)
                    if skill in skill_groups:
                        skill_groups[skill].append(inspector)
                    else:
                        skill_groups[1].append(inspector)  # デフォルトはスキル1
            
            # スキル別の平均勤務時間を計算（公平な分散のため）
            skill_avg_hours = {}
            for skill_level, inspectors in skill_groups.items():
                if not inspectors:
                    continue
                skill_hours = [self.inspector_work_hours.get(insp['コード'], 0.0) for insp in inspectors]
                if skill_hours:
                    skill_avg_hours[skill_level] = sum(skill_hours) / len(skill_hours)
                else:
                    skill_avg_hours[skill_level] = 0.0
            
            # 各グループ内で公平性指標を保持（実務条件を満たした候補の同点ブレーカーとして使用）
            for skill_level, inspectors in skill_groups.items():
                if not inspectors:
                    continue
                for order_index, insp in enumerate(inspectors):
                    code = insp['コード']
                    assignment_count = self.inspector_assignment_count.get(code, 0)
                    total_hours = self.inspector_work_hours.get(code, 0.0)
                    last_assignment = self.inspector_last_assignment.get(code, pd.Timestamp.min)
                    # 未使用検査員を優先的に考慮（割り当て回数が0の場合）
                    is_unused = (assignment_count == 0)
                    # スキル別平均からの偏差を計算（小さい方が良い）
                    avg_hours = skill_avg_hours.get(skill_level, 0.0)
                    deviation_from_avg = abs(total_hours - avg_hours)
                    # product_hoursが辞書の場合は0.0にフォールバック
                    projected_hours = insp.get('__projected_product_hours')
                    if projected_hours is None:
                        product_hours_dict = self.inspector_product_hours.get(code, {})
                        if isinstance(product_hours_dict, dict):
                            # 辞書の場合は、該当品番の時間を取得（品番が不明な場合は0.0）
                            if product_number and product_number in product_hours_dict:
                                product_hours = product_hours_dict[product_number]
                            else:
                                product_hours = 0.0
                        else:
                            product_hours = product_hours_dict if isinstance(product_hours_dict, (int, float)) else 0.0
                    else:
                        product_hours = projected_hours if isinstance(projected_hours, (int, float)) else 0.0
                    product_limit_penalty = 1 if insp.get('over_product_limit', False) else 0
                    near_limit_penalty = 1 if insp.get('__near_product_limit', False) else 0
                    product_assignment_count = insp.get('__product_assignment_count', 0)
                    insp['__fairness_priority'] = (
                        product_limit_penalty,
                        near_limit_penalty,
                        product_assignment_count,
                        product_hours,
                        total_hours,
                        assignment_count,
                        last_assignment,
                        deviation_from_avg,
                        is_unused,
                    )
                    insp['__candidate_order'] = order_index
            
            # 利用可能な検査員の総数を確認
            total_available = sum(len(inspectors) for inspectors in skill_groups.values())
            if total_available < required_count:
                self.log_message(f"警告: 利用可能な検査員数 {total_available}人 が要求人数 {required_count}人 より少ないため、新製品チームを追加で検索します")
                
                # 新製品チームを追加で検索
                new_team_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                if new_team_inspectors:
                    # 新製品チームをスキルグループに追加
                    skill_groups['new'] = new_team_inspectors
                    total_available = sum(len(inspectors) for inspectors in skill_groups.values())
                    self.log_message(f"新製品チーム追加後: 利用可能な検査員数 {total_available}人")
                
                # それでも足りない場合は可能な限り選択
                if total_available < required_count:
                    self.log_message(f"最終警告: 利用可能な検査員数 {total_available}人 が要求人数 {required_count}人 より少ないため、可能な限り選択します")
                    required_count = total_available
            
            # スキル組み合わせロジックを適用
            selected_inspectors = []
            
            # 【改善】複数の検査員を選択する際、各検査員を選択する前に10%超過をチェックするための一時的な状態を追跡
            temp_daily_assignments = {}
            for code in self.inspector_daily_assignments:
                temp_daily_assignments[code] = {}
                for date_key, hours in self.inspector_daily_assignments[code].items():
                    temp_daily_assignments[code][date_key] = hours
            
            if required_count == 1:
                # 1人の場合は公平性を最優先に選択（バランス重視版）
                # 【追加】優先順位: 0)固定検査員を最優先（登録済み品番の特別処置）, 1)未使用検査員優先, 2)総勤務時間が少ない, 3)スキルレベル, 4)割り当て回数が少ない, 5)4時間上限に近い場合は優先度を下げる
                all_inspectors_with_priority = []
                for skill_level, inspectors in skill_groups.items():
                    for insp in inspectors:
                        # 辞書形式でない場合はスキップ
                        if not isinstance(insp, dict):
                            self.log_message(f"警告: 検査員データが辞書形式ではありません: {type(insp)}", debug=True)
                            continue
                        try:
                            code = insp.get('コード', '')
                            if not code:
                                continue
                            assignment_count = self.inspector_assignment_count.get(code, 0)
                            total_hours = self.inspector_work_hours.get(code, 0.0)
                            last_assignment = self.inspector_last_assignment.get(code, pd.Timestamp.min)
                            # pd.Timestampを比較可能な形式に変換（タイムスタンプ値を使用）
                            last_assignment_key = last_assignment.value if isinstance(last_assignment, pd.Timestamp) else (last_assignment if last_assignment != pd.Timestamp.min else 0)
                            is_unused = (assignment_count == 0)
                            near_limit = insp.get('__near_product_limit', False)  # 4時間上限に近い場合は優先度を下げる
                            is_fixed = insp.get('__is_fixed', False)  # 固定検査員フラグ
                            
                            # 固定検査員を最優先し、その他の公平性指標を考慮
                            # product_hoursが辞書の場合は0.0にフォールバック
                            projected_hours = insp.get('__projected_product_hours')
                            if projected_hours is None:
                                product_hours_dict = self.inspector_product_hours.get(code, {})
                                if isinstance(product_hours_dict, dict):
                                    # 辞書の場合は、該当品番の時間を取得（品番が不明な場合は0.0）
                                    if product_number and product_number in product_hours_dict:
                                        product_hours = product_hours_dict[product_number]
                                    else:
                                        product_hours = 0.0
                                else:
                                    product_hours = product_hours_dict if isinstance(product_hours_dict, (int, float)) else 0.0
                            else:
                                product_hours = projected_hours if isinstance(projected_hours, (int, float)) else 0.0
                            product_limit_penalty = 1 if insp.get('over_product_limit', False) else 0
                            product_assignment_count = insp.get('__product_assignment_count', 0)
                            priority = (
                                not is_fixed,  # False=固定検査員を優先
                                product_limit_penalty,  # 4時間上限を超える場合は最終手段
                                near_limit,  # 4時間上限に近い場合は優先度を下げる
                                product_assignment_count,  # 同一品番での割当回数
                                product_hours,  # 品番単位の累計時間が少ない順
                                not is_unused,  # False=未使用の検査員を優先
                                total_hours,   # 一日の総作業時間が少ない順
                                skill_order_map.get(skill_level, 99),  # スキルレベルの優先度
                                assignment_count,  # 割当回数が少ない順
                                last_assignment_key  # 直近の割当が古い順（比較可能な形式に変換）
                            )
                            all_inspectors_with_priority.append((priority, insp))
                        except (KeyError, TypeError, AttributeError) as e:
                            self.log_message(f"警告: 検査員データの処理中にエラーが発生しました: {e}, 検査員: {insp}", debug=True)
                            continue
                
                all_inspectors_with_priority.sort(key=self._priority_sort_key)
                if all_inspectors_with_priority:
                    # 【改善】10%超過をチェックしてから選択
                    selected_inspector = None
                    for priority, insp in all_inspectors_with_priority:
                        code = insp['コード']
                        daily_hours_temp = temp_daily_assignments.get(code, {}).get(current_date, 0.0)
                        max_hours = self.get_inspector_max_hours(code, inspector_master_df)
                        if relax_work_hours:
                            # 緩和モード: 20%超過まで許容
                            allowed_max_hours = max_hours * (1.0 + SAME_DAY_WORK_HOURS_OVERRUN_RATE)
                        else:
                            allowed_max_hours = self._apply_work_hours_overrun(max_hours)
                        
                        # 緩和モードの場合は20%超過まで許容、通常モードは10%超過まで
                        if daily_hours_temp + divided_time > allowed_max_hours - WORK_HOURS_BUFFER:
                            continue
                        
                        selected_inspector = insp
                        break  # 最初に見つかった適切な検査員を選択
                    
                    if selected_inspector:
                        if selected_inspector.get('__is_fixed', False):
                            self.log_message(f"  固定検査員 '{selected_inspector['氏名']}' を優先的に選択しました（登録済み品番の特別処置）")
                        selected_inspectors.append(selected_inspector)
                        # 一時的な状態を更新
                        code = selected_inspector['コード']
                        if code not in temp_daily_assignments:
                            temp_daily_assignments[code] = {}
                        if current_date not in temp_daily_assignments[code]:
                            temp_daily_assignments[code][current_date] = 0.0
                        temp_daily_assignments[code][current_date] += divided_time

            elif required_count == 2:
                # 2人の場合の組み合わせロジック（10%超過をチェックしながら選択）
                selected_inspectors = self.select_two_inspectors_with_skill_combination(
                    skill_groups, product_number, divided_time, current_date, inspector_master_df, temp_daily_assignments,
                    relax_work_hours=relax_work_hours  # 緩和モードを引き継ぐ
                )
            
            elif required_count == 3:
                # 3人の場合の組み合わせロジック（10%超過をチェックしながら選択）
                selected_inspectors = self.select_three_inspectors_with_skill_combination(
                    skill_groups, product_number, divided_time, current_date, inspector_master_df, temp_daily_assignments,
                    relax_work_hours=relax_work_hours  # 緩和モードを引き継ぐ
                )
            
            else:
                # 4人以上の場合は公平な割り当て（バランス重視版）
                # 【追加】固定検査員を優先的に選択（登録済み品番の特別処置）
                all_inspectors_with_priority = []
                for skill_level, inspectors in skill_groups.items():
                    for insp in inspectors:
                        # 辞書形式でない場合はスキップ
                        if not isinstance(insp, dict):
                            self.log_message(f"警告: 検査員データが辞書形式ではありません: {type(insp)}", debug=True)
                            continue
                        try:
                            code = insp.get('コード', '')
                            if not code:
                                continue
                            assignment_count = self.inspector_assignment_count.get(code, 0)
                            total_hours = self.inspector_work_hours.get(code, 0.0)
                            last_assignment = self.inspector_last_assignment.get(code, pd.Timestamp.min)
                            # pd.Timestampを比較可能な形式に変換（タイムスタンプ値を使用）
                            last_assignment_key = last_assignment.value if isinstance(last_assignment, pd.Timestamp) else (last_assignment if last_assignment != pd.Timestamp.min else 0)
                            is_unused = (assignment_count == 0)
                            near_limit = insp.get('__near_product_limit', False)  # 4時間上限に近い場合は優先度を下げる
                            is_fixed = insp.get('__is_fixed', False)  # 固定検査員フラグ
                            
                            # 固定検査員を最優先し、その他の公平性指標を考慮
                            # product_hoursが辞書の場合は0.0にフォールバック
                            projected_hours = insp.get('__projected_product_hours')
                            if projected_hours is None:
                                product_hours_dict = self.inspector_product_hours.get(code, {})
                                if isinstance(product_hours_dict, dict):
                                    # 辞書の場合は、該当品番の時間を取得（品番が不明な場合は0.0）
                                    if product_number and product_number in product_hours_dict:
                                        product_hours = product_hours_dict[product_number]
                                    else:
                                        product_hours = 0.0
                                else:
                                    product_hours = product_hours_dict if isinstance(product_hours_dict, (int, float)) else 0.0
                            else:
                                product_hours = projected_hours if isinstance(projected_hours, (int, float)) else 0.0
                            product_limit_penalty = 1 if insp.get('over_product_limit', False) else 0
                            product_assignment_count = insp.get('__product_assignment_count', 0)
                            priority = (
                                not is_fixed,  # False=固定検査員を優先
                                product_limit_penalty,  # 4時間上限を超える場合は最終手段
                                near_limit,  # 4時間上限に近い場合は優先度を下げる
                                product_assignment_count,  # 同一品番の割当回数
                                product_hours,  # 品番単位の累計時間が少ない順
                                not is_unused,  # False=未使用の検査員を優先
                                total_hours,   # 一日の総作業時間が少ない順
                                skill_order_map.get(skill_level, 99),  # スキルレベルの優先度
                                assignment_count,  # 割当回数が少ない順
                                last_assignment_key  # 直近の割当が古い順（比較可能な形式に変換）
                            )
                            all_inspectors_with_priority.append((priority, insp))
                        except (KeyError, TypeError, AttributeError) as e:
                            self.log_message(f"警告: 検査員データの処理中にエラーが発生しました: {e}, 検査員: {insp}", debug=True)
                            continue
                
                all_inspectors_with_priority.sort(key=self._priority_sort_key)
                # 特例: 一ロットで検査員が5名以上必要になる場合、5名に制限
                max_count = min(5, required_count)
                # 【改善】10%超過をチェックしながら選択
                for priority, insp in all_inspectors_with_priority:
                    if len(selected_inspectors) >= max_count:
                        break
                    code = insp['コード']
                    daily_hours_temp = temp_daily_assignments.get(code, {}).get(current_date, 0.0)
                    max_hours = self.get_inspector_max_hours(code, inspector_master_df)
                    if relax_work_hours:
                        # 緩和モード: 20%超過まで許容
                        allowed_max_hours = max_hours * (1.0 + SAME_DAY_WORK_HOURS_OVERRUN_RATE)
                    else:
                        allowed_max_hours = self._apply_work_hours_overrun(max_hours)
                    
                    # 緩和モードの場合は20%超過まで許容、通常モードは10%超過まで
                    if daily_hours_temp + divided_time > allowed_max_hours - WORK_HOURS_BUFFER:
                        continue
                    
                    selected_inspectors.append(insp)
                    # 一時的な状態を更新
                    if code not in temp_daily_assignments:
                        temp_daily_assignments[code] = {}
                    if current_date not in temp_daily_assignments[code]:
                        temp_daily_assignments[code][current_date] = 0.0
                    temp_daily_assignments[code][current_date] += divided_time
                
                # 固定検査員が選択された場合のログ
                for insp in selected_inspectors:
                    if insp.get('__is_fixed', False):
                        self.log_message(f"  固定検査員 '{insp['氏名']}' を優先的に選択しました（登録済み品番の特別処置）")
            
            # 選択された検査員の履歴を更新（10%超過をチェック）
            for insp in selected_inspectors:
                code = insp['コード']
                
                # 【追加】割り当て実行前に、10%超過をチェック（緩和モード対応）
                daily_hours_before = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                max_hours = self.get_inspector_max_hours(code, inspector_master_df)
                if relax_work_hours:
                    # 緩和モード: 20%超過まで許容
                    allowed_max_hours = max_hours * (1.0 + SAME_DAY_WORK_HOURS_OVERRUN_RATE)
                else:
                    allowed_max_hours = self._apply_work_hours_overrun(max_hours)
                
                # 【追加】動的バッファを計算
                assignment_count_for_buffer = self.inspector_assignment_count.get(code, 0)
                dynamic_buffer = self._get_dynamic_work_hours_buffer(code, current_date, assignment_count_for_buffer)
                
                # 緩和モードの場合は20%超過まで許容、通常モードは10%超過まで
                threshold = allowed_max_hours - dynamic_buffer
                exceeded = daily_hours_before + divided_time > threshold
                exceeded_by = max(0.0, (daily_hours_before + divided_time) - threshold) if exceeded else 0.0
                
                # 【追加】バッファ使用状況を記録
                self._record_buffer_usage(daily_hours_before, allowed_max_hours, dynamic_buffer, exceeded, exceeded_by)
                
                if exceeded:
                    mode_text = "緩和モードでも20%超過" if relax_work_hours else "10%超過"
                    self.log_message(
                        f"警告: 検査員 '{insp['氏名']}' ({code}) は{mode_text}を超えるため、割り当てをスキップします "
                        f"(今日: {daily_hours_before:.1f}h + {divided_time:.1f}h > {allowed_max_hours:.1f}h - {dynamic_buffer:.2f}h)",
                        level='warning'
                    )
                    continue
                
                self.inspector_assignment_count[code] += 1
                self.inspector_last_assignment[code] = current_time
                self.inspector_work_hours[code] += divided_time
                self.inspector_daily_assignments[code][current_date] += divided_time
                
                # ログ出力
                count = self.inspector_assignment_count.get(code, 0)
                skill_info = f"スキル: {insp['スキル']}" if not insp.get('is_new_team', False) else "新製品チーム"
                self.log_message(f"検査員 '{insp['氏名']}' ({skill_info}, 割り当て回数: {count}) を選択")
                insp.pop('__fairness_priority', None)
                insp.pop('__candidate_order', None)
            
            return selected_inspectors
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            self.log_message(f"スキル組み合わせ選択中にエラーが発生しました: {str(e)}", level='error')
            self.log_message(f"エラー詳細: {error_detail}", debug=True)
            return []
    
    def select_two_inspectors_with_skill_combination(
        self,
        skill_groups: Dict[str, List[Dict[str, Any]]],
        product_number: Optional[str] = None,
        divided_time: float = 0.0,
        current_date: Optional[date] = None,
        inspector_master_df: Optional[pd.DataFrame] = None,
        temp_daily_assignments: Optional[Dict[str, Dict[date, float]]] = None,
        relax_work_hours: bool = False
    ) -> List[Dict[str, Any]]:
        """
        2人の検査員をスキル組み合わせ考慮で選択（バランス重視版）
        
        Args:
            skill_groups: スキルレベル別の検査員グループ
            product_number: 品番（オプション）
        
        Returns:
            選択された検査員リスト
        """
        try:
            selected = []
            skill_order_map = {1: 0, 2: 1, 3: 2, 'new': 3}
            
            # 利用可能な検査員の総数を確認
            total_available = sum(len(inspectors) for inspectors in skill_groups.values())
            self.log_message(f"2人選択: 利用可能な検査員総数 {total_available}人")
            for skill_level, inspectors in skill_groups.items():
                self.log_message(f"  スキル{skill_level}: {len(inspectors)}人")
            
            if total_available < 2:
                self.log_message(f"警告: 2人選択要求だが、利用可能な検査員は {total_available}人 のみ")
                # 利用可能な分だけ選択
                for skill_level, inspectors in skill_groups.items():
                    for inspector in inspectors:
                        if len(selected) >= total_available:
                            break
                        selected.append(inspector)
                        self.log_message(f"  選択: {inspector['氏名']} (スキル: {inspector.get('スキル', '新製品')})")
                return selected
            
            # バランス重視の選択ロジック: 【追加】固定検査員 > 未使用検査員 > 総勤務時間のバランス > スキルレベル
            all_candidates = []
            for skill_level, inspectors in skill_groups.items():
                for insp in inspectors:
                    code = insp['コード']
                    assignment_count = self.inspector_assignment_count.get(code, 0)
                    total_hours = self.inspector_work_hours.get(code, 0.0)
                    last_assignment = self.inspector_last_assignment.get(code, pd.Timestamp.min)
                    is_unused = (assignment_count == 0)
                    is_fixed = insp.get('__is_fixed', False)  # 固定検査員フラグ
                    
                    # 【追加】優先順位: 0)固定検査員を最優先（登録済み品番の特別処置）, 1)未使用検査員優先, 2)総勤務時間が少ない, 3)スキルレベル(1>2>3>new、1が最高スキル), 4)割り当て回数が少ない, 5)4時間上限に近い場合は優先度を下げる
                    near_limit = insp.get('__near_product_limit', False)  # 4時間上限に近い場合は優先度を下げる
                    # product_hoursが辞書の場合は0.0にフォールバック
                    projected_hours = insp.get('__projected_product_hours')
                    if projected_hours is None:
                        product_hours_dict = self.inspector_product_hours.get(code, {})
                        if isinstance(product_hours_dict, dict):
                            # 辞書の場合は、該当品番の時間を取得（品番が不明な場合は0.0）
                            if product_number and product_number in product_hours_dict:
                                product_hours = product_hours_dict[product_number]
                            else:
                                product_hours = 0.0
                        else:
                            product_hours = product_hours_dict if isinstance(product_hours_dict, (int, float)) else 0.0
                    else:
                        product_hours = projected_hours if isinstance(projected_hours, (int, float)) else 0.0
                    product_limit_penalty = 1 if insp.get('over_product_limit', False) else 0
                    product_assignment_count = insp.get('__product_assignment_count', 0)
                    priority = (
                        not is_fixed,  # False=固定検査員を最優先（登録済み品番の特別処置）
                        product_limit_penalty,  # 4時間上限を超える場合は最終手段
                        near_limit,  # 4時間上限に近い場合は優先度を下げる（False < True）
                        product_assignment_count,  # 同一品番の割当回数
                        product_hours,  # 品番ごとの累計時間
                        not is_unused,  # False=未使用を優先
                        total_hours,   # 総勤務時間が少ない順
                        skill_order_map.get(skill_level, 99),  # スキル1を優先
                        assignment_count,  # 割り当て回数が少ない順
                        last_assignment  # 最後の割り当てが古い順
                    )
                    all_candidates.append((priority, insp, skill_level))
            
            # 【改善】10%超過をチェックしながら選択するためのヘルパー関数
            def check_and_add_inspector_two(insp, temp_assignments=None):
                """10%超過をチェックしてから検査員を追加（緩和モード対応）"""
                if temp_assignments is None:
                    temp_assignments = temp_daily_assignments if temp_daily_assignments is not None else {}
                if current_date is None or inspector_master_df is None or divided_time <= 0:
                    # パラメータが不足している場合は、チェックをスキップ
                    return True
                
                code = insp['コード']
                daily_hours_temp = temp_assignments.get(code, {}).get(current_date, 0.0)
                max_hours = self.get_inspector_max_hours(code, inspector_master_df)
                if relax_work_hours:
                    # 緩和モード: 20%超過まで許容
                    allowed_max_hours = max_hours * (1.0 + SAME_DAY_WORK_HOURS_OVERRUN_RATE)
                else:
                    allowed_max_hours = self._apply_work_hours_overrun(max_hours)
                
                # 緩和モードの場合は20%超過まで許容、通常モードは10%超過まで
                if daily_hours_temp + divided_time > allowed_max_hours - WORK_HOURS_BUFFER:
                    return False
                
                # 一時的な状態を更新
                if code not in temp_assignments:
                    temp_assignments[code] = {}
                if current_date not in temp_assignments[code]:
                    temp_assignments[code][current_date] = 0.0
                temp_assignments[code][current_date] += divided_time
                return True
            
            # スキル1がいる場合、バランスを考慮しつつスキル1を1人含める組み合わせを探す
            if skill_groups[1]:
                # スキル1の候補から最適な1人を選択（固定検査員 > 未使用・時間バランスを優先）
                skill1_candidates = [(p, i, sl) for p, i, sl in all_candidates if sl == 1]
                if skill1_candidates:
                    skill1_candidates.sort(key=self._priority_sort_key)
                    # 【改善】10%超過をチェックしながら選択
                    for priority, insp, sl in skill1_candidates:
                        if check_and_add_inspector_two(insp, temp_daily_assignments):
                            best_skill1 = insp
                            selected.append(best_skill1)
                            code = best_skill1['コード']
                            fixed_mark = " [固定検査員]" if best_skill1.get('__is_fixed', False) else ""
                            self.log_message(f"  スキル1選択: {best_skill1['氏名']}{fixed_mark} (総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
                            break
                    
                    if len(selected) == 1:
                        # 2人目を選択：固定検査員 > スキル3がいる場合は優先的に組み合わせる（教育のため）
                        remaining_candidates = [(p, i, sl) for p, i, sl in all_candidates if i != selected[0]]
                        if remaining_candidates:
                            # 固定検査員を優先的に探す
                            fixed_candidates = [(p, i, sl) for p, i, sl in remaining_candidates if i.get('__is_fixed', False)]
                            if fixed_candidates:
                                # 固定検査員がいる場合、優先的に選択（バランスを考慮してソート）
                                fixed_candidates.sort(key=self._priority_sort_key)
                                for priority, insp, sl in fixed_candidates:
                                    if check_and_add_inspector_two(insp, temp_daily_assignments):
                                        selected.append(insp)
                                        code = insp['コード']
                                        self.log_message(f"  固定検査員選択（登録済み品番の特別処置）: {insp['氏名']} (総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
                                        break
                            else:
                                # スキル3の候補を優先的に探す
                                skill3_candidates = [(p, i, sl) for p, i, sl in remaining_candidates if sl == 3]
                                if skill3_candidates:
                                    # スキル3がいる場合、優先的に選択（バランスを考慮してソート）
                                    skill3_candidates.sort(key=self._priority_sort_key)
                                    for priority, insp, sl in skill3_candidates:
                                        if check_and_add_inspector_two(insp, temp_daily_assignments):
                                            selected.append(insp)
                                            code = insp['コード']
                                            self.log_message(f"  スキル3選択（教育のため）: {insp['氏名']} (総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
                                            break
                                else:
                                    # スキル3がいない場合、バランスを考慮して選択
                                    remaining_candidates.sort(key=self._priority_sort_key)
                                    for priority, insp, sl in remaining_candidates:
                                        if check_and_add_inspector_two(insp, temp_daily_assignments):
                                            selected.append(insp)
                                            code = insp['コード']
                                            self.log_message(f"  2人目選択: {insp['氏名']} (総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
                                            break
            else:
                # スキル1がいない場合、バランスを最優先に2人選択（固定検査員を優先）
                all_candidates.sort(key=self._priority_sort_key)
                for priority, insp, sl in all_candidates:
                    if len(selected) >= 2:
                        break
                    if check_and_add_inspector_two(insp, temp_daily_assignments):
                        selected.append(insp)
                        code = insp['コード']
                        skill_info = f"スキル{sl}" if sl != 'new' else "新製品"
                        fixed_mark = " [固定検査員]" if insp.get('__is_fixed', False) else ""
                        self.log_message(f"  選択{len(selected)}: {insp['氏名']}{fixed_mark} ({skill_info}, 総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
            
            return selected
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            self.log_message(f"2人選択中にエラーが発生しました: {str(e)}", level='error')
            self.log_message(f"エラー詳細: {error_detail}", debug=True)
            return []
    
    def select_three_inspectors_with_skill_combination(
        self,
        skill_groups: Dict[str, List[Dict[str, Any]]],
        product_number: Optional[str] = None,
        divided_time: float = 0.0,
        current_date: Optional[date] = None,
        inspector_master_df: Optional[pd.DataFrame] = None,
        temp_daily_assignments: Optional[Dict[str, Dict[date, float]]] = None,
        relax_work_hours: bool = False
    ) -> List[Dict[str, Any]]:
        """
        3人の検査員をスキル組み合わせ考慮で選択（バランス重視版）
        
        Args:
            skill_groups: スキルレベル別の検査員グループ
            product_number: 品番（オプション）
        
        Returns:
            選択された検査員リスト
        """
        try:
            selected = []
            skill_order_map = {1: 0, 2: 1, 3: 2, 'new': 3}
            
            # 利用可能な検査員の総数を確認
            total_available = sum(len(inspectors) for inspectors in skill_groups.values())
            if total_available < 3:
                self.log_message(f"警告: 3人選択要求だが、利用可能な検査員は {total_available}人 のみ")
                # 利用可能な分だけ選択
                for skill_level, inspectors in skill_groups.items():
                    for inspector in inspectors:
                        if len(selected) >= total_available:
                            break
                        selected.append(inspector)
                return selected
            
            # バランス重視の選択ロジック: 【追加】固定検査員 > 未使用検査員 > 総勤務時間のバランス > スキルレベル
            all_candidates = []
            for skill_level, inspectors in skill_groups.items():
                for insp in inspectors:
                    code = insp['コード']
                    assignment_count = self.inspector_assignment_count.get(code, 0)
                    total_hours = self.inspector_work_hours.get(code, 0.0)
                    last_assignment = self.inspector_last_assignment.get(code, pd.Timestamp.min)
                    is_unused = (assignment_count == 0)
                    is_fixed = insp.get('__is_fixed', False)  # 固定検査員フラグ
                    
                    # 【追加】優先順位: 0)固定検査員を最優先（登録済み品番の特別処置）, 1)未使用検査員優先, 2)総勤務時間が少ない, 3)スキルレベル(1>2>3>new、1が最高スキル), 4)割り当て回数が少ない, 5)4時間上限に近い場合は優先度を下げる
                    near_limit = insp.get('__near_product_limit', False)  # 4時間上限に近い場合は優先度を下げる
                    # product_hoursが辞書の場合は0.0にフォールバック
                    projected_hours = insp.get('__projected_product_hours')
                    if projected_hours is None:
                        product_hours_dict = self.inspector_product_hours.get(code, {})
                        if isinstance(product_hours_dict, dict):
                            # 辞書の場合は、該当品番の時間を取得（品番が不明な場合は0.0）
                            if product_number and product_number in product_hours_dict:
                                product_hours = product_hours_dict[product_number]
                            else:
                                product_hours = 0.0
                        else:
                            product_hours = product_hours_dict if isinstance(product_hours_dict, (int, float)) else 0.0
                    else:
                        product_hours = projected_hours if isinstance(projected_hours, (int, float)) else 0.0
                    product_limit_penalty = 1 if insp.get('over_product_limit', False) else 0
                    product_assignment_count = insp.get('__product_assignment_count', 0)
                    priority = (
                        not is_fixed,  # False=固定検査員を最優先（登録済み品番の特別処置）
                        product_limit_penalty,  # 4時間上限を超える場合は最終手段
                        near_limit,  # 4時間上限に近い場合は優先度を下げる（False < True）
                        product_assignment_count,  # 同一品番の割当回数
                        product_hours,  # 品番単位の累計時間
                        not is_unused,  # False=未使用を優先
                        total_hours,   # 総勤務時間が少ない順
                        skill_order_map.get(skill_level, 99),  # スキル1を優先
                        assignment_count,  # 割り当て回数が少ない順
                        last_assignment  # 最後の割り当てが古い順
                    )
                    all_candidates.append((priority, insp, skill_level))
            
            # 【改善】10%超過をチェックしながら選択するためのヘルパー関数
            def check_and_add_inspector_three(insp, temp_assignments=None):
                """10%超過をチェックしてから検査員を追加（緩和モード対応）"""
                if temp_assignments is None:
                    temp_assignments = temp_daily_assignments if temp_daily_assignments is not None else {}
                if current_date is None or inspector_master_df is None or divided_time <= 0:
                    # パラメータが不足している場合は、チェックをスキップ
                    return True
                
                code = insp['コード']
                daily_hours_temp = temp_assignments.get(code, {}).get(current_date, 0.0)
                max_hours = self.get_inspector_max_hours(code, inspector_master_df)
                # 外側のスコープからrelax_work_hoursを取得（クロージャ）
                if relax_work_hours:
                    # 緩和モード: 20%超過まで許容
                    allowed_max_hours = max_hours * (1.0 + SAME_DAY_WORK_HOURS_OVERRUN_RATE)
                else:
                    allowed_max_hours = self._apply_work_hours_overrun(max_hours)
                
                # 緩和モードの場合は20%超過まで許容、通常モードは10%超過まで
                if daily_hours_temp + divided_time > allowed_max_hours - WORK_HOURS_BUFFER:
                    return False
                
                # 一時的な状態を更新
                if code not in temp_assignments:
                    temp_assignments[code] = {}
                if current_date not in temp_assignments[code]:
                    temp_assignments[code][current_date] = 0.0
                temp_assignments[code][current_date] += divided_time
                return True
            
            # スキル1がいる場合、バランスを考慮しつつスキル1を1人含める組み合わせを探す
            if skill_groups[1]:
                # スキル1の候補から最適な1人を選択（固定検査員 > 未使用・時間バランスを優先）
                skill1_candidates = [(p, i, sl) for p, i, sl in all_candidates if sl == 1]
                if skill1_candidates:
                    skill1_candidates.sort(key=self._priority_sort_key)
                    # 【改善】10%超過をチェックしながら選択
                    for priority, insp, sl in skill1_candidates:
                        if check_and_add_inspector_three(insp, temp_daily_assignments):
                            best_skill1 = insp
                            selected.append(best_skill1)
                            code = best_skill1['コード']
                            fixed_mark = " [固定検査員]" if best_skill1.get('__is_fixed', False) else ""
                            self.log_message(f"  スキル1選択: {best_skill1['氏名']}{fixed_mark} (総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
                            break
                    
                    if len(selected) == 1:
                        # 残り2人を選択：固定検査員 > スキル3がいる場合は優先的に組み合わせる（教育のため）
                        remaining_candidates = [(p, i, sl) for p, i, sl in all_candidates if i != selected[0]]
                        if remaining_candidates:
                            # 固定検査員を優先的に探す
                            fixed_candidates = [(p, i, sl) for p, i, sl in remaining_candidates if i.get('__is_fixed', False)]
                            if fixed_candidates:
                                # 固定検査員がいる場合、優先的に選択（バランスを考慮してソート）
                                fixed_candidates.sort(key=self._priority_sort_key)
                                # 【改善】10%超過をチェックしながら選択
                                for priority, insp, sl in fixed_candidates:
                                    if check_and_add_inspector_three(insp, temp_daily_assignments):
                                        selected.append(insp)
                                        code = insp['コード']
                                        self.log_message(f"  固定検査員選択（登録済み品番の特別処置）: {insp['氏名']} (総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
                                        break
                                
                                if len(selected) == 2:
                                    # 残り1人を選択（固定検査員以外から、固定検査員 > スキル3 > バランスを考慮）
                                    remaining_after_fixed = [(p, i, sl) for p, i, sl in remaining_candidates if i != selected[1]]
                                    if remaining_after_fixed:
                                        # 残りの固定検査員を優先的に探す
                                        remaining_fixed = [(p, i, sl) for p, i, sl in remaining_after_fixed if i.get('__is_fixed', False)]
                                        if remaining_fixed:
                                            remaining_fixed.sort(key=self._priority_sort_key)
                                            for priority, insp, sl in remaining_fixed:
                                                if check_and_add_inspector_three(insp, temp_daily_assignments):
                                                    selected.append(insp)
                                                    code = insp['コード']
                                                    self.log_message(f"  固定検査員選択（登録済み品番の特別処置）: {insp['氏名']} (総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
                                                    break
                                        else:
                                            # スキル3の候補を優先的に探す
                                            skill3_candidates = [(p, i, sl) for p, i, sl in remaining_after_fixed if sl == 3]
                                            if skill3_candidates:
                                                skill3_candidates.sort(key=self._priority_sort_key)
                                                for priority, insp, sl in skill3_candidates:
                                                    if check_and_add_inspector_three(insp, temp_daily_assignments):
                                                        selected.append(insp)
                                                        code = insp['コード']
                                                        self.log_message(f"  スキル3選択（教育のため）: {insp['氏名']} (総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
                                                        break
                                            else:
                                                remaining_after_fixed.sort(key=self._priority_sort_key)
                                                for priority, insp, sl in remaining_after_fixed:
                                                    if check_and_add_inspector_three(insp, temp_daily_assignments):
                                                        selected.append(insp)
                                                        code = insp['コード']
                                                        skill_info = f"スキル{sl}" if sl != 'new' else "新製品"
                                                        self.log_message(f"  3人目選択: {insp['氏名']} ({skill_info}, 総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
                                                        break
                        else:
                            # スキル3の候補を優先的に探す
                            skill3_candidates = [(p, i, sl) for p, i, sl in remaining_candidates if sl == 3]
                            if skill3_candidates:
                                # スキル3がいる場合、優先的に1人選択（バランスを考慮してソート）
                                skill3_candidates.sort(key=self._priority_sort_key)
                                # 【改善】10%超過をチェックしながら選択
                                for priority, insp, sl in skill3_candidates:
                                    if check_and_add_inspector_three(insp, temp_daily_assignments):
                                        selected.append(insp)
                                        code = insp['コード']
                                        self.log_message(f"  スキル3選択（教育のため）: {insp['氏名']} (総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
                                        break
                                
                                if len(selected) == 2:
                                # 残り1人を選択（スキル1とスキル3以外から、固定検査員 > バランスを考慮）
                                    remaining_after_skill3 = [(p, i, sl) for p, i, sl in remaining_candidates if i != selected[1]]
                                if remaining_after_skill3:
                                    # 固定検査員を優先的に探す
                                    remaining_fixed = [(p, i, sl) for p, i, sl in remaining_after_skill3 if i.get('__is_fixed', False)]
                                    if remaining_fixed:
                                        remaining_fixed.sort(key=self._priority_sort_key)
                                        for priority, insp, sl in remaining_fixed:
                                            if check_and_add_inspector_three(insp, temp_daily_assignments):
                                                selected.append(insp)
                                                code = insp['コード']
                                                self.log_message(f"  固定検査員選択（登録済み品番の特別処置）: {insp['氏名']} (総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
                                                break
                                    else:
                                        remaining_after_skill3.sort(key=self._priority_sort_key)
                                        for priority, insp, sl in remaining_after_skill3:
                                            if check_and_add_inspector_three(insp, temp_daily_assignments):
                                                selected.append(insp)
                                                code = insp['コード']
                                                skill_info = f"スキル{sl}" if sl != 'new' else "新製品"
                                                self.log_message(f"  3人目選択: {insp['氏名']} ({skill_info}, 総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
                                                break
                            else:
                                # スキル3がいない場合、バランスを考慮して2人選択（固定検査員を優先）
                                remaining_candidates.sort(key=self._priority_sort_key)
                                for priority, insp, sl in remaining_candidates:
                                    if len(selected) >= 3:
                                        break
                                    if check_and_add_inspector_three(insp, temp_daily_assignments):
                                        selected.append(insp)
                                        code = insp['コード']
                                        skill_info = f"スキル{sl}" if sl != 'new' else "新製品"
                                        fixed_mark = " [固定検査員]" if insp.get('__is_fixed', False) else ""
                                        self.log_message(f"  選択{len(selected)}: {insp['氏名']}{fixed_mark} ({skill_info}, 総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
            else:
                # スキル1がいない場合、バランスを最優先に3人選択（固定検査員を優先）
                all_candidates.sort(key=self._priority_sort_key)
                for priority, insp, sl in all_candidates:
                    if len(selected) >= 3:
                        break
                    if check_and_add_inspector_three(insp, temp_daily_assignments):
                        selected.append(insp)
                        code = insp['コード']
                        skill_info = f"スキル{sl}" if sl != 'new' else "新製品"
                        fixed_mark = " [固定検査員]" if insp.get('__is_fixed', False) else ""
                        self.log_message(f"  選択{len(selected)}: {insp['氏名']}{fixed_mark} ({skill_info}, 総勤務時間: {self.inspector_work_hours.get(code, 0.0):.1f}h, 割当回数: {self.inspector_assignment_count.get(code, 0)})")
            
            return selected
            
        except Exception as e:
            self.log_message(f"3人選択中にエラーが発生しました: {str(e)}")
            return []

    def _priority_sort_key(self, candidate_tuple: Tuple[object, Dict[str, Any], Any]) -> Tuple:
        """
        ソートキーを生成するメソッド
        
        Args:
            candidate_tuple: (priority, inspector, ...) の形式のタプル
        
        Returns:
            ソート可能なタプル
        """
        try:
            priority = candidate_tuple[0]
            inspector = candidate_tuple[1] if len(candidate_tuple) > 1 else {}
        except (IndexError, TypeError):
            # タプル形式でない場合は、デフォルト値を返す
            return (999, ('', ''))

        # 同点ブレーカー（検査員コード昇順）: 同じ優先度の候補が複数ある場合でも割当がブレないようにする
        raw_code = inspector.get('コード', '') if isinstance(inspector, dict) else ''
        try:
            code_str = str(raw_code).strip()
        except Exception:
            code_str = ''
        try:
            code_int = int(code_str) if code_str.isdigit() else None
        except Exception:
            code_int = None
        code_key: Tuple[int, Any] = (0, code_int) if code_int is not None else (1, code_str)

        # priorityの型に応じて適切なソートキーを生成
        if isinstance(priority, tuple):
            return priority + (code_key,)
        if isinstance(priority, list):
            return tuple(priority) + (code_key,)
        if isinstance(priority, dict):
            # 辞書の場合は、ソート可能な形式に変換（キーと値の組み合わせをソート）
            try:
                # 辞書のキーと値をソート可能な形式に変換
                dict_items = sorted(priority.items())
                dict_key = tuple((k, v) for k, v in dict_items)
                return dict_key + (code_key,)
            except Exception:
                # 変換に失敗した場合は文字列表現を使用
                return (repr(priority), code_key)
        # その他の型（int, float, str, bool, Noneなど）はそのまま使用
        try:
            # 比較可能な型かチェック
            if priority is None:
                return (999, code_key)
            return (priority, code_key)
        except TypeError:
            # 比較できない型の場合は文字列表現を使用
            return (repr(priority), code_key)

    def filter_available_inspectors(
        self,
        available_inspectors: List[Dict[str, Any]],
        divided_time: float,
        inspector_master_df: pd.DataFrame,
        product_number: str,
        relax_work_hours: bool = False,
        process_name_context: Optional[str] = None,
        ignore_product_limit: bool = False
    ) -> List[Dict[str, Any]]:
        """
        勤務時間と品番上限を考慮して利用可能な検査員をフィルタリングする（第1パスは緩和版）。
        
        Args:
            available_inspectors: 利用可能な検査員リスト
            divided_time: 分割検査時間
            inspector_master_df: 検査員マスタ
            product_number: 品番
            relax_work_hours: 勤務時間チェックを緩和するか
        
        Returns:
            フィルタリングされた検査員リスト
        """
        try:
            filtered_inspectors = []
            excluded_by_work_hours = []  # 勤務時間で除外された検査員
            excluded_by_product_limit = []  # 4時間上限で除外された検査員
            excluded_by_vacation = []  # 休暇で除外された検査員
            current_date = pd.Timestamp.now().date()

            for inspector in available_inspectors:
                inspector_code = inspector['コード']
                inspector_name = inspector['氏名']
                inspector_entry = inspector.copy()

                # 【追加】休暇情報をチェック（終日休みの場合は除外）
                vacation_info = self.get_vacation_info(inspector_name)
                if vacation_info:
                    code = vacation_info.get("code", "")
                    work_status = vacation_info.get("work_status", "")
                    
                    # 終日休みの場合は除外
                    if code in ["休", "出", "当"]:
                        interpretation = vacation_info.get("interpretation", "")
                        excluded_by_vacation.append(f"{inspector_name}({inspector_code})")
                        self.log_message(
                            f"検査員 '{inspector_name}' は終日休暇のため除外 "
                            f"(休暇コード: {code}, 解釈: {interpretation})",
                            debug=True
                        )
                        continue

                # 現在の日付での累積勤務時間を取得
                daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
                additional_hours = divided_time

                # 勤務時間上限を算出（検査員マスタベース、休暇情報を考慮）
                max_daily_hours = self.get_inspector_max_hours(inspector_code, inspector_master_df)
                allowed_max_hours = self._apply_work_hours_overrun(max_daily_hours)
                
                # 実質勤務時間が0以下の場合は除外
                if max_daily_hours <= 0:
                    excluded_by_work_hours.append(f"{inspector_name}({inspector_code}): 調整後勤務時間0時間以下")
                    warning_key = (f"調整後勤務時間0時間", inspector_name)
                    if warning_key not in self.logged_warnings:
                        self.log_message(
                            f"警告: 検査員 '{inspector_name}' の調整後勤務時間が0時間以下です - 除外します",
                            level='warning'
                        )
                        self.logged_warnings.add(warning_key)
                    continue

                # 改善ポイント: 定数を使用
                # 勤務時間チェック（WORK_HOURS_BUFFERの余裕を確保）
                # 緩和モードの場合は20%超過まで許容（当日洗浄上がり品の制約緩和時など）
                if relax_work_hours:
                    # 緩和モード: 20%超過まで許容（SAME_DAY_WORK_HOURS_OVERRUN_RATEを使用）
                    relaxed_allowed_max_hours = max_daily_hours * (1.0 + SAME_DAY_WORK_HOURS_OVERRUN_RATE)
                    if daily_hours + additional_hours > relaxed_allowed_max_hours - WORK_HOURS_BUFFER:
                        excluded_by_work_hours.append(f"{inspector_name}({inspector_code}): {daily_hours:.1f}h+{additional_hours:.1f}h>{relaxed_allowed_max_hours:.1f}h-{WORK_HOURS_BUFFER:.2f}h")
                        self.log_message(
                            f"検査員 '{inspector['氏名']}' は緩和モードでも20%超過を超えるため除外 "
                            f"(今日: {daily_hours:.1f}h + {additional_hours:.1f}h > {relaxed_allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h = {relaxed_allowed_max_hours - WORK_HOURS_BUFFER:.1f}h)",
                            debug=True
                        )
                        continue
                else:
                    # 通常モード: 10%超過+バッファを厳格にチェック
                    if daily_hours + additional_hours > allowed_max_hours - WORK_HOURS_BUFFER:
                        excluded_by_work_hours.append(f"{inspector_name}({inspector_code}): {daily_hours:.1f}h+{additional_hours:.1f}h>{allowed_max_hours:.1f}h-{WORK_HOURS_BUFFER:.2f}h")
                        self.log_message(
                            f"検査員 '{inspector['氏名']}' は10%超過を超えるため除外 "
                            f"(今日: {daily_hours:.1f}h + {additional_hours:.1f}h > {allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h = {allowed_max_hours - WORK_HOURS_BUFFER:.1f}h)",
                            debug=True
                        )
                        continue

                # 改善ポイント: 4時間上限ルールの2段階化
                # ドラフトフェーズ：4.5h未満までは許容（4.0h超は over_product_limit=True を設定）
                # 最適化フェーズ：ここで4.0h遵守へ是正。置換不可能な場合のみ未割当へ戻す。
                if not ignore_product_limit:
                    product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                    projected_hours = product_hours + divided_time
                    
                    # ドラフトフェーズでの許容上限チェック（4.5h未満まで許容）
                    if projected_hours >= PRODUCT_LIMIT_DRAFT_THRESHOLD:
                        excluded_by_product_limit.append(f"{inspector_name}({inspector_code}): {product_hours:.1f}h+{divided_time:.1f}h={projected_hours:.1f}h>={PRODUCT_LIMIT_DRAFT_THRESHOLD}h")
                        self.log_message(
                            f"検査員 '{inspector['氏名']}' は品番 {product_number} の累計が {product_hours:.1f}h で、"
                            f"追加すると {projected_hours:.1f}h となるため（{PRODUCT_LIMIT_DRAFT_THRESHOLD}h以上）今回は除外します",
                            debug=True
                        )
                        continue

                    # 4.0h超過の場合はフラグを設定（ドラフトフェーズでは許容、最適化フェーズで是正）
                    inspector_entry['over_product_limit'] = projected_hours > PRODUCT_LIMIT_HARD_THRESHOLD
                else:
                    # ignore_product_limit=Trueの場合、4時間上限チェックをスキップ
                    inspector_entry['over_product_limit'] = False
                    # ログ出力用にprojected_hoursを計算（チェックは行わない）
                    product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                    projected_hours = product_hours + divided_time
                filtered_inspectors.append(inspector_entry)
                self.log_message(
                    f"検査員 '{inspector['氏名']}' は利用可能 "
                    f"(今日: {daily_hours:.1f}h + {additional_hours:.1f}h = {daily_hours + additional_hours:.1f}h, "
                    f"最大勤務時間: {max_daily_hours:.1f}h, 品番累計予定: {projected_hours:.1f}h)",
                    debug=True
                )

            # 除外された検査員の詳細をログ出力（未割当ロット再処理時など、入力数が多い場合）
            input_count = len(available_inspectors)
            filtered_count = len(filtered_inspectors)
            mode_str = "緩和モード（20%超過許容）" if relax_work_hours else "通常モード（10%超過許容）"
            if input_count > 0 and filtered_count == 0:
                # 全て除外された場合、詳細ログを出力
                if excluded_by_vacation:
                    self.log_message(f"filter_available_inspectors: 休暇で除外: {', '.join(excluded_by_vacation[:3])}{'...' if len(excluded_by_vacation) > 3 else ''}", debug=True)
                if excluded_by_work_hours:
                    self.log_message(f"filter_available_inspectors: 勤務時間で除外（{mode_str}）: {', '.join(excluded_by_work_hours[:3])}{'...' if len(excluded_by_work_hours) > 3 else ''}", debug=True)
                if excluded_by_product_limit:
                    self.log_message(f"filter_available_inspectors: 4時間上限で除外: {', '.join(excluded_by_product_limit[:3])}{'...' if len(excluded_by_product_limit) > 3 else ''}", debug=True)
                self.log_message(f"filter_available_inspectors: 品番 {product_number} の入力 {input_count}名が全て除外されました（休暇: {len(excluded_by_vacation)}名、勤務時間: {len(excluded_by_work_hours)}名、4時間上限: {len(excluded_by_product_limit)}名、{mode_str}, ignore_product_limit={ignore_product_limit}）", level='warning')

            # 【追加】固定検査員を優先的に配置
            fixed_inspector_names = self._collect_fixed_inspector_names(product_number, process_name_context)
            if fixed_inspector_names:
                # 固定検査員とそれ以外に分離
                fixed_inspectors = []
                other_inspectors = []
                
                for inspector in filtered_inspectors:
                    inspector_name = inspector['氏名']
                    if inspector_name in fixed_inspector_names:
                        fixed_inspectors.append(inspector)
                    else:
                        other_inspectors.append(inspector)
                
                # 固定検査員を優先的にリストの先頭に配置
                filtered_inspectors = fixed_inspectors + other_inspectors
                if fixed_inspectors:
                    self.log_message(f"固定検査員を優先配置（フィルタ後）: {len(fixed_inspectors)}名を先頭に配置")

            return filtered_inspectors

        except Exception as e:
            self.log_message(f"検査員フィルタリング中にエラーが発生しました: {str(e)}")
            return available_inspectors
    
    def set_vacation_data(
        self,
        vacation_data: Dict[str, Dict[str, Any]],
        target_date: date,
        inspector_master_df: Optional[pd.DataFrame] = None
    ) -> None:
        """
        休暇情報を設定する
        
        Args:
            vacation_data: {従業員名: 休暇情報辞書} の形式の辞書
            target_date: 対象日付
            inspector_master_df: 検査員マスタDataFrame（別名マッピング用）
        """
        self.vacation_data = vacation_data
        self.vacation_date = target_date
        
        # 検査員マスタの「休暇予定表の別名」列を考慮してマッピングを作成
        self.inspector_name_to_vacation = {}
        
        if inspector_master_df is not None and '#氏名' in inspector_master_df.columns:
            # 別名列がある場合はそれを使用
            if '休暇予定表の別名' in inspector_master_df.columns:
                # 列インデックスを事前に取得（itertuples()で高速化）
                name_col_idx = inspector_master_df.columns.get_loc('#氏名')
                alias_col_idx = inspector_master_df.columns.get_loc('休暇予定表の別名')
                
                for row_tuple in inspector_master_df.itertuples(index=False):
                    inspector_name = row_tuple[name_col_idx]
                    alias_name = row_tuple[alias_col_idx] if alias_col_idx < len(row_tuple) else ''
                    
                    # 別名が設定されている場合は別名で検索、なければ氏名で検索
                    vacation_name = alias_name.strip() if pd.notna(alias_name) and alias_name.strip() else inspector_name
                    
                    if vacation_name in vacation_data:
                        self.inspector_name_to_vacation[inspector_name] = vacation_data[vacation_name]
                        self.log_message(f"検査員 '{inspector_name}' の休暇情報をマッピング（別名: '{vacation_name}'）")
                    elif inspector_name in vacation_data:
                        self.inspector_name_to_vacation[inspector_name] = vacation_data[inspector_name]
            else:
                # 別名列がない場合は氏名で直接マッピング
                for inspector_name in inspector_master_df['#氏名']:
                    if inspector_name in vacation_data:
                        self.inspector_name_to_vacation[inspector_name] = vacation_data[inspector_name]
        else:
            # 検査員マスタがない場合は直接マッピング
            self.inspector_name_to_vacation = vacation_data.copy()
        
        self.log_message(f"休暇情報を設定しました: {len(self.inspector_name_to_vacation)}名、対象日: {target_date}")
    
    def get_vacation_info(self, inspector_name: str) -> Optional[Dict[str, Any]]:
        """
        検査員の休暇情報を取得する
        
        Args:
            inspector_name: 検査員名
        
        Returns:
            dict: 休暇情報辞書（休暇でない場合はNone）
        """
        return self.inspector_name_to_vacation.get(inspector_name)

    def _remove_inspector_from_same_day_sets(
        self,
        product_number: str,
        product_name: str,
        inspector_code: str
    ) -> None:
        """当日洗浄制約のセットから検査員を除外して最新状態にする"""
        if not inspector_code:
            return

        if product_number:
            existing = self.same_day_cleaning_inspectors.get(product_number)
            if existing and inspector_code in existing:
                existing.discard(inspector_code)
                if not existing:
                    del self.same_day_cleaning_inspectors[product_number]

        if product_name:
            product_name_str = str(product_name).strip()
            existing_by_name = self.same_day_cleaning_inspectors_by_product_name.get(product_name_str)
            if existing_by_name and inspector_code in existing_by_name:
                existing_by_name.discard(inspector_code)
                if not existing_by_name:
                    del self.same_day_cleaning_inspectors_by_product_name[product_name_str]
    
    def is_inspector_on_vacation(self, inspector_name: str) -> bool:
        """
        検査員が休暇中かどうかを判定する
        
        Args:
            inspector_name: 検査員名
        
        Returns:
            bool: 休暇中の場合はTrue
        """
        vacation_info = self.get_vacation_info(inspector_name)
        if not vacation_info:
            return False
        
        work_status = vacation_info.get("work_status")
        return work_status == "休み"
    
    def get_inspector_max_hours(
        self,
        inspector_code: str,
        inspector_master_df: pd.DataFrame
    ) -> float:
        """
        検査員の最大勤務時間を取得（検査員マスタから、休暇情報を考慮）
        
        Args:
            inspector_code: 検査員コード
            inspector_master_df: 検査員マスタのDataFrame
        
        Returns:
            最大勤務時間（時間単位）
            inspector_master_df: 検査員マスタDataFrame
        
        Returns:
            float: 実質的な最大勤務時間（時間単位）
        """
        try:
            inspector_info = inspector_master_df[inspector_master_df['#ID'] == inspector_code]
            if not inspector_info.empty:
                inspector_data = inspector_info.iloc[0]
                start_time = inspector_data['開始時刻']
                end_time = inspector_data['終了時刻']
                inspector_name = inspector_data['#氏名']
                
                if pd.notna(start_time) and pd.notna(end_time):
                    try:
                        # 時刻文字列を時間に変換
                        if isinstance(start_time, str):
                            start_hour = float(start_time.split(':')[0]) + float(start_time.split(':')[1]) / 60.0
                            start_time_str = start_time
                        else:
                            start_hour = start_time.hour + start_time.minute / 60.0
                            start_time_str = f"{start_time.hour:02d}:{start_time.minute:02d}"
                            
                        if isinstance(end_time, str):
                            end_hour = float(end_time.split(':')[0]) + float(end_time.split(':')[1]) / 60.0
                            end_time_str = end_time
                        else:
                            end_hour = end_time.hour + end_time.minute / 60.0
                            end_time_str = f"{end_time.hour:02d}:{end_time.minute:02d}"
                        
                        # 基本勤務時間を計算
                        max_daily_hours = end_hour - start_hour
                        
                        # 休憩時間（12:15～13:00）を含む場合は1時間を差し引く
                        if start_hour <= 12.25 and end_hour >= 13.0:
                            max_daily_hours -= 1.0
                        
                        # 【追加】休暇情報を考慮して不在時間を差し引く
                        vacation_info = self.get_vacation_info(inspector_name)
                        if vacation_info:
                            from app.services.vacation_schedule_service import calculate_vacation_absence_hours
                            absence_hours = calculate_vacation_absence_hours(
                                vacation_info, start_time_str, end_time_str
                            )
                            max_daily_hours -= absence_hours
                            
                            if absence_hours > 0:
                                code = vacation_info.get("code", "")
                                interpretation = vacation_info.get("interpretation", "")
                                vacation_key = (
                                    inspector_name,
                                    code,
                                    interpretation,
                                    self.vacation_date.isoformat() if self.vacation_date else None
                                )
                                if vacation_key not in self.logged_vacation_messages:
                                    self.logged_vacation_messages.add(vacation_key)
                                    self.log_message(
                                        f"検査員 '{inspector_name}' の休暇を考慮: "
                                        f"基本勤務時間 {max_daily_hours + absence_hours:.1f}h - "
                                        f"不在時間 {absence_hours:.1f}h = "
                                        f"実質勤務時間 {max_daily_hours:.1f}h "
                                        f"(休暇コード: {code}, {interpretation})"
                                    )
                        
                        return max(0.0, max_daily_hours)
                    except Exception as e:
                        self.log_message(f"勤務時間計算エラー ({inspector_name}): {str(e)}", level='warning')
                        return 8.0
                else:
                    return 8.0
            else:
                return 8.0
        except Exception as e:
            self.log_message(f"最大勤務時間取得エラー: {str(e)}", level='warning')
            return 8.0

    def print_assignment_statistics(
        self,
        inspector_master_df: Optional[pd.DataFrame] = None
    ) -> None:
        """
        割り当て統計を表示
        
        Args:
            inspector_master_df: 検査員マスタのDataFrame（オプション）
        """
        try:
            if not self.inspector_assignment_count:
                self.log_message("割り当て統計: まだ割り当てがありません")
                return
            
            self.log_message("検査員割り当て統計")
            
            # 割り当て回数でソート
            sorted_assignments = sorted(self.inspector_assignment_count.items(), 
                                      key=lambda x: x[1], reverse=True)
            
            total_assignments = sum(self.inspector_assignment_count.values())
            inspector_count = len(self.inspector_assignment_count)
            average_assignments = total_assignments / inspector_count if inspector_count > 0 else 0
            
            self.log_message(f"割り当て実績: {inspector_count}名 / 合計{total_assignments}回 / 平均{average_assignments:.1f}回")
            
            # 警告がある検査員を収集
            warning_inspectors = []
            
            # 各検査員の割り当て回数と勤務時間を確認
            for inspector_code, count in sorted_assignments:
                work_hours = self.inspector_work_hours.get(inspector_code, 0.0)
                daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(pd.Timestamp.now().date(), 0.0)
                
                # 検査員マスタから最大勤務時間を取得
                if inspector_master_df is not None:
                    max_hours = self.get_inspector_max_hours(inspector_code, inspector_master_df)
                    allowed_max_hours = self._apply_work_hours_overrun(max_hours)
                    # 警告対象をチェック
                    if work_hours > allowed_max_hours * 0.8:  # 80%超過で警告
                        warning_inspectors.append(
                            (inspector_code, count, work_hours, max_hours, allowed_max_hours, daily_hours, work_hours > allowed_max_hours)
                        )
                else:
                    # 検査員マスタがない場合
                    if work_hours > 6.0:  # 6時間超過で警告
                        max_hours = 8.0
                        allowed_max_hours = self._apply_work_hours_overrun(max_hours)
                        warning_inspectors.append((inspector_code, count, work_hours, max_hours, allowed_max_hours, daily_hours, work_hours > allowed_max_hours))
            
            # デバッグモードでない場合は警告がある検査員のみ詳細表示
            if not self.debug_mode:
                if warning_inspectors:
                    # 多すぎると読みにくいので上位のみ表示
                    warning_inspectors_sorted = sorted(warning_inspectors, key=lambda x: (x[6], x[2]), reverse=True)
                    self.log_message(f"警告対象の検査員: {len(warning_inspectors_sorted)}名（上位10名まで表示）")
                    for inspector_code, count, work_hours, max_hours, allowed_max_hours, daily_hours, is_over in warning_inspectors_sorted[:10]:
                        if is_over:
                            # 10%超過許容後の最大時間を超えている場合
                            excess_overrun = work_hours - allowed_max_hours
                            excess_base = work_hours - max_hours
                            status = f"（10%超過超過: {excess_overrun:.1f}h, 基本超過: {excess_base:.1f}h）"
                        else:
                            status = f"（80%超: {work_hours:.1f}h/{allowed_max_hours:.1f}h, 基本: {max_hours:.1f}h）"
                        self.log_message(f"  {inspector_code}: {count}回 (勤務時間: {work_hours:.1f}h/許容最大: {allowed_max_hours:.1f}h, 基本: {max_hours:.1f}h, 今日: {daily_hours:.1f}h){status}")
                else:
                    self.log_message("警告対象の検査員: 0名（正常範囲内）")
            else:
                # デバッグモード: 全員の詳細を表示
                self.log_message("")
                self.log_message("【詳細情報（デバッグモード）】:")
                for inspector_code, count in sorted_assignments:
                    work_hours = self.inspector_work_hours.get(inspector_code, 0.0)
                    daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(pd.Timestamp.now().date(), 0.0)
                    
                if inspector_master_df is not None:
                    max_hours = self.get_inspector_max_hours(inspector_code, inspector_master_df)
                    allowed_max_hours = self._apply_work_hours_overrun(max_hours)
                    status = ""
                    if work_hours > allowed_max_hours:
                        status = f" ⚠️ {allowed_max_hours:.1f}h超過"
                    elif work_hours > allowed_max_hours * 0.8:
                        status = f" ⚠️ {allowed_max_hours:.1f}hの80%超過"
                    self.log_message(f"  {inspector_code}: {count}回 (総勤務時間: {work_hours:.1f}h/{allowed_max_hours:.1f}h, 今日: {daily_hours:.1f}h){status}")
                else:
                    status = ""
                    if work_hours > 8.0:
                        status = " ⚠️ 8時間超過"
                    elif work_hours > 6.0:
                        status = " ⚠️ 6時間超過"
                    self.log_message(f"  {inspector_code}: {count}回 (総勤務時間: {work_hours:.1f}h, 今日: {daily_hours:.1f}h){status}")
            
            # 偏り度を計算
            max_count = max(self.inspector_assignment_count.values())
            min_count = min(self.inspector_assignment_count.values())
            imbalance = max_count - min_count
            
            self.log_message(f"割り当て回数: 最大{max_count}回 / 最小{min_count}回 / 偏り度{imbalance}回")
            
            if imbalance <= 1:
                self.log_message("判定: 偏り小")
            elif imbalance <= 2:
                self.log_message("判定: 偏りあり（軽微）")
            else:
                self.log_message("判定: 偏りあり")
            
        except Exception as e:
            self.log_message(f"統計表示中にエラーが発生しました: {str(e)}", level='error')
    
    def print_detailed_kpi_statistics(
        self,
        result_df: pd.DataFrame,
        inspector_master_df: pd.DataFrame,
        skill_master_df: pd.DataFrame
    ) -> None:
        """
        改善ポイント: 最終ログ出力の拡充
        
        以下のKPIを毎回出力する:
        - 未割当ロット総数と assignability_status ごとの件数
        - 理論上割当可能（残時間合計≧必要時間）だが未成立ロット数
        - 4.0h超過→置換で解消できた件数／率
        - 偏り是正フェーズの swap 実施率
        - 各検査員の勤務時間平均・分散・変動係数（CV）
        """
        try:
            self.log_message("")
            self.log_message("KPI統計")
             
            # 1. 未割当ロット総数と assignability_status ごとの件数
            if 'assignability_status' in result_df.columns:
                status_counts = result_df['assignability_status'].value_counts().to_dict()
                unassigned_total = sum(count for status, count in status_counts.items() 
                                     if status not in ['fully_assigned', 'capacity_shortage_resolved', 'skill_mismatch_resolved'])
                
                if unassigned_total > 0:
                    self.log_message(f"未割当ロット総数: {unassigned_total}件")
                    if self.debug_mode:
                        self.log_message("【assignability_status ごとの件数】:")
                        for status, count in sorted(status_counts.items()):
                            self.log_message(f"  - {status}: {count}件")
                    else:
                        # 通常モード: 未割当のstatusのみ表示
                        unassigned_statuses = {status: count for status, count in status_counts.items() 
                                             if status not in ['fully_assigned', 'capacity_shortage_resolved', 'skill_mismatch_resolved']}
                        if unassigned_statuses:
                            self.log_message("未割当のstatus別件数:")
                            for status, count in sorted(unassigned_statuses.items()):
                                self.log_message(f"  - {status}: {count}件")
                else:
                    self.log_message("未割当ロット総数: 0件（すべて割り当て完了）")
            else:
                self.log_message("未割当ロット総数: assignability_status列が見つかりません")
            
            # 2. 理論上割当可能（残時間合計≧必要時間）だが未成立ロット数
            if 'available_capacity_hours' in result_df.columns and '検査時間' in result_df.columns:
                theoretical_possible = result_df[
                    (result_df['available_capacity_hours'] >= result_df['検査時間']) &
                    (result_df['assignability_status'].isin(['logic_conflict', 'partial_assigned', 'capacity_shortage_partial', 'skill_mismatch_partial']))
                ]
                if len(theoretical_possible) > 0:
                    self.log_message(f"理論上割当可能だが未成立ロット数: {len(theoretical_possible)}件")
                else:
                    self.log_message("理論上割当可能だが未成立ロット数: 0件")
            else:
                if self.debug_mode:
                    self.log_message("理論上割当可能だが未成立ロット数: 必要な列が見つかりません")
            
            # 3. 4.0h超過→置換で解消できた件数／率
            # (relaxed_product_limit_assignmentsに含まれるが、最終的に4.0h以下になった件数)
            resolved_over_limit = 0
            total_over_limit = len(self.relaxed_product_limit_assignments)
            if total_over_limit > 0:
                for inspector_code, product_number in self.relaxed_product_limit_assignments:
                    product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                    if product_hours <= self.product_limit_hard_threshold:
                        resolved_over_limit += 1
                resolution_rate = (resolved_over_limit / total_over_limit * 100) if total_over_limit > 0 else 0.0
                self.log_message(f"4.0h超過→置換で解消: {resolved_over_limit}件 / {total_over_limit}件 ({resolution_rate:.1f}%)")
            else:
                if self.debug_mode:
                    self.log_message("4.0h超過→置換で解消: 0件（該当なし）")
            
            # 4. 偏り是正フェーズの swap 実施率
            # (fix_single_violationでswapが実行された件数 / 総違反件数)
            if hasattr(self, 'violation_count') and self.violation_count > 0:
                swap_count = getattr(self, 'swap_count', 0)
                swap_rate = (swap_count / self.violation_count * 100) if self.violation_count > 0 else 0.0
                self.log_message(f"偏り是正フェーズのswap実施率: {swap_count}/{self.violation_count} = {swap_rate:.1f}%")
            else:
                if self.debug_mode:
                    self.log_message("偏り是正フェーズのswap実施率: 違反件数が0件のため計算不可")
            
            # 5. 各検査員の勤務時間平均・分散・変動係数（CV）
            if inspector_master_df is not None and self.inspector_daily_assignments:
                current_date = pd.Timestamp.now().date()
                work_hours_list = []
                for inspector_code in self.inspector_daily_assignments:
                    daily_hours = self.inspector_daily_assignments[inspector_code].get(current_date, 0.0)
                    if daily_hours > 0:
                        work_hours_list.append(daily_hours)
                
                if work_hours_list:
                    mean_hours = np.mean(work_hours_list)
                    std_hours = np.std(work_hours_list)
                    cv = (std_hours / mean_hours * 100) if mean_hours > 0 else 0.0
                    
                    self.log_message("検査員勤務時間統計")
                    self.log_message(f"  - 平均: {mean_hours:.2f}h")
                    if self.debug_mode:
                        self.log_message(f"  - 標準偏差: {std_hours:.2f}h")
                        self.log_message(f"  - 変動係数(CV): {cv:.2f}%")
                    else:
                        # 通常モード: 変動係数のみ表示（分散の目安）
                        if cv > 30:
                            self.log_message(f"  - 変動係数(CV): {cv:.2f}%（分散が大きい）")
                        else:
                            self.log_message(f"  - 変動係数(CV): {cv:.2f}%（分散は適切）")
                else:
                    if self.debug_mode:
                        self.log_message("検査員勤務時間統計: データなし")
            else:
                if self.debug_mode:
                    self.log_message("検査員勤務時間統計: 検査員マスタまたは履歴データなし")

            # 6. 制約緩和（頻出ログを集計して要約）
            if not self.debug_mode and self._suppressed_relax_assign_total:
                top = sorted(self._suppressed_relax_assign_by_product.items(), key=lambda kv: (-kv[1], kv[0]))[:5]
                top_str = ", ".join([f"{pn}={cnt}" for pn, cnt in top])
                self.log_message(f"制約緩和で割当: {self._suppressed_relax_assign_total}件（上位: {top_str}）")
             
            self.log_message("=" * 60)
            self.log_message("")
             
        except Exception as e:
            self.log_message(f"KPI統計表示中にエラーが発生しました: {str(e)}", level='error')
    
    def optimize_assignments(
        self,
        result_df: pd.DataFrame,
        inspector_master_df: pd.DataFrame,
        skill_master_df: pd.DataFrame,
        show_skill_values: bool = False,
        process_master_df: Optional[pd.DataFrame] = None,
        inspection_target_keywords: Optional[List[str]] = None
    ) -> pd.DataFrame:
        """
        全体最適化：勤務時間超過の是正と偏りの調整
        
        Args:
            result_df: 割当結果のDataFrame
            inspector_master_df: 検査員マスタデータ
            skill_master_df: スキルマスタデータ
            show_skill_values: スキル値を表示するかどうか
            process_master_df: 工程マスタデータ（オプション）
            inspection_target_keywords: 検査対象キーワードリスト（オプション）
        
        Returns:
            最適化された割当結果のDataFrame
        """
        try:
            # 【高速化】検査員マスタのインデックスを構築
            self._build_inspector_index(inspector_master_df)

            # 【高速化】検査員名 -> 検査員ID のローカル辞書を作成（DataFrame生成を避ける）
            inspector_name_to_id: Dict[str, Any] = {}
            try:
                for name_key, row in self.inspector_name_to_row.items():
                    if row is None:
                        continue
                    inspector_id = row.get('#ID')
                    if pd.notna(inspector_id):
                        inspector_name_to_id[str(name_key).strip()] = inspector_id
            except Exception:
                inspector_name_to_id = {}

            # 【高速化】固定検査員判定のローカルキャッシュ（本関数内では不変を前提）
            fixed_names_cache: Dict[Tuple[str, str], Set[str]] = {}
            on_vacation_cache: Dict[str, bool] = {}

            def _is_on_vacation_cached(inspector_name: str) -> bool:
                cached = on_vacation_cache.get(inspector_name)
                if cached is not None:
                    return cached
                value = self.is_inspector_on_vacation(inspector_name)
                on_vacation_cache[inspector_name] = value
                return value

            def _is_fixed_inspector_for_lot_cached(
                product_number_val: Any,
                process_name_context_val: Optional[str],
                inspector_name_val: Any,
            ) -> bool:
                if not product_number_val or pd.isna(product_number_val):
                    return False
                if not inspector_name_val or pd.isna(inspector_name_val):
                    return False

                product_number_str = str(product_number_val).strip()
                inspector_name_str = str(inspector_name_val).strip()
                if not product_number_str or not inspector_name_str:
                    return False

                if '(' in inspector_name_str:
                    inspector_name_str = inspector_name_str.split('(')[0].strip()
                if not inspector_name_str:
                    return False

                if _is_on_vacation_cached(inspector_name_str):
                    return False

                process_key = str(process_name_context_val).strip() if process_name_context_val is not None else ''
                cache_key = (product_number_str, process_key)
                fixed_names = fixed_names_cache.get(cache_key)
                if fixed_names is None:
                    fixed_names = set(self._collect_fixed_inspector_names(product_number_str, process_key))
                    fixed_names_cache[cache_key] = fixed_names
                return inspector_name_str in fixed_names

            # 【高速化】スキルマスタに存在する品番セット（新規品判定）
            skill_product_values: Set[Any] = set()
            skill_allowed_inspectors_by_product: Dict[Any, Set[Any]] = {}
            try:
                if skill_master_df is not None and not skill_master_df.empty:
                    skill_product_values = set(skill_master_df.iloc[:, 0].dropna().tolist())
                    # 品番 -> 検査員コード集合（スキルチェック高速化）
                    for row in skill_master_df.itertuples(index=False):
                        if len(row) < 2:
                            continue
                        prod = row[0]
                        code = row[1]
                        if pd.isna(prod) or pd.isna(code):
                            continue
                        skill_allowed_inspectors_by_product.setdefault(prod, set()).add(code)
            except Exception:
                skill_product_values = set()
                skill_allowed_inspectors_by_product = {}
             
            self.log_message("全体最適化フェーズ0: result_dfから実際の割り当てを再計算")
            perf_logger = loguru_logger.bind(channel="PERF")
            _t_perf_phase0 = perf_counter()
             
            # 最優先ルール: 出荷予定日の古い順にソート（処理の最初に必ず実行）
            # 出荷予定日を変換（当日洗浄品は文字列として保持）
            result_df['出荷予定日'] = result_df['出荷予定日'].apply(self._convert_shipping_date)
            
            current_date = pd.Timestamp.now().date()
            
            # ソート用のキー関数: 新しい優先順位に従う
            def get_next_business_day(date_val):
                """翌営業日を取得（金曜日の場合は翌週の月曜日）"""
                weekday = date_val.weekday()  # 0=月曜日, 4=金曜日
                if weekday == 4:  # 金曜日
                    return date_val + timedelta(days=3)  # 翌週の月曜日
                else:
                    return date_val + timedelta(days=1)  # 翌日
            
            next_business_day = get_next_business_day(current_date)
            
            def sort_key(val):
                if pd.isna(val):
                    return (5, None)  # 最後に
                val_str = str(val).strip()
                
                # 1. 当日の日付（優先度0）
                try:
                    date_val = pd.to_datetime(val, errors='coerce')
                    if pd.notna(date_val):
                        date_date = date_val.date()
                        if date_date == current_date:
                            return (0, date_val)
                except:
                    pass
                
                # 2. 当日洗浄上がり品（優先度1）
                if (val_str == "当日洗浄上がり品" or 
                    val_str == "当日洗浄品" or
                    "当日洗浄" in val_str):
                    return (1, val_str)
                
                # 3. 先行検査品（優先度2）
                if (val_str == "先行検査" or
                    val_str == "当日先行検査"):
                    return (2, val_str)
                
                # 4. 翌日または翌営業日（優先度3）
                try:
                    date_val = pd.to_datetime(val, errors='coerce')
                    if pd.notna(date_val):
                        date_date = date_val.date()
                        if date_date == next_business_day:
                            return (3, date_val)
                except:
                    pass
                
                # 5. それ以降の日付（優先度4）
                try:
                    date_val = pd.to_datetime(val, errors='coerce')
                    if pd.notna(date_val):
                        return (4, date_val)
                except:
                    pass
                
                return (5, val_str)  # その他文字列
            
            # ソートキーを追加
            result_df['_sort_key'] = result_df['出荷予定日'].apply(sort_key)
            result_df = result_df.sort_values('_sort_key', na_position='last').reset_index(drop=True)
            result_df = result_df.drop(columns=['_sort_key'], errors='ignore')
            
            self.log_message("最適化処理開始前に出荷予定日の古い順でソートしました（最優先ルール）")
            
            # result_dfから実際の割り当てを読み取って、履歴を再計算（正確な状態を把握）
            # まず、分割検査時間を実際の検査員数で再計算
            result_cols_pre = {col: idx for idx, col in enumerate(result_df.columns)}
            for row_idx, row in enumerate(result_df.itertuples(index=False)):
                index = result_df.index[row_idx]
                inspection_time = row[result_cols_pre.get('検査時間', -1)] if '検査時間' in result_cols_pre else 0.0
                if inspection_time == -1 or inspection_time == 0:
                    continue
                
                # 実際に割り当てられた検査員数をカウント
                actual_inspector_count = 0
                for i in range(1, 6):
                    inspector_col = f'検査員{i}'
                    inspector_col_idx = result_cols_pre.get(inspector_col, -1)
                    if inspector_col_idx != -1:
                        inspector_name = row[inspector_col_idx]
                        if pd.notna(inspector_name) and str(inspector_name).strip() != '':
                            actual_inspector_count += 1
                
                # 分割検査時間を再計算: 検査時間 ÷ 実際の分割した検査人数
                if actual_inspector_count > 0:
                    actual_divided_time = inspection_time / actual_inspector_count
                    result_df.at[index, '分割検査時間'] = round(actual_divided_time, 1)
            
            self.inspector_daily_assignments = {}
            self.inspector_work_hours = {}
            self.inspector_product_hours = {}
            
            # 列名のインデックスマップを作成（itertuples用）
            result_cols = {col: idx for idx, col in enumerate(result_df.columns)}
            
            for row_idx, row in enumerate(result_df.itertuples(index=False)):
                index = result_df.index[row_idx]
                product_number = row[result_cols['品番']]
                divided_time = row[result_cols.get('分割検査時間', -1)] if '分割検査時間' in result_cols else 0.0
                if divided_time == -1:
                    divided_time = 0.0
                inspection_time = row[result_cols.get('検査時間', -1)] if '検査時間' in result_cols else divided_time
                if inspection_time == -1:
                    inspection_time = divided_time
                
                # 各検査員の割り当てを確認
                for i in range(1, 6):
                    inspector_col = f'検査員{i}'
                    inspector_col_idx = result_cols.get(inspector_col, -1)
                    if inspector_col_idx != -1:
                        inspector_name_raw = row[inspector_col_idx]
                        if pd.notna(inspector_name_raw) and str(inspector_name_raw).strip() != '':
                            inspector_name = str(inspector_name_raw).strip()
                            # スキル値や(新)を除去
                            if '(' in inspector_name:
                                inspector_name = inspector_name.split('(')[0].strip()
                            
                            if not inspector_name:
                                continue
                            
                            # 検査員コードを取得（辞書 -> フォールバック）
                            inspector_code = inspector_name_to_id.get(inspector_name)
                            if inspector_code is None:
                                inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                if not inspector_info.empty:
                                    inspector_code = inspector_info.iloc[0]['#ID']
                                else:
                                    continue
                                
                                # 履歴を初期化
                                if inspector_code not in self.inspector_daily_assignments:
                                    self.inspector_daily_assignments[inspector_code] = {}
                                if current_date not in self.inspector_daily_assignments[inspector_code]:
                                    self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                                if inspector_code not in self.inspector_work_hours:
                                    self.inspector_work_hours[inspector_code] = 0.0
                                if inspector_code not in self.inspector_product_hours:
                                    self.inspector_product_hours[inspector_code] = {}
                                if product_number not in self.inspector_product_hours[inspector_code]:
                                    self.inspector_product_hours[inspector_code][product_number] = 0.0
                                
                                # 履歴を累積
                                self.inspector_daily_assignments[inspector_code][current_date] += divided_time
                                self.inspector_work_hours[inspector_code] += divided_time
                                self.inspector_product_hours[inspector_code][product_number] += divided_time
                    else:
                        continue
            
            self.log_message("履歴の再計算が完了しました")
            perf_logger.debug(
                "PERF {}: {:.1f} ms",
                "inspector_assignment.optimize.phase0.rebuild_histories",
                (perf_counter() - _t_perf_phase0) * 1000.0,
            )
             
            # 各検査員の最大勤務時間を取得
            _t_perf_max_hours = perf_counter()
            inspector_max_hours = {}
            inspector_cols = {col: idx for idx, col in enumerate(inspector_master_df.columns)}
            for row in inspector_master_df.itertuples(index=False):
                code = row[inspector_cols['#ID']]
                max_hours = self.get_inspector_max_hours(code, inspector_master_df)
                inspector_max_hours[code] = self._apply_work_hours_overrun(max_hours)
            perf_logger.debug(
                "PERF {}: {:.1f} ms",
                "inspector_assignment.optimize.inspector_max_hours",
                (perf_counter() - _t_perf_max_hours) * 1000.0,
            )
            
            # 改善ポイント: フェーズ間スラッシング防止用のタブーリストを初期化
            self.tabu_list = {}
            
            # フェーズ1: 勤務時間超過と同一品番の時間上限超過を検出・是正（繰り返し処理）
            self.log_message(f"全体最適化フェーズ1: 勤務時間超過と同一品番{self.product_limit_hard_threshold:.1f}時間超過の検出と是正を開始")
            _t_perf_phase1_total = perf_counter()
             
            # 【改善】最大繰り返し回数を調整可能にする（環境変数で設定可能）
            try:
                max_iterations = int(os.getenv("PHASE1_MAX_ITERATIONS", "10").strip() or "10")
            except Exception:
                max_iterations = 10  # デフォルトは10回
            max_iterations = max(1, min(max_iterations, 50))  # 1回以上50回以下に制限
            
            # 【追加】収束判定用の変数
            previous_violation_count = None
            convergence_stable_iterations = 0  # 違反件数が変化しない連続回数
            CONVERGENCE_STABLE_THRESHOLD = 2  # 2回連続で違反件数が変化しない場合に収束と判定
            
            # 【追加】フェーズ1の効果測定メトリクス
            phase1_metrics = {
                'iterations': [],
                'violation_counts': [],
                'overworked_counts': [],
                'product_limit_counts': [],
                'resolved_counts': [],
                'convergence_reason': None,
            }
            
            iteration = 0

            # 列名のインデックスマップ（フェーズ1以降で繰り返し使う）
            sorted_cols = {col: idx for idx, col in enumerate(result_df.columns)}
            process_name_idx = sorted_cols.get('現在工程名', -1)
            divided_time_idx = sorted_cols.get('分割検査時間', -1)
            inspection_time_idx = sorted_cols.get('検査時間', -1)
            inspector_col_idxs = [sorted_cols.get(f'検査員{i}', -1) for i in range(1, 6)]
            two_weeks_later = current_date + timedelta(days=14)
            shipping_date_date_cache: Dict[int, Optional[date]] = {}
            violation_date_cache: Dict[int, date] = {}

            def _get_shipping_date_date_cached(row_index: int) -> Optional[date]:
                cached = shipping_date_date_cache.get(row_index, None)
                if row_index in shipping_date_date_cache:
                    return cached

                shipping_date = result_df.at[row_index, '出荷予定日'] if row_index < len(result_df) else None
                if shipping_date is None or pd.isna(shipping_date):
                    shipping_date_date_cache[row_index] = None
                    return None

                try:
                    if isinstance(shipping_date, pd.Timestamp):
                        value = shipping_date.date()
                    elif isinstance(shipping_date, str):
                        parsed = pd.to_datetime(shipping_date, errors='coerce')
                        value = parsed.date() if pd.notna(parsed) else None
                    else:
                        value = shipping_date.date() if hasattr(shipping_date, 'date') else shipping_date
                    shipping_date_date_cache[row_index] = value
                    return value
                except Exception:
                    shipping_date_date_cache[row_index] = None
                    return None
             
            while iteration < max_iterations:
                _t_perf_iter_total = perf_counter()
                iteration += 1
                self.log_message(f"是正処理 イテレーション {iteration}")
                
                # 改善ポイント: タブーリストの更新（古いエントリを削除）
                old_tabu_size = len(self.tabu_list)
                self.tabu_list = {idx: count - 1 for idx, count in self.tabu_list.items() if count > 1}
                removed_count = old_tabu_size - len(self.tabu_list)
                if removed_count > 0:
                    self.log_message(f"タブーリスト更新: {removed_count}件のエントリが期限切れで削除されました（残り: {len(self.tabu_list)}件）", debug=True)
                
                violations_found = False
                overworked_assignments = []
                product_limit_violations = []

                # 出荷予定日の変換とソートは最初に一度だけ実施済み（出荷予定日は最適化中に変化しないため）
                # 以降の処理は常に「出荷予定日順（最優先ルール）」を維持した result_df を前提とする。
                result_df_sorted = result_df
                
                _t_perf_iter_detect = perf_counter()
                for row_idx, row in enumerate(result_df_sorted.itertuples(index=False)):
                    index = result_df_sorted.index[row_idx]
                    product_number = row[sorted_cols['品番']]
                    process_name_context = row[process_name_idx] if process_name_idx != -1 else ''
                    if pd.isna(process_name_context):
                        process_name_context = ''
                    process_name_context_str = str(process_name_context).strip()

                    divided_time = row[divided_time_idx] if divided_time_idx != -1 else 0.0
                    if divided_time == -1:
                        divided_time = 0.0
                    inspection_time = row[inspection_time_idx] if inspection_time_idx != -1 else divided_time
                    if inspection_time == -1:
                        inspection_time = divided_time
                    
                    # 各検査員の割り当てを確認
                    for i, inspector_col_idx in enumerate(inspector_col_idxs, start=1):
                        if inspector_col_idx == -1:
                            continue
                        inspector_name_raw = row[inspector_col_idx]
                        if pd.notna(inspector_name_raw) and str(inspector_name_raw).strip() != '':
                            inspector_name = str(inspector_name_raw).strip()
                            if '(' in inspector_name:
                                inspector_name = inspector_name.split('(')[0].strip()

                            if not inspector_name:
                                continue

                            # 固定検査員が割当済みのロットは、勤務時間/同一品番上限の違反検出・是正対象から除外する
                            if _is_fixed_inspector_for_lot_cached(product_number, process_name_context_str, inspector_name):
                                # 【追加】保護の履歴追跡
                                self.fixed_inspector_protection_metrics['total_protections'] += 1
                                self.fixed_inspector_protection_metrics['protection_by_phase']['phase1_violation_detection'] = \
                                    self.fixed_inspector_protection_metrics['protection_by_phase'].get('phase1_violation_detection', 0) + 1
                                self.fixed_inspector_protection_metrics['protection_by_reason']['violation_detection_exclusion'] = \
                                    self.fixed_inspector_protection_metrics['protection_by_reason'].get('violation_detection_exclusion', 0) + 1
                                self.fixed_inspector_protection_metrics['protected_lots'].add(index)
                                self.fixed_inspector_protection_metrics['protected_inspectors'].add(inspector_name)
                                
                                protection_history_entry = {
                                    'lot_index': index,
                                    'product_number': product_number,
                                    'inspector_name': inspector_name,
                                    'phase': 'phase1',
                                    'reason': 'violation_detection_exclusion',
                                    'process_name': process_name_context_str,
                                }
                                self.fixed_inspector_protection_metrics['protection_history'].append(protection_history_entry)
                                if len(self.fixed_inspector_protection_metrics['protection_history']) > 100:
                                    self.fixed_inspector_protection_metrics['protection_history'] = \
                                        self.fixed_inspector_protection_metrics['protection_history'][-100:]
                                continue

                            # 検査員コードを取得（辞書 -> フォールバック）
                            inspector_code = inspector_name_to_id.get(inspector_name)
                            if inspector_code is None:
                                inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                if not inspector_info.empty:
                                    inspector_code = inspector_info.iloc[0]['#ID']
                                else:
                                    continue
                                     
                                    # 現在の履歴を取得
                                    daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
                                    max_hours = inspector_max_hours.get(inspector_code, 8.0)
                                    product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                                    
                                    # 【修正】勤務時間超過をチェック（10%超過まで許容）
                                    # 10%超過を考慮した許容最大勤務時間を計算
                                    allowed_max_hours = max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                    # 実際の勤務時間が10%超過を超えている場合のみ違反とする
                                    if daily_hours > allowed_max_hours - WORK_HOURS_BUFFER:
                                        excess = daily_hours - allowed_max_hours
                                        overworked_assignments.append((index, inspector_code, inspector_name, excess, divided_time, product_number, inspection_time, i))
                                        violations_found = True
                                        self.log_message(f"⚠️ 勤務時間超過: 検査員 '{inspector_name}' (コード: {inspector_code}) {daily_hours:.1f}h > {allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h (超過: {excess:.1f}h, 品番: {product_number}, ロットインデックス: {index})", level='warning')
                                    
                                    # 改善ポイント: 最適化フェーズでの設定時間上限チェック（厳格）
                                    if product_hours > self.product_limit_hard_threshold:
                                        excess = product_hours - self.product_limit_hard_threshold
                                        product_limit_violations.append((index, inspector_code, inspector_name, excess, divided_time, product_number, inspection_time, i))
                                        violations_found = True
                                        self.log_message(f"⚠️ 同一品番{self.product_limit_hard_threshold:.1f}時間超過: 検査員 '{inspector_name}' (コード: {inspector_code}) 品番 {product_number} {product_hours:.1f}h > {self.product_limit_hard_threshold:.1f}h (超過: {excess:.1f}h, ロットインデックス: {index})", level='warning')
                perf_logger.debug(
                    "PERF {}: {:.1f} ms",
                    f"inspector_assignment.optimize.phase1.iter{iteration}.detect_violations",
                    (perf_counter() - _t_perf_iter_detect) * 1000.0,
                )
                
                # 【追加】違反件数を記録
                current_violation_count = len(overworked_assignments) + len(product_limit_violations)
                phase1_metrics['iterations'].append(iteration)
                phase1_metrics['violation_counts'].append(current_violation_count)
                phase1_metrics['overworked_counts'].append(len(overworked_assignments))
                phase1_metrics['product_limit_counts'].append(len(product_limit_violations))
                
                # 【追加】収束判定：違反件数が変化しない場合
                if previous_violation_count is not None:
                    if current_violation_count == previous_violation_count:
                        convergence_stable_iterations += 1
                        if convergence_stable_iterations >= CONVERGENCE_STABLE_THRESHOLD:
                            phase1_metrics['convergence_reason'] = 'stable_violations'
                            self.log_message(
                                f"フェーズ1: 収束判定 - 違反件数が{CONVERGENCE_STABLE_THRESHOLD}回連続で変化しませんでした "
                                f"（{iteration}回目のイテレーション、違反件数: {current_violation_count}件）",
                                level='info',
                            )
                            # 【追加】収束しない原因を分析
                            if current_violation_count > 0:
                                self._analyze_phase1_non_convergence(
                                    overworked_assignments, product_limit_violations,
                                    result_df_sorted, inspector_master_df, inspector_max_hours
                                )
                            result_df = result_df_sorted
                            break
                    else:
                        convergence_stable_iterations = 0
                
                previous_violation_count = current_violation_count
                
                # 違反が見つからない場合は終了
                if not violations_found:
                    phase1_metrics['convergence_reason'] = 'no_violations'
                    phase1_metrics['resolved_counts'].append(current_violation_count)
                    self.log_message(f"全てのルール違反が解消されました（{iteration}回目のイテレーションで完了）")
                    result_df = result_df_sorted
                    perf_logger.debug(
                        "PERF {}: {:.1f} ms",
                        f"inspector_assignment.optimize.phase1.iter{iteration}.total",
                        (perf_counter() - _t_perf_iter_total) * 1000.0,
                    )
                    break
                
                # 違反を是正（当日洗浄上がり品を優先）
                all_violations = overworked_assignments + product_limit_violations
                # 重複を除去（同じロットの複数の違反を1つにまとめる）
                unique_violations = {}
                for violation in all_violations:
                    index = violation[0]
                    if index not in unique_violations:
                        unique_violations[index] = violation
                    else:
                        # 既存の違反より新しい違反の方が重要度が高い場合は置き換え
                        existing = unique_violations[index]
                        if violation[3] > existing[3]:  # excess値が大きい方
                            unique_violations[index] = violation
                
                # 【改善】当日洗浄上がり品の違反を分離
                same_day_cleaning_violations = []
                other_violations = []
                
                for violation in unique_violations.values():
                    violation_index = violation[0]
                    if violation_index < len(result_df_sorted):
                        shipping_date_raw = result_df_sorted.at[violation_index, '出荷予定日']
                        shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                        is_same_day_cleaning = (
                            shipping_date_str == "当日洗浄上がり品" or
                            shipping_date_str == "当日洗浄品" or
                            "当日洗浄" in shipping_date_str
                        )
                        
                        if is_same_day_cleaning:
                            same_day_cleaning_violations.append(violation)
                        else:
                            other_violations.append(violation)
                    else:
                        other_violations.append(violation)
                
                def violation_priority(violation):
                    """違反の優先順位を計算（ソート用）"""
                    try:
                        violation_index = violation[0]
                        if violation_index < len(result_df_sorted):
                            violation_date = violation_date_cache.get(violation_index)
                            if violation_date is None:
                                shipping_date_raw = result_df_sorted.at[violation_index, '出荷予定日']
                                if pd.notna(shipping_date_raw):
                                    if isinstance(shipping_date_raw, pd.Timestamp):
                                        violation_date = shipping_date_raw.date()
                                    elif isinstance(shipping_date_raw, str):
                                        parsed = pd.to_datetime(shipping_date_raw, errors='coerce')
                                        violation_date = parsed.date() if pd.notna(parsed) else pd.Timestamp.max.date()
                                    else:
                                        try:
                                            violation_date = shipping_date_raw.date() if hasattr(shipping_date_raw, 'date') else shipping_date_raw
                                        except Exception:
                                            violation_date = pd.Timestamp.max.date()
                                else:
                                    violation_date = pd.Timestamp.max.date()
                                violation_date_cache[violation_index] = violation_date
                        else:
                            violation_date = pd.Timestamp.min.date()
                        inspector_code = violation[1]
                        product_number = violation[5]
                        relaxed_flag = (inspector_code, product_number) in self.relaxed_product_limit_assignments
                        return (0 if relaxed_flag else 1, violation_date, violation_index)
                    except Exception as e:
                        # エラー発生時は優先度を最低にしてソートを継続
                        self.log_message(f"violation_priority計算エラー: {str(e)} (ロットインデックス: {violation[0]})", level='warning')
                        return (2, pd.Timestamp.max.date(), violation[0])
                
                # 当日洗浄上がり品の違反を優先的に処理
                same_day_cleaning_violations_sorted = sorted(same_day_cleaning_violations, key=violation_priority)
                other_violations_sorted = sorted(other_violations, key=violation_priority)
                violations_to_fix = same_day_cleaning_violations_sorted + other_violations_sorted
                
                if same_day_cleaning_violations:
                    self.log_message(f"当日洗浄上がり品の違反 {len(same_day_cleaning_violations)}件を優先的に処理します", level='info')

                try:
                    sorted_violations = violations_to_fix
                except Exception as e:
                    # ソートエラーが発生した場合は、インデックス順でソート
                    self.log_message(f"違反のソート中にエラーが発生しました: {str(e)}。インデックス順で処理します。", level='warning')
                    sorted_violations = sorted(unique_violations.values(), key=lambda v: v[0])
                
                self.log_message(f"違反ロット数: {len(sorted_violations)}件を是正します")
                
                # 各違反を是正
                _t_perf_iter_fix = perf_counter()
                fixed_any = False
                fixed_indices = set()
                for violation in sorted_violations:
                    index, inspector_code, inspector_name, excess, divided_time, product_number, inspection_time, inspector_col_num = violation
                    # 既に是正済みのロットはスキップ
                    if index in fixed_indices:
                        continue
                    
                    # 改善ポイント: フェーズ間スラッシング防止 - タブーリストに含まれるロットはスキップ
                    if index in self.tabu_list:
                        # 【追加】タブーリストによるスキップを記録
                        self.tabu_list_metrics['total_skips'] += 1
                        continue

                    # 登録済み品番の先行検査×固定検査員ロットは最適化フェーズで動かさない（固定維持）
                    if self._is_locked_fixed_preinspection_lot(result_df_sorted, index):
                        # 【追加】保護の履歴追跡
                        self.fixed_inspector_protection_metrics['total_protections'] += 1
                        self.fixed_inspector_protection_metrics['protection_by_phase']['phase1_reassignment_skip'] = \
                            self.fixed_inspector_protection_metrics['protection_by_phase'].get('phase1_reassignment_skip', 0) + 1
                        self.fixed_inspector_protection_metrics['protection_by_reason']['reassignment_skip'] = \
                            self.fixed_inspector_protection_metrics['protection_by_reason'].get('reassignment_skip', 0) + 1
                        self.fixed_inspector_protection_metrics['protected_lots'].add(index)
                        
                        # 固定検査員名を取得
                        fixed_inspector_name = None
                        for i in range(1, 6):
                            inspector_col = f'検査員{i}'
                            if inspector_col in result_df_sorted.columns:
                                inspector_value = result_df_sorted.at[index, inspector_col]
                                if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                                    inspector_name = str(inspector_value).strip()
                                    if '(' in inspector_name:
                                        inspector_name = inspector_name.split('(')[0].strip()
                                    if inspector_name:
                                        product_number = result_df_sorted.at[index, '品番']
                                        process_name = result_df_sorted.at[index, '現在工程名'] if '現在工程名' in result_df_sorted.columns else ''
                                        if self._is_fixed_inspector_for_lot(product_number, process_name, inspector_name):
                                            fixed_inspector_name = inspector_name
                                            self.fixed_inspector_protection_metrics['protected_inspectors'].add(inspector_name)
                                            break
                        
                        protection_history_entry = {
                            'lot_index': index,
                            'product_number': result_df_sorted.at[index, '品番'] if '品番' in result_df_sorted.columns else '',
                            'inspector_name': fixed_inspector_name or 'unknown',
                            'phase': 'phase1',
                            'reason': 'reassignment_skip',
                            'process_name': str(result_df_sorted.at[index, '現在工程名']).strip() if '現在工程名' in result_df_sorted.columns else '',
                        }
                        self.fixed_inspector_protection_metrics['protection_history'].append(protection_history_entry)
                        if len(self.fixed_inspector_protection_metrics['protection_history']) > 100:
                            self.fixed_inspector_protection_metrics['protection_history'] = \
                                self.fixed_inspector_protection_metrics['protection_history'][-100:]
                        continue
                    
                    # 改善ポイント: 新規品（出荷予定日指定日数以内）は保護対象として移動対象外にする
                    shipping_date_date = _get_shipping_date_date_cached(index)
                    protection_threshold_date = current_date + timedelta(days=NEW_PRODUCT_PROTECTION_DAYS)
                    is_within_protection_period = bool(
                        shipping_date_date is not None and 
                        shipping_date_date <= protection_threshold_date
                    )

                    is_new_product = product_number not in skill_product_values
                    if NEW_PRODUCT_PROTECTION_ENABLED and is_new_product and is_within_protection_period:
                        # 【追加】保護の履歴追跡
                        self.new_product_protection_metrics['total_protections'] += 1
                        self.new_product_protection_metrics['protection_by_phase']['phase1'] = \
                            self.new_product_protection_metrics['protection_by_phase'].get('phase1', 0) + 1
                        self.new_product_protection_metrics['protected_lots'].add(index)
                        
                        protection_history_entry = {
                            'lot_index': index,
                            'product_number': product_number,
                            'violation_type': 'reassignment_prevention',
                            'phase': 'phase1',
                            'shipping_date': shipping_date_date,
                            'protection_days': NEW_PRODUCT_PROTECTION_DAYS,
                        }
                        self.new_product_protection_metrics['protection_history'].append(protection_history_entry)
                        if len(self.new_product_protection_metrics['protection_history']) > 100:
                            self.new_product_protection_metrics['protection_history'] = \
                                self.new_product_protection_metrics['protection_history'][-100:]
                        
                        self.log_message(
                            f"⚠️ 新規品（出荷予定日{NEW_PRODUCT_PROTECTION_DAYS}日以内）のため保護: "
                            f"ロットインデックス {index} (品番: {product_number}) - 違反があっても再割当てをスキップします",
                            level='warning'
                        )
                        continue
                    
                    self.log_message(f"違反是正を試みます: ロットインデックス {index}, 検査員 {inspector_name}, 品番 {product_number}")
                    fixed = self.fix_single_violation(
                        index, inspector_code, inspector_name, divided_time, product_number, inspection_time, inspector_col_num,
                        result_df_sorted, inspector_master_df, skill_master_df, inspector_max_hours, current_date, show_skill_values
                    )
                    if fixed:
                        fixed_any = True
                        fixed_indices.add(index)
                        # 改善ポイント: フェーズ間スラッシング防止 - 再配置されたロットをタブーリストに追加
                        # 【追加】再配置回数を記録
                        self.tabu_list_metrics['lot_reassignment_counts'][index] = self.tabu_list_metrics['lot_reassignment_counts'].get(index, 0) + 1
                        # 【追加】スラッシング検出（同じロットが複数回再配置されている場合）
                        if self.tabu_list_metrics['lot_reassignment_counts'][index] >= 3:
                            if index not in self.tabu_list_metrics['thrashing_detected']:
                                self.tabu_list_metrics['thrashing_detected'].append(index)
                                self.log_message(
                                    f"⚠️ スラッシング検出: ロットインデックス {index} が{self.tabu_list_metrics['lot_reassignment_counts'][index]}回再配置されています",
                                    level='warning',
                                )
                        self.tabu_list[index] = TABU_LIST_MAX_ITERATIONS
                        self.tabu_list_metrics['total_additions'] += 1
                        self.log_message(f"✅ 違反是正成功: ロットインデックス {index} (タブーリストに追加)")
                
                # 【追加】最初の是正処理での解決件数を記録
                if fixed_any:
                    phase1_metrics['resolved_counts'].append(len(fixed_indices))
                else:
                    phase1_metrics['resolved_counts'].append(0)
                
                if not fixed_any and len(sorted_violations) > 0:
                    # 是正できなかった違反がある場合は、出荷予定日が古いロットを優先的に再割り当てを試みる
                    unresolved_violations = [v for v in sorted_violations if v[0] not in fixed_indices]
                    if unresolved_violations:
                        self.log_message(f"⚠️ 是正できなかった違反が {len(unresolved_violations)}件あります", level='warning')
                        
                        # 出荷予定日が古いロットを優先的に再割り当てを試みる
                        violations_with_date = []
                        for violation in unresolved_violations:
                            index = violation[0]
                            if index < len(result_df_sorted):
                                shipping_date_raw = result_df_sorted.at[index, '出荷予定日']
                                if pd.notna(shipping_date_raw):
                                    # pd.Timestampをdatetime.dateに変換して比較エラーを防ぐ
                                    if isinstance(shipping_date_raw, pd.Timestamp):
                                        shipping_date = shipping_date_raw.date()
                                    elif isinstance(shipping_date_raw, str):
                                        shipping_date_str = str(shipping_date_raw).strip()
                                        # 「当日洗浄上がり品」「先行検査」などの文字列の場合は最優先として扱う
                                        if (shipping_date_str == "当日洗浄上がり品" or
                                            shipping_date_str == "当日洗浄品" or
                                            "当日洗浄" in shipping_date_str or
                                            shipping_date_str == "先行検査" or
                                            shipping_date_str == "当日先行検査"):
                                            shipping_date = pd.Timestamp.min.date()  # 最優先として扱う
                                        else:
                                            # 日付文字列の場合は変換を試みる
                                            try:
                                                shipping_date_parsed = pd.to_datetime(shipping_date_raw, errors='coerce')
                                                if pd.notna(shipping_date_parsed):
                                                    shipping_date = shipping_date_parsed.date()
                                                else:
                                                    shipping_date = pd.Timestamp.max.date()
                                            except Exception as e:
                                                logger.debug(f"出荷日の変換でエラーが発生しました（デフォルト値を使用）: {e}")
                                                shipping_date = pd.Timestamp.max.date()
                                    else:
                                        # その他の型（datetime.date等）の場合はそのまま使用
                                        try:
                                            shipping_date = shipping_date_raw.date() if hasattr(shipping_date_raw, 'date') else shipping_date_raw
                                        except Exception as e:
                                            logger.debug(f"出荷日の変換でエラーが発生しました（デフォルト値を使用）: {e}")
                                            shipping_date = pd.Timestamp.max.date()
                                else:
                                    shipping_date = pd.Timestamp.max.date()
                            else:
                                shipping_date = pd.Timestamp.max.date()
                            violations_with_date.append((violation, shipping_date))
                        
                        # 出荷予定日の古い順にソート（既にソートされているが、念のため）
                        violations_with_date.sort(key=lambda x: self._normalize_shipping_date(x[1]))
                        
                        # 出荷予定日が古いロットから順に再割り当てを試みる
                        re_resolved_count = 0
                        for violation, shipping_date in violations_with_date:
                            index, inspector_code, inspector_name, excess, divided_time, product_number, inspection_time, inspector_col_num = violation
                            
                            self.log_message(f"⚠️ ロットインデックス {index} (品番: {product_number}, 出荷予定日: {shipping_date}) の再是正を試みます")
                            
                            # 再是正を試みる
                            re_fixed = self.fix_single_violation(
                                index, inspector_code, inspector_name, divided_time, product_number, inspection_time, inspector_col_num,
                                result_df_sorted, inspector_master_df, skill_master_df, inspector_max_hours, current_date, show_skill_values
                            )
                            
                            if re_fixed:
                                re_resolved_count += 1
                                fixed_indices.add(index)
                                self.log_message(f"✅ ロットインデックス {index} の再是正に成功しました")
                            else:
                                # 再是正できなかった場合は未割当にする
                                self.clear_assignment(result_df_sorted, index)
                                self.log_message(f"⚠️ ロットインデックス {index} (品番: {product_number}) を未割当にしました")
                        
                        self.log_message(f"再是正結果: {re_resolved_count}件是正、{len(unresolved_violations) - re_resolved_count}件未割当")
                    
                    # 【追加】解決件数を記録
                    total_resolved = len(fixed_indices) + re_resolved_count
                    phase1_metrics['resolved_counts'].append(total_resolved)
                    
                    result_df = result_df_sorted
                    # 【改善】分割検査時間を実際の検査員数で再計算（是正できなかった違反の処理後）
                    sorted_cols_recalc_pre = {col: idx for idx, col in enumerate(result_df_sorted.columns)}
                    for row_idx, row in enumerate(result_df_sorted.itertuples(index=False)):
                        index = result_df_sorted.index[row_idx]
                        inspection_time = row[sorted_cols_recalc_pre.get('検査時間', -1)] if '検査時間' in sorted_cols_recalc_pre else 0.0
                        if inspection_time == -1 or inspection_time == 0:
                            continue
                        
                        # 実際に割り当てられた検査員数をカウント
                        actual_inspector_count = 0
                        for i in range(1, 6):
                            inspector_col = f'検査員{i}'
                            inspector_col_idx = sorted_cols_recalc_pre.get(inspector_col, -1)
                            if inspector_col_idx != -1:
                                inspector_name = row[inspector_col_idx]
                                if pd.notna(inspector_name) and str(inspector_name).strip() != '':
                                    actual_inspector_count += 1
                        
                        # 分割検査時間を再計算: 検査時間 ÷ 実際の分割した検査人数
                        if actual_inspector_count > 0:
                            actual_divided_time = inspection_time / actual_inspector_count
                            result_df_sorted.at[index, '分割検査時間'] = round(actual_divided_time, 1)
                    # 履歴再計算は、イテレーションの最後に1回だけ実行する（重複を避ける）
                    # continue は削除し、履歴再計算をイテレーションの最後に統合

                perf_logger.debug(
                    "PERF {}: {:.1f} ms",
                    f"inspector_assignment.optimize.phase1.iter{iteration}.fix_violations",
                    (perf_counter() - _t_perf_iter_fix) * 1000.0,
                )
                
                # 【改善】履歴を再計算（是正後の状態を反映）- 是正できなかった違反の処理後も含めて1回だけ実行
                _t_perf_iter_rebuild = perf_counter()
                self.inspector_daily_assignments = {}
                self.inspector_work_hours = {}
                self.inspector_product_hours = {}
                
                sorted_cols_recalc2 = {col: idx for idx, col in enumerate(result_df_sorted.columns)}
                for row_idx, row in enumerate(result_df_sorted.itertuples(index=False)):
                    index = result_df_sorted.index[row_idx]
                    product_number = row[sorted_cols_recalc2['品番']]
                    divided_time = row[sorted_cols_recalc2.get('分割検査時間', -1)] if '分割検査時間' in sorted_cols_recalc2 else 0.0
                    if divided_time == -1:
                        divided_time = 0.0
                    inspection_time = row[sorted_cols_recalc2.get('検査時間', -1)] if '検査時間' in sorted_cols_recalc2 else divided_time
                    if inspection_time == -1:
                        inspection_time = divided_time
                    
                    for i in range(1, 6):
                        inspector_col = f'検査員{i}'
                        inspector_col_idx = sorted_cols_recalc2.get(inspector_col, -1)
                        if inspector_col_idx != -1:
                            inspector_name_raw = row[inspector_col_idx]
                            if pd.notna(inspector_name_raw) and str(inspector_name_raw).strip() != '':
                                inspector_name = str(inspector_name_raw).strip()
                                if '(' in inspector_name:
                                    inspector_name = inspector_name.split('(')[0].strip()
                                if not inspector_name:
                                    continue

                                inspector_code = inspector_name_to_id.get(inspector_name)
                                if inspector_code is None:
                                    inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                    if not inspector_info.empty:
                                        inspector_code = inspector_info.iloc[0]['#ID']
                                    else:
                                        continue
                                    
                                    if inspector_code not in self.inspector_daily_assignments:
                                        self.inspector_daily_assignments[inspector_code] = {}
                                    if current_date not in self.inspector_daily_assignments[inspector_code]:
                                        self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                                    if inspector_code not in self.inspector_work_hours:
                                        self.inspector_work_hours[inspector_code] = 0.0
                                    if inspector_code not in self.inspector_product_hours:
                                        self.inspector_product_hours[inspector_code] = {}
                                    if product_number not in self.inspector_product_hours[inspector_code]:
                                        self.inspector_product_hours[inspector_code][product_number] = 0.0
                                    
                                    self.inspector_daily_assignments[inspector_code][current_date] += divided_time
                                    self.inspector_work_hours[inspector_code] += divided_time
                                    self.inspector_product_hours[inspector_code][product_number] += divided_time
                
                result_df = result_df_sorted
                perf_logger.debug(
                    "PERF {}: {:.1f} ms",
                    f"inspector_assignment.optimize.phase1.iter{iteration}.rebuild_histories",
                    (perf_counter() - _t_perf_iter_rebuild) * 1000.0,
                )
                perf_logger.debug(
                    "PERF {}: {:.1f} ms",
                    f"inspector_assignment.optimize.phase1.iter{iteration}.total",
                    (perf_counter() - _t_perf_iter_total) * 1000.0,
                )

            # 【追加】最大繰り返し回数に達した場合の処理
            if iteration >= max_iterations:
                phase1_metrics['convergence_reason'] = 'max_iterations'
                self.log_message(
                    f"フェーズ1: 最大繰り返し回数（{max_iterations}回）に達しました。"
                    f"最終違反件数: {len(overworked_assignments) + len(product_limit_violations)}件",
                    level='warning',
                )
                # 【追加】収束しない原因を分析
                if violations_found:
                    self._analyze_phase1_non_convergence(
                        overworked_assignments, product_limit_violations,
                        result_df_sorted, inspector_master_df, inspector_max_hours
                    )
            
            # 【追加】フェーズ1の効果測定メトリクスを出力
            if phase1_metrics['iterations']:
                total_violations_resolved = phase1_metrics['violation_counts'][0] - phase1_metrics['violation_counts'][-1] if len(phase1_metrics['violation_counts']) > 1 else 0
                self.log_message(
                    f"フェーズ1効果測定メトリクス: "
                    f"実行イテレーション数={len(phase1_metrics['iterations'])}, "
                    f"初期違反件数={phase1_metrics['violation_counts'][0] if phase1_metrics['violation_counts'] else 0}, "
                    f"最終違反件数={phase1_metrics['violation_counts'][-1] if phase1_metrics['violation_counts'] else 0}, "
                    f"解消した違反件数={total_violations_resolved}, "
                    f"収束理由={phase1_metrics['convergence_reason']}, "
                    f"各イテレーションの違反件数={phase1_metrics['violation_counts']}",
                    debug=True,
                )
            
            # 【追加】タブーリストの効果測定メトリクスを出力
            if self.tabu_list_metrics['total_additions'] > 0 or self.tabu_list_metrics['total_skips'] > 0:
                thrashing_count = len(self.tabu_list_metrics['thrashing_detected'])
                avg_reassignments = (
                    sum(self.tabu_list_metrics['lot_reassignment_counts'].values()) / len(self.tabu_list_metrics['lot_reassignment_counts'])
                    if self.tabu_list_metrics['lot_reassignment_counts'] else 0
                )
                max_reassignments = (
                    max(self.tabu_list_metrics['lot_reassignment_counts'].values())
                    if self.tabu_list_metrics['lot_reassignment_counts'] else 0
                )
                self.log_message(
                    f"タブーリスト効果測定メトリクス: "
                    f"追加回数={self.tabu_list_metrics['total_additions']}, "
                    f"スキップ回数={self.tabu_list_metrics['total_skips']}, "
                    f"スキップ率={self.tabu_list_metrics['total_skips'] / max(1, self.tabu_list_metrics['total_additions']) * 100:.1f}%, "
                    f"スラッシング検出ロット数={thrashing_count}, "
                    f"平均再配置回数={avg_reassignments:.2f}, "
                    f"最大再配置回数={max_reassignments}, "
                    f"現在のタブーリストサイズ={len(self.tabu_list)}",
                    debug=True,
                )
                if thrashing_count > 0:
                    self.log_message(
                        f"スラッシング検出ロット: {self.tabu_list_metrics['thrashing_detected'][:10]}",  # 最初の10件のみ
                        level='warning',
                    )
            
            # 【追加】同一品番の割当回数制約の効果測定メトリクスを出力
            if self.product_assignment_metrics['total_checks'] > 0:
                max_assignments_reached_count = sum(
                    len(products) for products in self.product_assignment_metrics['max_assignments_reached'].values()
                )
                self.log_message(
                    f"同一品番の割当回数制約効果測定メトリクス: "
                    f"制約チェック回数={self.product_assignment_metrics['total_checks']}, "
                    f"制約違反回数={self.product_assignment_metrics['constraint_violations']}, "
                    f"緩和条件適用回数={self.product_assignment_metrics['relaxed_assignments']}, "
                    f"最大割当回数到達数={max_assignments_reached_count}, "
                    f"通常上限={MAX_ASSIGNMENTS_PER_PRODUCT}回, "
                    f"緩和上限={MAX_ASSIGNMENTS_PER_PRODUCT_RELAXED}回",
                    debug=True,
                )
                if self.product_assignment_metrics['relaxed_assignments'] > 0:
                    self.log_message(
                        f"緩和条件適用の詳細（最初の5件）: {self.product_assignment_metrics['relaxation_reasons'][:5]}",
                        debug=True,
                    )
            
            # 【追加】新規品の保護条件の効果測定メトリクスを出力
            if self.new_product_protection_metrics['total_protections'] > 0:
                self.log_message(
                    f"新規品の保護条件効果測定メトリクス: "
                    f"保護有効={self.new_product_protection_metrics['protection_enabled']}, "
                    f"保護期間={self.new_product_protection_metrics['protection_days']}日, "
                    f"総保護回数={self.new_product_protection_metrics['total_protections']}, "
                    f"違反タイプ別={dict(self.new_product_protection_metrics['protection_by_violation_type'])}, "
                    f"フェーズ別={dict(self.new_product_protection_metrics['protection_by_phase'])}, "
                    f"保護ロット数={len(self.new_product_protection_metrics['protected_lots'])}, "
                    f"保護履歴件数={len(self.new_product_protection_metrics['protection_history'])}",
                    debug=True,
                )
                if self.new_product_protection_metrics['protection_history']:
                    # 最初の5件の履歴を出力
                    self.log_message(
                        f"保護履歴（最初の5件）: {self.new_product_protection_metrics['protection_history'][:5]}",
                        debug=True,
                    )
            
            # 【追加】固定検査員の保護条件の効果測定メトリクスを出力
            if self.fixed_inspector_protection_metrics['total_protections'] > 0:
                self.log_message(
                    f"固定検査員の保護条件効果測定メトリクス: "
                    f"総保護回数={self.fixed_inspector_protection_metrics['total_protections']}, "
                    f"フェーズ別={dict(self.fixed_inspector_protection_metrics['protection_by_phase'])}, "
                    f"保護理由別={dict(self.fixed_inspector_protection_metrics['protection_by_reason'])}, "
                    f"保護ロット数={len(self.fixed_inspector_protection_metrics['protected_lots'])}, "
                    f"保護検査員数={len(self.fixed_inspector_protection_metrics['protected_inspectors'])}, "
                    f"保護履歴件数={len(self.fixed_inspector_protection_metrics['protection_history'])}",
                    debug=True,
                )
                if self.fixed_inspector_protection_metrics['protection_history']:
                    # 最初の5件の履歴を出力
                    self.log_message(
                        f"保護履歴（最初の5件）: {self.fixed_inspector_protection_metrics['protection_history'][:5]}",
                        debug=True,
                    )
            
            perf_logger.debug(
                "PERF {}: {:.1f} ms",
                "inspector_assignment.optimize.phase1.total",
                (perf_counter() - _t_perf_phase1_total) * 1000.0,
            )
            
            # フェーズ1.5: 最終違反チェック（是正が完全に機能したか確認）
            self.log_message("全体最適化フェーズ1.5: 最終違反チェックを開始")
            
            # 最終的な履歴を再計算
            self.inspector_daily_assignments = {}
            self.inspector_work_hours = {}
            self.inspector_product_hours = {}
            
            final_cols = {col: idx for idx, col in enumerate(result_df.columns)}
            for row_idx, row in enumerate(result_df.itertuples(index=False)):
                index = result_df.index[row_idx]
                product_number = row[final_cols['品番']]
                divided_time = row[final_cols.get('分割検査時間', -1)] if '分割検査時間' in final_cols else 0.0
                if divided_time == -1:
                    divided_time = 0.0
                
                for i in range(1, 6):
                    inspector_col = f'検査員{i}'
                    inspector_col_idx = final_cols.get(inspector_col, -1)
                    if inspector_col_idx != -1:
                        inspector_name_raw = row[inspector_col_idx]
                        if pd.notna(inspector_name_raw) and str(inspector_name_raw).strip() != '':
                            inspector_name = str(inspector_name_raw).strip()
                            if '(' in inspector_name:
                                inspector_name = inspector_name.split('(')[0].strip()
                            if not inspector_name:
                                continue

                            inspector_code = inspector_name_to_id.get(inspector_name)
                            if inspector_code is None:
                                inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                if not inspector_info.empty:
                                    inspector_code = inspector_info.iloc[0]['#ID']
                                else:
                                    continue
                                
                            # 履歴を初期化（inspector_codeが取得できた場合も実行）
                                if inspector_code not in self.inspector_daily_assignments:
                                    self.inspector_daily_assignments[inspector_code] = {}
                                if current_date not in self.inspector_daily_assignments[inspector_code]:
                                    self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                                if inspector_code not in self.inspector_work_hours:
                                    self.inspector_work_hours[inspector_code] = 0.0
                                if inspector_code not in self.inspector_product_hours:
                                    self.inspector_product_hours[inspector_code] = {}
                                if product_number not in self.inspector_product_hours[inspector_code]:
                                    self.inspector_product_hours[inspector_code][product_number] = 0.0
                                
                            # 履歴を累積
                                self.inspector_daily_assignments[inspector_code][current_date] += divided_time
                                self.inspector_work_hours[inspector_code] += divided_time
                                self.inspector_product_hours[inspector_code][product_number] += divided_time
            
            # 最終違反チェック
            final_violations = []
            final_cols = {col: idx for idx, col in enumerate(result_df.columns)}
            for row_idx, row in enumerate(result_df.itertuples(index=False)):
                index = result_df.index[row_idx]
                product_number = row[final_cols['品番']]
                process_name_context = row[final_cols.get('現在工程名', -1)] if '現在工程名' in final_cols else ''
                divided_time = row[final_cols.get('分割検査時間', -1)] if '分割検査時間' in final_cols else 0.0
                if divided_time == -1:
                    divided_time = 0.0
                
                for i in range(1, 6):
                    inspector_col = f'検査員{i}'
                    inspector_col_idx = final_cols.get(inspector_col, -1)
                    if inspector_col_idx != -1:
                        inspector_name_raw = row[inspector_col_idx]
                        if pd.notna(inspector_name_raw) and str(inspector_name_raw).strip() != '':
                            inspector_name = str(inspector_name_raw).strip()
                            if '(' in inspector_name:
                                inspector_name = inspector_name.split('(')[0].strip()
                            if not inspector_name:
                                continue

                            # 固定検査員が割当済みのロットは、勤務時間/同一品番上限の最終違反チェック対象から除外する
                            if _is_fixed_inspector_for_lot_cached(product_number, str(process_name_context).strip(), inspector_name):
                                # 【追加】保護の履歴追跡
                                self.fixed_inspector_protection_metrics['total_protections'] += 1
                                self.fixed_inspector_protection_metrics['protection_by_phase']['phase1.5_violation_check'] = \
                                    self.fixed_inspector_protection_metrics['protection_by_phase'].get('phase1.5_violation_check', 0) + 1
                                self.fixed_inspector_protection_metrics['protection_by_reason']['violation_check_exclusion'] = \
                                    self.fixed_inspector_protection_metrics['protection_by_reason'].get('violation_check_exclusion', 0) + 1
                                self.fixed_inspector_protection_metrics['protected_lots'].add(index)
                                self.fixed_inspector_protection_metrics['protected_inspectors'].add(inspector_name)
                                
                                protection_history_entry = {
                                    'lot_index': index,
                                    'product_number': product_number,
                                    'inspector_name': inspector_name,
                                    'phase': 'phase1.5',
                                    'reason': 'violation_check_exclusion',
                                    'process_name': str(process_name_context).strip(),
                                }
                                self.fixed_inspector_protection_metrics['protection_history'].append(protection_history_entry)
                                if len(self.fixed_inspector_protection_metrics['protection_history']) > 100:
                                    self.fixed_inspector_protection_metrics['protection_history'] = \
                                        self.fixed_inspector_protection_metrics['protection_history'][-100:]
                                continue
                            
                            inspector_code = inspector_name_to_id.get(inspector_name)
                            if inspector_code is None:
                                inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                if not inspector_info.empty:
                                    inspector_code = inspector_info.iloc[0]['#ID']
                                else:
                                    continue
                                daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
                                max_hours = inspector_max_hours.get(inspector_code, 8.0)
                                product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                                
                                # 【修正】当日洗浄上がり品かどうかを判定して適切な超過率を適用
                                shipping_date_val = row[final_cols.get('出荷予定日', -1)] if '出荷予定日' in final_cols else None
                                is_same_day_cleaning = self._should_force_assign_same_day(shipping_date_val)
                                
                                if is_same_day_cleaning:
                                    # 当日洗浄上がり品: 20%超過まで許容
                                    allowed_max_hours = max_hours * (1.0 + SAME_DAY_WORK_HOURS_OVERRUN_RATE)
                                else:
                                    # 通常品: 10%超過まで許容
                                    allowed_max_hours = max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                
                                # 最終チェック: 許容上限を超えている場合は違反
                                if daily_hours > allowed_max_hours - WORK_HOURS_BUFFER:
                                    excess = daily_hours - allowed_max_hours
                                    final_violations.append((index, inspector_code, inspector_name, "勤務時間超過", daily_hours, allowed_max_hours))
                                    mode_str = "緩和モード（20%超過許容）" if is_same_day_cleaning else "通常モード（10%超過許容）"
                                    self.log_message(f"❌ 最終チェック: 勤務時間超過が残っています - 検査員 '{inspector_name}' {daily_hours:.1f}h > {allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h ({mode_str}, 超過: {excess:.1f}h, ロット {index})", level='warning')
                                
                                # 改善ポイント: 最適化フェーズでの4時間上限チェック（厳格）
                                # ただし、最終検証では4.2h未満まで許容（代替検査員が見つからない場合の保護）
                                if product_hours > PRODUCT_LIMIT_FINAL_TOLERANCE:
                                    final_violations.append((index, inspector_code, inspector_name, f"同一品番{self.product_limit_hard_threshold:.1f}時間超過", product_hours, self.product_limit_hard_threshold))
                                    self.log_message(f"❌ 最終チェック: 同一品番{self.product_limit_hard_threshold:.1f}時間超過が残っています - 検査員 '{inspector_name}' 品番 {product_number} {product_hours:.1f}h > {PRODUCT_LIMIT_FINAL_TOLERANCE}h (ロット {index})", level='warning')
                                elif product_hours > self.product_limit_hard_threshold:
                                    # 4.0h超4.2h未満の場合は、警告のみで違反リストには追加しない（許容）
                                    self.log_message(f"⚠️ 最終チェック: 同一品番4時間をわずかに超過していますが許容します - 検査員 '{inspector_name}' 品番 {product_number} {product_hours:.1f}h (ロット {index})", level='warning')
                                    # relaxed_product_limit_assignmentsに追加して保護
                                    self.relaxed_product_limit_assignments.add((inspector_code, product_number))
            
            if final_violations:
                self.log_message(f"⚠️ 警告: {len(final_violations)}件の違反が最終チェックで検出されました", level='warning')
                
                # 出荷予定日が古いロットを優先的に再割り当てを試みる
                # 違反を検出されたロットを出荷予定日順にソート
                violations_with_date = []
                for violation in final_violations:
                    index = violation[0]
                    row_series = result_df.iloc[index]
                    shipping_date_raw = row_series['出荷予定日'] if '出荷予定日' in row_series.index else pd.Timestamp.max
                    product_number = row_series['品番'] if '品番' in row_series.index else ''
                    
                    # 出荷予定日が「当日洗浄上がり品」の場合は文字列として保持
                    shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                    is_same_day_cleaning = (
                        shipping_date_str == "当日洗浄上がり品" or 
                        shipping_date_str == "当日洗浄品" or
                        "当日洗浄" in shipping_date_str or
                        shipping_date_str == "先行検査" or
                        shipping_date_str == "当日先行検査"
                    )
                    
                    if is_same_day_cleaning:
                        # 当日洗浄品の場合は文字列として保持
                        shipping_date = "当日洗浄上がり品"
                    else:
                        shipping_date = shipping_date_raw
                    
                    # 新製品かどうかを判定
                    skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                    is_new_product = skill_rows.empty
                    
                    # 2週間以内の出荷予定日かどうかを判定（当日洗浄品の場合はスキップ）
                    is_within_two_weeks = False
                    if not is_same_day_cleaning and pd.notna(shipping_date):
                        try:
                            if isinstance(shipping_date, pd.Timestamp):
                                shipping_date_date = shipping_date.date()
                            elif isinstance(shipping_date, str):
                                shipping_date_date = pd.to_datetime(shipping_date, errors='coerce').date()
                                if pd.isna(shipping_date_date):
                                    shipping_date_date = None
                            else:
                                shipping_date_date = shipping_date.date() if hasattr(shipping_date, 'date') else shipping_date
                            
                            if shipping_date_date is not None:
                                two_weeks_later = current_date + timedelta(days=14)
                                is_within_two_weeks = shipping_date_date <= two_weeks_later
                        except Exception as e:
                            self.log_message(f"出荷予定日の比較エラー: {str(e)} (ロットインデックス: {index})", level='warning')
                            is_within_two_weeks = False
                    
                    violations_with_date.append((violation, shipping_date, is_new_product, product_number, is_within_two_weeks))
                
                # 出荷予定日の古い順にソート（新製品はさらに優先）
                violations_with_date.sort(key=lambda x: (self._normalize_shipping_date(x[1]), not x[2]))  # 出荷予定日順、新製品を優先
                
                # 出荷予定日が古いロットから順に再割り当てを試みる
                resolved_count = 0
                for violation, shipping_date, is_new_product, product_number, is_within_two_weeks in violations_with_date:
                    index = violation[0]
                    row_series = result_df.iloc[index]
                    inspection_time = row_series['検査時間'] if '検査時間' in row_series.index else 0
                    lot_quantity = row_series['ロット数量'] if 'ロット数量' in row_series.index else 0

                    # 登録済み品番の先行検査×固定検査員ロットは最適化フェーズで動かさない（固定維持）
                    if self._is_locked_fixed_preinspection_lot(result_df, index):
                        continue
                    
                    # 改善ポイント: 2週間以内の新規品は保護（違反があっても許容）
                    if is_new_product and is_within_two_weeks:
                        inspector_code = violation[1]
                        violation_type = violation[3]
                        
                        # 違反内容に応じて保護処理
                        if violation_type == "同一品番4時間超過":
                            current_product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                            # 4.2h未満の場合は許容
                            if current_product_hours <= PRODUCT_LIMIT_FINAL_TOLERANCE:
                                self.log_message(f"⚠️ 新規品 {product_number} (出荷予定日: {shipping_date}) は2週間以内のため保護します - 同一品番4時間超過を許容（{current_product_hours:.1f}h ≤ {PRODUCT_LIMIT_FINAL_TOLERANCE}h）", level='warning')
                                self.relaxed_product_limit_assignments.add((inspector_code, product_number))
                                resolved_count += 1
                                continue
                            else:
                                # 4.2h超過でも保護する（優先度が高いため）
                                self.log_message(f"⚠️ 新規品 {product_number} (出荷予定日: {shipping_date}) は2週間以内のため保護します - 同一品番4時間超過を許容（{current_product_hours:.1f}h > {PRODUCT_LIMIT_FINAL_TOLERANCE}h）", level='warning')
                                self.relaxed_product_limit_assignments.add((inspector_code, product_number))
                                resolved_count += 1
                                continue
                        elif violation_type == "勤務時間超過":
                            # 勤務時間超過でも保護する（優先度が高いため）
                            daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
                            max_hours = inspector_max_hours.get(inspector_code, 8.0)
                            self.log_message(f"⚠️ 新規品 {product_number} (出荷予定日: {shipping_date}) は2週間以内のため保護します - 勤務時間超過を許容（{daily_hours:.1f}h > {max_hours:.1f}h）", level='warning')
                            resolved_count += 1
                            continue
                    
                    self.log_message(f"⚠️ ロットインデックス {index} (品番: {product_number}, 出荷予定日: {shipping_date}, 新製品: {is_new_product}) の違反を是正します（{violation[3]}）")
                    
                    # 新製品かつ出荷予定日が古いロットの場合は、強制的に新製品チームで再割り当てを試みる
                    if is_new_product:
                        self.log_message(f"🔵 新製品ロットのため、新製品チームで強制的に再割り当てを試みます")
                        # 一旦割り当てをクリアして、新製品チームで再割り当て
                        self.clear_assignment(result_df, index)
                        # 新製品チームを取得
                        new_product_team = self.get_new_product_team_inspectors(inspector_master_df)
                        if new_product_team:
                            # 必要な検査員人数を計算
                            # 特例: 一ロットで検査員が5名以上必要になる場合、5名に制限
                            if inspection_time <= 0:
                                required_inspectors = 1
                            else:
                                # 通常の計算（設定時間で割る、最低2人）
                                if inspection_time <= self.required_inspectors_threshold:
                                    required_inspectors = 1
                                else:
                                    calculated_inspectors = max(2, int(inspection_time / self.required_inspectors_threshold) + 1)
                                    # 5名以上になる場合は5名に制限（特例）
                                    required_inspectors = min(5, calculated_inspectors)
                            divided_time = inspection_time / required_inspectors
                            
                            # 新製品チームから検査員を選択
                            assigned_inspectors = self.select_inspectors(
                                new_product_team, required_inspectors, divided_time, 
                                inspector_master_df, product_number, is_new_product=True
                            )
                            
                            if assigned_inspectors:
                                # 割り当てを設定
                                result_df.at[index, '検査員人数'] = len(assigned_inspectors)
                                # 分割検査時間の計算: 検査時間 ÷ 実際の分割した検査人数
                                actual_divided_time = inspection_time / len(assigned_inspectors)
                                result_df.at[index, '分割検査時間'] = round(actual_divided_time, 1)
                                for i, inspector in enumerate(assigned_inspectors, 1):
                                    if i <= 5:
                                        result_df.at[index, f'検査員{i}'] = inspector.get('氏名', '')
                                result_df.at[index, 'チーム情報'] = f"新製品チーム({len(assigned_inspectors)}人)"
                                result_df.at[index, 'assignability_status'] = 'assigned'
                                resolved_count += 1
                                self.log_message(f"✅ 新製品チームでロットインデックス {index} を再割り当てしました")
                                # 履歴を更新（select_inspectors内で既に更新されているが、確認のため）
                                continue
                    
                    # 通常の違反是正処理を試みる
                    divided_time_for_fix = row_series['分割検査時間'] if '分割検査時間' in row_series.index else 0.0
                    violation_resolved = self.fix_single_violation(
                        index, violation[1], violation[2], 
                        divided_time_for_fix, product_number, inspection_time,
                        None, result_df, inspector_master_df, skill_master_df,
                        inspector_max_hours, current_date, show_skill_values
                    )
                    
                    if violation_resolved:
                        resolved_count += 1
                        self.log_message(f"✅ ロットインデックス {index} の違反を是正しました")
                    else:
                        # 是正できなかった場合の処理
                        # 同一品番4時間超過の場合、超過時間が少ない場合は許容する
                        violation_type = violation[3]
                        if violation_type == "同一品番4時間超過":
                            # 違反を起こしている検査員の同一品番累計時間を確認
                            inspector_code = violation[1]
                            current_product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                            
                            # 4.2h未満の場合は許容（代替検査員が見つからない場合の保護）
                            if current_product_hours <= PRODUCT_LIMIT_FINAL_TOLERANCE:
                                self.log_message(f"⚠️ 同一品番4時間をわずかに超過していますが許容します（{current_product_hours:.1f}h ≤ {PRODUCT_LIMIT_FINAL_TOLERANCE}h）- 検査員 '{violation[2]}' 品番 {product_number} (ロット {index})", level='warning')
                                # relaxed_product_limit_assignmentsに追加して保護
                                self.relaxed_product_limit_assignments.add((inspector_code, product_number))
                                # 割り当てを維持（未割当にしない）
                                resolved_count += 1
                                continue
                        
                        # 当日洗浄品の場合は割り当てを維持（優先順位2のため保護）
                        # violations_with_dateから取得したshipping_dateが文字列「当日洗浄上がり品」の場合をチェック
                        shipping_date_str = str(shipping_date).strip() if pd.notna(shipping_date) else ''
                        # 元のデータフレームからも確認（念のため）
                        original_shipping_date = row_series.get('出荷予定日', '') if '出荷予定日' in row_series.index else ''
                        original_shipping_date_str = str(original_shipping_date).strip() if pd.notna(original_shipping_date) else ''
                        
                        is_same_day_cleaning = (
                            shipping_date_str == "当日洗浄上がり品" or 
                            shipping_date_str == "当日洗浄品" or
                            "当日洗浄" in shipping_date_str or
                            shipping_date_str == "先行検査" or
                            shipping_date_str == "当日先行検査" or
                            original_shipping_date_str == "当日洗浄上がり品" or 
                            original_shipping_date_str == "当日洗浄品" or
                            "当日洗浄" in original_shipping_date_str or
                            original_shipping_date_str == "先行検査" or
                            original_shipping_date_str == "当日先行検査"
                        )
                        if is_same_day_cleaning:
                            self.log_message(f"⚠️ 当日洗浄品のため、ルール違反があっても割り当てを維持します（品番: {product_number}, 出荷予定日: {shipping_date_str or original_shipping_date_str}）", level='warning')
                            # 割り当てを維持（未割当にしない）
                            resolved_count += 1
                        # 出荷予定日が最も古い新製品ロットの場合は割り当てを維持
                        elif is_new_product:
                            min_shipping_date = min((self._normalize_shipping_date(v[1]) for v in violations_with_date), default=pd.Timestamp.max)
                            if self._normalize_shipping_date(shipping_date) == min_shipping_date:
                                self.log_message(f"⚠️ 出荷予定日が最も古い新製品ロットのため、ルール違反があっても割り当てを維持します（品番: {product_number}）", level='warning')
                                # 割り当てを維持（未割当にしない）
                                resolved_count += 1
                            else:
                                # 是正できなかった場合は未割当にする
                                self.clear_assignment(result_df, index)
                                self.log_message(f"⚠️ ロットインデックス {index} を未割当にしました（{violation[3]}）", level='warning')
                        else:
                            # 是正できなかった場合は未割当にする
                            self.clear_assignment(result_df, index)
                            self.log_message(f"⚠️ ロットインデックス {index} を未割当にしました（{violation[3]}）", level='warning')
                
                self.log_message(f"違反是正結果: {resolved_count}件是正、{len(final_violations) - resolved_count}件未割当")
                
                # 未割当後の履歴を再計算
                self.inspector_daily_assignments = {}
                self.inspector_work_hours = {}
                self.inspector_product_hours = {}
                
                recalc_cols = {col: idx for idx, col in enumerate(result_df.columns)}
                for row_idx, row in enumerate(result_df.itertuples(index=False)):
                    index = result_df.index[row_idx]
                    product_number = row[recalc_cols['品番']]
                    divided_time = row[recalc_cols.get('分割検査時間', -1)] if '分割検査時間' in recalc_cols else 0.0
                    if divided_time == -1:
                        divided_time = 0.0
                    
                    for i in range(1, 6):
                        inspector_col = f'検査員{i}'
                        inspector_col_idx = recalc_cols.get(inspector_col, -1)
                        if inspector_col_idx != -1:
                            inspector_name_raw = row[inspector_col_idx]
                            if pd.notna(inspector_name_raw) and str(inspector_name_raw).strip() != '':
                                inspector_name = str(inspector_name_raw).strip()
                                if '(' in inspector_name:
                                    inspector_name = inspector_name.split('(')[0].strip()
                                if not inspector_name:
                                    continue
                                
                                inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                if not inspector_info.empty:
                                    inspector_code = inspector_info.iloc[0]['#ID']
                                    
                                    if inspector_code not in self.inspector_daily_assignments:
                                        self.inspector_daily_assignments[inspector_code] = {}
                                    if current_date not in self.inspector_daily_assignments[inspector_code]:
                                        self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                                    if inspector_code not in self.inspector_work_hours:
                                        self.inspector_work_hours[inspector_code] = 0.0
                                    if inspector_code not in self.inspector_product_hours:
                                        self.inspector_product_hours[inspector_code] = {}
                                    if product_number not in self.inspector_product_hours[inspector_code]:
                                        self.inspector_product_hours[inspector_code][product_number] = 0.0
                                    
                                    self.inspector_daily_assignments[inspector_code][current_date] += divided_time
                                    self.inspector_work_hours[inspector_code] += divided_time
                                    self.inspector_product_hours[inspector_code][product_number] += divided_time
            else:
                self.log_message("✅ 最終チェック: 全てのルール違反が解消されました")
                # 【修正】違反が検出されなかった場合でも、フェーズ2で偏り是正を行うために履歴を再計算する
                self.inspector_daily_assignments = {}
                self.inspector_work_hours = {}
                self.inspector_product_hours = {}
                
                recalc_cols = {col: idx for idx, col in enumerate(result_df.columns)}
                for row_idx, row in enumerate(result_df.itertuples(index=False)):
                    index = result_df.index[row_idx]
                    product_number = row[recalc_cols['品番']]
                    divided_time = row[recalc_cols.get('分割検査時間', -1)] if '分割検査時間' in recalc_cols else 0.0
                    if divided_time == -1:
                        divided_time = 0.0
                    
                    for i in range(1, 6):
                        inspector_col = f'検査員{i}'
                        inspector_col_idx = recalc_cols.get(inspector_col, -1)
                        if inspector_col_idx != -1:
                            inspector_name_raw = row[inspector_col_idx]
                            if pd.notna(inspector_name_raw) and str(inspector_name_raw).strip() != '':
                                inspector_name = str(inspector_name_raw).strip()
                                if '(' in inspector_name:
                                    inspector_name = inspector_name.split('(')[0].strip()
                                if not inspector_name:
                                    continue
                                
                                inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                if not inspector_info.empty:
                                    inspector_code = inspector_info.iloc[0]['#ID']
                                    
                                    if inspector_code not in self.inspector_daily_assignments:
                                        self.inspector_daily_assignments[inspector_code] = {}
                                    if current_date not in self.inspector_daily_assignments[inspector_code]:
                                        self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                                    if inspector_code not in self.inspector_work_hours:
                                        self.inspector_work_hours[inspector_code] = 0.0
                                    if inspector_code not in self.inspector_product_hours:
                                        self.inspector_product_hours[inspector_code] = {}
                                    if product_number not in self.inspector_product_hours[inspector_code]:
                                        self.inspector_product_hours[inspector_code][product_number] = 0.0
                                    
                                    self.inspector_daily_assignments[inspector_code][current_date] += divided_time
                                    self.inspector_work_hours[inspector_code] += divided_time
                                    self.inspector_product_hours[inspector_code][product_number] += divided_time
            
            # フェーズ2: 偏りの是正（総勤務時間の偏りを調整）
            self.log_message("全体最適化フェーズ2: 偏りの是正を開始")
            _t_perf_phase2_total = perf_counter()
            
            # 平均勤務時間を計算
            total_hours = sum(self.inspector_work_hours.values())
            active_inspectors = [code for code in self.inspector_work_hours.keys() if self.inspector_work_hours.get(code, 0.0) > 0]
            
            self.log_message(f"偏り是正準備: inspector_work_hoursの総数={len(self.inspector_work_hours)}, アクティブ検査員数={len(active_inspectors)}, 総勤務時間={total_hours:.1f}h")
            
            if active_inspectors:
                avg_hours = total_hours / len(active_inspectors)
                max_hours_val = max(self.inspector_work_hours.values())
                min_hours_val = min([self.inspector_work_hours.get(code, 0.0) for code in active_inspectors])
                imbalance = max_hours_val - min_hours_val
                
                self.log_message(f"偏り分析: 平均 {avg_hours:.1f}h, 最大 {max_hours_val:.1f}h, 最小 {min_hours_val:.1f}h, 偏り {imbalance:.1f}h")
                
                # 偏りが大きい場合（平均の3%以上）、調整を試みる（より積極的に調整）
                imbalance_threshold = avg_hours * 0.03
                if imbalance > imbalance_threshold and len(active_inspectors) > 1:
                    self.log_message(f"偏りが大きいため調整を試みます (閾値: {imbalance_threshold:.1f}h, 実際: {imbalance:.1f}h)")
                    
                    # 多忙な検査員から余裕のある検査員へ一部を移動
                    # （ただし出荷予定日の順序は維持）
                    
                    # 【変更】新規品対応チームメンバーも偏り是正の対象に含める
                    # 新規品対応チームメンバーのコードリストを取得
                    new_team_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                    new_team_codes = {insp['コード'] for insp in new_team_inspectors}
                    
                    # 全検査員（新規品対応チームメンバーを含む）から偏りを計算
                    # 新規品対応チームメンバーで総検査時間が0の検査員も含める
                    for code in new_team_codes:
                        if code not in self.inspector_work_hours:
                            self.inspector_work_hours[code] = 0.0
                    
                    # 【改善】偏り是正の閾値を厳しくする（偏りを最小限にするため、より積極的に調整）
                    # 多忙: 平均を超える検査員、余裕: 平均未満の検査員（偏りを最小限にする）
                    # 【追加】10%超過を超える検査員を優先的に処理するため、まず10%超過を超える検査員を抽出
                    over_loaded = []
                    over_loaded_10pct_overrun = []  # 10%超過を超える検査員（優先処理）
                    for code, hours in self.inspector_work_hours.items():
                        if hours > avg_hours:
                            max_hours = inspector_max_hours.get(code, 8.0)
                            allowed_max_hours = max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                            daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                            excess = daily_hours - (allowed_max_hours - WORK_HOURS_BUFFER)
                            if excess > 0:
                                # 10%超過を超える検査員（優先処理）
                                over_loaded_10pct_overrun.append((code, hours, excess))
                            else:
                                # 通常の多忙な検査員
                                over_loaded.append((code, hours))
                    
                    # 10%超過を超える検査員を超過量が多い順にソート
                    over_loaded_10pct_overrun.sort(key=lambda x: x[2], reverse=True)
                    # 通常の多忙な検査員を勤務時間が多い順にソート
                    over_loaded.sort(key=lambda x: x[1], reverse=True)
                    # 10%超過を超える検査員を先頭に配置
                    over_loaded = [(code, hours) for code, hours, _ in over_loaded_10pct_overrun] + over_loaded
                    
                    under_loaded = [(code, hours) for code, hours in self.inspector_work_hours.items() 
                                    if hours < avg_hours]  # 平均未満の検査員を余裕ありとする（偏りを最小限に）
                    
                    if over_loaded and under_loaded:
                        overrun_count = len(over_loaded_10pct_overrun)
                        self.log_message(f"調整対象: 多忙 {len(over_loaded)}人（うち10%超過: {overrun_count}人）, 余裕あり {len(under_loaded)}人")
                        
                        # 出荷予定日の古い順にソート（順序を維持）
                        # 出荷予定日を変換（当日洗浄品は文字列として保持）
                        result_df['出荷予定日'] = result_df['出荷予定日'].apply(self._convert_shipping_date)
                        
                        # ソート用のキー関数: 新しい優先順位に従う
                        current_date = pd.Timestamp.now().date()
                        
                        def get_next_business_day(date_val):
                            """翌営業日を取得（金曜日の場合は翌週の月曜日）"""
                            weekday = date_val.weekday()  # 0=月曜日, 4=金曜日
                            if weekday == 4:  # 金曜日
                                return date_val + timedelta(days=3)  # 翌週の月曜日
                            else:
                                return date_val + timedelta(days=1)  # 翌日
                        
                        next_business_day = get_next_business_day(current_date)
                        
                        def sort_key(val):
                            if pd.isna(val):
                                return (5, None)  # 最後に
                            val_str = str(val).strip()
                            
                            # 1. 当日の日付（優先度0）
                            try:
                                date_val = pd.to_datetime(val, errors='coerce')
                                if pd.notna(date_val):
                                    date_date = date_val.date()
                                    if date_date == current_date:
                                        return (0, date_val)
                            except:
                                pass
                            
                            # 2. 当日洗浄上がり品（優先度1）
                            if (val_str == "当日洗浄上がり品" or
                                val_str == "当日洗浄品" or
                                "当日洗浄" in val_str):
                                return (1, val_str)
                            
                            # 3. 先行検査品（優先度2）
                            if (val_str == "先行検査" or
                                val_str == "当日先行検査"):
                                return (2, val_str)
                            
                            # 4. 翌日または翌営業日（優先度3）
                            try:
                                date_val = pd.to_datetime(val, errors='coerce')
                                if pd.notna(date_val):
                                    date_date = date_val.date()
                                    if date_date == next_business_day:
                                        return (3, date_val)
                            except:
                                pass
                            
                            # 5. それ以降の日付（優先度4）
                            try:
                                date_val = pd.to_datetime(val, errors='coerce')
                                if pd.notna(date_val):
                                    return (4, date_val)
                            except:
                                pass
                            
                            return (5, val_str)  # その他文字列
                        
                        # ソートキーを追加（出荷予定日順）
                        result_df['_sort_key'] = result_df['出荷予定日'].apply(sort_key)
                        
                        # 指示日のソートキーを追加（FIFO用：指示日が古い順）
                        if '指示日' in result_df.columns:
                            def instruction_date_sort_key_for_bias(val):
                                if pd.isna(val):
                                    return pd.Timestamp.max  # 最後に
                                try:
                                    date_val = pd.to_datetime(val, errors='coerce')
                                    if pd.notna(date_val):
                                        return date_val
                                except:
                                    pass
                                return pd.Timestamp.max
                            result_df['_instruction_date_sort_key_bias'] = result_df['指示日'].apply(instruction_date_sort_key_for_bias)
                        else:
                            result_df['_instruction_date_sort_key_bias'] = pd.Timestamp.max
                        
                        # 出荷予定日順にソートし、同じ出荷予定日の場合は指示日が古い順（FIFO）でソート
                        result_df_sorted = result_df.sort_values(
                            ['_sort_key', '_instruction_date_sort_key_bias', '品番'],
                            ascending=[True, True, True],
                            na_position='last'
                        ).reset_index(drop=True)
                        result_df_sorted = result_df_sorted.drop(columns=['_sort_key', '_instruction_date_sort_key_bias'], errors='ignore')
                        
                        # 【重要】FIFO（先入先出）の維持: result_df_sortedは出荷予定日の古い順にソート済み
                        # 同じ出荷予定日の場合は指示日が古い順（FIFO）でソートされている
                        # この順序でロットを処理するため、偏り是正処理でもFIFOが維持される
                        
                        # 【改善】複数回のパスで偏り是正を実行（収束するまで、最大5回）
                        max_passes = 5
                        total_reassignment_count = 0
                        
                        # 【追加】偏り是正の効果測定メトリクス
                        bias_correction_metrics = {
                            'initial_imbalance': imbalance,
                            'pass_imbalances': [],
                            'pass_std_deviations': [],
                            'pass_improvements': [],
                            'total_reassignments': 0,
                        }
                        
                        # 初期状態の標準偏差を計算
                        initial_hours_list = [self.inspector_work_hours.get(code, 0.0) for code in active_inspectors]
                        if len(initial_hours_list) > 1:
                            import statistics
                            initial_std_dev = statistics.stdev(initial_hours_list)
                            bias_correction_metrics['initial_std_deviation'] = initial_std_dev
                        else:
                            bias_correction_metrics['initial_std_deviation'] = 0.0
                        
                        previous_imbalance = imbalance
                        previous_std_dev = bias_correction_metrics['initial_std_deviation']
                        
                        # 【改善】無限ループ防止: 10%超過を検出した検査員の履歴を初期化（偏り是正フェーズ全体で追跡）
                        if not hasattr(self, '_overrun_inspector_history'):
                            self._overrun_inspector_history = {}
                        else:
                            # 偏り是正フェーズ開始時に履歴をリセット
                            self._overrun_inspector_history.clear()
                        
                        for pass_num in range(max_passes):
                            # 【追加】各パスの開始時に、inspector_daily_assignmentsを再計算して正確な勤務時間を反映
                            # 偏り是正の段階で、複数の再割当が連続して発生すると、累積的に超過が発生する可能性があるため
                            if pass_num > 0:  # 最初のパス以外は再計算
                                self.log_message(f"偏り是正パス{pass_num + 1}: inspector_daily_assignmentsを再計算中...", debug=True)
                                # result_df_sortedから実際の割り当てを読み取って、inspector_daily_assignmentsを再計算
                                # まず、分割検査時間を実際の検査員数で再計算
                                result_cols_bias = {col: idx for idx, col in enumerate(result_df_sorted.columns)}
                                for row_idx, row in enumerate(result_df_sorted.itertuples(index=False)):
                                    index = result_df_sorted.index[row_idx]
                                    inspection_time = row[result_cols_bias.get('検査時間', -1)] if '検査時間' in result_cols_bias else 0.0
                                    if inspection_time == -1 or inspection_time == 0:
                                        continue
                                    
                                    # 実際に割り当てられた検査員数をカウント
                                    actual_inspector_count = 0
                                    for i in range(1, 6):
                                        inspector_col = f'検査員{i}'
                                        inspector_col_idx = result_cols_bias.get(inspector_col, -1)
                                        if inspector_col_idx != -1:
                                            inspector_name = row[inspector_col_idx]
                                            if pd.notna(inspector_name) and str(inspector_name).strip() != '':
                                                actual_inspector_count += 1
                                    
                                    # 分割検査時間を再計算: 検査時間 ÷ 実際の分割した検査人数
                                    if actual_inspector_count > 0:
                                        actual_divided_time = inspection_time / actual_inspector_count
                                        result_df_sorted.at[index, '分割検査時間'] = round(actual_divided_time, 1)
                                
                                # inspector_daily_assignmentsを再計算
                                self.inspector_daily_assignments = {}
                                self.inspector_work_hours = {}
                                self.inspector_product_hours = {}
                                
                                for row_idx, row in enumerate(result_df_sorted.itertuples(index=False)):
                                    index = result_df_sorted.index[row_idx]
                                    product_number = row[result_cols_bias['品番']]
                                    divided_time = row[result_cols_bias.get('分割検査時間', -1)] if '分割検査時間' in result_cols_bias else 0.0
                                    if divided_time == -1:
                                        divided_time = 0.0
                                    
                                    # 各検査員の割り当てを確認
                                    for i in range(1, 6):
                                        inspector_col = f'検査員{i}'
                                        inspector_col_idx = result_cols_bias.get(inspector_col, -1)
                                        if inspector_col_idx != -1:
                                            inspector_name_raw = row[inspector_col_idx]
                                            if pd.notna(inspector_name_raw) and str(inspector_name_raw).strip() != '':
                                                inspector_name = str(inspector_name_raw).strip()
                                                # スキル値や(新)を除去
                                                if '(' in inspector_name:
                                                    inspector_name = inspector_name.split('(')[0].strip()
                                                
                                                if not inspector_name:
                                                    continue
                                                
                                                # 検査員コードを取得
                                                inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                                if not inspector_info.empty:
                                                    inspector_code = inspector_info.iloc[0]['#ID']
                                                    
                                                    # 履歴を初期化
                                                    if inspector_code not in self.inspector_daily_assignments:
                                                        self.inspector_daily_assignments[inspector_code] = {}
                                                    if current_date not in self.inspector_daily_assignments[inspector_code]:
                                                        self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                                                    if inspector_code not in self.inspector_work_hours:
                                                        self.inspector_work_hours[inspector_code] = 0.0
                                                    if inspector_code not in self.inspector_product_hours:
                                                        self.inspector_product_hours[inspector_code] = {}
                                                    if product_number not in self.inspector_product_hours[inspector_code]:
                                                        self.inspector_product_hours[inspector_code][product_number] = 0.0
                                                    
                                                    # 履歴を累積
                                                    self.inspector_daily_assignments[inspector_code][current_date] += divided_time
                                                    self.inspector_work_hours[inspector_code] += divided_time
                                                    self.inspector_product_hours[inspector_code][product_number] += divided_time
                                
                                self.log_message(f"偏り是正パス{pass_num + 1}: inspector_daily_assignmentsの再計算が完了しました", debug=True)
                            
                            # 各パスで偏りを再計算
                            total_hours_pass = sum(self.inspector_work_hours.values())
                            active_inspectors_pass = [code for code in self.inspector_work_hours.keys() 
                                                      if self.inspector_work_hours.get(code, 0.0) > 0]
                            if not active_inspectors_pass:
                                break
                            
                            avg_hours_pass = total_hours_pass / len(active_inspectors_pass)
                            max_hours_pass = max(self.inspector_work_hours.values())
                            min_hours_pass = min([self.inspector_work_hours.get(code, 0.0) for code in active_inspectors_pass])
                            imbalance_pass = max_hours_pass - min_hours_pass
                            
                            # 【追加】標準偏差を計算（効果測定メトリクス）
                            hours_list_pass = [self.inspector_work_hours.get(code, 0.0) for code in active_inspectors_pass]
                            if len(hours_list_pass) > 1:
                                import statistics
                                std_dev_pass = statistics.stdev(hours_list_pass)
                            else:
                                std_dev_pass = 0.0
                            
                            # 【追加】改善率を計算（前のパスと比較）
                            if pass_num > 0:
                                improvement_ratio = (previous_imbalance - imbalance_pass) / previous_imbalance if previous_imbalance > 0 else 0.0
                                std_improvement_ratio = (previous_std_dev - std_dev_pass) / previous_std_dev if previous_std_dev > 0 else 0.0
                            else:
                                improvement_ratio = (imbalance - imbalance_pass) / imbalance if imbalance > 0 else 0.0
                                std_improvement_ratio = (bias_correction_metrics['initial_std_deviation'] - std_dev_pass) / bias_correction_metrics['initial_std_deviation'] if bias_correction_metrics['initial_std_deviation'] > 0 else 0.0
                            
                            # メトリクスを記録
                            bias_correction_metrics['pass_imbalances'].append(imbalance_pass)
                            bias_correction_metrics['pass_std_deviations'].append(std_dev_pass)
                            bias_correction_metrics['pass_improvements'].append(improvement_ratio)
                            
                            # 偏りが閾値以下になったら終了
                            if imbalance_pass <= imbalance_threshold:
                                self.log_message(
                                    f"偏り是正パス{pass_num + 1}: 偏りが閾値以下になりました（{imbalance_pass:.1f}h <= {imbalance_threshold:.1f}h, "
                                    f"標準偏差: {std_dev_pass:.2f}h, 改善率: {improvement_ratio * 100:.1f}%）"
                                )
                                break
                            
                            # 【追加】早期終了判定：改善率が閾値未満の場合（0.5%未満）
                            MIN_IMPROVEMENT_RATIO = 0.005  # 0.5%
                            if pass_num > 0 and improvement_ratio < MIN_IMPROVEMENT_RATIO:
                                self.log_message(
                                    f"偏り是正パス{pass_num + 1}: 改善率が閾値未満のため早期終了します "
                                    f"（改善率: {improvement_ratio * 100:.2f}% < {MIN_IMPROVEMENT_RATIO * 100:.1f}%, "
                                    f"偏り: {imbalance_pass:.1f}h, 標準偏差: {std_dev_pass:.2f}h）",
                                    level='info',
                                )
                                break
                            
                            # 多忙/余裕のリストを再計算（偏りを最小限にするため、より敏感に検出）
                            over_loaded_pass = [(code, hours) for code, hours in self.inspector_work_hours.items() 
                                                if hours > avg_hours_pass]  # 平均を超える検査員を多忙とする（偏りを最小限に）
                            under_loaded_pass = [(code, hours) for code, hours in self.inspector_work_hours.items() 
                                                 if hours < avg_hours_pass]  # 平均未満の検査員を余裕ありとする（偏りを最小限に）
                            
                            if not over_loaded_pass or not under_loaded_pass:
                                self.log_message(f"偏り是正パス{pass_num + 1}: 調整対象がなくなりました")
                                break
                            
                            self.log_message(
                                f"偏り是正パス{pass_num + 1}: 平均 {avg_hours_pass:.1f}h, 最大 {max_hours_pass:.1f}h, "
                                f"最小 {min_hours_pass:.1f}h, 偏り {imbalance_pass:.1f}h, "
                                f"標準偏差: {std_dev_pass:.2f}h, 改善率: {improvement_ratio * 100:.2f}%"
                            )
                            
                            # 前のパスの値を更新（次のパスでの改善率計算用）
                            previous_imbalance = imbalance_pass
                            previous_std_dev = std_dev_pass
                            
                            # 【改善】10%超過を超える検査員を優先的に処理するため、ソートキーを変更
                            # 優先順位: 1) 10%超過を超える検査員（超過量が多い順）、2) それ以外（勤務時間が多い順）
                            def sort_key_overloaded(x):
                                code, hours = x
                                max_hours = inspector_max_hours.get(code, 8.0)
                                allowed_max_hours = max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                                excess = daily_hours - (allowed_max_hours - WORK_HOURS_BUFFER)
                                # 10%超過を超える場合は優先度を上げる（超過量が多い順）
                                if excess > 0:
                                    return (-1, -excess)  # 負の値でソート（大きい順）
                                else:
                                    return (1, -hours)  # 通常は勤務時間が多い順
                            
                            over_loaded_pass.sort(key=sort_key_overloaded)
                            # 余裕のある検査員を勤務時間の少ない順にソート
                            under_loaded_pass.sort(key=lambda x: x[1])
                        
                        # 再割当て回数を制限（無限ループを防ぐ）
                        max_reassignments_per_pass = 50
                        reassignment_count = 0
                        
                        # 【追加】このパスで実行された再割当の履歴を記録（10%超過検出時に元に戻すため）
                        pass_reassignment_history = []
                        
                        # 各多忙な検査員について、割り当てられたロットを確認
                        for overloaded_code, overloaded_hours in over_loaded_pass:
                            if reassignment_count >= max_reassignments_per_pass:
                                break
                            
                            # この検査員が割り当てられているロットを取得（出荷予定日順、FIFO維持）
                            # result_df_sortedは既に出荷予定日の古い順にソートされているため、
                            # この順序で処理することでFIFOが維持される
                            assigned_lots = []
                            self.log_message(f"偏り是正パス{pass_num + 1}: 多忙な検査員 {overloaded_code} ({overloaded_hours:.1f}h) の割り当てロットを検索中...", debug=True)
                            # 列インデックスを事前に取得（高速化：itertuples()を使用）
                            product_col_idx = result_df_sorted.columns.get_loc('品番')
                            divided_time_col_idx = result_df_sorted.columns.get_loc('分割検査時間')
                            inspector_col_indices = [result_df_sorted.columns.get_loc(f'検査員{i}') for i in range(1, 6)]
                            
                            # 出荷予定日の古い順（FIFO）で処理
                            for row_tuple in result_df_sorted.itertuples(index=True):
                                index = row_tuple[0]  # インデックス
                                product_number = row_tuple[product_col_idx + 1]  # itertuplesはインデックスを含むため+1
                                divided_time = row_tuple[divided_time_col_idx + 1] if divided_time_col_idx < len(row_tuple) - 1 else 0.0
                                
                                # このロットにこの検査員が含まれているか確認
                                for i in range(1, 6):
                                    inspector_col_idx = inspector_col_indices[i - 1]
                                    inspector_value = row_tuple[inspector_col_idx + 1] if inspector_col_idx < len(row_tuple) - 1 else None
                                    if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                                        inspector_name = str(inspector_value).strip()
                                        if '(' in inspector_name:
                                            inspector_name = inspector_name.split('(')[0].strip()
                                        
                                        if not inspector_name:
                                            continue
                                        
                                        inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                        if not inspector_info.empty:
                                            lot_inspector_code = inspector_info.iloc[0]['#ID']
                                            if lot_inspector_code == overloaded_code:
                                                # Seriesとして扱うために元の行を取得
                                                row = result_df_sorted.loc[index]
                                                # rowがSeriesであることを確認
                                                if not isinstance(row, pd.Series):
                                                    # 文字列の場合はスキップ
                                                    continue
                                                assigned_lots.append((index, product_number, divided_time, i, row))
                                                break
                            
                                self.log_message(f"偏り是正パス{pass_num + 1}: 多忙な検査員 {overloaded_code} の割り当てロット数: {len(assigned_lots)}件", debug=True)
                            
                            # 各ロットについて、余裕のある検査員への再割当てを試みる
                            for lot_index, product_number, divided_time, inspector_col_num, row in assigned_lots:
                                if reassignment_count >= max_reassignments_per_pass:
                                    break
                                
                                # rowがSeriesであることを確認（エラー防止）
                                if not isinstance(row, pd.Series):
                                    # rowがSeriesでない場合は、result_df_sortedから再取得
                                    try:
                                        row = result_df_sorted.loc[lot_index]
                                    except (KeyError, IndexError):
                                        continue
                                
                                # 【改善】タブーリストに含まれるロットはスキップ（フェーズ間スラッシング防止）
                                if lot_index in self.tabu_list:
                                    self.log_message(
                                        f"偏り是正: ロットインデックス {lot_index} (品番: {product_number}) はタブーリストに含まれるため再割当てをスキップします（フェーズ間スラッシング防止）",
                                        debug=True,
                                    )
                                    continue

                                # 登録済み品番の先行検査×固定検査員ロットは最適化フェーズで動かさない（固定維持）
                                if self._is_locked_fixed_preinspection_lot(result_df_sorted, lot_index):
                                    continue
                                
                                # 【追加】3D025-G4960の初回割当を維持：最適化フェーズで再割当てをスキップ
                                if str(product_number).strip() == "3D025-G4960":
                                    self.log_message(
                                        f"偏り是正: 品番 '3D025-G4960' の初回割当を維持するため再割当てをスキップします（ロットインデックス: {lot_index}）",
                                        debug=True,
                                    )
                                    continue
                                
                                # 【追加】固定検査員を保護：このロットに固定検査員が割り当てられている場合は再割当てをスキップ
                                lot_process_name = str(row.get('現在工程名', '') or '').strip()
                                fixed_inspector_names = self._collect_fixed_inspector_names(product_number, lot_process_name)
                                if fixed_inspector_names:
                                    # 現在割り当てられている検査員名を取得
                                    current_inspector_value = row.get(f'検査員{inspector_col_num}', '')
                                    if pd.notna(current_inspector_value) and str(current_inspector_value).strip() != '':
                                        current_inspector_name = str(current_inspector_value).strip()
                                        if '(' in current_inspector_name:
                                            current_inspector_name = current_inspector_name.split('(')[0].strip()
                                        
                                        if current_inspector_name in fixed_inspector_names:
                                            # 固定検査員が割り当てられている場合は再割当てをスキップ
                                            # 【追加】保護の履歴追跡
                                            self.fixed_inspector_protection_metrics['total_protections'] += 1
                                            self.fixed_inspector_protection_metrics['protection_by_phase']['phase2_bias_correction'] = \
                                                self.fixed_inspector_protection_metrics['protection_by_phase'].get('phase2_bias_correction', 0) + 1
                                            self.fixed_inspector_protection_metrics['protection_by_reason']['bias_correction_skip'] = \
                                                self.fixed_inspector_protection_metrics['protection_by_reason'].get('bias_correction_skip', 0) + 1
                                            self.fixed_inspector_protection_metrics['protected_lots'].add(lot_index)
                                            self.fixed_inspector_protection_metrics['protected_inspectors'].add(current_inspector_name)
                                            
                                            protection_history_entry = {
                                                'lot_index': lot_index,
                                                'product_number': product_number,
                                                'inspector_name': current_inspector_name,
                                                'phase': 'phase2',
                                                'reason': 'bias_correction_skip',
                                                'process_name': lot_process_name,
                                            }
                                            self.fixed_inspector_protection_metrics['protection_history'].append(protection_history_entry)
                                            if len(self.fixed_inspector_protection_metrics['protection_history']) > 100:
                                                self.fixed_inspector_protection_metrics['protection_history'] = \
                                                    self.fixed_inspector_protection_metrics['protection_history'][-100:]
                                            
                                            once_key = (
                                                "bias_skip_fixed_inspector",
                                                str(product_number).strip(),
                                                str(current_inspector_name).strip(),
                                            )
                                            if once_key not in self.logged_warnings:
                                                self.logged_warnings.add(once_key)
                                                self.log_message(
                                                    f"偏り是正: 品番 '{product_number}' の固定検査員 '{current_inspector_name}' は保護のため再割当てをスキップします",
                                                    debug=True,
                                                )
                                            continue
                                
                                # 再割当て可能かチェック（出荷予定日が古い順に処理）
                                process_number = row.get('現在工程番号', '')
                                
                                # 当日洗浄上がり品かどうかを判定
                                shipping_date_raw = row.get('出荷予定日', None)
                                shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                                is_same_day_cleaning_lot = (
                                    shipping_date_str == "当日洗浄上がり品" or
                                    shipping_date_str == "当日洗浄品" or
                                    "当日洗浄" in shipping_date_str
                                )
                                
                                # スキルマスタに登録があるか確認
                                skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                                is_new_product = skill_rows.empty
                                
                                # 新規品で出荷予定日が2週間以内の場合は、再割当てを避ける（保護）
                                if is_new_product:
                                    shipping_date = row.get('出荷予定日', None)
                                    if pd.notna(shipping_date):
                                        shipping_date = pd.to_datetime(shipping_date, errors='coerce')
                                        if pd.notna(shipping_date):
                                            shipping_date_date = shipping_date.date()
                                            # 本日から2週間以内の出荷予定日かどうかを判定
                                            two_weeks_later = current_date + timedelta(days=14)
                                            if shipping_date_date <= two_weeks_later:
                                                # 新規品で出荷予定日が2週間以内の場合は再割当てをスキップ
                                                once_key = (
                                                    "bias_skip_new_product",
                                                    str(product_number).strip(),
                                                    str(shipping_date_date),
                                                )
                                                if once_key not in self.logged_warnings:
                                                    self.logged_warnings.add(once_key)
                                                    self.log_message(
                                                        f"偏り是正: 新規品 {product_number} (出荷予定日: {shipping_date_date}) は保護のため再割当てをスキップします",
                                                        debug=True,
                                                    )
                                                continue
                                
                                # 利用可能な検査員を取得
                                shipping_date = row.get('出荷予定日', None)
                                if is_new_product:
                                    available_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                                else:
                                    # 【変更】新規品対応チームメンバーも通常の品番に割り当て可能にするため、allow_new_team_fallbackをTrueに変更
                                    available_inspectors = self.get_available_inspectors(
                                        product_number, process_number, skill_master_df, inspector_master_df,
                                        shipping_date=shipping_date, allow_new_team_fallback=True,
                                        process_master_df=process_master_df, inspection_target_keywords=inspection_target_keywords,
                                        process_name_context=lot_process_name
                                    )
                                
                                if not available_inspectors:
                                    self.log_message(
                                        f"偏り是正: ロットインデックス {lot_index} (品番: {product_number}) は利用可能な検査員がいないため再割当てをスキップします",
                                        debug=True,
                                    )
                                    continue
                                
                                # 現在のロットの他の検査員を取得（再割当て時に除外するため）
                                current_codes = []
                                for i in range(1, 6):
                                    inspector_col = f'検査員{i}'
                                    if pd.notna(row.get(inspector_col)) and str(row[inspector_col]).strip() != '':
                                        inspector_name_check = str(row[inspector_col]).strip()
                                        if '(' in inspector_name_check:
                                            inspector_name_check = inspector_name_check.split('(')[0].strip()
                                        if not inspector_name_check:
                                            continue
                                        inspector_info_check = self._get_inspector_by_name(inspector_name_check, inspector_master_df)
                                        if not inspector_info_check.empty:
                                            current_codes.append(inspector_info_check.iloc[0]['#ID'])
                                
                                # 余裕のある検査員の中から、条件を満たす候補を探す
                                replacement_candidates = []
                                self.log_message(
                                    f"偏り是正: ロットインデックス {lot_index} (品番: {product_number}) の再割当て候補を検索中... (利用可能検査員数: {len(available_inspectors)})",
                                    debug=True,
                                )
                                for insp in available_inspectors:
                                    candidate_code = insp['コード']
                                    candidate_name = insp['氏名']
                                    
                                    # 【追加】固定検査員を保護：置き換え先として固定検査員を選択しない
                                    # （固定検査員は既に優先的に割り当てられているため、他のロットから奪うべきではない）
                                    if fixed_inspector_names and candidate_name in fixed_inspector_names:
                                        # 固定検査員は置き換え先候補から除外
                                        continue
                                    
                                    # 既に割り当てられている人は除外
                                    if candidate_code in current_codes:
                                        continue
                                    
                                    # 当日洗浄上がり品のロットの場合、既にこの品番または同じ品名の他の品番に割り当てられた検査員を除外（品番単位・品名単位の制約）
                                    if is_same_day_cleaning_lot:
                                        # この品番に既に割り当てられた検査員をチェック（品番単位）
                                        if product_number in self.same_day_cleaning_inspectors:
                                            if candidate_code in self.same_day_cleaning_inspectors[product_number]:
                                                # 既にこの品番に割り当てられている場合は除外
                                                continue
                                        
                                        # 【改善】品名単位の制約もチェック
                                        product_name = row.get('品名', '')
                                        product_name_str = str(product_name).strip() if pd.notna(product_name) else ''
                                        if product_name_str:
                                            # 同じ品名の他の品番に既に割り当てられた検査員をチェック（品名単位）
                                            if product_name_str in self.same_day_cleaning_inspectors_by_product_name:
                                                if candidate_code in self.same_day_cleaning_inspectors_by_product_name[product_name_str]:
                                                    # 既に同じ品名の他の品番に割り当てられている場合は除外
                                                    continue
                                    
                                    # 多忙な人（平均以上）への再割当ては避ける（偏りを最小限に）
                                    candidate_total_hours = self.inspector_work_hours.get(candidate_code, 0.0)
                                    # 【改善】偏り是正のため、平均以下を優先（偏りを最小限にする）
                                    if candidate_total_hours > avg_hours_pass:
                                        continue
                                        
                                        # 勤務時間制約をチェック（偏り是正では厳格にチェック）
                                        candidate_max_hours = inspector_max_hours.get(candidate_code, 8.0)
                                        candidate_daily_hours = self.inspector_daily_assignments.get(candidate_code, {}).get(current_date, 0.0)
                                        # 偏り是正では、10%超過まで許容するが、それ以上は許容しない（WORK_HOURS_BUFFERを考慮）
                                        allowed_max_hours = candidate_max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                        if candidate_daily_hours + divided_time > allowed_max_hours - WORK_HOURS_BUFFER:
                                            self.log_message(
                                                f"偏り是正: 候補検査員 {candidate_name} ({candidate_code}) は勤務時間超過のため除外します（現在: {candidate_daily_hours:.1f}h + {divided_time:.1f}h > {allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h）",
                                                debug=False,  # 通常ログレベルで出力
                                            )
                                            continue  # 勤務時間超過のため除外
                                    
                                    # 改善ポイント: 最適化フェーズでの4時間上限チェック（厳格）
                                    candidate_product_hours = self.inspector_product_hours.get(candidate_code, {}).get(product_number, 0.0)
                                    if candidate_product_hours + divided_time > self.product_limit_hard_threshold:
                                        continue
                                    
                                    # 候補として追加（総勤務時間が少ない順に優先）
                                        # _priority_sort_keyは(priority, inspector_dict, ...)の形式を期待しているため、順序を合わせる
                                        replacement_candidates.append((candidate_total_hours, insp, candidate_code))
                                
                                # 最も総勤務時間が少ない候補を選択
                                if replacement_candidates:
                                    # 違反件数をカウント
                                    self.violation_count += 1
                                    
                                    replacement_candidates.sort(key=self._priority_sort_key)
                                    _, replacement_inspector, new_code = replacement_candidates[0]
                                    
                                    # 再割当てを実行
                                    # 元の検査員名を取得
                                    old_inspector_name = None
                                    for i in range(1, 6):
                                        inspector_col = f'検査員{i}'
                                        if pd.notna(row.get(inspector_col)) and str(row[inspector_col]).strip() != '':
                                            inspector_name_check = str(row[inspector_col]).strip()
                                            if '(' in inspector_name_check:
                                                inspector_name_check = inspector_name_check.split('(')[0].strip()
                                            if not inspector_name_check:
                                                continue
                                            inspector_info_check = self._get_inspector_by_name(inspector_name_check, inspector_master_df)
                                            if not inspector_info_check.empty:
                                                if inspector_info_check.iloc[0]['#ID'] == overloaded_code:
                                                    old_inspector_name = inspector_name_check
                                                    break
                                    
                                    if old_inspector_name:
                                        # 新しい検査員名を設定
                                        if show_skill_values:
                                            if replacement_inspector.get('is_new_team', False):
                                                replacement_name = f"{replacement_inspector['氏名']}(新)"
                                            else:
                                                replacement_name = f"{replacement_inspector['氏名']}({replacement_inspector['スキル']})"
                                        else:
                                            if replacement_inspector.get('is_new_team', False):
                                                replacement_name = f"{replacement_inspector['氏名']}(新)"
                                            else:
                                                replacement_name = replacement_inspector['氏名']
                                        
                                            # 結果データフレームで該当する検査員を置き換え（一時的に）
                                        result_df_sorted.at[lot_index, f'検査員{inspector_col_num}'] = replacement_name
                                        
                                        # 【改善】分割検査時間を再計算（再割当て後の実際の検査員数で）
                                        inspection_time_for_recalc = row.get('検査時間', 0.0)
                                        if inspection_time_for_recalc > 0:
                                            # 実際に割り当てられた検査員数をカウント
                                            actual_inspector_count_recalc = 0
                                            for i in range(1, 6):
                                                inspector_col_recalc = f'検査員{i}'
                                                inspector_value_recalc = result_df_sorted.at[lot_index, inspector_col_recalc] if inspector_col_recalc in result_df_sorted.columns else ''
                                                if pd.notna(inspector_value_recalc) and str(inspector_value_recalc).strip() != '':
                                                    actual_inspector_count_recalc += 1
                                            
                                            # 分割検査時間を再計算: 検査時間 ÷ 実際の分割した検査人数
                                            if actual_inspector_count_recalc > 0:
                                                actual_divided_time_recalc = inspection_time_for_recalc / actual_inspector_count_recalc
                                                # 再計算後の分割検査時間を使用
                                                divided_time = actual_divided_time_recalc
                                        
                                            # 【追加】再割当実行前に、分割検査時間を再計算した後、10%超過をチェック
                                            # 偏り是正の段階で、複数の再割当が連続して発生すると、累積的に超過が発生する可能性があるため
                                            current_new_daily_hours = self.inspector_daily_assignments.get(new_code, {}).get(current_date, 0.0)
                                            new_max_hours_check = self.get_inspector_max_hours(new_code, inspector_master_df)
                                            new_allowed_max_hours_check = new_max_hours_check * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                            
                                            # 再割当実行前に、現在の勤務時間 + 再計算後の割り当て時間が10%超過を超える場合はスキップ
                                            if current_new_daily_hours + divided_time > new_allowed_max_hours_check - WORK_HOURS_BUFFER:
                                                self.log_message(
                                                    f"偏り是正: 再割当実行前に10%超過を検出 - 検査員 {replacement_inspector['氏名']} ({new_code}) は再割当をスキップします（現在: {current_new_daily_hours:.1f}h + {divided_time:.1f}h > {new_allowed_max_hours_check:.1f}h - {WORK_HOURS_BUFFER:.2f}h）",
                                                    debug=True,
                                                )
                                                # 元の検査員に戻す
                                                result_df_sorted.at[lot_index, f'検査員{inspector_col_num}'] = old_inspector_name
                                                continue
                                            
                                            # 分割検査時間を更新
                                            result_df_sorted.at[lot_index, '分割検査時間'] = round(divided_time, 1)
                                            
                                            # 【重要】元の検査員から時間を引く前に、新しい検査員への追加が可能かチェック
                                            # 新しい検査員の現在の勤務時間を取得
                                            if new_code not in self.inspector_daily_assignments:
                                                self.inspector_daily_assignments[new_code] = {}
                                            if current_date not in self.inspector_daily_assignments[new_code]:
                                                self.inspector_daily_assignments[new_code][current_date] = 0.0
                                            
                                            new_daily_hours_before = self.inspector_daily_assignments[new_code][current_date]
                                            new_max_hours = self.get_inspector_max_hours(new_code, inspector_master_df)
                                            new_allowed_max_hours = new_max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                            
                                            # 10%超過を超える場合は、割り当て可能な時間を調整
                                            if new_daily_hours_before + divided_time > new_allowed_max_hours - WORK_HOURS_BUFFER:
                                                # 10%超過を超えない範囲で調整
                                                adjusted_divided_time = max(0.0, new_allowed_max_hours - WORK_HOURS_BUFFER - new_daily_hours_before)
                                                if adjusted_divided_time <= 0:
                                                    # 割り当て不可の場合は、再割当をスキップ（元の検査員に戻す必要はない、まだ変更していないため）
                                                    self.log_message(
                                                        f"偏り是正: 検査員 {replacement_inspector['氏名']} ({new_code}) は10%超過を超えるため、再割当てをスキップします（現在: {new_daily_hours_before:.1f}h + {divided_time:.1f}h > {new_allowed_max_hours:.1f}h）",
                                                        debug=True,
                                                    )
                                                    # 結果データフレームを元に戻す（一時的に変更していたため）
                                                    result_df_sorted.at[lot_index, f'検査員{inspector_col_num}'] = old_inspector_name
                                                    continue
                                                divided_time = adjusted_divided_time
                                                # 分割検査時間も更新
                                                result_df_sorted.at[lot_index, '分割検査時間'] = round(divided_time, 1)
                                        
                                        # 履歴を更新（元の検査員から時間を引く）
                                        old_daily = self.inspector_daily_assignments.get(overloaded_code, {}).get(current_date, 0.0)
                                        old_total = self.inspector_work_hours.get(overloaded_code, 0.0)
                                        self.inspector_daily_assignments[overloaded_code][current_date] = max(0.0, old_daily - divided_time)
                                        self.inspector_work_hours[overloaded_code] = max(0.0, old_total - divided_time)
                                        
                                        # 品番別累計時間も更新
                                        if overloaded_code in self.inspector_product_hours:
                                            if product_number in self.inspector_product_hours[overloaded_code]:
                                                self.inspector_product_hours[overloaded_code][product_number] = max(
                                                    0.0, 
                                                    self.inspector_product_hours[overloaded_code][product_number] - divided_time
                                                )
                                        
                                            # 新しい検査員に時間を追加（10%超過チェックは既に実施済み）
                                        self.inspector_daily_assignments[new_code][current_date] += divided_time
                                        if new_code not in self.inspector_work_hours:
                                            self.inspector_work_hours[new_code] = 0.0
                                        self.inspector_work_hours[new_code] += divided_time
                                        
                                        # 品番別累計時間も更新
                                        if new_code not in self.inspector_product_hours:
                                            self.inspector_product_hours[new_code] = {}
                                        self.inspector_product_hours[new_code][product_number] = (
                                            self.inspector_product_hours[new_code].get(product_number, 0.0) + divided_time
                                        )
                                        
                                        # 当日洗浄上がり品のロットの場合、same_day_cleaning_inspectorsを更新（品番単位・品名単位）
                                        if is_same_day_cleaning_lot:
                                            # 品名を取得
                                            product_name = row.get('品名', '')
                                            product_name_str = str(product_name).strip() if pd.notna(product_name) else ''
                                            
                                            # 元の検査員がこの品番に割り当てられていた場合、削除（品番単位）
                                            if product_number in self.same_day_cleaning_inspectors:
                                                self.same_day_cleaning_inspectors[product_number].discard(overloaded_code)
                                            
                                            # 【改善】品名単位の制約も更新
                                            if product_name_str:
                                                # 元の検査員がこの品名に割り当てられていた場合、削除（品名単位）
                                                if product_name_str in self.same_day_cleaning_inspectors_by_product_name:
                                                    self.same_day_cleaning_inspectors_by_product_name[product_name_str].discard(overloaded_code)
                                                # 新しい検査員をこの品名に割り当てられた検査員として記録（品名単位）
                                                self.same_day_cleaning_inspectors_by_product_name.setdefault(product_name_str, set()).add(new_code)
                                            
                                            # 新しい検査員をこの品番に割り当てられた検査員として記録（品番単位）
                                            self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(new_code)
                                        
                                        # チーム情報を更新
                                        self.update_team_info(result_df_sorted, lot_index, inspector_master_df, show_skill_values)
                                        
                                        # swap成功時にカウント
                                        self.swap_count += 1
                                        
                                        reassignment_count += 1
                                        
                                        # 【追加】このパスで実行された再割当の履歴を記録
                                        pass_reassignment_history.append({
                                            'lot_index': lot_index,
                                            'old_inspector_code': overloaded_code,
                                            'old_inspector_name': old_inspector_name,
                                            'new_inspector_code': new_code,
                                            'new_inspector_name': replacement_inspector['氏名'],
                                            'product_number': product_number,
                                            'divided_time': divided_time,
                                            'inspector_col_num': inspector_col_num,
                                            'old_daily': old_daily,
                                            'old_total': old_total,
                                            'is_same_day_cleaning_lot': is_same_day_cleaning_lot,
                                            'product_name': row.get('品名', '') if is_same_day_cleaning_lot else '',
                                        })
                                            
                                        self.log_message(
                                            f"偏り是正: '{old_inspector_name}' ({overloaded_hours:.1f}h) → "
                                            f"'{replacement_inspector['氏名']}' ({self.inspector_work_hours[new_code]:.1f}h) "
                                            f"(品番: {product_number}, 出荷予定日: {row['出荷予定日']})"
                                        )
                                        
                                        # 【追加】再割当実行後、10%超過を即座に検証
                                        new_daily_hours_after = self.inspector_daily_assignments.get(new_code, {}).get(current_date, 0.0)
                                        new_max_hours_after = self.get_inspector_max_hours(new_code, inspector_master_df)
                                        new_allowed_max_hours_after = new_max_hours_after * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                        if new_daily_hours_after > new_allowed_max_hours_after - WORK_HOURS_BUFFER:
                                                self.log_message(
                                                    f"⚠️ 偏り是正: 再割当実行後に10%超過を検出 - 検査員 {replacement_inspector['氏名']} ({new_code}) "
                                                    f"現在: {new_daily_hours_after:.1f}h > {new_allowed_max_hours_after:.1f}h - {WORK_HOURS_BUFFER:.2f}h "
                                                    f"（ロットインデックス: {lot_index}）",
                                                    level='warning',
                                                )
                                                # 【修正】10%超過を超える場合は、割り当てを元に戻す
                                                # 新しい検査員から時間を引く
                                                self.inspector_daily_assignments[new_code][current_date] = max(0.0, new_daily_hours_after - divided_time)
                                                self.inspector_work_hours[new_code] = max(0.0, self.inspector_work_hours.get(new_code, 0.0) - divided_time)
                                                if new_code in self.inspector_product_hours and product_number in self.inspector_product_hours[new_code]:
                                                    self.inspector_product_hours[new_code][product_number] = max(
                                                        0.0,
                                                        self.inspector_product_hours[new_code][product_number] - divided_time
                                                    )
                                                
                                                # 元の検査員に時間を戻す
                                                self.inspector_daily_assignments[overloaded_code][current_date] = old_daily
                                                self.inspector_work_hours[overloaded_code] = old_total
                                                if overloaded_code in self.inspector_product_hours:
                                                    if product_number in self.inspector_product_hours[overloaded_code]:
                                                        self.inspector_product_hours[overloaded_code][product_number] = (
                                                            self.inspector_product_hours[overloaded_code].get(product_number, 0.0) + divided_time
                                                        )
                                                
                                                # 結果データフレームを元に戻す
                                                result_df_sorted.at[lot_index, f'検査員{inspector_col_num}'] = old_inspector_name
                                                result_df_sorted.at[lot_index, '分割検査時間'] = row.get('分割検査時間', divided_time)
                                                
                                                self.log_message(
                                                    f"偏り是正: 10%超過を超えるため、割り当てを元に戻しました（ロットインデックス: {lot_index}）",
                                                    debug=True,
                                                )
                                                continue
                                        
                                        # 再割当て後、多忙な検査員のリストを更新
                                        overloaded_hours = self.inspector_work_hours.get(overloaded_code, 0.0)
                                        if overloaded_hours <= avg_hours_pass * 1.1:
                                            # この検査員はもう多忙ではないので終了
                                            break
                        
                            total_reassignment_count += reassignment_count
                            if reassignment_count > 0:
                                self.log_message(f"偏り是正パス{pass_num + 1}: {reassignment_count}件の再割当てを実行しました")
                            
                            # 【追加】各パス終了時に、inspector_daily_assignmentsを再計算して10%超過を検証
                            self.log_message(f"偏り是正パス{pass_num + 1}: パス終了時にinspector_daily_assignmentsを再計算して検証中...", debug=True)
                            # result_df_sortedから実際の割り当てを読み取って、inspector_daily_assignmentsを再計算
                            result_cols_bias_verify = {col: idx for idx, col in enumerate(result_df_sorted.columns)}
                            inspector_daily_assignments_verify = {}
                            
                            for row_idx, row in enumerate(result_df_sorted.itertuples(index=False)):
                                index = result_df_sorted.index[row_idx]
                                product_number = row[result_cols_bias_verify['品番']]
                                divided_time = row[result_cols_bias_verify.get('分割検査時間', -1)] if '分割検査時間' in result_cols_bias_verify else 0.0
                                if divided_time == -1:
                                    divided_time = 0.0
                                
                                # 各検査員の割り当てを確認
                                for i in range(1, 6):
                                    inspector_col = f'検査員{i}'
                                    inspector_col_idx = result_cols_bias_verify.get(inspector_col, -1)
                                    if inspector_col_idx != -1:
                                        inspector_name_raw = row[inspector_col_idx]
                                        if pd.notna(inspector_name_raw) and str(inspector_name_raw).strip() != '':
                                            inspector_name = str(inspector_name_raw).strip()
                                            if '(' in inspector_name:
                                                inspector_name = inspector_name.split('(')[0].strip()
                                            
                                            if not inspector_name:
                                                continue
                                            
                                            # 検査員コードを取得
                                            inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                            if not inspector_info.empty:
                                                inspector_code = inspector_info.iloc[0]['#ID']
                                                
                                                # 履歴を初期化
                                                if inspector_code not in inspector_daily_assignments_verify:
                                                    inspector_daily_assignments_verify[inspector_code] = {}
                                                if current_date not in inspector_daily_assignments_verify[inspector_code]:
                                                    inspector_daily_assignments_verify[inspector_code][current_date] = 0.0
                                                
                                                # 履歴を累積
                                                inspector_daily_assignments_verify[inspector_code][current_date] += divided_time
                            
                            # 10%超過を検証
                            overrun_detected = False
                            overrun_inspector_codes = set()  # 10%超過を検出した検査員コードのセット
                            for inspector_code, daily_dict in inspector_daily_assignments_verify.items():
                                daily_hours = daily_dict.get(current_date, 0.0)
                                max_hours = self.get_inspector_max_hours(inspector_code, inspector_master_df)
                                allowed_max_hours = max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                if daily_hours > allowed_max_hours - WORK_HOURS_BUFFER:
                                    inspector_info = inspector_master_df[inspector_master_df['#ID'] == inspector_code]
                                    inspector_name = inspector_info.iloc[0]['#氏名'] if not inspector_info.empty else inspector_code
                                    self.log_message(
                                        f"⚠️ 偏り是正パス{pass_num + 1}: 10%超過を検出 - 検査員 '{inspector_name}' ({inspector_code}) "
                                        f"{daily_hours:.1f}h > {allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h",
                                        level='warning',
                                    )
                                    overrun_detected = True
                                    overrun_inspector_codes.add(inspector_code)
                            
                            if overrun_detected:
                                # 【改善】10%超過を検出した場合、10%超過を引き起こす再割当のみを元に戻す
                                # これにより、偏り是正の効果を維持しながら、10%超過を防ぐ
                                reverted_count = 0
                                for history_item in reversed(pass_reassignment_history):
                                    new_code = history_item['new_inspector_code']
                                    # 10%超過を検出した検査員に関連する再割当のみを元に戻す
                                    if new_code not in overrun_inspector_codes:
                                        continue
                                    
                                    reverted_count += 1
                                    lot_index = history_item['lot_index']
                                    old_code = history_item['old_inspector_code']
                                    old_name = history_item['old_inspector_name']
                                    new_code = history_item['new_inspector_code']
                                    new_name = history_item['new_inspector_name']
                                    product_number = history_item['product_number']
                                    divided_time = history_item['divided_time']
                                    inspector_col_num = history_item['inspector_col_num']
                                    old_daily = history_item['old_daily']
                                    old_total = history_item['old_total']
                                    is_same_day_cleaning_lot = history_item['is_same_day_cleaning_lot']
                                    product_name = history_item['product_name']
                                    
                                    # 新しい検査員から時間を引く
                                    if new_code in self.inspector_daily_assignments and current_date in self.inspector_daily_assignments[new_code]:
                                        self.inspector_daily_assignments[new_code][current_date] = max(0.0, self.inspector_daily_assignments[new_code][current_date] - divided_time)
                                    if new_code in self.inspector_work_hours:
                                        self.inspector_work_hours[new_code] = max(0.0, self.inspector_work_hours[new_code] - divided_time)
                                    if new_code in self.inspector_product_hours and product_number in self.inspector_product_hours[new_code]:
                                        self.inspector_product_hours[new_code][product_number] = max(
                                            0.0,
                                            self.inspector_product_hours[new_code][product_number] - divided_time
                                        )
                                    
                                    # 元の検査員に時間を戻す
                                    if old_code not in self.inspector_daily_assignments:
                                        self.inspector_daily_assignments[old_code] = {}
                                    if current_date not in self.inspector_daily_assignments[old_code]:
                                        self.inspector_daily_assignments[old_code][current_date] = 0.0
                                    self.inspector_daily_assignments[old_code][current_date] = old_daily
                                    self.inspector_work_hours[old_code] = old_total
                                    if old_code not in self.inspector_product_hours:
                                        self.inspector_product_hours[old_code] = {}
                                    if product_number not in self.inspector_product_hours[old_code]:
                                        self.inspector_product_hours[old_code][product_number] = 0.0
                                    self.inspector_product_hours[old_code][product_number] = (
                                        self.inspector_product_hours[old_code].get(product_number, 0.0) + divided_time
                                    )
                                    
                                    # 当日洗浄上がり品のロットの場合、same_day_cleaning_inspectorsを更新（品番単位・品名単位）
                                    if is_same_day_cleaning_lot:
                                        # 新しい検査員を削除（品番単位）
                                        if product_number in self.same_day_cleaning_inspectors:
                                            self.same_day_cleaning_inspectors[product_number].discard(new_code)
                                        
                                        # 品名単位の制約も更新
                                        if product_name:
                                            product_name_str = str(product_name).strip() if pd.notna(product_name) else ''
                                            if product_name_str and product_name_str in self.same_day_cleaning_inspectors_by_product_name:
                                                self.same_day_cleaning_inspectors_by_product_name[product_name_str].discard(new_code)
                                        
                                        # 元の検査員を追加（品番単位）
                                        self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(old_code)
                                        
                                        # 品名単位の制約も更新
                                        if product_name:
                                            product_name_str = str(product_name).strip() if pd.notna(product_name) else ''
                                            if product_name_str:
                                                self.same_day_cleaning_inspectors_by_product_name.setdefault(product_name_str, set()).add(old_code)
                                    
                                    # 結果データフレームを元に戻す
                                    result_df_sorted.at[lot_index, f'検査員{inspector_col_num}'] = old_name
                                    
                                    self.log_message(
                                        f"偏り是正パス{pass_num + 1}: 10%超過を引き起こす再割当を元に戻しました - '{new_name}' → '{old_name}' (ロットインデックス: {lot_index})",
                                        debug=True,
                                    )
                                
                                # 元に戻した再割当を履歴リストから削除
                                pass_reassignment_history = [h for h in pass_reassignment_history if h != history_item]
                                
                                if reverted_count > 0:
                                    self.log_message(
                                        f"偏り是正パス{pass_num + 1}: 10%超過を検出したため、{reverted_count}件の再割当を元に戻しました（偏り是正の効果を維持）",
                                        level='warning',
                                    )
                                
                                # inspector_daily_assignmentsを再計算した値で更新（元に戻した後の状態）
                                self.log_message(f"偏り是正パス{pass_num + 1}: inspector_daily_assignmentsを再計算した値で更新しました", debug=True)
                                
                                # 【追加】10%超過を検出した検査員から、他の検査員へ再割当を行う処理
                                # 偏り是正の段階で既に存在していた10%超過を解消するため
                                if overrun_inspector_codes:
                                    # 【改善】無限ループ防止: 同じ検査員が何度も10%超過を検出された場合の追跡
                                    if not hasattr(self, '_overrun_inspector_history'):
                                        self._overrun_inspector_history = {}  # {inspector_code: count}
                                    
                                    self.log_message(
                                        f"偏り是正パス{pass_num + 1}: 10%超過を検出した検査員から、他の検査員へ再割当を行います",
                                        level='warning',
                                    )
                                    
                                    # 10%超過を検出した検査員のロットを探して、他の検査員へ再割当を行う
                                    for overrun_code in overrun_inspector_codes:
                                        # 【改善】無限ループ防止: 同じ検査員が3回以上10%超過を検出された場合、スキップ
                                        if self._overrun_inspector_history.get(overrun_code, 0) >= 3:
                                            inspector_info_overrun_skip = inspector_master_df[inspector_master_df['#ID'] == overrun_code]
                                            if not inspector_info_overrun_skip.empty:
                                                overrun_name_skip = inspector_info_overrun_skip.iloc[0]['#氏名']
                                                self.log_message(
                                                    f"偏り是正パス{pass_num + 1}: 検査員 '{overrun_name_skip}' ({overrun_code}) は3回以上10%超過を検出されているため、再割当をスキップします（無限ループ防止）",
                                                    level='warning',
                                                )
                                            continue
                                        inspector_info_overrun = inspector_master_df[inspector_master_df['#ID'] == overrun_code]
                                        if inspector_info_overrun.empty:
                                            continue
                                        overrun_name = inspector_info_overrun.iloc[0]['#氏名']
                                        overrun_max_hours = self.get_inspector_max_hours(overrun_code, inspector_master_df)
                                        overrun_allowed_max_hours = overrun_max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                        overrun_current_hours = inspector_daily_assignments_verify.get(overrun_code, {}).get(current_date, 0.0)
                                        overrun_excess = overrun_current_hours - (overrun_allowed_max_hours - WORK_HOURS_BUFFER)
                                        
                                        if overrun_excess <= 0:
                                            continue
                                        
                                        # この検査員が割り当てられているロットを探す
                                        # 【改善】10%超過を解消しやすいロット（時間が短い、制約が少ない）を優先的に処理するため、まずロットを収集してソート
                                        overrun_lots = []
                                        for row_idx, row in enumerate(result_df_sorted.itertuples(index=False)):
                                            lot_index_overrun = result_df_sorted.index[row_idx]
                                            product_number_overrun = row[result_cols_bias_verify['品番']]
                                            divided_time_overrun = row[result_cols_bias_verify.get('分割検査時間', -1)] if '分割検査時間' in result_cols_bias_verify else 0.0
                                            if divided_time_overrun == -1:
                                                divided_time_overrun = 0.0
                                            
                                            # この検査員が割り当てられているか確認
                                            found_overrun = False
                                            inspector_col_num_overrun = None
                                            for i in range(1, 6):
                                                inspector_col_overrun = f'検査員{i}'
                                                inspector_col_idx_overrun = result_cols_bias_verify.get(inspector_col_overrun, -1)
                                                if inspector_col_idx_overrun != -1:
                                                    inspector_name_raw_overrun = row[inspector_col_idx_overrun]
                                                    if pd.notna(inspector_name_raw_overrun) and str(inspector_name_raw_overrun).strip() != '':
                                                        inspector_name_overrun = str(inspector_name_raw_overrun).strip()
                                                        if '(' in inspector_name_overrun:
                                                            inspector_name_overrun = inspector_name_overrun.split('(')[0].strip()
                                                        
                                                        if inspector_name_overrun == overrun_name:
                                                            found_overrun = True
                                                            inspector_col_num_overrun = i
                                                            break
                                            
                                            if found_overrun:
                                                row_overrun = result_df_sorted.loc[lot_index_overrun]
                                                is_same_day_cleaning_lot_overrun = row_overrun.get('当日洗浄上がり品', False)
                                                # 制約の少ないロット（当日洗浄上がり品でない、時間が短い）を優先
                                                priority = 0 if is_same_day_cleaning_lot_overrun else 1  # 当日洗浄上がり品でない方が優先
                                                overrun_lots.append((priority, divided_time_overrun, lot_index_overrun, row_idx, row, inspector_col_num_overrun))
                                        
                                        # 優先順位: 1) 制約が少ない（当日洗浄上がり品でない）、2) 時間が短い順
                                        overrun_lots.sort(key=lambda x: (x[0], x[1]))
                                        
                                        reassignment_count_overrun = 0
                                        for priority, divided_time_overrun, lot_index_overrun, row_idx, row, inspector_col_num_overrun in overrun_lots:
                                            lot_index_overrun = result_df_sorted.index[row_idx]
                                            product_number_overrun = row[result_cols_bias_verify['品番']]
                                            divided_time_overrun = row[result_cols_bias_verify.get('分割検査時間', -1)] if '分割検査時間' in result_cols_bias_verify else 0.0
                                            if divided_time_overrun == -1:
                                                divided_time_overrun = 0.0
                                            
                                            # この検査員が割り当てられているか確認
                                            found_overrun = False
                                            inspector_col_num_overrun = None
                                            for i in range(1, 6):
                                                inspector_col_overrun = f'検査員{i}'
                                                inspector_col_idx_overrun = result_cols_bias_verify.get(inspector_col_overrun, -1)
                                                if inspector_col_idx_overrun != -1:
                                                    inspector_name_raw_overrun = row[inspector_col_idx_overrun]
                                                    if pd.notna(inspector_name_raw_overrun) and str(inspector_name_raw_overrun).strip() != '':
                                                        inspector_name_overrun = str(inspector_name_raw_overrun).strip()
                                                        if '(' in inspector_name_overrun:
                                                            inspector_name_overrun = inspector_name_overrun.split('(')[0].strip()
                                                        
                                                        if inspector_name_overrun == overrun_name:
                                                            found_overrun = True
                                                            inspector_col_num_overrun = i
                                                            break
                                            
                                            if not found_overrun:
                                                continue
                                            
                                            # ロット情報を取得
                                            row_overrun = result_df_sorted.loc[lot_index_overrun]
                                            process_number_overrun = row_overrun.get('工程番号', None)
                                            shipping_date_overrun = row_overrun.get('出荷予定日', None)
                                            is_same_day_cleaning_lot_overrun = row_overrun.get('当日洗浄上がり品', False)
                                            
                                            # 利用可能な検査員を取得
                                            available_inspectors_overrun = self.get_available_inspectors(
                                                product_number_overrun,
                                                process_number_overrun,
                                                skill_master_df,
                                                inspector_master_df,
                                                shipping_date=shipping_date_overrun,
                                                allow_new_team_fallback=True,
                                                process_master_df=process_master_df,
                                                inspection_target_keywords=inspection_target_keywords,
                                            )
                                            
                                            if not available_inspectors_overrun:
                                                continue
                                            
                                            # 余裕のある検査員をフィルタリング（10%超過を解消するため、緩和しない）
                                            replacement_candidates_overrun = []
                                            for insp_overrun in available_inspectors_overrun:
                                                candidate_code_overrun = insp_overrun['コード']
                                                candidate_name_overrun = insp_overrun['氏名']
                                                
                                                # 10%超過を検出した検査員を除外
                                                if candidate_code_overrun == overrun_code:
                                                    continue
                                                
                                                # 既に割り当てられている人は除外
                                                current_codes_overrun = []
                                                for i in range(1, 6):
                                                    inspector_col_overrun = f'検査員{i}'
                                                    inspector_value_overrun = row_overrun.get(inspector_col_overrun, '')
                                                    if pd.notna(inspector_value_overrun) and str(inspector_value_overrun).strip() != '':
                                                        inspector_name_check_overrun = str(inspector_value_overrun).strip()
                                                        if '(' in inspector_name_check_overrun:
                                                            inspector_name_check_overrun = inspector_name_check_overrun.split('(')[0].strip()
                                                        if inspector_name_check_overrun:
                                                            inspector_info_check_overrun = self._get_inspector_by_name(inspector_name_check_overrun, inspector_master_df)
                                                            if not inspector_info_check_overrun.empty:
                                                                current_codes_overrun.append(inspector_info_check_overrun.iloc[0]['#ID'])
                                                
                                                if candidate_code_overrun in current_codes_overrun:
                                                    continue
                                                
                                                # 当日洗浄上がり品のロットの場合、制約をチェック
                                                if is_same_day_cleaning_lot_overrun:
                                                    if product_number_overrun in self.same_day_cleaning_inspectors:
                                                        if candidate_code_overrun in self.same_day_cleaning_inspectors[product_number_overrun]:
                                                            continue
                                                    
                                                    product_name_overrun = row_overrun.get('品名', '')
                                                    product_name_str_overrun = str(product_name_overrun).strip() if pd.notna(product_name_overrun) else ''
                                                    if product_name_str_overrun:
                                                        if product_name_str_overrun in self.same_day_cleaning_inspectors_by_product_name:
                                                            if candidate_code_overrun in self.same_day_cleaning_inspectors_by_product_name[product_name_str_overrun]:
                                                                continue
                                                
                                                # 10%超過をチェック
                                                candidate_daily_hours_overrun = self.inspector_daily_assignments.get(candidate_code_overrun, {}).get(current_date, 0.0)
                                                candidate_max_hours_overrun = self.get_inspector_max_hours(candidate_code_overrun, inspector_master_df)
                                                candidate_allowed_max_hours_overrun = candidate_max_hours_overrun * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                                
                                                if candidate_daily_hours_overrun + divided_time_overrun > candidate_allowed_max_hours_overrun - WORK_HOURS_BUFFER:
                                                    continue
                                                
                                                # 品番上限をチェック
                                                if candidate_code_overrun in self.inspector_product_hours:
                                                    product_hours_overrun = self.inspector_product_hours[candidate_code_overrun].get(product_number_overrun, 0.0)
                                                    if product_hours_overrun + divided_time_overrun > 4.0:
                                                        continue
                                                
                                                replacement_candidates_overrun.append(insp_overrun)
                                            
                                            if not replacement_candidates_overrun:
                                                continue
                                            
                                            # 【改善】余裕のある検査員を選択（許容最大時間までの余裕が大きい順）
                                            # 10%超過を解消するため、より余裕がある検査員を優先的に選択
                                            def sort_key_candidate(x):
                                                candidate_code = x['コード']
                                                candidate_daily_hours = self.inspector_daily_assignments.get(candidate_code, {}).get(current_date, 0.0)
                                                candidate_max_hours = self.get_inspector_max_hours(candidate_code, inspector_master_df)
                                                candidate_allowed_max_hours = candidate_max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                                # 許容最大時間までの余裕
                                                remaining_capacity = candidate_allowed_max_hours - WORK_HOURS_BUFFER - candidate_daily_hours - divided_time_overrun
                                                # 余裕が大きい順（負の値は除外されているので、正の値のみ）
                                                return -remaining_capacity  # 負の値でソート（大きい順）
                                            
                                            replacement_candidates_overrun_sorted = sorted(
                                                replacement_candidates_overrun,
                                                key=sort_key_candidate
                                            )
                                            
                                            replacement_inspector_overrun = replacement_candidates_overrun_sorted[0]
                                            new_code_overrun = replacement_inspector_overrun['コード']
                                            new_name_overrun = replacement_inspector_overrun['氏名']
                                            
                                            # 再割当を実行
                                            # 元の検査員から時間を引く
                                            if overrun_code in self.inspector_daily_assignments and current_date in self.inspector_daily_assignments[overrun_code]:
                                                self.inspector_daily_assignments[overrun_code][current_date] = max(0.0, self.inspector_daily_assignments[overrun_code][current_date] - divided_time_overrun)
                                            if overrun_code in self.inspector_work_hours:
                                                self.inspector_work_hours[overrun_code] = max(0.0, self.inspector_work_hours[overrun_code] - divided_time_overrun)
                                            if overrun_code in self.inspector_product_hours and product_number_overrun in self.inspector_product_hours[overrun_code]:
                                                self.inspector_product_hours[overrun_code][product_number_overrun] = max(
                                                    0.0,
                                                    self.inspector_product_hours[overrun_code][product_number_overrun] - divided_time_overrun
                                                )
                                            
                                            # 新しい検査員に時間を追加
                                            if new_code_overrun not in self.inspector_daily_assignments:
                                                self.inspector_daily_assignments[new_code_overrun] = {}
                                            if current_date not in self.inspector_daily_assignments[new_code_overrun]:
                                                self.inspector_daily_assignments[new_code_overrun][current_date] = 0.0
                                            self.inspector_daily_assignments[new_code_overrun][current_date] += divided_time_overrun
                                            if new_code_overrun not in self.inspector_work_hours:
                                                self.inspector_work_hours[new_code_overrun] = 0.0
                                            self.inspector_work_hours[new_code_overrun] += divided_time_overrun
                                            if new_code_overrun not in self.inspector_product_hours:
                                                self.inspector_product_hours[new_code_overrun] = {}
                                            self.inspector_product_hours[new_code_overrun][product_number_overrun] = (
                                                self.inspector_product_hours[new_code_overrun].get(product_number_overrun, 0.0) + divided_time_overrun
                                            )
                                            
                                            # 結果データフレームを更新
                                            if show_skill_values:
                                                if replacement_inspector_overrun.get('is_new_team', False):
                                                    replacement_name_overrun = f"{new_name_overrun}(新)"
                                                else:
                                                    replacement_name_overrun = f"{new_name_overrun}({replacement_inspector_overrun['スキル']})"
                                            else:
                                                if replacement_inspector_overrun.get('is_new_team', False):
                                                    replacement_name_overrun = f"{new_name_overrun}(新)"
                                                else:
                                                    replacement_name_overrun = new_name_overrun
                                            
                                            result_df_sorted.at[lot_index_overrun, f'検査員{inspector_col_num_overrun}'] = replacement_name_overrun
                                            
                                            # 当日洗浄上がり品のロットの場合、same_day_cleaning_inspectorsを更新（品番単位・品名単位）
                                            if is_same_day_cleaning_lot_overrun:
                                                # 品名を取得
                                                product_name_overrun = row_overrun.get('品名', '')
                                                product_name_str_overrun = str(product_name_overrun).strip() if pd.notna(product_name_overrun) else ''
                                                
                                                # 元の検査員を削除（品番単位）
                                                if product_number_overrun in self.same_day_cleaning_inspectors:
                                                    old_inspectors_product = self.same_day_cleaning_inspectors[product_number_overrun].copy()
                                                    self.same_day_cleaning_inspectors[product_number_overrun].discard(overrun_code)
                                                    self.log_message(
                                                        f"偏り是正パス{pass_num + 1}: 10%超過解消のため再割当 - 当日洗浄上がり品の制約更新（品番単位）: "
                                                        f"品番 '{product_number_overrun}' から検査員 '{overrun_name}' ({overrun_code}) を削除",
                                                        debug=True,
                                                    )
                                                
                                                # 品名単位の制約も更新
                                                if product_name_str_overrun:
                                                    if product_name_str_overrun in self.same_day_cleaning_inspectors_by_product_name:
                                                        old_inspectors_name = self.same_day_cleaning_inspectors_by_product_name[product_name_str_overrun].copy()
                                                        self.same_day_cleaning_inspectors_by_product_name[product_name_str_overrun].discard(overrun_code)
                                                        self.log_message(
                                                            f"偏り是正パス{pass_num + 1}: 10%超過解消のため再割当 - 当日洗浄上がり品の制約更新（品名単位）: "
                                                            f"品名 '{product_name_str_overrun}' から検査員 '{overrun_name}' ({overrun_code}) を削除",
                                                            debug=True,
                                                        )
                                                
                                                # 新しい検査員を追加（品番単位）
                                                if product_number_overrun not in self.same_day_cleaning_inspectors:
                                                    self.same_day_cleaning_inspectors[product_number_overrun] = set()
                                                self.same_day_cleaning_inspectors[product_number_overrun].add(new_code_overrun)
                                                self.log_message(
                                                    f"偏り是正パス{pass_num + 1}: 10%超過解消のため再割当 - 当日洗浄上がり品の制約更新（品番単位）: "
                                                    f"品番 '{product_number_overrun}' に検査員 '{new_name_overrun}' ({new_code_overrun}) を追加",
                                                    debug=True,
                                                )
                                                
                                                # 品名単位の制約も更新
                                                if product_name_str_overrun:
                                                    if product_name_str_overrun not in self.same_day_cleaning_inspectors_by_product_name:
                                                        self.same_day_cleaning_inspectors_by_product_name[product_name_str_overrun] = set()
                                                    self.same_day_cleaning_inspectors_by_product_name[product_name_str_overrun].add(new_code_overrun)
                                                    self.log_message(
                                                        f"偏り是正パス{pass_num + 1}: 10%超過解消のため再割当 - 当日洗浄上がり品の制約更新（品名単位）: "
                                                        f"品名 '{product_name_str_overrun}' に検査員 '{new_name_overrun}' ({new_code_overrun}) を追加",
                                                        debug=True,
                                                    )
                                            
                                            # チーム情報を更新
                                            self.update_team_info(result_df_sorted, lot_index_overrun, inspector_master_df, show_skill_values)
                                            
                                            # 【改善】再割当後に履歴を更新
                                            if overrun_code in self.inspector_daily_assignments:
                                                if current_date in self.inspector_daily_assignments[overrun_code]:
                                                    self.inspector_daily_assignments[overrun_code][current_date] -= divided_time_overrun
                                                    if self.inspector_daily_assignments[overrun_code][current_date] < 0:
                                                        self.inspector_daily_assignments[overrun_code][current_date] = 0.0
                                            if overrun_code in self.inspector_work_hours:
                                                self.inspector_work_hours[overrun_code] -= divided_time_overrun
                                                if self.inspector_work_hours[overrun_code] < 0:
                                                    self.inspector_work_hours[overrun_code] = 0.0
                                            if overrun_code in self.inspector_product_hours:
                                                if product_number_overrun in self.inspector_product_hours[overrun_code]:
                                                    self.inspector_product_hours[overrun_code][product_number_overrun] -= divided_time_overrun
                                                    if self.inspector_product_hours[overrun_code][product_number_overrun] < 0:
                                                        self.inspector_product_hours[overrun_code][product_number_overrun] = 0.0
                                            
                                            if new_code_overrun not in self.inspector_daily_assignments:
                                                self.inspector_daily_assignments[new_code_overrun] = {}
                                            if current_date not in self.inspector_daily_assignments[new_code_overrun]:
                                                self.inspector_daily_assignments[new_code_overrun][current_date] = 0.0
                                            if new_code_overrun not in self.inspector_work_hours:
                                                self.inspector_work_hours[new_code_overrun] = 0.0
                                            if new_code_overrun not in self.inspector_product_hours:
                                                self.inspector_product_hours[new_code_overrun] = {}
                                            if product_number_overrun not in self.inspector_product_hours[new_code_overrun]:
                                                self.inspector_product_hours[new_code_overrun][product_number_overrun] = 0.0
                                            
                                            self.inspector_daily_assignments[new_code_overrun][current_date] += divided_time_overrun
                                            self.inspector_work_hours[new_code_overrun] += divided_time_overrun
                                            self.inspector_product_hours[new_code_overrun][product_number_overrun] += divided_time_overrun
                                            
                                            reassignment_count_overrun += 1
                                            total_reassignment_count += 1
                                            
                                            self.log_message(
                                                f"偏り是正パス{pass_num + 1}: 10%超過解消のため再割当 - '{overrun_name}' → '{new_name_overrun}' (ロットインデックス: {lot_index_overrun}, 品番: {product_number_overrun})",
                                                debug=True,
                                            )
                                            
                                            # 超過分が解消されたか確認（履歴を更新した後の値を使用）
                                            overrun_current_hours_after = self.inspector_daily_assignments.get(overrun_code, {}).get(current_date, 0.0)
                                            
                                            if overrun_current_hours_after <= overrun_allowed_max_hours - WORK_HOURS_BUFFER:
                                                self.log_message(
                                                    f"偏り是正パス{pass_num + 1}: 検査員 '{overrun_name}' ({overrun_code}) の10%超過が解消されました ({overrun_current_hours_after:.1f}h <= {overrun_allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h)",
                                                    level='info',
                                                )
                                                # 【改善】10%超過が解消された場合、カウントをリセットして次のパスでも再割当を試みられるようにする
                                                if overrun_code in self._overrun_inspector_history:
                                                    self._overrun_inspector_history[overrun_code] = 0
                                                break
                                            
                                            # 【改善】10%超過が解消されない場合、最大10件まで再割当を試みる
                                            if reassignment_count_overrun >= 10:
                                                self.log_message(
                                                    f"偏り是正パス{pass_num + 1}: 検査員 '{overrun_name}' ({overrun_code}) の10%超過解消のため、最大10件まで再割当を試みましたが、完全には解消されませんでした（現在: {overrun_current_hours_after:.1f}h > {overrun_allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h）",
                                                    level='warning',
                                                )
                                                # 【改善】再割当を試みたが解消できなかった場合、カウントを増やす
                                                if overrun_code not in self._overrun_inspector_history:
                                                    self._overrun_inspector_history[overrun_code] = 0
                                                self._overrun_inspector_history[overrun_code] += 1
                                                break
                                        
                                        # 【改善】再割当を試みたが解消できなかった場合、カウントを増やす
                                        if reassignment_count_overrun > 0:
                                            # 最終的な超過状態を確認
                                            overrun_final_hours = self.inspector_daily_assignments.get(overrun_code, {}).get(current_date, 0.0)
                                            if overrun_final_hours > overrun_allowed_max_hours - WORK_HOURS_BUFFER:
                                                # 解消されていない場合、カウントを増やす（解消された場合は既にリセット済み）
                                                if overrun_code not in self._overrun_inspector_history or self._overrun_inspector_history[overrun_code] == 0:
                                                    # カウントがない場合（解消されなかった場合）のみ増やす
                                                    if overrun_code not in self._overrun_inspector_history:
                                                        self._overrun_inspector_history[overrun_code] = 0
                                                    self._overrun_inspector_history[overrun_code] += 1
                                        
                                        if reassignment_count_overrun > 0:
                                            self.log_message(
                                                f"偏り是正パス{pass_num + 1}: 10%超過解消のため、{reassignment_count_overrun}件の再割当を実行しました",
                                                level='info',
                                            )
                        
                        # 結果を更新
                        result_df = result_df_sorted
                        
                        # 最終的な偏りを再計算
                        total_hours_after = sum(self.inspector_work_hours.values())
                        active_inspectors_after = [code for code in self.inspector_work_hours.keys() 
                                                   if self.inspector_work_hours.get(code, 0.0) > 0]
                        if active_inspectors_after:
                            avg_hours_after = total_hours_after / len(active_inspectors_after)
                            max_hours_after = max(self.inspector_work_hours.values())
                            min_hours_after = min([self.inspector_work_hours.get(code, 0.0) for code in active_inspectors_after])
                            imbalance_after = max_hours_after - min_hours_after
                            
                            # 【追加】最終的な標準偏差を計算
                            hours_list_after = [self.inspector_work_hours.get(code, 0.0) for code in active_inspectors_after]
                            if len(hours_list_after) > 1:
                                import statistics
                                std_dev_after = statistics.stdev(hours_list_after)
                            else:
                                std_dev_after = 0.0
                            
                            # 【追加】総合的な改善率を計算
                            total_improvement_ratio = (imbalance - imbalance_after) / imbalance if imbalance > 0 else 0.0
                            std_improvement_ratio = (bias_correction_metrics['initial_std_deviation'] - std_dev_after) / bias_correction_metrics['initial_std_deviation'] if bias_correction_metrics['initial_std_deviation'] > 0 else 0.0
                            
                            # メトリクスを更新
                            bias_correction_metrics['total_reassignments'] = total_reassignment_count
                            bias_correction_metrics['final_imbalance'] = imbalance_after
                            bias_correction_metrics['final_std_deviation'] = std_dev_after
                            bias_correction_metrics['total_improvement_ratio'] = total_improvement_ratio
                            bias_correction_metrics['std_improvement_ratio'] = std_improvement_ratio
                            
                            self.log_message(
                                f"偏り是正完了: 平均 {avg_hours_after:.1f}h, 最大 {max_hours_after:.1f}h, "
                                f"最小 {min_hours_after:.1f}h, 偏り {imbalance_after:.1f}h "
                                f"(改善: {imbalance - imbalance_after:.1f}h, 改善率: {total_improvement_ratio * 100:.1f}%), "
                                f"標準偏差: {std_dev_after:.2f}h (改善: {bias_correction_metrics['initial_std_deviation'] - std_dev_after:.2f}h, "
                                f"改善率: {std_improvement_ratio * 100:.1f}%), "
                                f"総再割当て数: {total_reassignment_count}件, 実行パス数: {pass_num + 1}",
                                level='info',
                            )
                            
                            # 【追加】詳細メトリクスをdebugログに記録
                            self.log_message(
                                f"偏り是正詳細メトリクス: "
                                f"初期偏り={bias_correction_metrics['initial_imbalance']:.1f}h, "
                                f"最終偏り={bias_correction_metrics['final_imbalance']:.1f}h, "
                                f"初期標準偏差={bias_correction_metrics['initial_std_deviation']:.2f}h, "
                                f"最終標準偏差={bias_correction_metrics['final_std_deviation']:.2f}h, "
                                f"各パスの偏り={[f'{im:.1f}' for im in bias_correction_metrics['pass_imbalances']]}, "
                                f"各パスの標準偏差={[f'{sd:.2f}' for sd in bias_correction_metrics['pass_std_deviations']]}, "
                                f"各パスの改善率={[f'{ir*100:.2f}%' for ir in bias_correction_metrics['pass_improvements']]}",
                                debug=True,
                            )
                            
                            # 【追加】バッファ効果測定メトリクスを出力
                            if self.buffer_usage_metrics['total_checks'] > 0:
                                buffer_exceeded_rate = (self.buffer_usage_metrics['buffer_exceeded_count'] / self.buffer_usage_metrics['total_checks']) * 100
                                avg_exceeded_by = sum(self.buffer_usage_metrics['buffer_exceeded_by']) / len(self.buffer_usage_metrics['buffer_exceeded_by']) if self.buffer_usage_metrics['buffer_exceeded_by'] else 0.0
                                max_exceeded_by = max(self.buffer_usage_metrics['buffer_exceeded_by']) if self.buffer_usage_metrics['buffer_exceeded_by'] else 0.0
                                
                                self.log_message(
                                    f"勤務時間バッファ効果測定メトリクス: "
                                    f"総チェック回数={self.buffer_usage_metrics['total_checks']}, "
                                    f"バッファ超過回数={self.buffer_usage_metrics['buffer_exceeded_count']} "
                                    f"({buffer_exceeded_rate:.1f}%), "
                                    f"平均超過量={avg_exceeded_by:.3f}h, "
                                    f"最大超過量={max_exceeded_by:.3f}h, "
                                    f"動的バッファ調整回数={self.buffer_usage_metrics['dynamic_buffer_adjustments']}",
                                    debug=True,
                                )
                        
                            self.log_message(f"偏り是正: 合計{total_reassignment_count}件の再割当てを実行しました")
            
            perf_logger.debug(
                "PERF {}: {:.1f} ms",
                "inspector_assignment.optimize.phase2.total",
                (perf_counter() - _t_perf_phase2_total) * 1000.0,
            )

            # フェーズ2.5: 偏り是正後の最終検証（勤務時間超過の再チェック）
            self.log_message("全体最適化フェーズ2.5: 偏り是正後の最終検証を開始")
            _t_perf_phase2_5_total = perf_counter()
            
            # 履歴を再計算
            self.inspector_daily_assignments = {}
            self.inspector_work_hours = {}
            self.inspector_product_hours = {}
            
            # 列インデックスを事前に取得（高速化：itertuples()を使用）
            product_col_idx = result_df.columns.get_loc('品番')
            divided_time_col_idx = result_df.columns.get_loc('分割検査時間')
            inspector_col_indices = [result_df.columns.get_loc(f'検査員{i}') for i in range(1, 6)]
            
            for row_tuple in result_df.itertuples(index=True):
                index = row_tuple[0]  # インデックス
                product_number = row_tuple[product_col_idx + 1]  # itertuplesはインデックスを含むため+1
                divided_time = row_tuple[divided_time_col_idx + 1] if divided_time_col_idx < len(row_tuple) - 1 else 0.0
                
                for i in range(1, 6):
                    inspector_col_idx = inspector_col_indices[i - 1]
                    inspector_value = row_tuple[inspector_col_idx + 1] if inspector_col_idx < len(row_tuple) - 1 else None
                    if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                        inspector_name = str(inspector_value).strip()
                        if '(' in inspector_name:
                            inspector_name = inspector_name.split('(')[0].strip()
                        if not inspector_name:
                            continue

                        inspector_code = inspector_name_to_id.get(inspector_name)
                        if inspector_code is None:
                            inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                            if not inspector_info.empty:
                                inspector_code = inspector_info.iloc[0]['#ID']
                            else:
                                continue
                            
                            if inspector_code not in self.inspector_daily_assignments:
                                self.inspector_daily_assignments[inspector_code] = {}
                            if current_date not in self.inspector_daily_assignments[inspector_code]:
                                self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                            if inspector_code not in self.inspector_work_hours:
                                self.inspector_work_hours[inspector_code] = 0.0
                            if inspector_code not in self.inspector_product_hours:
                                self.inspector_product_hours[inspector_code] = {}
                            if product_number not in self.inspector_product_hours[inspector_code]:
                                self.inspector_product_hours[inspector_code][product_number] = 0.0
                            
                            self.inspector_daily_assignments[inspector_code][current_date] += divided_time
                            self.inspector_work_hours[inspector_code] += divided_time
                            self.inspector_product_hours[inspector_code][product_number] += divided_time
            
            # 【修正】フェーズ2.5で10%超過を検出する処理を、違反検出処理の前に実行する
            # （違反が検出されなかった場合でも、10%超過を検出した検査員から他の検査員へ再割当を行う）
            overrun_inspector_codes_phase2_5 = set()
            for inspector_code, daily_dict in self.inspector_daily_assignments.items():
                daily_hours = daily_dict.get(current_date, 0.0)
                max_hours = self.get_inspector_max_hours(inspector_code, inspector_master_df)
                allowed_max_hours = max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                if daily_hours > allowed_max_hours - WORK_HOURS_BUFFER:
                    inspector_info = inspector_master_df[inspector_master_df['#ID'] == inspector_code]
                    inspector_name = inspector_info.iloc[0]['#氏名'] if not inspector_info.empty else inspector_code
                    self.log_message(
                        f"⚠️ フェーズ2.5: 10%超過を検出 - 検査員 '{inspector_name}' ({inspector_code}) "
                        f"{daily_hours:.1f}h > {allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h",
                        level='warning',
                    )
                    overrun_inspector_codes_phase2_5.add(inspector_code)
            
            if overrun_inspector_codes_phase2_5:
                self.log_message(
                    f"フェーズ2.5: 10%超過を検出した検査員から、他の検査員へ再割当を行います",
                    level='warning',
                )
                
                # 10%超過を検出した検査員のロットを探して、他の検査員へ再割当を行う
                for overrun_code in overrun_inspector_codes_phase2_5:
                    inspector_info_overrun = inspector_master_df[inspector_master_df['#ID'] == overrun_code]
                    if inspector_info_overrun.empty:
                        continue
                    overrun_name = inspector_info_overrun.iloc[0]['#氏名']
                    overrun_max_hours = self.get_inspector_max_hours(overrun_code, inspector_master_df)
                    overrun_allowed_max_hours = overrun_max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                    overrun_current_hours = self.inspector_daily_assignments.get(overrun_code, {}).get(current_date, 0.0)
                    overrun_excess = overrun_current_hours - (overrun_allowed_max_hours - WORK_HOURS_BUFFER)
                    
                    if overrun_excess <= 0:
                        continue
                    
                    self.log_message(
                        f"フェーズ2.5: 検査員 '{overrun_name}' ({overrun_code}) の超過分 {overrun_excess:.1f}h を他の検査員へ再割当します",
                        level='warning',
                    )
                    
                    # この検査員が割り当てられているロットを取得（出荷予定日順、FIFO維持）
                    assigned_lots_overrun = []
                    for row_idx, row in enumerate(result_df.itertuples(index=True)):
                        index_overrun = row[0]
                        product_number_overrun = row[product_col_idx + 1]
                        divided_time_overrun = row[divided_time_col_idx + 1] if divided_time_col_idx < len(row) - 1 else 0.0
                        
                        for i in range(1, 6):
                            inspector_col_idx_overrun = inspector_col_indices[i - 1]
                            inspector_value_overrun = row[inspector_col_idx_overrun + 1] if inspector_col_idx_overrun < len(row) - 1 else None
                            if pd.notna(inspector_value_overrun) and str(inspector_value_overrun).strip() != '':
                                inspector_name_overrun = str(inspector_value_overrun).strip()
                                if '(' in inspector_name_overrun:
                                    inspector_name_overrun = inspector_name_overrun.split('(')[0].strip()
                                if not inspector_name_overrun:
                                    continue
                                
                                inspector_info_check = self._get_inspector_by_name(inspector_name_overrun, inspector_master_df)
                                if not inspector_info_check.empty:
                                    lot_inspector_code = inspector_info_check.iloc[0]['#ID']
                                    if lot_inspector_code == overrun_code:
                                        row_series = result_df.iloc[index_overrun]
                                        assigned_lots_overrun.append({
                                            'index': index_overrun,
                                            'product_number': product_number_overrun,
                                            'divided_time': divided_time_overrun,
                                            'inspection_time': row_series.get('検査時間', 0),
                                            'shipping_date': row_series.get('出荷予定日', pd.Timestamp.max),
                                            'row': row_series
                                        })
                                        break
                    
                    # 【改善】10%超過を解消しやすいロット（時間が短い、制約が少ない）を優先的に処理
                    # 優先順位: 1) 制約が少ない（当日洗浄上がり品でない）、2) 時間が短い順、3) 出荷予定日順
                    for lot_info_overrun in assigned_lots_overrun:
                        row_series_overrun = lot_info_overrun['row']
                        is_same_day_cleaning_lot_overrun = row_series_overrun.get('当日洗浄上がり品', False)
                        lot_info_overrun['priority'] = 0 if is_same_day_cleaning_lot_overrun else 1  # 当日洗浄上がり品でない方が優先
                    
                    assigned_lots_overrun.sort(key=lambda x: (x['priority'], x['divided_time'], self._normalize_shipping_date(x['shipping_date'])))
                    
                    reassignment_count_overrun = 0
                    max_reassignments_overrun = 10  # 最大10件まで再割当を試みる
                    
                    for lot_info_overrun in assigned_lots_overrun:
                        if reassignment_count_overrun >= max_reassignments_overrun:
                            break
                        
                        lot_index_overrun = lot_info_overrun['index']
                        product_number_overrun = lot_info_overrun['product_number']
                        divided_time_overrun = lot_info_overrun['divided_time']
                        inspection_time_overrun = lot_info_overrun['inspection_time']
                        shipping_date_overrun = lot_info_overrun['shipping_date']
                        
                        # 固定検査員が割り当て済みのロットはスキップ
                        if self._is_locked_fixed_preinspection_lot(result_df, lot_index_overrun):
                            continue
                        
                        # 利用可能な検査員を取得
                        process_number_overrun = lot_info_overrun['row'].get('現在工程番号', '')
                        lot_process_name_overrun = str(lot_info_overrun['row'].get('現在工程名', '') or '').strip()
                        skill_rows_overrun = skill_master_df[skill_master_df.iloc[:, 0] == product_number_overrun]
                        is_new_product_overrun = skill_rows_overrun.empty
                        available_inspectors_overrun = self.get_available_inspectors(
                            product_number_overrun, process_number_overrun, skill_master_df, inspector_master_df,
                            shipping_date=shipping_date_overrun, allow_new_team_fallback=is_new_product_overrun,
                            process_master_df=process_master_df, inspection_target_keywords=inspection_target_keywords,
                            process_name_context=lot_process_name_overrun
                        )
                        
                        if not available_inspectors_overrun and is_new_product_overrun:
                            available_inspectors_overrun = self.get_new_product_team_inspectors(inspector_master_df)
                        
                        if not available_inspectors_overrun:
                            continue
                        
                        # 10%超過を検出した検査員を除外
                        available_inspectors_overrun = [insp for insp in available_inspectors_overrun 
                                                       if insp['コード'] != overrun_code]
                        
                        if not available_inspectors_overrun:
                            continue
                        
                        # 余裕のある検査員を選択（平均未満の検査員を優先）
                        avg_hours_phase2_5 = sum(self.inspector_work_hours.values()) / len([c for c in self.inspector_work_hours.keys() if self.inspector_work_hours.get(c, 0.0) > 0]) if self.inspector_work_hours else 0.0
                        replacement_candidates_overrun = []
                        for insp_overrun in available_inspectors_overrun:
                            candidate_code_overrun = insp_overrun['コード']
                            candidate_total_hours_overrun = self.inspector_work_hours.get(candidate_code_overrun, 0.0)
                            
                            # 平均未満の検査員を優先
                            if candidate_total_hours_overrun > avg_hours_phase2_5:
                                continue
                            
                            # 勤務時間制約をチェック
                            candidate_max_hours_overrun = inspector_max_hours.get(candidate_code_overrun, 8.0)
                            candidate_daily_hours_overrun = self.inspector_daily_assignments.get(candidate_code_overrun, {}).get(current_date, 0.0)
                            allowed_max_hours_overrun = candidate_max_hours_overrun * (1.0 + WORK_HOURS_OVERRUN_RATE)
                            if candidate_daily_hours_overrun + divided_time_overrun > allowed_max_hours_overrun - WORK_HOURS_BUFFER:
                                continue
                            
                            # 4時間上限チェック
                            candidate_product_hours_overrun = self.inspector_product_hours.get(candidate_code_overrun, {}).get(product_number_overrun, 0.0)
                            if candidate_product_hours_overrun + divided_time_overrun > self.product_limit_hard_threshold:
                                continue
                            
                            replacement_candidates_overrun.append((candidate_total_hours_overrun, insp_overrun, candidate_code_overrun))
                        
                        if replacement_candidates_overrun:
                            # 【改善】余裕のある検査員を選択（許容最大時間までの余裕が大きい順）
                            # 10%超過を解消するため、より余裕がある検査員を優先的に選択
                            def sort_key_candidate_phase2_5(x):
                                _, insp, code = x
                                candidate_daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                                candidate_max_hours = inspector_max_hours.get(code, 8.0)
                                candidate_allowed_max_hours = candidate_max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                # 許容最大時間までの余裕
                                remaining_capacity = candidate_allowed_max_hours - WORK_HOURS_BUFFER - candidate_daily_hours - divided_time_overrun
                                # 余裕が大きい順（負の値は除外されているので、正の値のみ）
                                return -remaining_capacity  # 負の値でソート（大きい順）
                            
                            replacement_candidates_overrun.sort(key=sort_key_candidate_phase2_5)
                            _, replacement_inspector_overrun, new_code_overrun = replacement_candidates_overrun[0]
                            
                            # 再割当を実行
                            old_inspector_name_overrun = overrun_name
                            if show_skill_values:
                                if replacement_inspector_overrun.get('is_new_team', False):
                                    new_name_overrun = f"{replacement_inspector_overrun['氏名']}(新)"
                                else:
                                    new_name_overrun = f"{replacement_inspector_overrun['氏名']}({replacement_inspector_overrun['スキル']})"
                            else:
                                if replacement_inspector_overrun.get('is_new_team', False):
                                    new_name_overrun = f"{replacement_inspector_overrun['氏名']}(新)"
                                else:
                                    new_name_overrun = replacement_inspector_overrun['氏名']
                            
                            # 元の検査員を置き換え
                            for i in range(1, 6):
                                inspector_col = f'検査員{i}'
                                if inspector_col in result_df.columns:
                                    inspector_value_check = result_df.at[lot_index_overrun, inspector_col]
                                    if pd.notna(inspector_value_check) and str(inspector_value_check).strip() != '':
                                        inspector_name_check = str(inspector_value_check).strip()
                                        if '(' in inspector_name_check:
                                            inspector_name_check = inspector_name_check.split('(')[0].strip()
                                        if inspector_name_check == overrun_name:
                                            result_df.at[lot_index_overrun, inspector_col] = new_name_overrun
                                            break
                            
                            # 【改善】再割当後に履歴を更新
                            if overrun_code in self.inspector_daily_assignments:
                                if current_date in self.inspector_daily_assignments[overrun_code]:
                                    self.inspector_daily_assignments[overrun_code][current_date] -= divided_time_overrun
                                    if self.inspector_daily_assignments[overrun_code][current_date] < 0:
                                        self.inspector_daily_assignments[overrun_code][current_date] = 0.0
                            if overrun_code in self.inspector_work_hours:
                                self.inspector_work_hours[overrun_code] -= divided_time_overrun
                                if self.inspector_work_hours[overrun_code] < 0:
                                    self.inspector_work_hours[overrun_code] = 0.0
                            if overrun_code in self.inspector_product_hours:
                                if product_number_overrun in self.inspector_product_hours[overrun_code]:
                                    self.inspector_product_hours[overrun_code][product_number_overrun] -= divided_time_overrun
                                    if self.inspector_product_hours[overrun_code][product_number_overrun] < 0:
                                        self.inspector_product_hours[overrun_code][product_number_overrun] = 0.0
                            
                            if new_code_overrun not in self.inspector_daily_assignments:
                                self.inspector_daily_assignments[new_code_overrun] = {}
                            if current_date not in self.inspector_daily_assignments[new_code_overrun]:
                                self.inspector_daily_assignments[new_code_overrun][current_date] = 0.0
                            if new_code_overrun not in self.inspector_work_hours:
                                self.inspector_work_hours[new_code_overrun] = 0.0
                            if new_code_overrun not in self.inspector_product_hours:
                                self.inspector_product_hours[new_code_overrun] = {}
                            if product_number_overrun not in self.inspector_product_hours[new_code_overrun]:
                                self.inspector_product_hours[new_code_overrun][product_number_overrun] = 0.0
                            
                            self.inspector_daily_assignments[new_code_overrun][current_date] += divided_time_overrun
                            self.inspector_work_hours[new_code_overrun] += divided_time_overrun
                            self.inspector_product_hours[new_code_overrun][product_number_overrun] += divided_time_overrun
                            
                            reassignment_count_overrun += 1
                            self.log_message(
                                f"フェーズ2.5: 10%超過解消のため再割当 - '{old_inspector_name_overrun}' → '{new_name_overrun}' (ロットインデックス: {lot_index_overrun}, 品番: {product_number_overrun})",
                                level='info',
                            )
                            
                            # 超過分が解消されたか確認（履歴を更新した後の値を使用）
                            overrun_current_hours_after = self.inspector_daily_assignments.get(overrun_code, {}).get(current_date, 0.0)
                            
                            if overrun_current_hours_after <= overrun_allowed_max_hours - WORK_HOURS_BUFFER:
                                self.log_message(
                                    f"フェーズ2.5: 検査員 '{overrun_name}' ({overrun_code}) の10%超過が解消されました ({overrun_current_hours_after:.1f}h <= {overrun_allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h)",
                                    level='info',
                                )
                                break
                            
                            # 【改善】10%超過が解消されない場合、最大10件まで再割当を試みる
                            if reassignment_count_overrun >= 10:
                                self.log_message(
                                    f"フェーズ2.5: 検査員 '{overrun_name}' ({overrun_code}) の10%超過解消のため、最大10件まで再割当を試みましたが、完全には解消されませんでした（現在: {overrun_current_hours_after:.1f}h > {overrun_allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h）",
                                    level='warning',
                                )
                                break
                    
                    if reassignment_count_overrun > 0:
                        self.log_message(
                            f"フェーズ2.5: 10%超過解消のため、{reassignment_count_overrun}件の再割当を実行しました",
                            level='info',
                        )
            
            # 勤務時間超過を再チェック
            phase2_5_violations = []
            for row_tuple in result_df.itertuples(index=True):
                index = row_tuple[0]  # インデックス
                product_number = row_tuple[product_col_idx + 1]  # itertuplesはインデックスを含むため+1
                process_name_context = getattr(row_tuple, '現在工程名', '') if hasattr(row_tuple, '現在工程名') else ''
                divided_time = row_tuple[divided_time_col_idx + 1] if divided_time_col_idx < len(row_tuple) - 1 else 0.0
                
                for i in range(1, 6):
                    inspector_col_idx = inspector_col_indices[i - 1]
                    inspector_value = row_tuple[inspector_col_idx + 1] if inspector_col_idx < len(row_tuple) - 1 else None
                    if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                        inspector_name = str(inspector_value).strip()
                        if '(' in inspector_name:
                            inspector_name = inspector_name.split('(')[0].strip()
                        if not inspector_name:
                            continue

                        # 固定検査員が割当済みのロットは、勤務時間/同一品番上限の違反チェック対象から除外する
                        if _is_fixed_inspector_for_lot_cached(product_number, str(process_name_context).strip(), inspector_name):
                            # 【追加】保護の履歴追跡
                            self.fixed_inspector_protection_metrics['total_protections'] += 1
                            self.fixed_inspector_protection_metrics['protection_by_phase']['phase2.5_violation_check'] = \
                                self.fixed_inspector_protection_metrics['protection_by_phase'].get('phase2.5_violation_check', 0) + 1
                            self.fixed_inspector_protection_metrics['protection_by_reason']['violation_check_exclusion'] = \
                                self.fixed_inspector_protection_metrics['protection_by_reason'].get('violation_check_exclusion', 0) + 1
                            self.fixed_inspector_protection_metrics['protected_lots'].add(index)
                            self.fixed_inspector_protection_metrics['protected_inspectors'].add(inspector_name)
                            
                            protection_history_entry = {
                                'lot_index': index,
                                'product_number': product_number,
                                'inspector_name': inspector_name,
                                'phase': 'phase2.5',
                                'reason': 'violation_check_exclusion',
                                'process_name': str(process_name_context).strip(),
                            }
                            self.fixed_inspector_protection_metrics['protection_history'].append(protection_history_entry)
                            if len(self.fixed_inspector_protection_metrics['protection_history']) > 100:
                                self.fixed_inspector_protection_metrics['protection_history'] = \
                                    self.fixed_inspector_protection_metrics['protection_history'][-100:]
                            continue
                        
                        inspector_code = inspector_name_to_id.get(inspector_name)
                        if inspector_code is None:
                            inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                            if not inspector_info.empty:
                                inspector_code = inspector_info.iloc[0]['#ID']
                            else:
                                continue
                            daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
                            max_hours = inspector_max_hours.get(inspector_code, 8.0)
                            product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                            
                            # 【修正】当日洗浄上がり品かどうかを判定して適切な超過率を適用
                            shipping_date_val = getattr(row_tuple, '出荷予定日', None) if hasattr(row_tuple, '出荷予定日') else None
                            is_same_day_cleaning = self._should_force_assign_same_day(shipping_date_val)
                            
                            if is_same_day_cleaning:
                                # 当日洗浄上がり品: 20%超過まで許容
                                allowed_max_hours = max_hours * (1.0 + SAME_DAY_WORK_HOURS_OVERRUN_RATE)
                            else:
                                # 通常品: 10%超過まで許容
                                allowed_max_hours = max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                            
                            # フェーズ2.5検証: 許容上限を超えている場合は違反
                            if daily_hours > allowed_max_hours - WORK_HOURS_BUFFER:
                                excess = daily_hours - allowed_max_hours
                                phase2_5_violations.append((index, inspector_code, inspector_name, "勤務時間超過", daily_hours, allowed_max_hours))
                                mode_str = "緩和モード（20%超過許容）" if is_same_day_cleaning else "通常モード（10%超過許容）"
                                self.log_message(f"❌ フェーズ2.5検証: 勤務時間超過が検出されました - 検査員 '{inspector_name}' {daily_hours:.1f}h > {allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h ({mode_str}, 超過: {excess:.1f}h, ロット {index})", level='warning')
                            
                            # 改善ポイント: 最適化フェーズでの設定時間上限チェック（厳格）
                            if product_hours > self.product_limit_hard_threshold:
                                phase2_5_violations.append((index, inspector_code, inspector_name, f"同一品番{self.product_limit_hard_threshold:.1f}時間超過", product_hours, self.product_limit_hard_threshold))
                                self.log_message(f"❌ フェーズ2.5検証: 同一品番{self.product_limit_hard_threshold:.1f}時間超過が検出されました - 検査員 '{inspector_name}' 品番 {product_number} {product_hours:.1f}h > {self.product_limit_hard_threshold:.1f}h (ロット {index})")
            
            if phase2_5_violations:
                self.log_message(f"⚠️ 警告: フェーズ2.5検証で {len(phase2_5_violations)}件の違反が検出されました", level='warning')
                
                # 違反を品番ごとにグループ化
                violations_by_product = {}
                for violation in phase2_5_violations:
                    index = violation[0]
                    row = result_df.iloc[index]
                    product_number = row.get('品番', '')
                    if product_number not in violations_by_product:
                        violations_by_product[product_number] = []
                    violations_by_product[product_number].append((violation, row))
                
                # 同じ品番で複数の違反がある場合、まとめて再割当を試みる
                resolved_count = 0
                processed_indices = set()
                
                for product_number, product_violations in violations_by_product.items():
                    if len(product_violations) > 1:
                        # 同じ品番で複数の違反がある場合、まとめて処理
                        self.log_message(f"🔵 品番 {product_number} で {len(product_violations)}件の違反が検出されました。まとめて再割当を試みます")
                        
                        # 違反ロットを一度クリア（ただし、2週間以内の新規品は保護のためスキップ）
                        violation_indices = []
                        violation_lots = []
                        protected_indices = set()  # 保護されたロットのインデックス
                        for violation, row in product_violations:
                            index = violation[0]
                            shipping_date = row.get('出荷予定日', pd.Timestamp.max)

                            # 固定検査員が割り当て済みのロットは保護（未割当にしない）
                            if self._is_locked_fixed_preinspection_lot(result_df, index):
                                protected_indices.add(index)
                                resolved_count += 1
                                self.log_message(
                                    f"固定検査員のため保護: フェーズ3.5の再割当対象から除外します "
                                    f"(ロット {index}, 品番: {product_number})",
                                    level='warning'
                                )
                                continue
                            
                            # 指定日数以内の新規品かどうかを判定
                            is_new_product = product_number not in skill_product_values
                            is_within_protection_period = False
                            
                            if NEW_PRODUCT_PROTECTION_ENABLED and is_new_product and pd.notna(shipping_date):
                                try:
                                    if isinstance(shipping_date, pd.Timestamp):
                                        shipping_date_date = shipping_date.date()
                                    elif isinstance(shipping_date, str):
                                        shipping_date_date = pd.to_datetime(shipping_date, errors='coerce').date()
                                        if pd.isna(shipping_date_date):
                                            shipping_date_date = None
                                    else:
                                        shipping_date_date = shipping_date.date() if hasattr(shipping_date, 'date') else shipping_date
                                    
                                    if shipping_date_date is not None:
                                        protection_threshold_date = current_date + timedelta(days=NEW_PRODUCT_PROTECTION_DAYS)
                                        is_within_protection_period = shipping_date_date <= protection_threshold_date
                                except Exception as e:
                                    self.log_message(f"出荷予定日の比較エラー: {str(e)} (ロットインデックス: {index})", level='warning')
                                    is_within_protection_period = False
                            
                            # 指定日数以内の新規品は保護（未割当にしない）
                            if is_new_product and is_within_protection_period:
                                inspector_code = violation[1]
                                violation_type = violation[3]
                                
                                # 違反内容に応じて保護処理
                                if violation_type == "同一品番4時間超過":
                                    current_product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                                    self.log_message(f"⚠️ 新規品 {product_number} (出荷予定日: {shipping_date}) は2週間以内のため保護します - 同一品番4時間超過を許容（{current_product_hours:.1f}h）", level='warning')
                                    self.relaxed_product_limit_assignments.add((inspector_code, product_number))
                                    protected_indices.add(index)
                                    resolved_count += 1
                                    continue
                                elif violation_type == "勤務時間超過":
                                    daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
                                    max_hours = inspector_max_hours.get(inspector_code, 8.0)
                                    self.log_message(f"⚠️ 新規品 {product_number} (出荷予定日: {shipping_date}) は2週間以内のため保護します - 勤務時間超過を許容（{daily_hours:.1f}h > {max_hours:.1f}h）", level='warning')
                                    protected_indices.add(index)
                                    resolved_count += 1
                                    continue
                            
                            # 保護対象でない場合は、通常通りクリア
                            violation_indices.append(index)
                            normalized_shipping_date = self._normalize_shipping_date(shipping_date)
                            violation_lots.append({
                                'index': index,
                                'violation': violation,
                                'row': row,
                                'inspection_time': row.get('検査時間', 0),
                                'shipping_date': shipping_date,
                                'normalized_shipping_date': normalized_shipping_date
                            })
                            self.clear_assignment(result_df, index)
                            processed_indices.add(index)
                        
                        # 履歴を再計算（クリアしたロットと保護されたロットを除外）
                        self.inspector_daily_assignments = {}
                        self.inspector_work_hours = {}
                        self.inspector_product_hours = {}
                        
                        # 列インデックスを事前に取得（itertuples()で高速化）
                        prod_num_col_idx_p2_5_f = result_df.columns.get_loc('品番')
                        div_time_col_idx_p2_5_f = result_df.columns.get_loc('分割検査時間') if '分割検査時間' in result_df.columns else -1
                        inspector_col_indices_p2_5_f = {}
                        for i in range(1, 6):
                            col_name = f'検査員{i}'
                            if col_name in result_df.columns:
                                inspector_col_indices_p2_5_f[i] = result_df.columns.get_loc(col_name)
                        
                        for row_tuple in result_df.itertuples(index=True):
                            idx = row_tuple[0]  # インデックス
                            if idx in violation_indices or idx in protected_indices:
                                continue  # クリアしたロットと保護されたロットはスキップ
                            
                            prod_num = row_tuple[prod_num_col_idx_p2_5_f + 1]  # +1はインデックス分
                            div_time = row_tuple[div_time_col_idx_p2_5_f + 1] if div_time_col_idx_p2_5_f >= 0 and div_time_col_idx_p2_5_f + 1 < len(row_tuple) else 0.0
                            
                            for i in range(1, 6):
                                if i not in inspector_col_indices_p2_5_f:
                                    continue
                                inspector_col_idx = inspector_col_indices_p2_5_f[i]
                                inspector_value = row_tuple[inspector_col_idx + 1] if inspector_col_idx + 1 < len(row_tuple) else None
                                
                                if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                                    inspector_name = str(inspector_value).strip()
                                    if '(' in inspector_name:
                                        inspector_name = inspector_name.split('(')[0].strip()
                                    if not inspector_name:
                                        continue
                                    
                                    inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                    if not inspector_info.empty:
                                        inspector_code = inspector_info.iloc[0]['#ID']
                                        
                                        if inspector_code not in self.inspector_daily_assignments:
                                            self.inspector_daily_assignments[inspector_code] = {}
                                        if current_date not in self.inspector_daily_assignments[inspector_code]:
                                            self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                                        if inspector_code not in self.inspector_work_hours:
                                            self.inspector_work_hours[inspector_code] = 0.0
                                        if inspector_code not in self.inspector_product_hours:
                                            self.inspector_product_hours[inspector_code] = {}
                                        if prod_num not in self.inspector_product_hours[inspector_code]:
                                            self.inspector_product_hours[inspector_code][prod_num] = 0.0
                                        
                                        self.inspector_daily_assignments[inspector_code][current_date] += div_time
                                        self.inspector_work_hours[inspector_code] += div_time
                                        self.inspector_product_hours[inspector_code][prod_num] += div_time
                        
                        # 出荷予定日順にソートして再割当
                        violation_lots.sort(key=lambda x: x['normalized_shipping_date'])
                        
                        # 各ロットを再割当
                        for lot_info in violation_lots:
                            index = lot_info['index']
                            inspection_time = lot_info['inspection_time']
                            shipping_date = lot_info['shipping_date']
                            
                            self.log_message(f"⚠️ ロットインデックス {index} (品番: {product_number}, 出荷予定日: {shipping_date}) を再割当します")
                            
                            # スキルマスタから利用可能な検査員を取得
                            process_number = lot_info['row'].get('現在工程番号', '')
                            lot_process_name = str(lot_info['row'].get('現在工程名', '') or '').strip()
                            skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                            is_new_product = skill_rows.empty
                            available_inspectors = self.get_available_inspectors(
                                product_number, process_number, skill_master_df, inspector_master_df,
                                shipping_date=shipping_date, allow_new_team_fallback=is_new_product,
                                process_master_df=process_master_df, inspection_target_keywords=inspection_target_keywords,
                                process_name_context=lot_process_name
                            )
                            
                            if not available_inspectors and is_new_product:
                                available_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                            
                            if available_inspectors:
                                # 再割当を試みる
                                if inspection_time <= self.required_inspectors_threshold:
                                    required_inspectors = 1
                                else:
                                    calculated_inspectors = max(2, int(inspection_time / self.required_inspectors_threshold) + 1)
                                    # 5名以上になる場合は5名に制限（特例）
                                    required_inspectors = min(5, calculated_inspectors)
                                divided_time = inspection_time / required_inspectors
                                
                                # 利用可能な検査員から選択
                                assigned_inspectors = self.select_inspectors(
                                    available_inspectors, required_inspectors, divided_time,
                                    inspector_master_df, product_number,
                                    is_new_product=skill_rows.empty,
                                    ignore_product_limit=self._should_force_assign_same_day(shipping_date),
                                )
                                
                                if assigned_inspectors:
                                    # 割り当てを設定
                                    result_df.at[index, '検査員人数'] = len(assigned_inspectors)
                                    # 分割検査時間の計算: 検査時間 ÷ 実際の分割した検査人数
                                    actual_divided_time = inspection_time / len(assigned_inspectors)
                                    result_df.at[index, '分割検査時間'] = round(actual_divided_time, 1)
                                    for i, inspector in enumerate(assigned_inspectors, 1):
                                        if i <= 5:
                                            if show_skill_values:
                                                if inspector.get('is_new_team', False):
                                                    inspector_name = f"{inspector['氏名']}(新)"
                                                else:
                                                    inspector_name = f"{inspector['氏名']}({inspector['スキル']})"
                                            else:
                                                if inspector.get('is_new_team', False):
                                                    inspector_name = f"{inspector['氏名']}(新)"
                                                else:
                                                    inspector_name = inspector['氏名']
                                            result_df.at[index, f'検査員{i}'] = inspector_name
                                    
                                    result_df.at[index, 'assignability_status'] = 'assigned'
                                    resolved_count += 1
                                    self.log_message(f"✅ ロットインデックス {index} の再割当に成功しました")
                                else:
                                    self.log_message(f"⚠️ ロットインデックス {index} の再割当に失敗しました（利用可能な検査員が見つかりません）")
                            else:
                                    self.log_message(f"⚠️ ロットインデックス {index} の再割当に失敗しました（利用可能な検査員が0人）")
                        
                        # 履歴を再計算（再割当後の状態）
                        self.inspector_daily_assignments = {}
                        self.inspector_work_hours = {}
                        self.inspector_product_hours = {}
                        
                        # 列インデックスを事前に取得（itertuples()で高速化）
                        prod_num_col_idx_p3 = result_df.columns.get_loc('品番')
                        div_time_col_idx_p3 = result_df.columns.get_loc('分割検査時間') if '分割検査時間' in result_df.columns else -1
                        inspector_col_indices_p3 = {}
                        for i in range(1, 6):
                            col_name = f'検査員{i}'
                            if col_name in result_df.columns:
                                inspector_col_indices_p3[i] = result_df.columns.get_loc(col_name)
                        
                        for row_tuple in result_df.itertuples(index=False):
                            prod_num = row_tuple[prod_num_col_idx_p3]
                            div_time = row_tuple[div_time_col_idx_p3] if div_time_col_idx_p3 >= 0 and div_time_col_idx_p3 < len(row_tuple) else 0.0
                            
                            for i in range(1, 6):
                                if i not in inspector_col_indices_p3:
                                    continue
                                inspector_col_idx = inspector_col_indices_p3[i]
                                inspector_value = row_tuple[inspector_col_idx] if inspector_col_idx < len(row_tuple) else None
                                
                                if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                                    inspector_name = str(inspector_value).strip()
                                    if '(' in inspector_name:
                                        inspector_name = inspector_name.split('(')[0].strip()
                                    if not inspector_name:
                                        continue
                                    
                                    inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                    if not inspector_info.empty:
                                        inspector_code = inspector_info.iloc[0]['#ID']
                                        
                                        if inspector_code not in self.inspector_daily_assignments:
                                            self.inspector_daily_assignments[inspector_code] = {}
                                        if current_date not in self.inspector_daily_assignments[inspector_code]:
                                            self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                                        if inspector_code not in self.inspector_work_hours:
                                            self.inspector_work_hours[inspector_code] = 0.0
                                        if inspector_code not in self.inspector_product_hours:
                                            self.inspector_product_hours[inspector_code] = {}
                                        if prod_num not in self.inspector_product_hours[inspector_code]:
                                            self.inspector_product_hours[inspector_code][prod_num] = 0.0
                                        
                                        self.inspector_daily_assignments[inspector_code][current_date] += div_time
                                        self.inspector_work_hours[inspector_code] += div_time
                                        self.inspector_product_hours[inspector_code][prod_num] += div_time
                
                # 残りの違反（単独または処理済み以外）を個別に処理
                violations_with_date = []
                for violation in phase2_5_violations:
                    index = violation[0]
                    if index not in processed_indices:
                        row = result_df.iloc[index]
                        shipping_date_raw = row.get('出荷予定日', pd.Timestamp.max)
                        
                        # 【修正】出荷予定日を正規化（文字列の場合はpd.Timestamp.minに変換してソート可能にする）
                        if pd.notna(shipping_date_raw):
                            shipping_date_str = str(shipping_date_raw).strip()
                            # 当日洗浄上がり品などの文字列の場合は、最優先としてpd.Timestamp.minに変換
                            if (shipping_date_str == "当日洗浄上がり品" or
                                shipping_date_str == "当日洗浄品" or
                                "当日洗浄" in shipping_date_str or
                                shipping_date_str == "先行検査" or
                                shipping_date_str == "当日先行検査"):
                                shipping_date = pd.Timestamp.min  # 最優先として扱う
                            elif isinstance(shipping_date_raw, pd.Timestamp):
                                shipping_date = shipping_date_raw
                            elif isinstance(shipping_date_raw, str):
                                # 日付文字列の場合は変換を試みる
                                try:
                                    shipping_date = pd.to_datetime(shipping_date_raw)
                                except Exception as e:
                                    logger.debug(f"出荷日の変換でエラーが発生しました（デフォルト値を使用）: {e}")
                                    shipping_date = pd.Timestamp.min
                            else:
                                shipping_date = shipping_date_raw
                        else:
                            shipping_date = pd.Timestamp.max
                        
                        violations_with_date.append((violation, shipping_date))
                
                # 出荷予定日の古い順にソート
                violations_with_date.sort(key=lambda x: self._normalize_shipping_date(x[1]))
                
                # 出荷予定日が古いロットから順に再割り当てを試みる
                for violation, shipping_date in violations_with_date:
                    index = violation[0]
                    row = result_df.iloc[index]
                    product_number = row.get('品番', '')
                    inspection_time = row.get('検査時間', 0)

                    # 登録済み品番の先行検査×固定検査員ロットは最適化フェーズで動かさない（固定維持）
                    if self._is_locked_fixed_preinspection_lot(result_df, index):
                        continue
                    
                    self.log_message(
                        f"違反是正: ロットインデックス {index} (品番: {product_number}, 出荷予定日: {shipping_date}) ({violation[3]})",
                        level='debug',
                    )
                    
                    # 違反を是正する処理を試みる
                    violation_resolved = self.fix_single_violation(
                        index, violation[1], violation[2], 
                        row.get('分割検査時間', 0.0), product_number, inspection_time,
                        None, result_df, inspector_master_df, skill_master_df,
                        inspector_max_hours, current_date, show_skill_values
                    )
                    
                    if violation_resolved:
                        resolved_count += 1
                        self.log_message(f"違反是正完了: ロットインデックス {index}", level='debug')
                    else:
                        # 是正できなかった場合は未割当にする
                        if self._is_locked_fixed_preinspection_lot(result_df, index):
                            continue
                        self.clear_assignment(result_df, index)
                        self.log_message(f"⚠️ ロットインデックス {index} を未割当にしました（{violation[3]}）")
                
                self.log_message(f"フェーズ2.5違反是正結果: {resolved_count}件是正、{len(phase2_5_violations) - resolved_count}件未割当")
                
                # 未割当後の履歴を再計算
                self.inspector_daily_assignments = {}
                self.inspector_work_hours = {}
                self.inspector_product_hours = {}
                
                # 列インデックスを事前に取得（高速化：itertuples()を使用）
                product_col_idx = result_df.columns.get_loc('品番')
                divided_time_col_idx = result_df.columns.get_loc('分割検査時間')
                inspector_col_indices = [result_df.columns.get_loc(f'検査員{i}') for i in range(1, 6)]
                
                for row_tuple in result_df.itertuples(index=True):
                    index = row_tuple[0]  # インデックス
                    product_number = row_tuple[product_col_idx + 1]  # itertuplesはインデックスを含むため+1
                    divided_time = row_tuple[divided_time_col_idx + 1] if divided_time_col_idx < len(row_tuple) - 1 else 0.0
                    
                    for i in range(1, 6):
                        inspector_col_idx = inspector_col_indices[i - 1]
                        inspector_value = row_tuple[inspector_col_idx + 1] if inspector_col_idx < len(row_tuple) - 1 else None
                        if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                            inspector_name = str(inspector_value).strip()
                            if '(' in inspector_name:
                                inspector_name = inspector_name.split('(')[0].strip()
                            if not inspector_name:
                                continue

                            inspector_code = inspector_name_to_id.get(inspector_name)
                            if inspector_code is None:
                                inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                if not inspector_info.empty:
                                    inspector_code = inspector_info.iloc[0]['#ID']
                                else:
                                    continue
                                
                                if inspector_code not in self.inspector_daily_assignments:
                                    self.inspector_daily_assignments[inspector_code] = {}
                                if current_date not in self.inspector_daily_assignments[inspector_code]:
                                    self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                                if inspector_code not in self.inspector_work_hours:
                                    self.inspector_work_hours[inspector_code] = 0.0
                                if inspector_code not in self.inspector_product_hours:
                                    self.inspector_product_hours[inspector_code] = {}
                                if product_number not in self.inspector_product_hours[inspector_code]:
                                    self.inspector_product_hours[inspector_code][product_number] = 0.0
                                
                                self.inspector_daily_assignments[inspector_code][current_date] += divided_time
                                self.inspector_work_hours[inspector_code] += divided_time
                                self.inspector_product_hours[inspector_code][product_number] += divided_time

            perf_logger.debug(
                "PERF {}: {:.1f} ms",
                "inspector_assignment.optimize.phase2_5.total",
                (perf_counter() - _t_perf_phase2_5_total) * 1000.0,
            )

            # フェーズ3: 未割当ロットの再処理（出荷予定日順、新規品優先）
            self.log_message("全体最適化フェーズ3: 未割当ロットの再処理を開始")
            _t_perf_phase3_total = perf_counter()
            
            # 【修正】フェーズ3開始時に、inspector_daily_assignmentsを再計算して正確な勤務時間を反映
            self.log_message("フェーズ3: inspector_daily_assignmentsを再計算中...", debug=True)
            self.inspector_daily_assignments = {}
            self.inspector_work_hours = {}
            self.inspector_product_hours = {}
            
            # result_dfから実際の割り当てを読み取って、inspector_daily_assignmentsを再計算
            phase3_cols = {col: idx for idx, col in enumerate(result_df.columns)}
            for row_idx, row in enumerate(result_df.itertuples(index=False)):
                index = result_df.index[row_idx]
                product_number = row[phase3_cols['品番']]
                divided_time = row[phase3_cols.get('分割検査時間', -1)] if '分割検査時間' in phase3_cols else 0.0
                if divided_time == -1:
                    divided_time = 0.0
                
                # 実際に割り当てられた検査員を取得
                for i in range(1, 6):
                    col_name = f'検査員{i}'
                    if col_name not in phase3_cols:
                        continue
                    inspector_value = row[phase3_cols[col_name]]
                    if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                        inspector_name = str(inspector_value).strip()
                        if '(' in inspector_name:
                            inspector_name = inspector_name.split('(')[0].strip()
                        if not inspector_name:
                            continue
                        
                        inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                        if not inspector_info.empty:
                            inspector_code = inspector_info.iloc[0]['#ID']
                            current_date = pd.Timestamp.now().date()
                            
                            # inspector_daily_assignmentsを更新
                            if inspector_code not in self.inspector_daily_assignments:
                                self.inspector_daily_assignments[inspector_code] = {}
                            if current_date not in self.inspector_daily_assignments[inspector_code]:
                                self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                            self.inspector_daily_assignments[inspector_code][current_date] += divided_time
                            
                            # inspector_work_hoursを更新
                            if inspector_code not in self.inspector_work_hours:
                                self.inspector_work_hours[inspector_code] = 0.0
                            self.inspector_work_hours[inspector_code] += divided_time
                            
                            # inspector_product_hoursを更新
                            if inspector_code not in self.inspector_product_hours:
                                self.inspector_product_hours[inspector_code] = {}
                            if product_number not in self.inspector_product_hours[inspector_code]:
                                self.inspector_product_hours[inspector_code][product_number] = 0.0
                            self.inspector_product_hours[inspector_code][product_number] += divided_time
            
            self.log_message("フェーズ3: inspector_daily_assignmentsの再計算が完了しました", debug=True)
            
            # 【改善】フェーズ3開始時に、当日洗浄上がり品の制約を再構築（result_dfから実際の割り当てを反映）
            self.log_message("フェーズ3: 当日洗浄上がり品の制約を再構築します")
            self.same_day_cleaning_inspectors = {}
            self.same_day_cleaning_inspectors_by_product_name = {}
            
            # 列インデックスを事前に取得（itertuples()で高速化）
            prod_num_col_idx = result_df.columns.get_loc('品番')
            shipping_date_col_idx = result_df.columns.get_loc('出荷予定日') if '出荷予定日' in result_df.columns else -1
            product_name_col_idx = result_df.columns.get_loc('品名') if '品名' in result_df.columns else -1
            inspector_col_indices = {}
            for i in range(1, 6):
                col_name = f'検査員{i}'
                if col_name in result_df.columns:
                    inspector_col_indices[i] = result_df.columns.get_loc(col_name)
            
            for row_tuple in result_df.itertuples(index=True):
                index = row_tuple[0]  # インデックス
                product_number = row_tuple[prod_num_col_idx + 1]  # +1はインデックス分
                shipping_date_raw = row_tuple[shipping_date_col_idx + 1] if shipping_date_col_idx >= 0 and shipping_date_col_idx + 1 < len(row_tuple) else None
                shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                is_same_day_cleaning = (
                    shipping_date_str == "当日洗浄上がり品" or
                    shipping_date_str == "当日洗浄品" or
                    "当日洗浄" in shipping_date_str or
                    shipping_date_str == "先行検査" or
                    shipping_date_str == "当日先行検査"
                )
                
                if is_same_day_cleaning:
                    # 品名を取得
                    product_name = row_tuple[product_name_col_idx + 1] if product_name_col_idx >= 0 and product_name_col_idx + 1 < len(row_tuple) else ''
                    product_name_str = str(product_name).strip() if pd.notna(product_name) else ''
                    
                    # 実際に割り当てられている検査員を取得
                    for i in range(1, 6):
                        if i not in inspector_col_indices:
                            continue
                        inspector_col_idx = inspector_col_indices[i]
                        inspector_value = row_tuple[inspector_col_idx + 1] if inspector_col_idx + 1 < len(row_tuple) else None
                        
                        if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                            inspector_name = str(inspector_value).strip()
                            if '(' in inspector_name:
                                inspector_name = inspector_name.split('(')[0].strip()
                            if not inspector_name:
                                continue

                            inspector_code = inspector_name_to_id.get(inspector_name)
                            if inspector_code is None:
                                inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                if not inspector_info.empty:
                                    inspector_code = inspector_info.iloc[0]['#ID']
                                else:
                                    continue
                                # 品番単位の制約を更新
                                self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(inspector_code)
                                # 品名単位の制約を更新
                                if product_name_str:
                                    self.same_day_cleaning_inspectors_by_product_name.setdefault(product_name_str, set()).add(inspector_code)
            
            # 制約再構築の詳細ログ
            product_count = len(self.same_day_cleaning_inspectors)
            product_name_count = len(self.same_day_cleaning_inspectors_by_product_name)
            total_inspectors = sum(len(inspectors) for inspectors in self.same_day_cleaning_inspectors.values())
            total_inspectors_by_name = sum(len(inspectors) for inspectors in self.same_day_cleaning_inspectors_by_product_name.values())
            self.log_message(
                f"フェーズ3: 当日洗浄上がり品の制約を再構築しました "
                f"（品番数: {product_count}, 品名数: {product_name_count}, "
                f"品番単位の検査員割当数: {total_inspectors}, 品名単位の検査員割当数: {total_inspectors_by_name}）",
                debug=True,
            )
            
            # 【改善】アプローチ3: 当日洗浄上がり品の未割当ロットがある場合、優先度の低いロットから検査員を再割当て
            same_day_cleaning_unassigned = []
            # 列インデックスを事前に取得（itertuples()で高速化）
            shipping_date_col_idx_u = result_df.columns.get_loc('出荷予定日') if '出荷予定日' in result_df.columns else -1
            inspector_count_col_idx = result_df.columns.get_loc('検査員人数') if '検査員人数' in result_df.columns else -1
            
            for row_tuple in result_df.itertuples(index=True):
                index = row_tuple[0]  # インデックス
                shipping_date_raw = row_tuple[shipping_date_col_idx_u + 1] if shipping_date_col_idx_u >= 0 and shipping_date_col_idx_u + 1 < len(row_tuple) else None
                shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                is_same_day_cleaning = (
                    shipping_date_str == "当日洗浄上がり品" or
                    shipping_date_str == "当日洗浄品" or
                    "当日洗浄" in shipping_date_str or
                    shipping_date_str == "先行検査" or
                    shipping_date_str == "当日先行検査"
                )
                
                if is_same_day_cleaning:
                    inspector_count = row_tuple[inspector_count_col_idx + 1] if inspector_count_col_idx >= 0 and inspector_count_col_idx + 1 < len(row_tuple) else 0
                    if inspector_count == 0 or pd.isna(inspector_count) or inspector_count == 0:
                        same_day_cleaning_unassigned.append(index)
            
            # 当日洗浄上がり品の未割当ロットがある場合、優先度の低いロットから検査員を再割当て
            if same_day_cleaning_unassigned:
                self.log_message(
                    f"⚠️ 警告: 当日洗浄上がり品の未割当ロットが {len(same_day_cleaning_unassigned)}件あります。"
                    f"優先度の低いロットから検査員を再割当てします",
                    level='warning'
                )
                
                # 優先度の低いロット（出荷予定日が遠い、当日洗浄上がり品以外）を取得
                low_priority_lots = []
                current_date = pd.Timestamp.now().date()
                two_weeks_later = current_date + timedelta(days=14)
                
                # 列インデックスを事前に取得（itertuples()で高速化）
                shipping_date_col_idx_l = result_df.columns.get_loc('出荷予定日') if '出荷予定日' in result_df.columns else -1
                inspector_count_col_idx_l = result_df.columns.get_loc('検査員人数') if '検査員人数' in result_df.columns else -1
                
                for row_tuple in result_df.itertuples(index=True):
                    index = row_tuple[0]  # インデックス
                    if index in same_day_cleaning_unassigned:
                        continue  # 当日洗浄上がり品の未割当ロットは除外
                    
                    inspector_count = row_tuple[inspector_count_col_idx_l + 1] if inspector_count_col_idx_l >= 0 and inspector_count_col_idx_l + 1 < len(row_tuple) else 0
                    if inspector_count == 0 or pd.isna(inspector_count) or inspector_count == 0:
                        continue  # 未割当ロットは除外
                    
                    shipping_date_raw = row_tuple[shipping_date_col_idx_l + 1] if shipping_date_col_idx_l >= 0 and shipping_date_col_idx_l + 1 < len(row_tuple) else None
                    shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                    is_same_day_cleaning = (
                        shipping_date_str == "当日洗浄上がり品" or
                        shipping_date_str == "当日洗浄品" or
                        "当日洗浄" in shipping_date_str or
                        shipping_date_str == "先行検査" or
                        shipping_date_str == "当日先行検査"
                    )
                    
                    if not is_same_day_cleaning:
                        # 出荷予定日が2週間以上先のロットを優先度の低いロットとして追加
                        try:
                            shipping_date = pd.to_datetime(shipping_date_raw, errors='coerce')
                            if pd.notna(shipping_date):
                                shipping_date_date = shipping_date.date() if hasattr(shipping_date, 'date') else shipping_date
                                if shipping_date_date > two_weeks_later:
                                    # rowオブジェクトが必要な場合は、元のDataFrameから取得
                                    row = result_df.loc[index]
                                    low_priority_lots.append((index, row, shipping_date_date))
                        except Exception as e:
                            # 日付変換に失敗した場合はスキップ
                            pass
                
                # 出荷予定日が遠い順にソート（最も遠いロットから再割当て）
                low_priority_lots.sort(key=lambda x: x[2], reverse=True)
                
                # 当日洗浄上がり品の未割当ロットごとに、優先度の低いロットから検査員を再割当て
                for unassigned_index in same_day_cleaning_unassigned:
                    unassigned_row = result_df.iloc[unassigned_index]
                    unassigned_product_number = unassigned_row['品番']
                    unassigned_inspection_time = unassigned_row.get('検査時間', 0.0)
                    unassigned_process_number = unassigned_row.get('現在工程番号', '')
                    unassigned_product_name = unassigned_row.get('品名', '')
                    unassigned_product_name_str = str(unassigned_product_name).strip() if pd.notna(unassigned_product_name) else ''
                    shipping_date_raw = unassigned_row.get('出荷予定日', '')
                    shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                    relax_hours = self._should_relax_hours_for_lot(
                        unassigned_product_number, shipping_date_str, unassigned_inspection_time
                    )
                    is_same_day_unassigned_high_duration = self._is_same_day_high_duration(unassigned_inspection_time)

                    # 必要な検査員数を計算
                    threshold_for_calc = 2.5 if relax_hours else self.required_inspectors_threshold
                    required_inspectors = self._calc_required_inspectors(unassigned_inspection_time, threshold_for_calc)
                    if is_same_day_unassigned_high_duration:
                        required_inspectors = max(required_inspectors, 2)
                    
                    # 優先度の低いロットから検査員を取得
                    reassigned_inspectors = []
                    reassigned_codes = set()
                    
                    for low_priority_index, low_priority_row, low_priority_shipping_date in low_priority_lots:
                        if len(reassigned_inspectors) >= required_inspectors:
                            break
                        
                        # このロットに割り当てられている検査員を取得
                        low_priority_product_number = low_priority_row.get('品番', '')
                        low_priority_product_name = low_priority_row.get('品名', '')
                        low_priority_product_name_str = str(low_priority_product_name).strip() if pd.notna(low_priority_product_name) else ''
                        
                        for i in range(1, 6):
                            inspector_col = f'検査員{i}'
                            inspector_value = low_priority_row.get(inspector_col, '')
                            if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                                inspector_name = str(inspector_value).strip()
                                if '(' in inspector_name:
                                    inspector_name = inspector_name.split('(')[0].strip()
                                if not inspector_name:
                                    continue
                                    
                                inspector_code = inspector_name_to_id.get(inspector_name)
                                if inspector_code is None:
                                    inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                    if not inspector_info.empty:
                                        inspector_code = inspector_info.iloc[0]['#ID']
                                    else:
                                        continue
                                    
                                    # 既に再割当て済みの検査員はスキップ
                                    if inspector_code in reassigned_codes:
                                        continue
                                    
                                    # この検査員が当日洗浄上がり品に割り当て可能かチェック
                                    # 1. 品番単位の制約チェック
                                    already_assigned_to_this_product = self.same_day_cleaning_inspectors.get(unassigned_product_number, set())
                                    if inspector_code in already_assigned_to_this_product:
                                        continue
                                    
                                    # 2. 品名単位の制約チェック
                                    if unassigned_product_name_str:
                                        already_assigned_to_same_product_name = self.same_day_cleaning_inspectors_by_product_name.get(unassigned_product_name_str, set())
                                        if inspector_code in already_assigned_to_same_product_name:
                                            continue
                                    
                                    # 3. スキルチェック
                                    allowed_codes = skill_allowed_inspectors_by_product.get(unassigned_product_number)
                                    if allowed_codes is not None and inspector_code not in allowed_codes:
                                        continue
                                    
                                    # 4. 勤務時間チェック（10%超過を厳格にチェック）
                                    divided_time = unassigned_inspection_time / required_inspectors
                                    daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
                                    max_hours = inspector_max_hours.get(inspector_code, 8.0)
                                    # 10%超過を考慮した許容最大勤務時間を計算
                                    allowed_max_hours = max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                                    # 10%超過を超える場合は除外（WORK_HOURS_BUFFERを考慮）
                                    if daily_hours + divided_time > allowed_max_hours - WORK_HOURS_BUFFER:
                                        continue
                                    
                                    # 5. 4時間上限チェック
                                    product_hours = self.inspector_product_hours.get(inspector_code, {}).get(unassigned_product_number, 0.0)
                                    if product_hours + divided_time > PRODUCT_LIMIT_DRAFT_THRESHOLD:  # 4.5時間まで許容
                                        if not (relax_hours or is_same_day_unassigned_high_duration):
                                            continue
                                        inspector_entry = next((insp for insp in available_inspectors if insp['コード'] == inspector_code), None)
                                        if inspector_entry is not None:
                                            inspector_entry['__near_product_limit'] = True
                                    
                                    # 再割当て可能な検査員を追加
                                    reassigned_inspectors.append({
                                        'コード': inspector_code,
                                        '氏名': inspector_name,
                                        '割当時間': divided_time
                                    })
                                    reassigned_codes.add(inspector_code)
                                    
                                    # 優先度の低いロットからこの検査員を削除
                                    result_df.at[low_priority_index, inspector_col] = ''
                                    self.log_message(
                                        f"アプローチ3: 優先度の低いロット {low_priority_index} (品番: {low_priority_row['品番']}, 出荷予定日: {low_priority_shipping_date}) "
                                        f"から検査員 '{inspector_name}' (コード: {inspector_code}) を再割当てしました"
                                    )
                                    self._remove_inspector_from_same_day_sets(
                                        low_priority_product_number,
                                        low_priority_product_name_str,
                                        inspector_code
                                    )
                                    
                                    # 履歴からこの検査員の時間を引く
                                    if inspector_code in self.inspector_daily_assignments:
                                        if current_date in self.inspector_daily_assignments[inspector_code]:
                                            low_priority_divided_time = low_priority_row.get('分割検査時間', 0.0)
                                            self.inspector_daily_assignments[inspector_code][current_date] = max(0.0, self.inspector_daily_assignments[inspector_code][current_date] - low_priority_divided_time)
                                    if inspector_code in self.inspector_work_hours:
                                        low_priority_divided_time = low_priority_row.get('分割検査時間', 0.0)
                                        self.inspector_work_hours[inspector_code] = max(0.0, self.inspector_work_hours[inspector_code] - low_priority_divided_time)
                                    if inspector_code in self.inspector_product_hours:
                                        low_priority_product_number = low_priority_row.get('品番', '')
                                        if low_priority_product_number in self.inspector_product_hours[inspector_code]:
                                            low_priority_divided_time = low_priority_row.get('分割検査時間', 0.0)
                                            self.inspector_product_hours[inspector_code][low_priority_product_number] = max(0.0, self.inspector_product_hours[inspector_code][low_priority_product_number] - low_priority_divided_time)
                                    
                                    # 優先度の低いロットの検査員人数を更新
                                    low_priority_inspector_count = low_priority_row.get('検査員人数', 0)
                                    if low_priority_inspector_count > 0:
                                        result_df.at[low_priority_index, '検査員人数'] = max(0, low_priority_inspector_count - 1)
                                    
                                    # 優先度の低いロットの分割検査時間を再計算
                                    remaining_inspectors = result_df.at[low_priority_index, '検査員人数']
                                    if remaining_inspectors > 0:
                                        low_priority_inspection_time = low_priority_row.get('検査時間', 0.0)
                                        result_df.at[low_priority_index, '分割検査時間'] = round(low_priority_inspection_time / remaining_inspectors, 1)
                                    else:
                                        result_df.at[low_priority_index, '分割検査時間'] = 0.0
                                    
                                    if len(reassigned_inspectors) >= required_inspectors:
                                        break
                        
                        # このロットの検査員人数が0になった場合は、未割当にする
                        if result_df.at[low_priority_index, '検査員人数'] == 0:
                            self.log_message(
                                f"アプローチ3: 優先度の低いロット {low_priority_index} (品番: {low_priority_row['品番']}) は検査員が0人になったため未割当にしました"
                            )
                    
                    # 再割当てした検査員を当日洗浄上がり品に割り当て
                    if reassigned_inspectors:
                        self.log_message(
                            f"アプローチ3: 当日洗浄上がり品 {unassigned_product_number} に {len(reassigned_inspectors)}人の検査員を再割当てしました"
                        )
                        
                        # 検査員を割り当て
                        for i, inspector in enumerate(reassigned_inspectors[:5], 1):  # 最大5人まで
                            inspector_name = inspector['氏名']
                            inspector_code = inspector['コード']
                            divided_time = inspector['割当時間']
                            
                            result_df.at[unassigned_index, f'検査員{i}'] = inspector_name
                            
                            # 履歴を更新
                            if inspector_code not in self.inspector_daily_assignments:
                                self.inspector_daily_assignments[inspector_code] = {}
                            if current_date not in self.inspector_daily_assignments[inspector_code]:
                                self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                            if inspector_code not in self.inspector_work_hours:
                                self.inspector_work_hours[inspector_code] = 0.0
                            if inspector_code not in self.inspector_product_hours:
                                self.inspector_product_hours[inspector_code] = {}
                            if unassigned_product_number not in self.inspector_product_hours[inspector_code]:
                                self.inspector_product_hours[inspector_code][unassigned_product_number] = 0.0
                            
                            self.inspector_daily_assignments[inspector_code][current_date] += divided_time
                            self.inspector_work_hours[inspector_code] += divided_time
                            self.inspector_product_hours[inspector_code][unassigned_product_number] += divided_time
                            
                            # 当日洗浄上がり品の制約を更新
                            self.same_day_cleaning_inspectors.setdefault(unassigned_product_number, set()).add(inspector_code)
                            if unassigned_product_name_str:
                                self.same_day_cleaning_inspectors_by_product_name.setdefault(unassigned_product_name_str, set()).add(inspector_code)
                        
                        # 検査員人数と分割検査時間を更新
                        result_df.at[unassigned_index, '検査員人数'] = len(reassigned_inspectors)
                        result_df.at[unassigned_index, '分割検査時間'] = round(unassigned_inspection_time / len(reassigned_inspectors), 1)
                        
                        self.log_message(
                            f"アプローチ3: 当日洗浄上がり品 {unassigned_product_number} の割り当てが完了しました（検査員人数: {len(reassigned_inspectors)}人）"
                        )
                    else:
                        warning_key = ("same_day_cleaning.reassign_failed", str(unassigned_product_number).strip())
                        if warning_key not in self.logged_warnings:
                            self.logged_warnings.add(warning_key)
                            self.log_message(
                                f"⚠️ 警告: 当日洗浄上がり品 {unassigned_product_number} に再割当てできる検査員が見つかりませんでした",
                                level='warning'
                            )
            
            # 未割当のロットを取得（出荷予定日順）
            # 優先ロット（当日/当日洗浄/先行検査/3営業日以内）が未割当のまま残っている場合は、
            # それ以降の日付（3営業日より後）の割当を一旦解除して再処理する（近い日付を優先し、遠い日付を犠牲にする）
            try:
                if '出荷予定日' in result_df.columns and '検査員人数' in result_df.columns:
                    today_date = pd.Timestamp.now().normalize().date()

                    def next_business_day(date_val: date) -> date:
                        weekday = date_val.weekday()
                        if weekday == 4:  # Friday
                            return date_val + timedelta(days=3)
                        if weekday == 5:  # Saturday
                            return date_val + timedelta(days=2)
                        return date_val + timedelta(days=1)
                    
                    def add_business_days(start: date, business_days: int) -> date:
                        current = start
                        added = 0
                        while added < business_days:
                            current = next_business_day(current)
                            added += 1
                        return current

                    next_bd = next_business_day(today_date)
                    three_business_days_ahead = add_business_days(today_date, 3)
                    shipping_raw = result_df['出荷予定日']
                    shipping_str = shipping_raw.fillna('').astype(str).str.strip()
                    # format='mixed' を指定して、推測失敗時の element-wise 解析の警告を抑制する
                    shipping_dt = pd.to_datetime(shipping_raw, errors='coerce', format='mixed')
                    is_same_day_label = shipping_raw.apply(self._should_force_assign_same_day)

                    inspector_counts = pd.to_numeric(result_df['検査員人数'], errors='coerce').fillna(0).astype(int)
                    unassigned_mask = inspector_counts <= 0

                    # 3営業日以内のロットも優先的に処理する
                    is_near = (
                        is_same_day_label
                        | (shipping_str == "当日")
                        | (shipping_dt.notna() & (shipping_dt.dt.date <= three_business_days_ahead))
                    )
                    has_unassigned_near = bool((unassigned_mask & is_near).any())
                    if has_unassigned_near:
                        # 3営業日以内のロットを優先するため、3営業日より後のロットを解除対象とする
                        far_assigned_mask = (
                            (inspector_counts > 0)
                            & shipping_dt.notna()
                            & (shipping_dt.dt.date > three_business_days_ahead)
                        )
                        far_assigned_count = int(far_assigned_mask.sum())
                        if far_assigned_count > 0:
                            for i in range(1, 6):
                                col = f'検査員{i}'
                                if col in result_df.columns:
                                    result_df.loc[far_assigned_mask, col] = ''
                            result_df.loc[far_assigned_mask, '検査員人数'] = 0
                            if 'remaining_work_hours' in result_df.columns and '検査時間' in result_df.columns:
                                result_df.loc[far_assigned_mask, 'remaining_work_hours'] = pd.to_numeric(
                                    result_df.loc[far_assigned_mask, '検査時間'], errors='coerce'
                                ).fillna(0.0)
                            if 'assignability_status' in result_df.columns:
                                result_df.loc[far_assigned_mask, 'assignability_status'] = 'unassigned'
                            if 'チーム情報' in result_df.columns:
                                result_df.loc[far_assigned_mask, 'チーム情報'] = '未割当(優先ロット保護のため解除)'

                            self._rebuild_assignment_histories(result_df, inspector_master_df)
                            self.log_message(
                                f"優先ロット（3営業日以内）未割当のため、それ以降の日付（3営業日より後）の割当を一旦解除して再処理します: {far_assigned_count}件",
                                level='warning'
                            )
            except Exception as e:
                self.log_message(f"優先ロット保護のための割当解除でエラーが発生しました: {str(e)}", level='warning')

            unassigned_indices = []
            # 列インデックスを事前に取得（高速化：itertuples()を使用）
            inspector_count_col_idx = result_df.columns.get_loc('検査員人数')
            
            for row_tuple in result_df.itertuples(index=True):
                index = row_tuple[0]  # インデックス
                inspector_count = row_tuple[inspector_count_col_idx + 1] if inspector_count_col_idx < len(row_tuple) - 1 else 0
                if inspector_count == 0 or pd.isna(inspector_count) or inspector_count == 0:
                    unassigned_indices.append(index)
            
            if unassigned_indices:
                # 未割当ロット再処理前に、履歴を再構築して勤務時間を正しく反映
                # これにより、一旦外したロットの勤務時間が履歴から除外される
                self._rebuild_assignment_histories(result_df, inspector_master_df)
                self.log_message(f"未割当ロットが {len(unassigned_indices)}件見つかりました。再処理を開始します（履歴を再構築済み）")
                
                # 出荷予定日順にソート（元のインデックスを保持）
                # 当日洗浄上がり品を最優先に処理
                unassigned_df = result_df.loc[unassigned_indices].copy()
                unassigned_df['_original_index'] = unassigned_indices  # 元のインデックスを保持
                
                # 出荷予定日の優先順位を計算（当日洗浄上がり品を優先）
                def calculate_unassigned_priority(row):
                    shipping_date_raw = row.get('出荷予定日', None)
                    shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                    
                    # 当日洗浄上がり品・先行検査品は優先順位1（最優先）
                    if (shipping_date_str == "当日洗浄上がり品" or
                        shipping_date_str == "当日洗浄品" or
                        "当日洗浄" in shipping_date_str or
                        shipping_date_str == "先行検査" or
                        shipping_date_str == "当日先行検査"):
                        return (1, pd.Timestamp.min)  # 最優先
                    
                    # 日付の場合
                    try:
                        shipping_date = pd.to_datetime(shipping_date_raw, errors='coerce')
                        if pd.notna(shipping_date):
                            return (2, shipping_date)  # 日付順
                    except Exception:
                        pass
                    
                    return (3, pd.Timestamp.max)  # その他は最後
                
                unassigned_df['_priority'] = unassigned_df.apply(calculate_unassigned_priority, axis=1)
                unassigned_df['_priority_value'] = unassigned_df['_priority'].apply(lambda x: x[0])
                unassigned_df['_priority_date'] = unassigned_df['_priority'].apply(lambda x: x[1])
                
                # 新規品かどうかを判定
                def is_new_product_for_unassigned(row):
                    product_number = row['品番']
                    skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                    return skill_rows.empty
                
                unassigned_df['_is_new_product'] = unassigned_df.apply(is_new_product_for_unassigned, axis=1)
                
                # 指示日のソートキーを追加（FIFO用：指示日が古い順）
                if '指示日' in unassigned_df.columns:
                    def instruction_date_sort_key_for_unassigned(val):
                        if pd.isna(val):
                            return pd.Timestamp.max  # 最後に
                        try:
                            date_val = pd.to_datetime(val, errors='coerce')
                            if pd.notna(date_val):
                                return date_val
                        except:
                            pass
                        return pd.Timestamp.max
                    unassigned_df['_instruction_date_sort_key'] = unassigned_df['指示日'].apply(instruction_date_sort_key_for_unassigned)
                else:
                    unassigned_df['_instruction_date_sort_key'] = pd.Timestamp.max
                
                # 優先順位順にソート（当日洗浄上がり品 > 日付順 > その他、同じ優先順位内では同一品番内で指示日が古い順（FIFO））
                unassigned_df = unassigned_df.sort_values(
                    ['_priority_value', '_priority_date', '品番', '_instruction_date_sort_key', '_is_new_product'], 
                    ascending=[True, True, True, True, False],  # 優先順位値は昇順、日付は昇順、品番は昇順、指示日は昇順（FIFO）、新規品フラグは降順（Trueを先に）
                    na_position='last'
                ).reset_index(drop=True)
                
                # 出荷予定日を変換（当日洗浄品は文字列として保持）
                unassigned_df['出荷予定日'] = unassigned_df['出荷予定日'].apply(self._convert_shipping_date)
                
                # 各未割当ロットを再処理
                original_indices = unassigned_df['_original_index'].tolist()  # 元のインデックスを保存
                
                # ソート用の列を削除
                unassigned_df = unassigned_df.drop(columns=['_priority', '_priority_value', '_priority_date', '_is_new_product', '_instruction_date_sort_key', '_original_index'])
                
                # 各未割当ロットを再処理
                # 列インデックスを事前に取得（itertuples()で高速化）
                prod_num_col_idx_u = unassigned_df.columns.get_loc('品番')
                inspection_time_col_idx = unassigned_df.columns.get_loc('検査時間') if '検査時間' in unassigned_df.columns else -1
                process_num_col_idx = unassigned_df.columns.get_loc('現在工程番号') if '現在工程番号' in unassigned_df.columns else -1
                lot_qty_col_idx = unassigned_df.columns.get_loc('ロット数量') if 'ロット数量' in unassigned_df.columns else -1
                
                for row_tuple in unassigned_df.itertuples(index=True):
                    idx = row_tuple[0]  # インデックス
                    original_index = original_indices[idx]  # 元のインデックスを取得
                    product_number = row_tuple[prod_num_col_idx_u + 1]  # +1はインデックス分
                    inspection_time = row_tuple[inspection_time_col_idx + 1] if inspection_time_col_idx >= 0 and inspection_time_col_idx + 1 < len(row_tuple) else 0.0
                    process_number = row_tuple[process_num_col_idx + 1] if process_num_col_idx >= 0 and process_num_col_idx + 1 < len(row_tuple) else ''
                    lot_quantity = row_tuple[lot_qty_col_idx + 1] if lot_qty_col_idx >= 0 and lot_qty_col_idx + 1 < len(row_tuple) else 0
                    
                    # rowオブジェクトが必要な場合は、元のDataFrameから取得
                    row = unassigned_df.loc[idx]

                    lot_process_name = str(row.get('現在工程名', '') or '').strip()
                    
                    # ロット数量が0の場合は検査員を割り当てない
                    if lot_quantity == 0 or pd.isna(lot_quantity) or inspection_time == 0 or pd.isna(inspection_time):
                        self.log_message(f"未割当ロット再処理: ロット数量が0または検査時間が0のため、品番 {product_number} の検査員割り当てをスキップします")
                        continue
                    
                    # 【修正】各割り当て処理の前の再計算を削除
                    # select_inspectors_with_skill_combinationで更新されたinspector_daily_assignmentsを
                    # 次の割り当て処理で確実に使用するため、割り当て成功後の更新のみを維持
                    
                    # 必要人数を計算
                    if inspection_time <= self.required_inspectors_threshold:
                        required_inspectors = 1
                    else:
                        calculated_inspectors = max(2, int(inspection_time / self.required_inspectors_threshold) + 1)
                        # 5名以上になる場合は5名に制限（特例）
                        required_inspectors = min(5, calculated_inspectors)
                    
                    divided_time = inspection_time / required_inspectors
                    
                    # 新規品かどうかを判定
                    skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                    is_new_product = skill_rows.empty
                    
                    # 出荷予定日が近日（3営業日以内または2週間以内）かどうかを判定
                    shipping_date = row.get('出荷予定日', None)
                    is_near_shipping_date = False
                    if pd.notna(shipping_date):
                        shipping_date = pd.to_datetime(shipping_date, errors='coerce')
                        if pd.notna(shipping_date):
                            shipping_date_date = shipping_date.date()
                            current_date = pd.Timestamp.now().date()
                            
                            # 3営業日以内かどうかを判定
                            def add_business_days_local(start: date, business_days: int) -> date:
                                def next_business_day(date_val: date) -> date:
                                    weekday = date_val.weekday()
                                    if weekday == 4:  # Friday
                                        return date_val + timedelta(days=3)
                                    if weekday == 5:  # Saturday
                                        return date_val + timedelta(days=2)
                                    return date_val + timedelta(days=1)
                                current = start
                                added = 0
                                while added < business_days:
                                    current = next_business_day(current)
                                    added += 1
                                return current
                            
                            three_business_days_ahead = add_business_days_local(current_date, 3)
                            two_weeks_later = current_date + timedelta(days=14)
                            
                            # 3営業日以内または2週間以内の場合は制約を緩和
                            if shipping_date_date <= three_business_days_ahead or shipping_date_date <= two_weeks_later:
                                is_near_shipping_date = True
                                if shipping_date_date <= three_business_days_ahead:
                                    self.log_message(f"未割当ロット再処理: 品番 {product_number} の出荷予定日が3営業日以内（{shipping_date_date}）のため、制約を緩和して処理します")
                                else:
                                    self.log_message(f"未割当ロット再処理: 品番 {product_number} の出荷予定日が近日（{shipping_date_date}）のため、制約を緩和して処理します")
                    
                    # 当日洗浄上がり品かどうかを判定
                    shipping_date_raw = row.get('出荷予定日', None)
                    shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                    is_same_day_cleaning = (
                        shipping_date_str == "当日洗浄上がり品" or
                        shipping_date_str == "当日洗浄品" or
                        "当日洗浄" in shipping_date_str or
                        shipping_date_str == "先行検査" or
                        shipping_date_str == "当日先行検査"
                    )
                    
                    # 利用可能な検査員を取得
                    shipping_date = row.get('出荷予定日', None)
                    if is_new_product:
                        self.log_message(f"未割当ロット再処理: 品番 {product_number} は新規品です。新製品チームを優先的に取得します")
                        available_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                        # 新製品チームが見つからない場合、通常の検査員も候補に追加
                        if not available_inspectors or len(available_inspectors) < required_inspectors:
                            self.log_message(
                                f"未割当ロット再処理: 新規品 {product_number} の新製品チームが不足しているため、"
                                f"通常の検査員も候補に追加します（新製品チーム: {len(available_inspectors) if available_inspectors else 0}人, 必要: {required_inspectors}人）",
                                debug=True,
                            )
                            fallback_inspectors = self.get_available_inspectors(
                                product_number, process_number, skill_master_df, inspector_master_df,
                                shipping_date=shipping_date, allow_new_team_fallback=True,
                                process_master_df=process_master_df, inspection_target_keywords=inspection_target_keywords,
                                process_name_context=lot_process_name
                            )
                            if fallback_inspectors:
                                # 新製品チームと通常の検査員を統合（重複を除去）
                                new_team_codes = {insp['コード'] for insp in available_inspectors} if available_inspectors else set()
                                if not available_inspectors:
                                    available_inspectors = []
                                for insp in fallback_inspectors:
                                    if insp['コード'] not in new_team_codes:
                                        available_inspectors.append(insp)
                                self.log_message(
                                    f"未割当ロット再処理: 新規品 {product_number} の候補を統合しました（合計: {len(available_inspectors)}人）",
                                    debug=True,
                                )
                    else:
                        available_inspectors = self.get_available_inspectors(
                            product_number, process_number, skill_master_df, inspector_master_df,
                            shipping_date=shipping_date, allow_new_team_fallback=is_near_shipping_date,
                            process_master_df=process_master_df, inspection_target_keywords=inspection_target_keywords,
                            process_name_context=lot_process_name
                        )
                        # 通常の検査員が見つからない場合、新製品チームも候補に追加（3営業日以内の場合）
                        if (not available_inspectors or len(available_inspectors) < required_inspectors) and is_near_shipping_date:
                            self.log_message(
                                f"未割当ロット再処理: 品番 {product_number} の通常の検査員が不足しているため、"
                                f"新製品チームも候補に追加します（通常: {len(available_inspectors) if available_inspectors else 0}人, 必要: {required_inspectors}人）",
                                debug=True,
                            )
                            new_team_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                            if new_team_inspectors:
                                # 通常の検査員と新製品チームを統合（重複を除去）
                                existing_codes = {insp['コード'] for insp in available_inspectors} if available_inspectors else set()
                                if not available_inspectors:
                                    available_inspectors = []
                                for insp in new_team_inspectors:
                                    if insp['コード'] not in existing_codes:
                                        available_inspectors.append(insp)
                                self.log_message(
                                    f"未割当ロット再処理: 品番 {product_number} の候補を統合しました（合計: {len(available_inspectors)}人）",
                                    debug=True,
                                )
                    
                    if not available_inspectors:
                        if is_new_product:
                            self.log_message(f"警告: 新規品 {product_number} の新製品チームが見つかりません")
                        else:
                            self.log_message(f"警告: 品番 {product_number} の検査員が見つかりません")
                        continue
                    
                    # 【改善】当日洗浄上がり品の場合は、既に割り当て済みのロットを再処理しない（フェーズ1・2で適切に割り当てられたロットを保護）
                    # ただし、フェーズ3は未割当ロットのみを処理するため、このチェックは念のための保護処理
                    if is_same_day_cleaning:
                        # 既に割り当て済みかどうかをチェック（念のため）
                        inspector_count_check = row.get('検査員人数', 0)
                        if inspector_count_check > 0 and pd.notna(inspector_count_check):
                            # 既に割り当て済みの場合はスキップ（フェーズ1・2で適切に割り当てられたロットを保護）
                            self.log_message(f"未割当ロット再処理: 当日洗浄上がり品 {product_number} は既に割り当て済み（検査員人数: {inspector_count_check}）のためスキップします（フェーズ1・2の結果を保護）")
                            continue
                    
                    # 当日洗浄上がり品の場合は、既にこの品番または同じ品名の他の品番に割り当てられた検査員を除外（品番単位・品名単位の制約）
                    # ただし、候補が0人の場合は優先順位が高いため、制約を緩和して割り当てを試みる
                    if is_same_day_cleaning:
                        # この品番に既に割り当てられた検査員を取得（品番単位の制約）
                        shipping_date_val = row.get('出荷予定日', None)
                        force_assign_same_day = self._should_force_assign_same_day(shipping_date_val)
                        already_assigned_to_this_product = set() if force_assign_same_day else self.same_day_cleaning_inspectors.get(product_number, set())
                        
                        # 【修正】品名単位の制約も取得
                        product_name = row.get('品名', '')
                        product_name_str = str(product_name).strip() if pd.notna(product_name) else ''
                        already_assigned_to_same_product_name = set()
                        if (not force_assign_same_day) and product_name_str:
                            already_assigned_to_same_product_name = self.same_day_cleaning_inspectors_by_product_name.get(product_name_str, set())
                        
                        # 品番単位と品名単位の両方の制約を統合
                        excluded_codes_for_reprocessing = already_assigned_to_this_product | already_assigned_to_same_product_name
                        
                        filtered_candidates = []
                        excluded_candidates = []
                        for inspector_candidate in available_inspectors:
                            if inspector_candidate['コード'] in excluded_codes_for_reprocessing:
                                excluded_candidates.append(inspector_candidate)
                            else:
                                filtered_candidates.append(inspector_candidate)
                        
                        # 品番単位・品名単位の制約を適用
                        available_inspectors = filtered_candidates
                        excluded_count = len(excluded_candidates)
                        if excluded_count > 0:
                            constraint_types = []
                            if len(already_assigned_to_this_product) > 0:
                                constraint_types.append("品番単位")
                            if len(already_assigned_to_same_product_name) > 0:
                                constraint_types.append("品名単位")
                            constraint_msg = "・".join(constraint_types) if constraint_types else "制約"
                            self.log_message(
                                f"未割当ロット再処理: 当日洗浄上がり品/先行検査品 {product_number}: "
                                f"既に割り当てられた検査員 {excluded_count}人を除外しました（{constraint_msg}の制約）",
                                debug=True,
                            )
                        
                        # 候補が0人の場合は、優先順位が高いため制約を緩和して元の候補を使用
                        if len(available_inspectors) == 0:
                            product_key = product_name_str if product_name_str else product_number
                            if excluded_candidates and self._should_relax_same_day_same_name(product_key):
                                available_inspectors = excluded_candidates.copy()
                                attempts = self.same_day_same_name_relaxation_attempts.get(product_key, 0)
                                self.log_message(
                                    f"?? Warning: 当日洗浄品 {product_number} の同一品名制約候補を緩和中 ({attempts}/{MAX_SAME_DAY_SAME_NAME_RELAXATIONS})",
                                    level='warning'
                                )
                            else:
                                constraint_types = []
                                if len(already_assigned_to_this_product) > 0:
                                    constraint_types.append("品番単位")
                                if len(already_assigned_to_same_product_name) > 0:
                                    constraint_types.append("品名単位")
                                constraint_msg = "・".join(constraint_types) if constraint_types else "制約"
                                self.log_message(f"⚠️ 警告: 未割当ロット再処理: 当日洗浄上がり品 {product_number} の候補が0人ですが、優先順位が高いため{constraint_msg}の制約を緩和して再試行します", level='warning')
                                # 元の候補を再取得（フィルタリング前）
                            if is_new_product:
                                available_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                            else:
                                available_inspectors = self.get_available_inspectors(
                                    product_number, process_number, skill_master_df, inspector_master_df,
                                    shipping_date=shipping_date, allow_new_team_fallback=True,
                                    process_master_df=process_master_df, inspection_target_keywords=inspection_target_keywords,
                                    process_name_context=lot_process_name
                                )
                                self.log_message(f"未割当ロット再処理: 当日洗浄上がり品 {product_number}: 制約緩和後の候補数 {len(available_inspectors)}人")
                        elif len(available_inspectors) < required_inspectors:
                            self.log_message(f"⚠️ 警告: 未割当ロット再処理: 当日洗浄上がり品 {product_number} の候補が不足しています（必要: {required_inspectors}人、利用可能: {len(available_inspectors)}人）。可能な限り割り当てます。", level='warning')
                    
                    # 未割当ロット再処理時は、4時間上限を緩和して再試行
                    # まず通常の条件で試行
                    assigned_inspectors = self.select_inspectors(
                        available_inspectors, required_inspectors, divided_time,
                        inspector_master_df, product_number,
                        is_new_product=is_new_product,
                        ignore_product_limit=is_same_day_cleaning,
                    )
                    
                    # 割り当てできなかった場合、4時間上限を緩和（3.5h以上も許可）して再試行
                    # ただし、当日洗浄上がり品の場合は品番単位の制約を維持
                    if not assigned_inspectors:
                        self.log_message(f"未割当ロット再処理: 品番 {product_number} の通常条件での割り当てに失敗。4時間上限を緩和して再試行します")
                        # 4時間上限を緩和した候補を取得
                        # 未割当ロット再処理時は、以前の振分け結果（inspector_product_hours）を考慮しつつ、
                        # 4時間上限を緩和（PRODUCT_LIMIT_FINAL_TOLERANCE=4.2h未満まで許容）して優先ロットの割り当てを優先
                        relaxed_candidates = []
                        excluded_by_constraint = []  # 制約で除外された検査員（当日洗浄上がり品の品番単位・品名単位の制約のみ）
                        excluded_by_limit = []  # 4時間上限で除外された検査員（4.2h超過の場合のみ除外）
                        # 当日洗浄上がり品の場合は、既にこの品番または同じ品名の他の品番に割り当てられた検査員を除外（品番単位・品名単位の制約）
                        excluded_codes_for_relaxed = set()
                        if is_same_day_cleaning:
                            already_assigned_to_this_product = self.same_day_cleaning_inspectors.get(product_number, set())
                            product_name = row.get('品名', '')
                            product_name_str = str(product_name).strip() if pd.notna(product_name) else ''
                            already_assigned_to_same_product_name = set()
                            if product_name_str:
                                already_assigned_to_same_product_name = self.same_day_cleaning_inspectors_by_product_name.get(product_name_str, set())
                            excluded_codes_for_relaxed = already_assigned_to_this_product | already_assigned_to_same_product_name
                            if excluded_codes_for_relaxed:
                                self.log_message(f"未割当ロット再処理: 品番 {product_number} の制約により {len(excluded_codes_for_relaxed)}名の検査員を除外（品番単位・品名単位の制約）", debug=True)
                        excluded_by_work_hours = []  # 勤務時間で除外された検査員
                        current_date = pd.Timestamp.now().date()
                        for insp in available_inspectors:
                            code = insp['コード']
                            # 当日洗浄上がり品の場合は、既にこの品番または同じ品名の他の品番に割り当てられた検査員を除外
                            if is_same_day_cleaning and code in excluded_codes_for_relaxed:
                                excluded_by_constraint.append(f"{insp['氏名']}({code})")
                                continue
                            
                            # 勤務時間チェック（緩和モード: 20%超過まで許容）
                            daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                            max_daily_hours = self.get_inspector_max_hours(code, inspector_master_df)
                            if max_daily_hours <= 0:
                                excluded_by_work_hours.append(f"{insp['氏名']}({code}): 調整後勤務時間0時間以下")
                                continue
                            # 緩和モード: 20%超過まで許容（SAME_DAY_WORK_HOURS_OVERRUN_RATEを使用）
                            relaxed_allowed_max_hours = max_daily_hours * (1.0 + SAME_DAY_WORK_HOURS_OVERRUN_RATE)
                            if daily_hours + divided_time > relaxed_allowed_max_hours - WORK_HOURS_BUFFER:
                                excluded_by_work_hours.append(f"{insp['氏名']}({code}): {daily_hours:.1f}h+{divided_time:.1f}h>{relaxed_allowed_max_hours:.1f}h-{WORK_HOURS_BUFFER:.2f}h")
                                continue
                            
                            # 未割当ロット再処理時は、以前の振分け結果を考慮しつつ、4時間上限を緩和（4.2h未満まで許容）
                            # これにより、ルールの一貫性を保ちつつ、優先ロットの割り当てを優先できる
                            current = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                            projected_hours = current + divided_time
                            # 未割当ロット再処理時は、PRODUCT_LIMIT_FINAL_TOLERANCE（4.2h）未満まで許容
                            # 4.2h超過の場合は除外（ルールの一貫性を保つ）
                            if projected_hours >= PRODUCT_LIMIT_FINAL_TOLERANCE:
                                excluded_by_limit.append({
                                    'name': insp['氏名'],
                                    'code': code,
                                    'current': current,
                                    'divided_time': divided_time,
                                    'projected': projected_hours
                                })
                                continue
                            # 3.5h以上4.2h未満の場合は警告フラグを付ける（参考情報として）
                            if projected_hours >= 3.5:
                                insp['__near_product_limit'] = True
                            # 4.2h未満の場合は緩和候補として追加
                            relaxed_candidates.append(insp)
                        
                        # 詳細なログを出力
                        if excluded_by_constraint:
                            self.log_message(f"未割当ロット再処理: 制約で除外された検査員: {', '.join(excluded_by_constraint[:5])}{'...' if len(excluded_by_constraint) > 5 else ''}", debug=True)
                        if excluded_by_limit:
                            limit_details = []
                            for excl in excluded_by_limit[:3]:
                                limit_details.append(f"{excl['name']}({excl['code']}): {excl['current']:.1f}h+{excl['divided_time']:.1f}h={excl['projected']:.1f}h（{PRODUCT_LIMIT_FINAL_TOLERANCE}h超過）")
                            self.log_message(f"未割当ロット再処理: 4時間上限で除外された検査員（{PRODUCT_LIMIT_FINAL_TOLERANCE}h超過）: {', '.join(limit_details)}{'...' if len(excluded_by_limit) > 3 else ''}", debug=True)
                        
                        if excluded_by_work_hours:
                            self.log_message(f"未割当ロット再処理: 勤務時間で除外された検査員: {', '.join(excluded_by_work_hours[:5])}{'...' if len(excluded_by_work_hours) > 5 else ''}", debug=True)
                        if relaxed_candidates:
                            self.log_message(f"未割当ロット再処理: 緩和候補 {len(relaxed_candidates)}名（制約除外: {len(excluded_by_constraint)}名、勤務時間除外: {len(excluded_by_work_hours)}名、上限除外: {len(excluded_by_limit)}名、緩和上限: {PRODUCT_LIMIT_FINAL_TOLERANCE}h未満）")
                            # 緩和条件で再試行（未割当ロット再処理時は4時間上限を緩和（4.2h未満まで許容）し、勤務時間制約も緩和）
                            # 当日洗浄上がり品の場合は品番単位・品名単位の制約を維持するため、is_same_day_cleaningを考慮
                            # 注意: 未割当ロット再処理時は、以前の振分け結果（inspector_product_hours）を考慮しつつ、
                            # 4時間上限を緩和（4.2h未満まで許容）して優先ロットの割り当てを優先
                            # ignore_product_limit=Trueにより、select_inspectors内での4時間上限チェックはスキップされるが、
                            # 緩和候補作成時には4.2h未満まで許容するチェックを実施しているため、ルールの一貫性が保たれる
                            self.log_message(f"未割当ロット再処理: 品番 {product_number} の緩和候補 {len(relaxed_candidates)}名をselect_inspectorsに渡します（relax_work_hours=True, ignore_product_limit=True）")
                            assigned_inspectors = self.select_inspectors(
                                relaxed_candidates, required_inspectors, divided_time,
                                inspector_master_df, product_number,
                                is_new_product=is_new_product,
                                relax_work_hours=True,  # 未割当ロット再処理時は勤務時間制約も緩和
                                ignore_product_limit=True,  # 未割当ロット再処理時は4時間上限チェックをスキップ（緩和候補作成時に4.2h未満まで許容済み）
                            )
                            if assigned_inspectors:
                                self.log_message(f"未割当ロット再処理: 品番 {product_number} の上限緩和条件で割り当て成功（{len(assigned_inspectors)}名）")
                        else:
                            self.log_message(f"警告: 未割当ロット再処理: 品番 {product_number} の緩和候補が0名（制約除外: {len(excluded_by_constraint)}名、上限除外: {len(excluded_by_limit)}名）")
                    
                    # まだ割り当てできず、かつ出荷予定日が近日の新規品の場合、勤務時間制約も緩和して再試行
                    # ただし、当日洗浄上がり品の場合は品番単位の制約を維持
                    if not assigned_inspectors and is_new_product and is_near_shipping_date:
                        self.log_message(f"未割当ロット再処理: 品番 {product_number} の出荷予定日が近日のため、勤務時間制約も緩和して再試行します")
                        # 勤務時間制約を緩和した候補を取得
                        relaxed_work_hours_candidates = []
                        current_date = pd.Timestamp.now().date()
                        # 当日洗浄上がり品の場合は、既に当日洗浄上がり品全体に割り当てられた検査員を除外（当日洗浄上がり品全体の制約）
                        already_assigned_to_same_day_cleaning = self.same_day_cleaning_inspectors if is_same_day_cleaning else set()
                        
                        for insp in available_inspectors:
                            code = insp['コード']
                            # 当日洗浄上がり品の場合は、既に当日洗浄上がり品全体に割り当てられた検査員を除外
                            if is_same_day_cleaning and code in already_assigned_to_same_day_cleaning:
                                continue
                            # 4時間上限チェック
                            current = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                            projected_hours = current + divided_time
                            if projected_hours >= 4.5:
                                continue
                            
                            # 【修正】10%超過を厳格にチェック（WORK_HOURS_BUFFERを考慮）
                            daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                            max_daily_hours = self.get_inspector_max_hours(code, inspector_master_df)
                            allowed_max_hours = self._apply_work_hours_overrun(max_daily_hours)
                            # 10%超過を超える場合は除外（WORK_HOURS_BUFFERを考慮）
                            if daily_hours + divided_time > allowed_max_hours - WORK_HOURS_BUFFER:
                                    self.log_message(
                                    f"未割当ロット再処理（勤務時間緩和）: 検査員 '{insp['氏名']}' は10%超過を超えるため除外 "
                                    f"(今日: {daily_hours:.1f}h + {divided_time:.1f}h > {allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h)",
                                        debug=True
                                    )
                                    continue
                            
                            if projected_hours >= 3.5:
                                insp['__near_product_limit'] = True
                            relaxed_work_hours_candidates.append(insp)
                        
                        if relaxed_work_hours_candidates:
                            # 勤務時間制約を緩和した候補で再試行
                            # relax_work_hours=Trueを指定してselect_inspectorsを呼ぶ
                            assigned_inspectors = self.select_inspectors(
                                relaxed_work_hours_candidates, required_inspectors, divided_time,
                                inspector_master_df, product_number,
                                is_new_product=is_new_product,
                                relax_work_hours=True,
                                ignore_product_limit=is_same_day_cleaning,
                            )
                            if assigned_inspectors:
                                # 【追加】勤務時間制約緩和の履歴追跡
                                self.same_day_relaxation_metrics['constraints_relaxed']['work_hours_constraint'] += 1
                                relaxation_history_entry = {
                                    'lot_index': original_index,
                                    'product_number': product_number,
                                    'reason': 'work_hours_relaxation',
                                    'inspection_time': inspection_time,
                                    'constraints_relaxed': ['work_hours_constraint'],
                                    'candidates_count': len(relaxed_work_hours_candidates),
                                }
                                self.same_day_relaxation_metrics['relaxation_history'].append(relaxation_history_entry)
                                if len(self.same_day_relaxation_metrics['relaxation_history']) > 100:
                                    self.same_day_relaxation_metrics['relaxation_history'] = \
                                        self.same_day_relaxation_metrics['relaxation_history'][-100:]
                                
                                self.log_message(f"未割当ロット再処理: 品番 {product_number} の勤務時間制約緩和条件で割り当て成功")
                    
                    # 当日洗浄上がり品の場合は、候補が不足していても制約を大幅に緩和して割り当てを試みる
                    if not assigned_inspectors and is_same_day_cleaning:
                        # 【追加】制約緩和の履歴追跡
                        relaxation_reason = 'same_day_cleaning_unassigned'
                        self.same_day_relaxation_metrics['total_relaxations'] += 1
                        self.same_day_relaxation_metrics['relaxation_by_reason'][relaxation_reason] = \
                            self.same_day_relaxation_metrics['relaxation_by_reason'].get(relaxation_reason, 0) + 1
                        
                        # 【追加】SAME_DAY_RELAXATION_THRESHOLD_HOURS によるチェック
                        self.same_day_relaxation_metrics['relaxation_threshold_checks'] += 1
                        if inspection_time >= SAME_DAY_RELAXATION_THRESHOLD_HOURS:
                            self.same_day_relaxation_metrics['relaxation_threshold_applied'] += 1
                        
                        self.log_message(f"未割当ロット再処理: 当日洗浄上がり品 {product_number} の通常条件での割り当てに失敗。制約を大幅に緩和して再試行します")
                        
                        # 【改善】制約を緩和する前に、現在のresult_dfから実際に割り当てられている検査員を確認
                        # 品番単位の制約を再構築（高速化：事前にフィルタリング）
                        already_assigned_to_this_product = set()
                        # 同じ品番のロットのみをフィルタリング（高速化）
                        same_product_df = result_df[
                            (result_df['品番'] == product_number) & 
                            (result_df.index != original_index)
                        ]
                        if not same_product_df.empty and '出荷予定日' in same_product_df.columns:
                            # 当日洗浄上がり品/先行検査品のみをフィルタリング
                            same_day_mask = (
                                same_product_df['出荷予定日'].astype(str).str.contains('当日洗浄|先行検査', na=False)
                            )
                            same_day_df = same_product_df[same_day_mask]
                            
                            # 検査員列から割り当て済み検査員を取得
                            for j in range(1, 6):
                                col_name = f'検査員{j}'
                                if col_name in same_day_df.columns:
                                    for inspector_value in same_day_df[col_name].dropna():
                                        inspector_name = str(inspector_value).strip()
                                        if inspector_name and '(' in inspector_name:
                                            inspector_name = inspector_name.split('(')[0].strip()
                                        if inspector_name:
                                            inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                            if not inspector_info.empty:
                                                already_assigned_to_this_product.add(inspector_info.iloc[0]['#ID'])
                        
                        # 品名単位の制約を再構築（高速化：事前にフィルタリング）
                        product_name = row.get('品名', '')
                        product_name_str = str(product_name).strip() if pd.notna(product_name) else ''
                        already_assigned_to_same_product_name = set()
                        if product_name_str and '品名' in result_df.columns:
                            # 同じ品名で異なる品番のロットのみをフィルタリング（高速化）
                            same_name_df = result_df[
                                (result_df['品名'].astype(str).str.strip() == product_name_str) &
                                (result_df['品番'] != product_number) &
                                (result_df.index != original_index)
                            ]
                            if not same_name_df.empty and '出荷予定日' in same_name_df.columns:
                                # 当日洗浄上がり品/先行検査品のみをフィルタリング
                                same_day_mask = (
                                    same_name_df['出荷予定日'].astype(str).str.contains('当日洗浄|先行検査', na=False)
                                )
                                same_day_name_df = same_name_df[same_day_mask]
                                
                                # 検査員列から割り当て済み検査員を取得
                                for j in range(1, 6):
                                    col_name = f'検査員{j}'
                                    if col_name in same_day_name_df.columns:
                                        for inspector_value in same_day_name_df[col_name].dropna():
                                            inspector_name = str(inspector_value).strip()
                                            if inspector_name and '(' in inspector_name:
                                                inspector_name = inspector_name.split('(')[0].strip()
                                            if inspector_name:
                                                inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                                if not inspector_info.empty:
                                                    already_assigned_to_same_product_name.add(inspector_info.iloc[0]['#ID'])
                        
                        # 当日洗浄上がり品の場合は、制約を大幅に緩和して候補を取得
                        # 候補が0人の場合は、元の候補を再取得（当日洗浄上がり品全体の制約を緩和）
                        if len(available_inspectors) == 0:
                            self.log_message(f"未割当ロット再処理: 当日洗浄上がり品 {product_number} の候補が0人のため、元の候補を再取得します（当日洗浄上がり品全体の制約を緩和）")
                            if is_new_product:
                                available_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                            else:
                                available_inspectors = self.get_available_inspectors(
                                    product_number, process_number, skill_master_df, inspector_master_df,
                                    shipping_date=shipping_date, allow_new_team_fallback=True,
                                    process_master_df=process_master_df, inspection_target_keywords=inspection_target_keywords,
                                    process_name_context=lot_process_name
                                )
                            self.log_message(f"未割当ロット再処理: 当日洗浄上がり品 {product_number}: 制約緩和後の候補数 {len(available_inspectors)}人")
                        
                        relaxed_same_day_candidates = []
                        current_date = pd.Timestamp.now().date()
                        
                        # 元のavailable_inspectorsから、当日洗浄上がり品の制約を緩和した候補を取得
                        # 候補が0人の場合は、品番単位・品名単位の制約も緩和（既に割り当てられた検査員も含める）
                        # 候補が0人でない場合は、品番単位・品名単位の制約を維持
                        use_original_candidates = len(available_inspectors) == 0
                        for insp in available_inspectors:
                            code = insp['コード']
                            # 候補が0人の場合は、品番単位・品名単位の制約を緩和（既に割り当てられた検査員も含める）
                            # それ以外の場合は、品番単位・品名単位の制約を維持
                            if not use_original_candidates:
                                # 品番単位の制約をチェック
                                if code in already_assigned_to_this_product:
                                    continue
                                # 【改善】品名単位の制約もチェック
                                if product_name_str and code in already_assigned_to_same_product_name:
                                    continue
                            
                            # 4時間上限を大幅に緩和（4.5hまで許容）
                            current = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                            projected_hours = current + divided_time
                            if projected_hours >= 4.5:
                                continue
                            
                            # 【修正】10%超過を厳格にチェック（WORK_HOURS_BUFFERを考慮）
                            daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                            max_daily_hours = self.get_inspector_max_hours(code, inspector_master_df)
                            allowed_max_hours = self._apply_work_hours_overrun(max_daily_hours)
                            # 10%超過を超える場合は除外（WORK_HOURS_BUFFERを考慮）
                            if daily_hours + divided_time > allowed_max_hours - WORK_HOURS_BUFFER:
                                self.log_message(
                                    f"未割当ロット再処理（当日洗浄上がり品制約緩和）: 検査員 '{insp['氏名']}' は10%超過を超えるため除外 "
                                    f"(今日: {daily_hours:.1f}h + {divided_time:.1f}h > {allowed_max_hours:.1f}h - {WORK_HOURS_BUFFER:.2f}h)",
                                    debug=True
                                )
                                continue
                            
                            relaxed_same_day_candidates.append(insp)
                        
                        if relaxed_same_day_candidates:
                            # 【追加】制約緩和の履歴追跡
                            constraints_relaxed = []
                            if use_original_candidates:
                                constraints_relaxed.append('product_name_constraint')
                                constraints_relaxed.append('product_number_constraint')
                                self.same_day_relaxation_metrics['constraints_relaxed']['product_name_constraint'] += 1
                                self.same_day_relaxation_metrics['constraints_relaxed']['product_number_constraint'] += 1
                            if is_same_day_cleaning:
                                constraints_relaxed.append('product_limit_constraint')
                                self.same_day_relaxation_metrics['constraints_relaxed']['product_limit_constraint'] += 1
                            constraints_relaxed.append('work_hours_constraint')
                            self.same_day_relaxation_metrics['constraints_relaxed']['work_hours_constraint'] += 1
                            
                            # 緩和履歴を記録
                            relaxation_history_entry = {
                                'lot_index': original_index,
                                'product_number': product_number,
                                'reason': relaxation_reason,
                                'inspection_time': inspection_time,
                                'constraints_relaxed': constraints_relaxed,
                                'candidates_count': len(relaxed_same_day_candidates),
                                'use_original_candidates': use_original_candidates,
                            }
                            self.same_day_relaxation_metrics['relaxation_history'].append(relaxation_history_entry)
                            # 履歴は最大100件まで保持
                            if len(self.same_day_relaxation_metrics['relaxation_history']) > 100:
                                self.same_day_relaxation_metrics['relaxation_history'] = \
                                    self.same_day_relaxation_metrics['relaxation_history'][-100:]
                            
                            # 緩和条件で再試行
                            assigned_inspectors = self.select_inspectors(
                                relaxed_same_day_candidates, required_inspectors, divided_time,
                                inspector_master_df, product_number,
                                is_new_product=is_new_product,
                                relax_work_hours=True,
                                ignore_product_limit=is_same_day_cleaning,
                            )
                            if assigned_inspectors:
                                self.log_message(
                                    f"未割当ロット再処理: 当日洗浄上がり品 {product_number} の制約緩和条件で割り当て成功 "
                                    f"(緩和された制約: {', '.join(constraints_relaxed)})"
                                )
                    
                    if assigned_inspectors:
                        # 割り当て成功
                        result_df.at[original_index, '検査員人数'] = len(assigned_inspectors)
                        # 分割検査時間の計算: 検査時間 ÷ 実際の分割した検査人数
                        actual_divided_time = inspection_time / len(assigned_inspectors)
                        result_df.at[original_index, '分割検査時間'] = round(actual_divided_time, 1)
                        result_df.at[original_index, 'remaining_work_hours'] = 0.0
                        result_df.at[original_index, 'assignability_status'] = 'fully_assigned'
                        
                        # 検査員名を設定
                        team_members = []
                        for i, inspector in enumerate(assigned_inspectors):
                            if i < 5:
                                if show_skill_values:
                                    if inspector.get('is_new_team', False):
                                        inspector_name = f"{inspector['氏名']}(新)"
                                    else:
                                        inspector_name = f"{inspector['氏名']}({inspector['スキル']})"
                                else:
                                    if inspector.get('is_new_team', False):
                                        inspector_name = f"{inspector['氏名']}(新)"
                                    else:
                                        inspector_name = inspector['氏名']
                                
                                result_df.at[original_index, f'検査員{i+1}'] = inspector_name
                                team_members.append(inspector['氏名'])
                                
                                # 履歴を更新
                                code = inspector['コード']
                                current_date = pd.Timestamp.now().date()
                                
                                # inspector_daily_assignmentsとinspector_work_hoursを更新
                                if code not in self.inspector_daily_assignments:
                                    self.inspector_daily_assignments[code] = {}
                                if current_date not in self.inspector_daily_assignments[code]:
                                    self.inspector_daily_assignments[code][current_date] = 0.0
                                self.inspector_daily_assignments[code][current_date] += divided_time
                                
                                if code not in self.inspector_work_hours:
                                    self.inspector_work_hours[code] = 0.0
                                self.inspector_work_hours[code] += divided_time
                                
                                # inspector_product_hoursを更新
                                if code not in self.inspector_product_hours:
                                    self.inspector_product_hours[code] = {}
                                self.inspector_product_hours[code][product_number] = (
                                    self.inspector_product_hours[code].get(product_number, 0.0) + divided_time
                                )
                        
                        # チーム情報を設定
                        if len(assigned_inspectors) > 1:
                            team_info = f"チーム: {', '.join(team_members)}"
                        else:
                            team_info = f"個人: {team_members[0] if team_members else ''}"
                        
                        result_df.at[original_index, 'チーム情報'] = team_info
                        over_limit_present = any(isinstance(insp, dict) and insp.get('over_product_limit') for insp in assigned_inspectors)
                        result_df.at[original_index, 'over_product_limit_flag'] = over_limit_present
                        if over_limit_present:
                            for insp in assigned_inspectors:
                                if isinstance(insp, dict) and insp.get('over_product_limit') and 'コード' in insp:
                                    self.relaxed_product_limit_assignments.add((insp['コード'], product_number))
                        
                        # 当日洗浄上がり品の検査員を追跡（重複を避けるため、当日洗浄上がり品全体、品番単位・品名単位）
                        if is_same_day_cleaning:
                            # 品名を取得
                            product_name = row.get('品名', '')
                            product_name_str = str(product_name).strip() if pd.notna(product_name) else ''
                            
                            for inspector in assigned_inspectors:
                                if isinstance(inspector, dict) and 'コード' in inspector:
                                    code = inspector['コード']
                                    self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(code)
                                    # 【追加】品名単位でも検査員を記録
                                    if product_name_str:
                                        self.same_day_cleaning_inspectors_by_product_name.setdefault(product_name_str, set()).add(code)
                            self.log_message(f"未割当ロット再処理: 当日洗浄上がり品/先行検査品 {product_number} の検査員を追跡しました（{len(assigned_inspectors)}人）")
                        
                        self.log_message(f"未割当ロット再処理成功: 品番 {product_number}, 出荷予定日 {row['出荷予定日']}, {len(assigned_inspectors)}人割り当て")
                    else:
                        # 未割当原因分析ログを追加
                        unassigned_reasons = []
                        
                        # 1. 利用可能な検査員が存在するか
                        if not available_inspectors:
                            if is_new_product:
                                unassigned_reasons.append("新製品チームが見つからない")
                            else:
                                unassigned_reasons.append("スキルマッチする検査員が見つからない")
                        else:
                            # 2. 制約による除外
                            if is_same_day_cleaning:
                                excluded_by_product = len(already_assigned_to_this_product) if 'already_assigned_to_this_product' in locals() else 0
                                excluded_by_name = len(already_assigned_to_same_product_name) if 'already_assigned_to_same_product_name' in locals() else 0
                                if excluded_by_product > 0 or excluded_by_name > 0:
                                    constraint_details = []
                                    if excluded_by_product > 0:
                                        constraint_details.append(f"品番単位制約で{excluded_by_product}人除外")
                                    if excluded_by_name > 0:
                                        constraint_details.append(f"品名単位制約で{excluded_by_name}人除外")
                                    unassigned_reasons.append(f"当日洗浄上がり品制約: {', '.join(constraint_details)}")
                            
                            # 3. 勤務時間制約
                            if 'relaxed_work_hours_candidates' in locals() and len(relaxed_work_hours_candidates) == 0:
                                unassigned_reasons.append("全候補が10%超過を超える")
                            
                            # 4. 同一品番4時間上限
                            if 'relaxed_candidates' in locals() and len(relaxed_candidates) == 0:
                                unassigned_reasons.append("全候補が同一品番4.5時間上限を超える")
                            
                            # 5. 必要人数不足
                            if len(available_inspectors) < required_inspectors:
                                unassigned_reasons.append(f"必要人数({required_inspectors}人)に対して候補不足({len(available_inspectors)}人)")
                        
                        reason_msg = "、".join(unassigned_reasons) if unassigned_reasons else "原因不明"
                        self.log_message(
                            f"警告: 未割当ロット {product_number} (出荷予定日: {row.get('出荷予定日', 'N/A')}, "
                            f"検査時間: {inspection_time:.1f}h, 必要人数: {required_inspectors}人) の再処理に失敗しました。"
                            f"原因: {reason_msg}（ルール違反を避けるため未割当のまま）",
                            level='warning'
                        )
                
                # 履歴を再計算（再処理後の状態を反映）
                self.inspector_daily_assignments = {}
                self.inspector_work_hours = {}
                self.inspector_product_hours = {}
                
                current_date_temp = pd.Timestamp.now().date()
                # 列インデックスを事前に取得（itertuples()で高速化）
                prod_num_col_idx_t = result_df.columns.get_loc('品番')
                div_time_col_idx_t = result_df.columns.get_loc('分割検査時間') if '分割検査時間' in result_df.columns else -1
                inspector_col_indices_t = {}
                for i in range(1, 6):
                    col_name = f'検査員{i}'
                    if col_name in result_df.columns:
                        inspector_col_indices_t[i] = result_df.columns.get_loc(col_name)
                
                for row_tuple in result_df.itertuples(index=False):
                    product_number = row_tuple[prod_num_col_idx_t]
                    divided_time = row_tuple[div_time_col_idx_t] if div_time_col_idx_t >= 0 and div_time_col_idx_t < len(row_tuple) else 0.0
                    
                    for i in range(1, 6):
                        if i not in inspector_col_indices_t:
                            continue
                        inspector_col_idx = inspector_col_indices_t[i]
                        inspector_value = row_tuple[inspector_col_idx] if inspector_col_idx < len(row_tuple) else None
                        
                        if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                            inspector_name = str(inspector_value).strip()
                            if '(' in inspector_name:
                                inspector_name = inspector_name.split('(')[0].strip()
                            if not inspector_name:
                                continue
                            
                            inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                            if not inspector_info.empty:
                                inspector_code = inspector_info.iloc[0]['#ID']
                                
                                if inspector_code not in self.inspector_daily_assignments:
                                    self.inspector_daily_assignments[inspector_code] = {}
                                if current_date_temp not in self.inspector_daily_assignments[inspector_code]:
                                    self.inspector_daily_assignments[inspector_code][current_date_temp] = 0.0
                                if inspector_code not in self.inspector_work_hours:
                                    self.inspector_work_hours[inspector_code] = 0.0
                                if inspector_code not in self.inspector_product_hours:
                                    self.inspector_product_hours[inspector_code] = {}
                                if product_number not in self.inspector_product_hours[inspector_code]:
                                    self.inspector_product_hours[inspector_code][product_number] = 0.0
                                
                                self.inspector_daily_assignments[inspector_code][current_date_temp] += divided_time
                                self.inspector_work_hours[inspector_code] += divided_time
                                self.inspector_product_hours[inspector_code][product_number] += divided_time
                
                self.log_message("未割当ロットの再処理が完了しました")
                
                # 【追加】当日洗浄上がり品の制約緩和条件の効果測定メトリクスを出力
                if self.same_day_relaxation_metrics['total_relaxations'] > 0:
                    self.log_message(
                        f"当日洗浄上がり品の制約緩和条件効果測定メトリクス: "
                        f"総緩和回数={self.same_day_relaxation_metrics['total_relaxations']}, "
                        f"緩和理由別={dict(self.same_day_relaxation_metrics['relaxation_by_reason'])}, "
                        f"緩和された制約: "
                        f"品名単位制約={self.same_day_relaxation_metrics['constraints_relaxed']['product_name_constraint']}回, "
                        f"品番単位制約={self.same_day_relaxation_metrics['constraints_relaxed']['product_number_constraint']}回, "
                        f"同一品番4.0h上限制約={self.same_day_relaxation_metrics['constraints_relaxed']['product_limit_constraint']}回, "
                        f"勤務時間制約={self.same_day_relaxation_metrics['constraints_relaxed']['work_hours_constraint']}回, "
                        f"緩和閾値チェック回数={self.same_day_relaxation_metrics['relaxation_threshold_checks']}, "
                        f"緩和閾値適用回数={self.same_day_relaxation_metrics['relaxation_threshold_applied']}, "
                        f"緩和履歴件数={len(self.same_day_relaxation_metrics['relaxation_history'])}",
                        debug=True,
                    )
                    if self.same_day_relaxation_metrics['relaxation_history']:
                        # 最初の5件の履歴を出力
                        self.log_message(
                            f"緩和履歴（最初の5件）: {self.same_day_relaxation_metrics['relaxation_history'][:5]}",
                            debug=True,
                        )
            else:
                self.log_message("未割当ロットはありませんでした")
            
            perf_logger.debug(
                "PERF {}: {:.1f} ms",
                "inspector_assignment.optimize.phase3.total",
                (perf_counter() - _t_perf_phase3_total) * 1000.0,
            )

            # フェーズ3.5: 未割当ロット再処理後の最終検証（勤務時間超過の再チェック）
            self.log_message("全体最適化フェーズ3.5: 未割当ロット再処理後の最終検証を開始")
            _t_perf_phase3_5_total = perf_counter()
            
            # 現在日付を取得
            current_date = pd.Timestamp.now().date()
            
            # 履歴を再計算
            self.inspector_daily_assignments = {}
            self.inspector_work_hours = {}
            self.inspector_product_hours = {}
            
            # 列インデックスを事前に取得（itertuples()で高速化）
            prod_num_col_idx = result_df.columns.get_loc('品番')
            div_time_col_idx = result_df.columns.get_loc('分割検査時間') if '分割検査時間' in result_df.columns else -1
            shipping_date_col_idx = result_df.columns.get_loc('出荷予定日') if '出荷予定日' in result_df.columns else -1
            inspector_col_indices = {}
            for i in range(1, 6):
                col_name = f'検査員{i}'
                if col_name in result_df.columns:
                    inspector_col_indices[i] = result_df.columns.get_loc(col_name)
            
            for row_tuple in result_df.itertuples(index=False):
                product_number = row_tuple[prod_num_col_idx]
                divided_time = row_tuple[div_time_col_idx] if div_time_col_idx >= 0 and div_time_col_idx < len(row_tuple) else 0.0
                
                # 出荷予定日から日付を取得（当日洗浄上がり品の場合は今日の日付を使用）
                lot_date = current_date  # デフォルトは今日の日付
                if shipping_date_col_idx >= 0 and shipping_date_col_idx < len(row_tuple):
                    shipping_date_raw = row_tuple[shipping_date_col_idx]
                    if pd.notna(shipping_date_raw):
                        shipping_date_str = str(shipping_date_raw).strip()
                        # 当日洗浄上がり品の場合は今日の日付を使用
                        if self._is_same_day_cleaning(shipping_date_str):
                            lot_date = current_date
                        else:
                            # 日付型に変換を試みる
                            try:
                                shipping_date_dt = pd.to_datetime(shipping_date_raw, errors='coerce')
                                if pd.notna(shipping_date_dt):
                                    # 異常な日付（例: 1677年など）をチェック
                                    if shipping_date_dt.year >= 1900 and shipping_date_dt.year <= 2100:
                                        lot_date = shipping_date_dt.date()
                                    else:
                                        # 異常な日付の場合は今日の日付を使用
                                        self.log_message(
                                            f"警告: 出荷予定日が異常な値です ({shipping_date_raw})。今日の日付を使用します。",
                                            level='warning'
                                        )
                                        lot_date = current_date
                                else:
                                    lot_date = current_date
                            except Exception:
                                # 変換に失敗した場合は今日の日付を使用
                                lot_date = current_date
                
                for i in range(1, 6):
                    if i not in inspector_col_indices:
                        continue
                    inspector_col_idx = inspector_col_indices[i]
                    inspector_value = row_tuple[inspector_col_idx] if inspector_col_idx < len(row_tuple) else None
                    
                    if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                        inspector_name = str(inspector_value).strip()
                        if '(' in inspector_name:
                            inspector_name = inspector_name.split('(')[0].strip()
                        if not inspector_name:
                            continue
                        
                        inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                        if not inspector_info.empty:
                            inspector_code = inspector_info.iloc[0]['#ID']
                            
                            if inspector_code not in self.inspector_daily_assignments:
                                self.inspector_daily_assignments[inspector_code] = {}
                            if lot_date not in self.inspector_daily_assignments[inspector_code]:
                                self.inspector_daily_assignments[inspector_code][lot_date] = 0.0
                            if inspector_code not in self.inspector_work_hours:
                                self.inspector_work_hours[inspector_code] = 0.0
                            if inspector_code not in self.inspector_product_hours:
                                self.inspector_product_hours[inspector_code] = {}
                            if product_number not in self.inspector_product_hours[inspector_code]:
                                self.inspector_product_hours[inspector_code][product_number] = 0.0
                            
                            self.inspector_daily_assignments[inspector_code][lot_date] += divided_time
                            self.inspector_work_hours[inspector_code] += divided_time
                            self.inspector_product_hours[inspector_code][product_number] += divided_time
            
            # 【改善】分割検査時間を再計算（実際の検査員数に基づいて）
            inspection_time_col_idx_recalc = result_df.columns.get_loc('検査時間') if '検査時間' in result_df.columns else -1
            inspector_col_indices_recalc = {}
            for i in range(1, 6):
                col_name = f'検査員{i}'
                if col_name in result_df.columns:
                    inspector_col_indices_recalc[i] = result_df.columns.get_loc(col_name)
            
            for row_tuple in result_df.itertuples(index=True):
                index = row_tuple[0]  # インデックス
                inspection_time = row_tuple[inspection_time_col_idx_recalc + 1] if inspection_time_col_idx_recalc >= 0 and inspection_time_col_idx_recalc + 1 < len(row_tuple) else 0.0
                
                if inspection_time == 0 or pd.isna(inspection_time):
                    continue
                
                # 実際に割り当てられた検査員数をカウント
                actual_inspector_count = 0
                for i in range(1, 6):
                    if i not in inspector_col_indices_recalc:
                        continue
                    inspector_col_idx = inspector_col_indices_recalc[i]
                    inspector_value = row_tuple[inspector_col_idx + 1] if inspector_col_idx + 1 < len(row_tuple) else None
                    
                    if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                        actual_inspector_count += 1
                
                # 分割検査時間を再計算: 検査時間 ÷ 実際の分割した検査人数
                if actual_inspector_count > 0:
                    actual_divided_time = inspection_time / actual_inspector_count
                    result_df.at[index, '分割検査時間'] = round(actual_divided_time, 1)
                else:
                    # 検査員が割り当てられていない場合は0
                    result_df.at[index, '分割検査時間'] = 0.0
            
            # 勤務時間超過を再チェック
            phase3_5_violations = []
            # 列インデックスを事前に取得（itertuples()で高速化）
            prod_num_col_idx_v = result_df.columns.get_loc('品番')
            div_time_col_idx_v = result_df.columns.get_loc('分割検査時間') if '分割検査時間' in result_df.columns else -1
            shipping_date_col_idx_v = result_df.columns.get_loc('出荷予定日') if '出荷予定日' in result_df.columns else -1
            process_name_col_idx_v = result_df.columns.get_loc('現在工程名') if '現在工程名' in result_df.columns else -1
            inspector_col_indices_v = {}
            for i in range(1, 6):
                col_name = f'検査員{i}'
                if col_name in result_df.columns:
                    inspector_col_indices_v[i] = result_df.columns.get_loc(col_name)
            
            for row_tuple in result_df.itertuples(index=True):
                index = row_tuple[0]  # インデックス
                product_number = row_tuple[prod_num_col_idx_v + 1]  # +1はインデックス分
                divided_time = row_tuple[div_time_col_idx_v + 1] if div_time_col_idx_v >= 0 and div_time_col_idx_v + 1 < len(row_tuple) else 0.0
                process_name_context = row_tuple[process_name_col_idx_v + 1] if process_name_col_idx_v >= 0 and process_name_col_idx_v + 1 < len(row_tuple) else ''
                
                for i in range(1, 6):
                    if i not in inspector_col_indices_v:
                        continue
                    inspector_col_idx = inspector_col_indices_v[i]
                    inspector_value = row_tuple[inspector_col_idx + 1] if inspector_col_idx + 1 < len(row_tuple) else None
                    
                    if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                        inspector_name = str(inspector_value).strip()
                        if '(' in inspector_name:
                            inspector_name = inspector_name.split('(')[0].strip()
                        if not inspector_name:
                            continue

                        # 固定検査員が割当済みのロットは、勤務時間/同一品番上限の違反チェック対象から除外する
                        if self._is_fixed_inspector_for_lot(product_number, str(process_name_context).strip(), inspector_name):
                            # 【追加】保護の履歴追跡
                            self.fixed_inspector_protection_metrics['total_protections'] += 1
                            self.fixed_inspector_protection_metrics['protection_by_phase']['phase3_violation_check'] = \
                                self.fixed_inspector_protection_metrics['protection_by_phase'].get('phase3_violation_check', 0) + 1
                            self.fixed_inspector_protection_metrics['protection_by_reason']['violation_check_exclusion'] = \
                                self.fixed_inspector_protection_metrics['protection_by_reason'].get('violation_check_exclusion', 0) + 1
                            self.fixed_inspector_protection_metrics['protected_lots'].add(index)
                            self.fixed_inspector_protection_metrics['protected_inspectors'].add(inspector_name)
                            
                            protection_history_entry = {
                                'lot_index': index,
                                'product_number': product_number,
                                'inspector_name': inspector_name,
                                'phase': 'phase3',
                                'reason': 'violation_check_exclusion',
                                'process_name': str(process_name_context).strip(),
                            }
                            self.fixed_inspector_protection_metrics['protection_history'].append(protection_history_entry)
                            if len(self.fixed_inspector_protection_metrics['protection_history']) > 100:
                                self.fixed_inspector_protection_metrics['protection_history'] = \
                                    self.fixed_inspector_protection_metrics['protection_history'][-100:]
                            continue
                        
                        inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                        if not inspector_info.empty:
                            inspector_code = inspector_info.iloc[0]['#ID']
                            
                            # 出荷予定日から日付を取得（当日洗浄上がり品の場合は今日の日付を使用）
                            lot_date_for_check = current_date  # デフォルトは今日の日付
                            if shipping_date_col_idx_v >= 0 and shipping_date_col_idx_v + 1 < len(row_tuple):
                                shipping_date_raw = row_tuple[shipping_date_col_idx_v + 1]
                                if pd.notna(shipping_date_raw):
                                    shipping_date_str = str(shipping_date_raw).strip()
                                    # 当日洗浄上がり品の場合は今日の日付を使用
                                    if self._is_same_day_cleaning(shipping_date_str):
                                        lot_date_for_check = current_date
                                    else:
                                        # 日付型に変換を試みる
                                        try:
                                            shipping_date_dt = pd.to_datetime(shipping_date_raw, errors='coerce')
                                            if pd.notna(shipping_date_dt):
                                                # 異常な日付（例: 1677年など）をチェック
                                                if shipping_date_dt.year >= 1900 and shipping_date_dt.year <= 2100:
                                                    lot_date_for_check = shipping_date_dt.date()
                                                else:
                                                    # 異常な日付の場合は今日の日付を使用
                                                    lot_date_for_check = current_date
                                            else:
                                                lot_date_for_check = current_date
                                        except Exception:
                                            # 変換に失敗した場合は今日の日付を使用
                                            lot_date_for_check = current_date
                            
                            daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(lot_date_for_check, 0.0)
                            max_hours = inspector_max_hours.get(inspector_code, 8.0)
                            # 10%超過まで許容
                            allowed_max_hours = max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                            product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                            
                            # 勤務時間超過をチェック（10%超過まで許容）
                            if daily_hours > allowed_max_hours - WORK_HOURS_BUFFER:
                                phase3_5_violations.append((index, inspector_code, inspector_name, "勤務時間超過", daily_hours, allowed_max_hours))
                                self.log_message(f"❌ フェーズ3.5検証: 勤務時間超過が検出されました - 検査員 '{inspector_name}' {daily_hours:.1f}h > {allowed_max_hours:.1f}h (ロット {index})")
                            
                            # 改善ポイント: 最適化フェーズでの設定時間上限チェック（厳格）
                            if product_hours > self.product_limit_hard_threshold:
                                phase3_5_violations.append((index, inspector_code, inspector_name, f"同一品番{self.product_limit_hard_threshold:.1f}時間超過", product_hours, self.product_limit_hard_threshold))
                                self.log_message(f"❌ フェーズ3.5検証: 同一品番{self.product_limit_hard_threshold:.1f}時間超過が検出されました - 検査員 '{inspector_name}' 品番 {product_number} {product_hours:.1f}h > {self.product_limit_hard_threshold:.1f}h (ロット {index})")
            
            # 【改善】当日洗浄上がり品の品番単位・品名単位の制約違反をチェック
            # 固定検査員が絡む違反は「許容違反(固定)」として別カテゴリに落とし、❌は本当に止めるべき異常に限定する
            same_day_cleaning_violations = []
            tolerated_same_day_cleaning_violations = []
            # 列インデックスを事前に取得（itertuples()で高速化）
            prod_num_col_idx_vc2 = result_df.columns.get_loc('品番')
            shipping_date_col_idx_vc2 = result_df.columns.get_loc('出荷予定日') if '出荷予定日' in result_df.columns else -1
            product_name_col_idx_vc2 = result_df.columns.get_loc('品名') if '品名' in result_df.columns else -1
            process_name_col_idx_vc2 = result_df.columns.get_loc('現在工程名') if '現在工程名' in result_df.columns else -1
            inspector_col_indices_vc2 = {}
            for i in range(1, 6):
                col_name = f'検査員{i}'
                if col_name in result_df.columns:
                    inspector_col_indices_vc2[i] = result_df.columns.get_loc(col_name)
            
            # 旧実装は row×row の二重ループになりがちで重いため、ここでは線形に集計して同じ判定を行う
            same_day_lot_by_index: Dict[int, Dict[str, Any]] = {}
            occurrences_by_product_code: Dict[Tuple[str, str], List[int]] = {}
            occurrences_by_name_code: Dict[Tuple[str, str], List[int]] = {}
            tolerated_keys: Set[Tuple[int, str, str]] = set()
            violation_keys: Set[Tuple[int, str, str]] = set()

            for row_tuple in result_df.itertuples(index=True):
                index = int(row_tuple[0])  # インデックス
                product_number = row_tuple[prod_num_col_idx_vc2 + 1]  # +1はインデックス分
                product_number_str = str(product_number).strip()
                shipping_date_raw = row_tuple[shipping_date_col_idx_vc2 + 1] if shipping_date_col_idx_vc2 >= 0 and shipping_date_col_idx_vc2 + 1 < len(row_tuple) else None
                shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                process_name_context = row_tuple[process_name_col_idx_vc2 + 1] if process_name_col_idx_vc2 >= 0 and process_name_col_idx_vc2 + 1 < len(row_tuple) else ''
                process_name_context_str = str(process_name_context).strip()

                is_same_day_cleaning = (
                    shipping_date_str == "当日洗浄上がり品"
                    or shipping_date_str == "当日洗浄品"
                    or "当日洗浄" in shipping_date_str
                    or shipping_date_str == "先行検査"
                    or shipping_date_str == "当日先行検査"
                )
                if not is_same_day_cleaning or not product_number_str:
                    continue

                product_name = row_tuple[product_name_col_idx_vc2 + 1] if product_name_col_idx_vc2 >= 0 and product_name_col_idx_vc2 + 1 < len(row_tuple) else ''
                product_name_str = str(product_name).strip() if pd.notna(product_name) else ''

                assigned_codes_in_lot: Set[str] = set()
                assigned_names_in_lot: Dict[str, str] = {}
                for i in range(1, 6):
                    if i not in inspector_col_indices_vc2:
                        continue
                    inspector_col_idx = inspector_col_indices_vc2[i]
                    inspector_value = row_tuple[inspector_col_idx + 1] if inspector_col_idx + 1 < len(row_tuple) else None
                    if pd.isna(inspector_value) or str(inspector_value).strip() == '':
                        continue
                    inspector_name = str(inspector_value).strip()
                    if '(' in inspector_name:
                        inspector_name = inspector_name.split('(')[0].strip()
                    if not inspector_name:
                        continue
                    inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                    if inspector_info.empty:
                        continue
                    inspector_code = str(inspector_info.iloc[0]['#ID']).strip()
                    if not inspector_code:
                        continue
                    assigned_codes_in_lot.add(inspector_code)
                    assigned_names_in_lot[inspector_code] = inspector_name

                    occurrences_by_product_code.setdefault((product_number_str, inspector_code), []).append(index)
                    if product_name_str:
                        occurrences_by_name_code.setdefault((product_name_str, inspector_code), []).append(index)

                lot_info = {
                    'index': index,
                    'product_number': product_number_str,
                    'product_name': product_name_str,
                    'process_name_context': process_name_context_str,
                    'assigned_codes': assigned_codes_in_lot,
                    'assigned_names': assigned_names_in_lot,
                }
                same_day_lot_by_index[index] = lot_info

            # 品番単位の制約違反（同一品番の複数ロットに同一検査員）
            for (product_number_str, inspector_code), indices in occurrences_by_product_code.items():
                if len(indices) < 2:
                    continue
                indices_sorted = sorted(set(indices))
                if len(indices_sorted) < 2:
                    continue
                min_idx = indices_sorted[0]
                second_idx = indices_sorted[1]
                for index in indices_sorted:
                    other_index = min_idx if index != min_idx else second_idx
                    lot_info = same_day_lot_by_index.get(index)
                    other_lot_info = same_day_lot_by_index.get(other_index)
                    if not lot_info or not other_lot_info:
                        continue
                    inspector_name = lot_info['assigned_names'].get(inspector_code, '')
                    if not inspector_name:
                        continue
                    is_fixed_current = self._is_fixed_inspector_for_lot(
                        lot_info['product_number'], lot_info['process_name_context'], inspector_name
                    )
                    is_fixed_other = self._is_fixed_inspector_for_lot(
                        other_lot_info['product_number'], other_lot_info['process_name_context'], inspector_name
                    )
                    violation_key = (index, inspector_code, inspector_name)
                    if is_fixed_current or is_fixed_other:
                        if violation_key not in tolerated_keys:
                            tolerated_keys.add(violation_key)
                            tolerated_same_day_cleaning_violations.append(
                                (index, inspector_code, inspector_name, "同一品番複数ロット同一検査員", product_number_str, other_index)
                            )
                            self.log_message(
                                f"⚠️ 許容違反(固定): フェーズ3.5検証: 当日洗浄上がり品の品番単位制約違反 - "
                                f"品番 {product_number_str} のロット {index} と {other_index} に同一検査員 '{inspector_name}' (コード: {inspector_code}) が割り当てられています（固定検査員維持）",
                                level='warning'
                            )
                    else:
                        if violation_key not in violation_keys:
                            violation_keys.add(violation_key)
                            same_day_cleaning_violations.append(
                                (index, inspector_code, inspector_name, "同一品番複数ロット同一検査員", product_number_str)
                            )
                            self.log_message(
                                f"❌ フェーズ3.5検証: 当日洗浄上がり品の品番単位制約違反が検出されました - "
                                f"品番 {product_number_str} のロット {index} と {other_index} に同一検査員 '{inspector_name}' (コード: {inspector_code}) が割り当てられています",
                                level='error'
                            )

            # 品名単位の制約違反（同一品名の異品番ロットに同一検査員）
            for (product_name_str, inspector_code), indices in occurrences_by_name_code.items():
                if not product_name_str or len(indices) < 2:
                    continue
                indices_sorted = sorted(set(indices))
                if len(indices_sorted) < 2:
                    continue

                # 品番ごとの最小indexを使い、「別品番で最小のother_index」を決める（旧実装の探索順と一致）
                first_index_per_product: Dict[str, int] = {}
                for idx in indices_sorted:
                    lot = same_day_lot_by_index.get(idx)
                    if not lot:
                        continue
                    pnum = lot['product_number']
                    first_index_per_product.setdefault(pnum, idx)

                for index in indices_sorted:
                    lot_info = same_day_lot_by_index.get(index)
                    if not lot_info:
                        continue
                    current_product = lot_info['product_number']
                    other_candidates = [v for k, v in first_index_per_product.items() if k != current_product]
                    if not other_candidates:
                        continue
                    other_index = min(other_candidates)
                    other_lot_info = same_day_lot_by_index.get(other_index)
                    if not other_lot_info:
                        continue
                    inspector_name = lot_info['assigned_names'].get(inspector_code, '')
                    if not inspector_name:
                        continue

                    is_fixed_current = self._is_fixed_inspector_for_lot(
                        lot_info['product_number'], lot_info['process_name_context'], inspector_name
                    )
                    is_fixed_other = self._is_fixed_inspector_for_lot(
                        other_lot_info['product_number'], other_lot_info['process_name_context'], inspector_name
                    )
                    violation_key = (index, inspector_code, inspector_name)
                    if is_fixed_current or is_fixed_other:
                        if violation_key not in tolerated_keys:
                            tolerated_keys.add(violation_key)
                            tolerated_same_day_cleaning_violations.append(
                                (
                                    index,
                                    inspector_code,
                                    inspector_name,
                                    "同一品名異品番同一検査員",
                                    current_product,
                                    other_lot_info['product_number'],
                                    other_index,
                                )
                            )
                            self.log_message(
                                f"⚠️ 許容違反(固定): フェーズ3.5検証: 当日洗浄上がり品の品名単位制約違反 - "
                                f"品名 '{product_name_str}' の品番 {current_product} (ロット {index}) と {other_lot_info['product_number']} (ロット {other_index}) に同一検査員 '{inspector_name}' (コード: {inspector_code}) が割り当てられています（固定検査員維持）",
                                level='warning'
                            )
                    else:
                        if violation_key not in violation_keys:
                            violation_keys.add(violation_key)
                            same_day_cleaning_violations.append(
                                (index, inspector_code, inspector_name, "同一品名異品番同一検査員", current_product, other_lot_info['product_number'])
                            )
                            self.log_message(
                                f"❌ フェーズ3.5検証: 当日洗浄上がり品の品名単位制約違反が検出されました - "
                                f"品名 '{product_name_str}' の品番 {current_product} (ロット {index}) と {other_lot_info['product_number']} (ロット {other_index}) に同一検査員 '{inspector_name}' (コード: {inspector_code}) が割り当てられています",
                                level='error'
                            )
            
            if tolerated_same_day_cleaning_violations:
                self.log_message(
                    f"⚠️ 許容違反(固定): フェーズ3.5検証で当日洗浄上がり品の制約違反が {len(tolerated_same_day_cleaning_violations)}件検出されました（固定検査員維持のため）",
                    level='warning'
                )

            if same_day_cleaning_violations:
                self.log_message(f"⚠️ 警告: フェーズ3.5検証で当日洗浄上がり品の制約違反が {len(same_day_cleaning_violations)}件検出されました", level='warning')
                
                # 【改善】違反を解消するための再割り当てを試行
                resolved_count = 0
                current_date = pd.Timestamp.now().date()
                
                for violation in same_day_cleaning_violations:
                    violation_index = violation[0]
                    violation_code = violation[1]
                    violation_name = violation[2]
                    violation_type = violation[3]
                    product_number = violation[4] if len(violation) > 4 else ''
                    
                    if violation_index >= len(result_df):
                        continue
                    
                    row = result_df.iloc[violation_index]
                    lot_process_name = str(row.get('現在工程名', '') or '').strip()
                    fixed_inspector_names = self._collect_fixed_inspector_names(product_number, lot_process_name)
                    
                    # 違反している検査員を特定
                    violating_inspector_col = None
                    for j in range(1, 6):
                        inspector_col = f'検査員{j}'
                        inspector_value = row.get(inspector_col, '')
                        if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                            inspector_name = str(inspector_value).strip()
                            if '(' in inspector_name:
                                inspector_name = inspector_name.split('(')[0].strip()
                            if inspector_name == violation_name:
                                violating_inspector_col = inspector_col
                                break
                    
                    if violating_inspector_col:
                        # 固定検査員は最適化/是正で外さない
                        if fixed_inspector_names and str(violation_name).strip() in fixed_inspector_names:
                            self.log_message(
                                f"固定検査員のため保護: 当日洗浄上がり品の制約違反是正で検査員を変更しません "
                                f"(ロット {violation_index}, 品番: {product_number}, 検査員: {violation_name})",
                                level='warning'
                            )
                            continue

                        # 代替検査員を探す
                        process_number = row.get('現在工程番号', None)
                        shipping_date = row.get('出荷予定日', None)
                        inspection_time = row.get('検査時間', 0.0)
                        divided_time = row.get('分割検査時間', inspection_time)
                        
                        # 既に割り当てられている検査員を取得
                        assigned_codes = set()
                        for j in range(1, 6):
                            col = f'検査員{j}'
                            val = row.get(col, '')
                            if pd.notna(val) and str(val).strip() != '':
                                name = str(val).strip()
                                if '(' in name:
                                    name = name.split('(')[0].strip()
                                info = self._get_inspector_by_name(name, inspector_master_df)
                                if not info.empty:
                                    assigned_codes.add(info.iloc[0]['#ID'])
                        
                        # 品番単位と品名単位の制約を取得
                        already_assigned_to_product = self.same_day_cleaning_inspectors.get(product_number, set())
                        product_name = row.get('品名', '')
                        product_name_str = str(product_name).strip() if pd.notna(product_name) else ''
                        already_assigned_to_product_name = set()
                        if product_name_str:
                            already_assigned_to_product_name = self.same_day_cleaning_inspectors_by_product_name.get(product_name_str, set())
                        
                        # 除外すべき検査員コード
                        excluded_codes = assigned_codes | already_assigned_to_product | already_assigned_to_product_name
                        
                        # 代替検査員を取得（新製品チームも含める）
                        # skill_product_valuesが定義されているか確認
                        try:
                            skill_product_values_local = skill_product_values
                        except NameError:
                            skill_product_values_local = set()
                            if skill_master_df is not None and not skill_master_df.empty:
                                skill_product_values_local = set(skill_master_df.iloc[:, 0].dropna().tolist())
                        
                        is_new_product = product_number not in skill_product_values_local
                        available_inspectors = self.get_available_inspectors(
                            product_number, process_number, skill_master_df, inspector_master_df,
                            shipping_date=shipping_date, allow_new_team_fallback=True,
                            process_master_df=process_master_df,
                            process_name_context=lot_process_name
                        )
                        
                        # 新製品チームも候補に追加
                        if not available_inspectors and is_new_product:
                            available_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                        
                        # 除外条件を満たす検査員を探す
                        replacement_found = False
                        inspector_max_hours = {}
                        for code in inspector_master_df['#ID'].unique():
                            base_hours = self.get_inspector_max_hours(code, inspector_master_df)
                            inspector_max_hours[code] = self._apply_work_hours_overrun(base_hours)
                        
                        # 第一段階: 制約を厳格に守る検査員を探す
                        for insp in available_inspectors:
                            code = insp['コード']
                            if code in excluded_codes:
                                continue
                            if code == violation_code:
                                continue
                            
                            # 勤務時間チェック（10%超過を厳格にチェック）
                            max_hours = inspector_max_hours.get(code, 8.0)
                            allowed_max_hours = max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
                            daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                            if daily_hours + divided_time > allowed_max_hours - WORK_HOURS_BUFFER:
                                continue
                            
                            # 4時間上限チェック
                            product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                            if product_hours + divided_time > self.product_limit_hard_threshold:
                                continue
                            
                            # 代替検査員が見つかった
                            replacement_found = True
                            skill_value = insp.get('スキル値', '')
                            display_name = f"{insp['氏名']}({skill_value})" if skill_value else insp['氏名']
                            result_df.at[violation_index, violating_inspector_col] = display_name
                            
                            # 履歴を更新（旧検査員から時間を引く）
                            if violation_code in self.inspector_daily_assignments:
                                if current_date in self.inspector_daily_assignments[violation_code]:
                                    self.inspector_daily_assignments[violation_code][current_date] = max(0.0, 
                                        self.inspector_daily_assignments[violation_code][current_date] - divided_time)
                            if violation_code in self.inspector_work_hours:
                                self.inspector_work_hours[violation_code] = max(0.0, 
                                    self.inspector_work_hours[violation_code] - divided_time)
                            if violation_code in self.inspector_product_hours:
                                if product_number in self.inspector_product_hours[violation_code]:
                                    self.inspector_product_hours[violation_code][product_number] = max(0.0,
                                        self.inspector_product_hours[violation_code][product_number] - divided_time)
                            
                            # 履歴を更新（新検査員に時間を追加）
                            if code not in self.inspector_daily_assignments:
                                self.inspector_daily_assignments[code] = {}
                            if current_date not in self.inspector_daily_assignments[code]:
                                self.inspector_daily_assignments[code][current_date] = 0.0
                            self.inspector_daily_assignments[code][current_date] += divided_time
                            
                            if code not in self.inspector_work_hours:
                                self.inspector_work_hours[code] = 0.0
                            self.inspector_work_hours[code] += divided_time
                            
                            if code not in self.inspector_product_hours:
                                self.inspector_product_hours[code] = {}
                            if product_number not in self.inspector_product_hours[code]:
                                self.inspector_product_hours[code][product_number] = 0.0
                            self.inspector_product_hours[code][product_number] += divided_time
                            
                            # 品番単位・品名単位の追跡を更新
                            if product_number in self.same_day_cleaning_inspectors:
                                old_inspectors_product = self.same_day_cleaning_inspectors[product_number].copy()
                                self.same_day_cleaning_inspectors[product_number].discard(violation_code)
                                self.same_day_cleaning_inspectors[product_number].add(code)
                                self.log_message(
                                    f"フェーズ3.5検証: 当日洗浄上がり品の制約更新（品番単位）: "
                                    f"品番 '{product_number}' から検査員 '{violation_name}' ({violation_code}) を削除、"
                                    f"検査員 '{insp['氏名']}' ({code}) を追加",
                                    debug=True,
                                )
                            else:
                                self.same_day_cleaning_inspectors[product_number] = {code}
                                self.log_message(
                                    f"フェーズ3.5検証: 当日洗浄上がり品の制約更新（品番単位）: "
                                    f"品番 '{product_number}' に検査員 '{insp['氏名']}' ({code}) を追加（新規）",
                                    debug=True,
                                )
                            
                            if product_name_str:
                                if product_name_str not in self.same_day_cleaning_inspectors_by_product_name:
                                    self.same_day_cleaning_inspectors_by_product_name[product_name_str] = set()
                                old_inspectors_name = self.same_day_cleaning_inspectors_by_product_name[product_name_str].copy()
                                self.same_day_cleaning_inspectors_by_product_name[product_name_str].discard(violation_code)
                                self.same_day_cleaning_inspectors_by_product_name[product_name_str].add(code)
                                self.log_message(
                                    f"フェーズ3.5検証: 当日洗浄上がり品の制約更新（品名単位）: "
                                    f"品名 '{product_name_str}' から検査員 '{violation_name}' ({violation_code}) を削除、"
                                    f"検査員 '{insp['氏名']}' ({code}) を追加",
                                    debug=True,
                                )
                            
                            self.log_message(
                                f"✅ 当日洗浄上がり品の制約違反を解消: ロット {violation_index} (品番: {product_number}, 品名: {product_name_str if product_name_str else 'N/A'}) "
                                f"の検査員 '{violation_name}' を '{insp['氏名']}' に変更しました（違反タイプ: {violation_type}）",
                                level='info'
                            )
                            resolved_count += 1
                            break
                        
                        # 第二段階: 制約を緩和して検査員を探す（第一段階で見つからない場合）
                        if not replacement_found:
                            self.log_message(
                                f"フェーズ3.5検証: 制約を厳格に守る検査員が見つからないため、制約を緩和して再検索します - "
                                f"ロット {violation_index} (品番: {product_number}, 違反タイプ: {violation_type})",
                                debug=True,
                            )
                            
                            # 品番単位の制約を緩和（品名単位の制約は維持）
                            relaxed_excluded_codes = assigned_codes | already_assigned_to_product_name
                            
                            for insp in available_inspectors:
                                code = insp['コード']
                                if code in relaxed_excluded_codes:
                                    continue
                                if code == violation_code:
                                    continue
                                
                                # 勤務時間チェック（緩和モード: 20%超過まで許容）
                                max_hours = inspector_max_hours.get(code, 8.0)
                                allowed_max_hours = max_hours * (1.0 + SAME_DAY_WORK_HOURS_OVERRUN_RATE)
                                daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                                if daily_hours + divided_time > allowed_max_hours - WORK_HOURS_BUFFER:
                                    continue
                                
                                # 4時間上限チェック（緩和: 4.2hまで許容）
                                product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                                if product_hours + divided_time > PRODUCT_LIMIT_FINAL_TOLERANCE:
                                    continue
                                
                                # 代替検査員が見つかった（緩和モード）
                                replacement_found = True
                                skill_value = insp.get('スキル値', '')
                                display_name = f"{insp['氏名']}({skill_value})" if skill_value else insp['氏名']
                                result_df.at[violation_index, violating_inspector_col] = display_name
                                
                                # 履歴を更新（旧検査員から時間を引く）
                                if violation_code in self.inspector_daily_assignments:
                                    if current_date in self.inspector_daily_assignments[violation_code]:
                                        self.inspector_daily_assignments[violation_code][current_date] = max(0.0, 
                                            self.inspector_daily_assignments[violation_code][current_date] - divided_time)
                                if violation_code in self.inspector_work_hours:
                                    self.inspector_work_hours[violation_code] = max(0.0, 
                                        self.inspector_work_hours[violation_code] - divided_time)
                                if violation_code in self.inspector_product_hours:
                                    if product_number in self.inspector_product_hours[violation_code]:
                                        self.inspector_product_hours[violation_code][product_number] = max(0.0,
                                            self.inspector_product_hours[violation_code][product_number] - divided_time)
                                
                                # 履歴を更新（新検査員に時間を追加）
                                if code not in self.inspector_daily_assignments:
                                    self.inspector_daily_assignments[code] = {}
                                if current_date not in self.inspector_daily_assignments[code]:
                                    self.inspector_daily_assignments[code][current_date] = 0.0
                                self.inspector_daily_assignments[code][current_date] += divided_time
                                
                                if code not in self.inspector_work_hours:
                                    self.inspector_work_hours[code] = 0.0
                                self.inspector_work_hours[code] += divided_time
                                
                                if code not in self.inspector_product_hours:
                                    self.inspector_product_hours[code] = {}
                                if product_number not in self.inspector_product_hours[code]:
                                    self.inspector_product_hours[code][product_number] = 0.0
                                self.inspector_product_hours[code][product_number] += divided_time
                                
                                # 品番単位・品名単位の追跡を更新（品番単位の制約は緩和されているため、品名単位のみ更新）
                                if product_name_str:
                                    if product_name_str not in self.same_day_cleaning_inspectors_by_product_name:
                                        self.same_day_cleaning_inspectors_by_product_name[product_name_str] = set()
                                    old_inspectors_name = self.same_day_cleaning_inspectors_by_product_name[product_name_str].copy()
                                    self.same_day_cleaning_inspectors_by_product_name[product_name_str].discard(violation_code)
                                    self.same_day_cleaning_inspectors_by_product_name[product_name_str].add(code)
                                    self.log_message(
                                        f"フェーズ3.5検証: 当日洗浄上がり品の制約更新（品名単位、緩和モード）: "
                                        f"品名 '{product_name_str}' から検査員 '{violation_name}' ({violation_code}) を削除、"
                                        f"検査員 '{insp['氏名']}' ({code}) を追加",
                                        debug=True,
                                    )
                                
                                self.log_message(
                                    f"✅ 当日洗浄上がり品の制約違反を解消（緩和モード）: ロット {violation_index} (品番: {product_number}, 品名: {product_name_str if product_name_str else 'N/A'}) "
                                    f"の検査員 '{violation_name}' を '{insp['氏名']}' に変更しました（違反タイプ: {violation_type}）",
                                level='info'
                            )
                            resolved_count += 1
                            break
                        
                        if not replacement_found:
                            # 代替検査員が見つからない場合は警告のみを出力して割り当てを維持
                            self.log_message(
                                f"⚠️ 警告: 当日洗浄上がり品の制約違反が検出されましたが、代替検査員が見つかりませんでした - "
                                f"ロット {violation_index} (品番: {product_number}, 違反タイプ: {violation_type})。最優先のため割り当てを維持します。",
                                level='warning'
                            )
                
                if resolved_count > 0:
                    self.log_message(f"✅ 当日洗浄上がり品の制約違反を {resolved_count}件解消しました", level='info')
            
            if phase3_5_violations:
                self.log_message(f"⚠️ 警告: フェーズ3.5検証で {len(phase3_5_violations)}件の違反が検出されました", level='warning')
                
                # 違反を品番ごとにグループ化
                violations_by_product = {}
                for violation in phase3_5_violations:
                    index = violation[0]
                    row = result_df.iloc[index]
                    # 当日洗浄上がり品（＋先行検査）は未割当を避けるため、
                    # 同一品番4.0h上限（product limit）違反は最終是正の対象から外す
                    violation_type = str(violation[3]) if len(violation) > 3 else ''
                    shipping_date_for_filter = row.get('出荷予定日', None)
                    if self._should_force_keep_assignment(shipping_date_for_filter) and "同一品番" in violation_type:
                        continue
                    product_number = row.get('品番', '')
                    if product_number not in violations_by_product:
                        violations_by_product[product_number] = []
                    violations_by_product[product_number].append((violation, row))
                
                # 同じ品番で複数の違反がある場合、まとめて再割当を試みる
                resolved_count = 0
                processed_indices = set()
                
                for product_number, product_violations in violations_by_product.items():
                    if len(product_violations) > 1:
                        # 同じ品番で複数の違反がある場合、まとめて処理
                        self.log_message(f"🔵 品番 {product_number} で {len(product_violations)}件の違反が検出されました。まとめて再割当を試みます")
                        
                        # 違反ロットを一度クリア（ただし、2週間以内の新規品は保護のためスキップ）
                        violation_indices = []
                        violation_lots = []
                        protected_indices = set()  # 保護されたロットのインデックス
                        for violation, row in product_violations:
                            index = violation[0]
                            shipping_date = row.get('出荷予定日', pd.Timestamp.max)
                            # 最優先ロットは未割当を避けるため、割当クリアを行わず保護する
                            if self._should_force_keep_assignment(shipping_date):
                                protected_indices.add(index)
                                continue
                            
                            # 3営業日以内のロットかどうかを判定
                            def add_business_days_local(start: date, business_days: int) -> date:
                                def next_business_day(date_val: date) -> date:
                                    weekday = date_val.weekday()
                                    if weekday == 4:  # Friday
                                        return date_val + timedelta(days=3)
                                    if weekday == 5:  # Saturday
                                        return date_val + timedelta(days=2)
                                    return date_val + timedelta(days=1)
                                current = start
                                added = 0
                                while added < business_days:
                                    current = next_business_day(current)
                                    added += 1
                                return current
                            
                            is_within_three_business_days = False
                            is_within_two_weeks = False
                            shipping_date_date = None
                            
                            if pd.notna(shipping_date):
                                try:
                                    if isinstance(shipping_date, pd.Timestamp):
                                        shipping_date_date = shipping_date.date()
                                    elif isinstance(shipping_date, str):
                                        shipping_date_date = pd.to_datetime(shipping_date, errors='coerce').date()
                                        if pd.isna(shipping_date_date):
                                            shipping_date_date = None
                                    else:
                                        shipping_date_date = shipping_date.date() if hasattr(shipping_date, 'date') else shipping_date
                                    
                                    if shipping_date_date is not None:
                                        three_business_days_ahead = add_business_days_local(current_date, 3)
                                        two_weeks_later = current_date + timedelta(days=14)
                                        is_within_three_business_days = shipping_date_date <= three_business_days_ahead
                                        is_within_two_weeks = shipping_date_date <= two_weeks_later
                                except Exception as e:
                                    self.log_message(f"出荷予定日の比較エラー: {str(e)} (ロットインデックス: {index})", level='warning')
                            
                            # 2週間以内の新規品かどうかを判定
                            skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                            is_new_product = skill_rows.empty
                            
                            # 3営業日以内または2週間以内のロットは保護（未割当にしない）
                            # 未割当ロット再処理と同様に、「近日（2週間以内）」のロットも保護する
                            if is_within_three_business_days or is_within_two_weeks:
                                inspector_code = violation[1]
                                violation_type = violation[3]
                                
                                # 違反内容に応じて保護処理
                                if violation_type == "同一品番4時間超過":
                                    current_product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                                    if is_within_three_business_days:
                                        self.log_message(f"⚠️ 品番 {product_number} (出荷予定日: {shipping_date_date}) は3営業日以内のため保護します - 同一品番4時間超過を許容（{current_product_hours:.1f}h）", level='warning')
                                    else:
                                        self.log_message(f"⚠️ 品番 {product_number} (出荷予定日: {shipping_date_date}) は2週間以内のため保護します - 同一品番4時間超過を許容（{current_product_hours:.1f}h）", level='warning')
                                    self.relaxed_product_limit_assignments.add((inspector_code, product_number))
                                    protected_indices.add(index)
                                    resolved_count += 1
                                    continue
                                elif violation_type == "勤務時間超過":
                                    daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
                                    max_hours = inspector_max_hours.get(inspector_code, 8.0)
                                    if is_within_three_business_days:
                                        self.log_message(f"⚠️ 品番 {product_number} (出荷予定日: {shipping_date_date}) は3営業日以内のため保護します - 勤務時間超過を許容（{daily_hours:.1f}h > {max_hours:.1f}h）", level='warning')
                                    else:
                                        self.log_message(f"⚠️ 品番 {product_number} (出荷予定日: {shipping_date_date}) は2週間以内のため保護します - 勤務時間超過を許容（{daily_hours:.1f}h > {max_hours:.1f}h）", level='warning')
                                    protected_indices.add(index)
                                    resolved_count += 1
                                    continue
                            
                            # 2週間以内の新規品は保護（未割当にしない）
                            if is_new_product and is_within_two_weeks:
                                inspector_code = violation[1]
                                violation_type = violation[3]
                                
                                # 違反内容に応じて保護処理
                                if violation_type == "同一品番4時間超過":
                                    current_product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                                    self.log_message(f"⚠️ 新規品 {product_number} (出荷予定日: {shipping_date}) は2週間以内のため保護します - 同一品番4時間超過を許容（{current_product_hours:.1f}h）", level='warning')
                                    self.relaxed_product_limit_assignments.add((inspector_code, product_number))
                                    protected_indices.add(index)
                                    resolved_count += 1
                                    continue
                                elif violation_type == "勤務時間超過":
                                    daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
                                    max_hours = inspector_max_hours.get(inspector_code, 8.0)
                                    self.log_message(f"⚠️ 新規品 {product_number} (出荷予定日: {shipping_date}) は2週間以内のため保護します - 勤務時間超過を許容（{daily_hours:.1f}h > {max_hours:.1f}h）", level='warning')
                                    protected_indices.add(index)
                                    resolved_count += 1
                                    continue
                            
                            # 保護対象でない場合は、通常通りクリア
                            violation_indices.append(index)
                            normalized_shipping_date = self._normalize_shipping_date(shipping_date)
                            violation_lots.append({
                                'index': index,
                                'violation': violation,
                                'row': row,
                                'inspection_time': row.get('検査時間', 0),
                                'shipping_date': shipping_date,
                                'normalized_shipping_date': normalized_shipping_date
                            })
                            self.clear_assignment(result_df, index)
                            processed_indices.add(index)
                        
                        # 履歴を再計算（クリアしたロットと保護されたロットを除外）
                        self.inspector_daily_assignments = {}
                        self.inspector_work_hours = {}
                        self.inspector_product_hours = {}
                        
                        # 列インデックスを事前に取得（itertuples()で高速化）
                        prod_num_col_idx_p3_5 = result_df.columns.get_loc('品番')
                        div_time_col_idx_p3_5 = result_df.columns.get_loc('分割検査時間') if '分割検査時間' in result_df.columns else -1
                        inspector_col_indices_p3_5 = {}
                        for i in range(1, 6):
                            col_name = f'検査員{i}'
                            if col_name in result_df.columns:
                                inspector_col_indices_p3_5[i] = result_df.columns.get_loc(col_name)
                        
                        for row_tuple in result_df.itertuples(index=True):
                            idx = row_tuple[0]  # インデックス
                            if idx in violation_indices or idx in protected_indices:
                                continue  # クリアしたロットと保護されたロットはスキップ
                            
                            prod_num = row_tuple[prod_num_col_idx_p3_5 + 1]  # +1はインデックス分
                            div_time = row_tuple[div_time_col_idx_p3_5 + 1] if div_time_col_idx_p3_5 >= 0 and div_time_col_idx_p3_5 + 1 < len(row_tuple) else 0.0
                            
                            for i in range(1, 6):
                                if i not in inspector_col_indices_p3_5:
                                    continue
                                inspector_col_idx = inspector_col_indices_p3_5[i]
                                inspector_value = row_tuple[inspector_col_idx + 1] if inspector_col_idx + 1 < len(row_tuple) else None
                                
                                if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                                    inspector_name = str(inspector_value).strip()
                                    if '(' in inspector_name:
                                        inspector_name = inspector_name.split('(')[0].strip()
                                    if not inspector_name:
                                        continue
                                    
                                    inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                    if not inspector_info.empty:
                                        inspector_code = inspector_info.iloc[0]['#ID']
                                        
                                        if inspector_code not in self.inspector_daily_assignments:
                                            self.inspector_daily_assignments[inspector_code] = {}
                                        if current_date not in self.inspector_daily_assignments[inspector_code]:
                                            self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                                        if inspector_code not in self.inspector_work_hours:
                                            self.inspector_work_hours[inspector_code] = 0.0
                                        if inspector_code not in self.inspector_product_hours:
                                            self.inspector_product_hours[inspector_code] = {}
                                        if prod_num not in self.inspector_product_hours[inspector_code]:
                                            self.inspector_product_hours[inspector_code][prod_num] = 0.0
                                        
                                        self.inspector_daily_assignments[inspector_code][current_date] += div_time
                                        self.inspector_work_hours[inspector_code] += div_time
                                        self.inspector_product_hours[inspector_code][prod_num] += div_time
                        
                        # 出荷予定日順にソートして再割当
                        violation_lots.sort(key=lambda x: x['normalized_shipping_date'])
                        
                        # 各ロットを再割当
                        for lot_info in violation_lots:
                            index = lot_info['index']
                            inspection_time = lot_info['inspection_time']
                            shipping_date = lot_info['shipping_date']
                            
                            self.log_message(f"⚠️ ロットインデックス {index} (品番: {product_number}, 出荷予定日: {shipping_date}) を再割当します")
                            
                            # スキルマスタから利用可能な検査員を取得
                            process_number = lot_info['row'].get('現在工程番号', '')
                            lot_process_name = str(lot_info['row'].get('現在工程名', '') or '').strip()
                            skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                            is_new_product = skill_rows.empty
                            available_inspectors = self.get_available_inspectors(
                                product_number, process_number, skill_master_df, inspector_master_df,
                                shipping_date=shipping_date, allow_new_team_fallback=is_new_product,
                                process_master_df=process_master_df, inspection_target_keywords=inspection_target_keywords,
                                process_name_context=lot_process_name
                            )
                            
                            if not available_inspectors and is_new_product:
                                available_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                            
                            if available_inspectors:
                                # 再割当を試みる
                                if inspection_time <= self.required_inspectors_threshold:
                                    required_inspectors = 1
                                else:
                                    calculated_inspectors = max(2, int(inspection_time / self.required_inspectors_threshold) + 1)
                                    # 5名以上になる場合は5名に制限（特例）
                                    required_inspectors = min(5, calculated_inspectors)
                                divided_time = inspection_time / required_inspectors
                                
                                # 利用可能な検査員から選択
                                assigned_inspectors = self.select_inspectors(
                                    available_inspectors, required_inspectors, divided_time,
                                    inspector_master_df, product_number,
                                    is_new_product=skill_rows.empty,
                                    ignore_product_limit=self._should_force_assign_same_day(shipping_date),
                                )
                                
                                if assigned_inspectors:
                                    # 割り当てを設定
                                    result_df.at[index, '検査員人数'] = len(assigned_inspectors)
                                    # 分割検査時間の計算: 検査時間 ÷ 実際の分割した検査人数
                                    actual_divided_time = inspection_time / len(assigned_inspectors)
                                    result_df.at[index, '分割検査時間'] = round(actual_divided_time, 1)
                                    for i, inspector in enumerate(assigned_inspectors, 1):
                                        if i <= 5:
                                            if show_skill_values:
                                                if inspector.get('is_new_team', False):
                                                    inspector_name = f"{inspector['氏名']}(新)"
                                                else:
                                                    inspector_name = f"{inspector['氏名']}({inspector['スキル']})"
                                            else:
                                                if inspector.get('is_new_team', False):
                                                    inspector_name = f"{inspector['氏名']}(新)"
                                                else:
                                                    inspector_name = inspector['氏名']
                                            result_df.at[index, f'検査員{i}'] = inspector_name
                                    
                                    result_df.at[index, 'assignability_status'] = 'assigned'
                                    resolved_count += 1
                                    self.log_message(f"✅ ロットインデックス {index} の再割当に成功しました")
                                else:
                                    self.log_message(f"⚠️ ロットインデックス {index} の再割当に失敗しました（利用可能な検査員が見つかりません）")
                            else:
                                    self.log_message(f"⚠️ ロットインデックス {index} の再割当に失敗しました（利用可能な検査員が0人）")
                        
                        # 履歴を再計算（再割当後の状態）
                        self.inspector_daily_assignments = {}
                        self.inspector_work_hours = {}
                        self.inspector_product_hours = {}
                        
                        # 列インデックスを事前に取得（itertuples()で高速化）
                        prod_num_col_idx = result_df.columns.get_loc('品番')
                        div_time_col_idx = result_df.columns.get_loc('分割検査時間') if '分割検査時間' in result_df.columns else -1
                        inspector_col_indices = {}
                        for i in range(1, 6):
                            col_name = f'検査員{i}'
                            if col_name in result_df.columns:
                                inspector_col_indices[i] = result_df.columns.get_loc(col_name)
                        
                        for row_tuple in result_df.itertuples(index=False):
                            prod_num = row_tuple[prod_num_col_idx]
                            div_time = row_tuple[div_time_col_idx] if div_time_col_idx >= 0 and div_time_col_idx < len(row_tuple) else 0.0
                            
                            for i in range(1, 6):
                                if i not in inspector_col_indices:
                                    continue
                                inspector_col_idx = inspector_col_indices[i]
                                inspector_value = row_tuple[inspector_col_idx] if inspector_col_idx < len(row_tuple) else None
                                
                                if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                                    inspector_name = str(inspector_value).strip()
                                    if '(' in inspector_name:
                                        inspector_name = inspector_name.split('(')[0].strip()
                                    if not inspector_name:
                                        continue
                                    
                                    inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                                    if not inspector_info.empty:
                                        inspector_code = inspector_info.iloc[0]['#ID']
                                        
                                        if inspector_code not in self.inspector_daily_assignments:
                                            self.inspector_daily_assignments[inspector_code] = {}
                                        if current_date not in self.inspector_daily_assignments[inspector_code]:
                                            self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                                        if inspector_code not in self.inspector_work_hours:
                                            self.inspector_work_hours[inspector_code] = 0.0
                                        if inspector_code not in self.inspector_product_hours:
                                            self.inspector_product_hours[inspector_code] = {}
                                        if prod_num not in self.inspector_product_hours[inspector_code]:
                                            self.inspector_product_hours[inspector_code][prod_num] = 0.0
                                        
                                        self.inspector_daily_assignments[inspector_code][current_date] += div_time
                                        self.inspector_work_hours[inspector_code] += div_time
                                        self.inspector_product_hours[inspector_code][prod_num] += div_time
                
                # 残りの違反（単独または処理済み以外）を個別に処理
                violations_with_date = []
                for violation in phase3_5_violations:
                    index = violation[0]
                    if index not in processed_indices:
                        row = result_df.iloc[index]
                        shipping_date_raw = row.get('出荷予定日', pd.Timestamp.max)
                        
                        # 【修正】出荷予定日を正規化（文字列の場合はpd.Timestamp.minに変換してソート可能にする）
                        if pd.notna(shipping_date_raw):
                            shipping_date_str = str(shipping_date_raw).strip()
                            # 当日洗浄上がり品などの文字列の場合は、最優先としてpd.Timestamp.minに変換
                            if (shipping_date_str == "当日洗浄上がり品" or
                                shipping_date_str == "当日洗浄品" or
                                "当日洗浄" in shipping_date_str or
                                shipping_date_str == "先行検査" or
                                shipping_date_str == "当日先行検査"):
                                shipping_date = pd.Timestamp.min  # 最優先として扱う
                            elif isinstance(shipping_date_raw, pd.Timestamp):
                                shipping_date = shipping_date_raw
                            elif isinstance(shipping_date_raw, str):
                                # 日付文字列の場合は変換を試みる
                                try:
                                    shipping_date = pd.to_datetime(shipping_date_raw)
                                except Exception as e:
                                    logger.debug(f"出荷日の変換でエラーが発生しました（デフォルト値を使用）: {e}")
                                    shipping_date = pd.Timestamp.min
                            else:
                                shipping_date = shipping_date_raw
                        else:
                            shipping_date = pd.Timestamp.max
                        
                        violations_with_date.append((violation, shipping_date))
                
                # 出荷予定日の古い順にソート
                violations_with_date.sort(key=lambda x: self._normalize_shipping_date(x[1]))
                
                # 出荷予定日が古いロットから順に再割り当てを試みる
                for violation, shipping_date in violations_with_date:
                    index = violation[0]
                    row = result_df.iloc[index]
                    product_number = row.get('品番', '')
                    inspection_time = row.get('検査時間', 0)
                    
                    self.log_message(f"⚠️ ロットインデックス {index} (品番: {product_number}, 出荷予定日: {shipping_date}) の違反を是正します（{violation[3]}）")
                    
                    # 当日洗浄上がり品かどうかを判定
                    shipping_date_str = str(shipping_date).strip() if pd.notna(shipping_date) else ''
                    is_same_day_cleaning = (
                        shipping_date_str == "当日洗浄上がり品" or 
                        shipping_date_str == "当日洗浄品" or
                        "当日洗浄" in shipping_date_str or
                        shipping_date_str == "先行検査" or
                        shipping_date_str == "当日先行検査"
                    )
                    
                    # 違反を是正する処理を試みる
                    if self._is_locked_fixed_preinspection_lot(result_df, index):
                        continue
                    violation_resolved = self.fix_single_violation(
                        index, violation[1], violation[2], 
                        row.get('分割検査時間', 0.0), product_number, inspection_time,
                        None, result_df, inspector_master_df, skill_master_df,
                        inspector_max_hours, current_date, show_skill_values
                    )
                    
                    if violation_resolved:
                        resolved_count += 1
                        self.log_message(f"✅ ロットインデックス {index} の違反を是正しました")
                    else:
                        # 当日洗浄上がり品の場合は割り当てを維持（優先順位が高いため保護）
                        if is_same_day_cleaning:
                            self.log_message(
                                f"当日洗浄上がり品のため、ルール違反があっても割り当てを維持します（品番: {product_number}, 出荷予定日: {shipping_date_str}）",
                                level='warning',
                            )
                            # 割り当てを維持（未割当にしない）
                            resolved_count += 1
                        else:
                            # 是正できなかった場合は未割当にする
                            self.clear_assignment(result_df, index)
                            self.log_message(f"未割当: ロットインデックス {index}（{violation[3]}）", level='warning')
                
                self.log_message(f"フェーズ3.5違反是正結果: {resolved_count}件是正、{len(phase3_5_violations) - resolved_count}件未割当")
                
                # 未割当後の履歴を再計算
                self.inspector_daily_assignments = {}
                self.inspector_work_hours = {}
                self.inspector_product_hours = {}
                
                # 列インデックスを事前に取得（itertuples()で高速化）
                prod_num_col_idx_u2 = result_df.columns.get_loc('品番')
                div_time_col_idx_u2 = result_df.columns.get_loc('分割検査時間') if '分割検査時間' in result_df.columns else -1
                inspector_col_indices_u2 = {}
                for i in range(1, 6):
                    col_name = f'検査員{i}'
                    if col_name in result_df.columns:
                        inspector_col_indices_u2[i] = result_df.columns.get_loc(col_name)
                
                for row_tuple in result_df.itertuples(index=True):
                    index = row_tuple[0]  # インデックス
                    product_number = row_tuple[prod_num_col_idx_u2 + 1]  # +1はインデックス分
                    divided_time = row_tuple[div_time_col_idx_u2 + 1] if div_time_col_idx_u2 >= 0 and div_time_col_idx_u2 + 1 < len(row_tuple) else 0.0
                    
                    for i in range(1, 6):
                        if i not in inspector_col_indices_u2:
                            continue
                        inspector_col_idx = inspector_col_indices_u2[i]
                        inspector_value = row_tuple[inspector_col_idx + 1] if inspector_col_idx + 1 < len(row_tuple) else None
                        
                        if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                            inspector_name = str(inspector_value).strip()
                            if '(' in inspector_name:
                                inspector_name = inspector_name.split('(')[0].strip()
                            if not inspector_name:
                                continue
                            
                            inspector_info = self._get_inspector_by_name(inspector_name, inspector_master_df)
                            if not inspector_info.empty:
                                inspector_code = inspector_info.iloc[0]['#ID']
                                
                                if inspector_code not in self.inspector_daily_assignments:
                                    self.inspector_daily_assignments[inspector_code] = {}
                                if current_date not in self.inspector_daily_assignments[inspector_code]:
                                    self.inspector_daily_assignments[inspector_code][current_date] = 0.0
                                if inspector_code not in self.inspector_work_hours:
                                    self.inspector_work_hours[inspector_code] = 0.0
                                if inspector_code not in self.inspector_product_hours:
                                    self.inspector_product_hours[inspector_code] = {}
                                if product_number not in self.inspector_product_hours[inspector_code]:
                                    self.inspector_product_hours[inspector_code][product_number] = 0.0
                                
                                self.inspector_daily_assignments[inspector_code][current_date] += divided_time
                                self.inspector_work_hours[inspector_code] += divided_time
                                self.inspector_product_hours[inspector_code][product_number] += divided_time
            else:
                self.log_message("フェーズ3.5検証: 全てのルール違反が解消されました")
            
            # 最適化後に全体のチーム情報を再計算（確実に一致させるため）
            perf_logger.debug(
                "PERF {}: {:.1f} ms",
                "inspector_assignment.optimize.phase3_5.total",
                (perf_counter() - _t_perf_phase3_5_total) * 1000.0,
            )

            self.log_message("全体最適化フェーズ4: チーム情報の再計算を開始")
            _t_perf_phase4_total = perf_counter()
            
            # 最終的に出荷予定日順にソート（最優先ルールの維持）
            # 出荷予定日を変換（当日洗浄品は文字列として保持）
            result_df['出荷予定日'] = result_df['出荷予定日'].apply(self._convert_shipping_date)
            
            # ソート用のキー関数: 新しい優先順位に従う
            current_date = pd.Timestamp.now().date()
            
            def get_next_business_day(date_val):
                """翌営業日を取得（金曜日の場合は翌週の月曜日）"""
                weekday = date_val.weekday()  # 0=月曜日, 4=金曜日
                if weekday == 4:  # 金曜日
                    return date_val + timedelta(days=3)  # 翌週の月曜日
                else:
                    return date_val + timedelta(days=1)  # 翌日
            
            next_business_day = get_next_business_day(current_date)
            
            # 3営業日先を計算
            def add_business_days_for_sort(start: date, days: int) -> date:
                result = start
                added = 0
                while added < days:
                    result += timedelta(days=1)
                    if result.weekday() < 5:  # 月〜金
                        added += 1
                return result
            three_business_days_ahead = add_business_days_for_sort(current_date, 3)
            
            def sort_key(val):
                if pd.isna(val):
                    return (5, None)  # 最後に
                val_str = str(val).strip()
                
                # 1. 当日の日付（優先度0）
                try:
                    date_val = pd.to_datetime(val, errors='coerce')
                    if pd.notna(date_val):
                        date_date = date_val.date()
                        if date_date == current_date:
                            return (0, date_val)
                except:
                    pass
                
                # 2. 当日洗浄上がり品（優先度1）
                if (val_str == "当日洗浄上がり品" or
                    val_str == "当日洗浄品" or
                    "当日洗浄" in val_str):
                    return (1, val_str)
                
                # 3. 先行検査品（優先度2）
                if (val_str == "先行検査" or
                    val_str == "当日先行検査"):
                    return (2, val_str)
                
                # 4. 今日より後〜3営業日以内（週末除外、優先度3）
                try:
                    date_val = pd.to_datetime(val, errors='coerce')
                    if pd.notna(date_val):
                        date_date = date_val.date()
                        if date_date > current_date and date_date <= three_business_days_ahead:
                            return (3, date_val)
                except:
                    pass
                
                # 5. それ以降の日付（優先度4、出荷予定日昇順でソート）
                try:
                    date_val = pd.to_datetime(val, errors='coerce')
                    if pd.notna(date_val):
                        return (4, date_val)
                except:
                    pass
                
                return (5, val_str)  # その他文字列
            
            # ソートキーを追加
            result_df['_sort_key'] = result_df['出荷予定日'].apply(sort_key)
            result_df = result_df.sort_values('_sort_key', na_position='last').reset_index(drop=True)
            result_df = result_df.drop(columns=['_sort_key'], errors='ignore')
            self.log_message("最終結果を出荷予定日の古い順でソートしました（最優先ルール）")
            
            # チーム情報を更新（未割当の理由を保持）
            # 列インデックスを事前に取得（itertuples()で高速化）
            inspector_count_col_idx_ti = result_df.columns.get_loc('検査員人数') if '検査員人数' in result_df.columns else -1
            team_info_col_idx = result_df.columns.get_loc('チーム情報') if 'チーム情報' in result_df.columns else -1
            status_col_idx = result_df.columns.get_loc('assignability_status') if 'assignability_status' in result_df.columns else -1
            prod_num_col_idx_ti = result_df.columns.get_loc('品番')
            inspection_time_col_idx_ti = result_df.columns.get_loc('検査時間') if '検査時間' in result_df.columns else -1
            lot_qty_col_idx_ti = result_df.columns.get_loc('ロット数量') if 'ロット数量' in result_df.columns else -1
            
            for row_tuple in result_df.itertuples(index=True):
                index = row_tuple[0]  # インデックス
                inspector_count = row_tuple[inspector_count_col_idx_ti + 1] if inspector_count_col_idx_ti >= 0 and inspector_count_col_idx_ti + 1 < len(row_tuple) else 0
                current_team_info = row_tuple[team_info_col_idx + 1] if team_info_col_idx >= 0 and team_info_col_idx + 1 < len(row_tuple) else ''
                
                # 検査員人数が0で、チーム情報が空または「未割当」のみの場合は詳細な理由を再設定
                if (inspector_count == 0 or pd.isna(inspector_count)) and (
                    pd.isna(current_team_info) or 
                    str(current_team_info).strip() == '' or 
                    str(current_team_info).strip() == '未割当'
                ):
                    # assignability_statusから理由を推測
                    status = row_tuple[status_col_idx + 1] if status_col_idx >= 0 and status_col_idx + 1 < len(row_tuple) else ''
                    product_number = row_tuple[prod_num_col_idx_ti + 1]  # +1はインデックス分
                    inspection_time = row_tuple[inspection_time_col_idx_ti + 1] if inspection_time_col_idx_ti >= 0 and inspection_time_col_idx_ti + 1 < len(row_tuple) else 0
                    lot_quantity = row_tuple[lot_qty_col_idx_ti + 1] if lot_qty_col_idx_ti >= 0 and lot_qty_col_idx_ti + 1 < len(row_tuple) else 0
                    
                    # rowオブジェクトが必要な場合は、元のDataFrameから取得
                    row = result_df.loc[index]
                    
                    if status == 'quantity_zero':
                        reason = "ロット数量0" if (lot_quantity == 0 or pd.isna(lot_quantity)) else "検査時間0"
                    elif status == 'capacity_shortage':
                        # スキルマスタに登録があるか確認
                        skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                        if skill_rows.empty:
                            reason = "スキルマスタ未登録"
                        else:
                            process_number = row.get('現在工程番号', '')
                            if process_number and str(process_number).strip() != '':
                                reason = f"工程番号'{process_number}'に一致するスキル情報なし"
                            else:
                                reason = "条件に合う検査員がいない"
                    elif status == 'logic_conflict':
                        reason = "勤務時間または4時間上限により全員除外"
                    elif status == 'final_product_limit_violation':
                        # 最終検証で違反検出された場合、元の理由を推測
                        # スキルマスタに登録があるか確認
                        skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                        if skill_rows.empty:
                            reason = "スキルマスタ未登録(最終検証で違反検出)"
                        else:
                            process_number = row.get('現在工程番号', '')
                            if process_number and str(process_number).strip() != '':
                                reason = f"工程番号'{process_number}'に一致するスキル情報なし(最終検証で違反検出)"
                            else:
                                reason = "条件に合う検査員がいない(最終検証で違反検出)"
                    else:
                        reason = "未割当理由不明"
                    
                    result_df.at[index, 'チーム情報'] = f'未割当({reason})'
                    self.log_message(f"チーム情報再設定: インデックス{index} (品番: {product_number}) → '未割当({reason})'")
                else:
                    # 通常のチーム情報更新
                    self.update_team_info(result_df, index, inspector_master_df, show_skill_values)

            # 3D025-G4960: 検査時間>=3.0h のロットは検査員2人割当を最終保証する
            # （最適化/是正の過程で1人/0人に戻るケースがあるため）
            try:
                target_product = "3D025-G4960"
                if {'品番', '検査時間', '検査員人数'}.issubset(result_df.columns):
                    product_mask = result_df['品番'].astype(str).str.strip() == target_product
                    inspection_times = pd.to_numeric(result_df['検査時間'], errors='coerce').fillna(0.0)
                    target_indices = result_df.index[product_mask].tolist()
                    enforced_count = 0
                    changed_count = 0
                    insufficient_candidates = 0

                    def _get_code(candidate: Dict[str, Any]) -> str:
                        return str(
                            candidate.get('コーチID', candidate.get('コーチE', candidate.get('コード', candidate.get('#ID', ''))))
                        ).strip()

                    current_date = pd.Timestamp.now().date()

                    def _cand_sort_key(candidate: Dict[str, Any]) -> tuple:
                        code = _get_code(candidate)
                        daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                        return (daily_hours, candidate.get('_fairness_score', 0), code)

                    def _code_key(code: str) -> Tuple[int, Any]:
                        try:
                            code_str = str(code).strip()
                        except Exception:
                            code_str = ''
                        try:
                            code_int = int(code_str) if code_str.isdigit() else None
                        except Exception:
                            code_int = None
                        return (0, code_int) if code_int is not None else (1, code_str)

                    def _normalize_name(value: Any) -> str:
                        if pd.isna(value):
                            return ''
                        name = str(value).strip()
                        if not name:
                            return ''
                        if '(' in name:
                            name = name.split('(')[0].strip()
                        return name

                    # 3D025-G4960: スキル保有者全員を候補として使用（工程番号に関係なく）
                    # スキルマスタから直接全員を取得して平準化に使用
                    candidate_by_name: Dict[str, Dict[str, Any]] = {}
                    
                    # スキルマスタから3D025-G4960のスキル保有者全員を取得（工程番号フィルタなし）
                    skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == target_product]
                    if not skill_rows.empty:
                        skill_columns = skill_master_df.columns[2:]
                        seen_codes = set()
                        
                        for _, skill_row in skill_rows.iterrows():
                            for col_name in skill_columns:
                                inspector_code = str(col_name).strip()
                                if not inspector_code or inspector_code in seen_codes:
                                    continue
                                skill_value = skill_row.get(col_name, None)
                                if pd.notna(skill_value) and str(skill_value).strip() in {'1', '2', '3'}:
                                    inspector_info = inspector_master_df[inspector_master_df['#ID'] == inspector_code]
                                    if inspector_info.empty:
                                        continue
                                    inspector_name = inspector_info.iloc[0]['#氏名']
                                    # 休暇中の検査員は除外
                                    if self.is_inspector_on_vacation(inspector_name):
                                        continue
                                    # 勤務時間が0以下の検査員は除外
                                    max_hours = self.get_inspector_max_hours(inspector_code, inspector_master_df)
                                    if max_hours <= 0:
                                        continue
                                    try:
                                        numeric_skill = int(str(skill_value).strip())
                                    except ValueError:
                                        numeric_skill = 1
                                    
                                    # 検査員マスタから詳細情報を取得
                                    inspector_row = inspector_info.iloc[0]
                                    candidate = {
                                        '氏名': inspector_name,
                                        'コード': inspector_code,
                                        'スキル': numeric_skill,
                                        'is_new_team': False,
                                        '開始時刻': inspector_row.get('開始時刻', ''),
                                        '終了時刻': inspector_row.get('終了時刻', ''),
                                        '残業可能時間': inspector_row.get('残業可能時間', 0.0),
                                    }
                                    name = _normalize_name(inspector_name)
                                    if name:
                                        existing = candidate_by_name.get(name)
                                        if existing is None:
                                            candidate_by_name[name] = candidate
                                        else:
                                            # 検査員コード昇順で代表を固定
                                            existing_code = _get_code(existing)
                                            new_code = _get_code(candidate)
                                            if _code_key(new_code) < _code_key(existing_code):
                                                candidate_by_name[name] = candidate
                                        seen_codes.add(inspector_code)
                    
                    # スキル保有者が見つからない場合は、従来の方法で候補を取得（フォールバック）
                    if not candidate_by_name:
                        self.log_message(f"⚠️ {target_product}: スキルマスタから候補が見つかりません。従来の方法で候補を取得します", level='warning')
                    for target_index in target_indices:
                        row = result_df.loc[target_index]
                        process_number = str(row.get('現在工程番号', '') or '').strip()
                        process_name_context = str(row.get('現在工程名', '') or '').strip()
                        shipping_date = row.get('出荷予定日', None)
                        candidates = self.get_available_inspectors(
                            target_product,
                            process_number,
                            skill_master_df,
                            inspector_master_df,
                            shipping_date=shipping_date,
                            allow_new_team_fallback=False,
                            process_master_df=process_master_df,
                            inspection_target_keywords=inspection_target_keywords,
                            process_name_context=process_name_context,
                        ) or []
                        for c in candidates:
                            name = _normalize_name(c.get('氏名', ''))
                            if not name:
                                continue
                            if self.is_inspector_on_vacation(name):
                                continue
                            existing = candidate_by_name.get(name)
                            if existing is None:
                                candidate_by_name[name] = c
                            else:
                                existing_code = _get_code(existing)
                                new_code = _get_code(c)
                                if _code_key(new_code) < _code_key(existing_code):
                                    candidate_by_name[name] = c

                    candidates_sorted = sorted(candidate_by_name.values(), key=_cand_sort_key)
                    candidate_names_ordered = [_normalize_name(c.get('氏名', '')) for c in candidates_sorted if _normalize_name(c.get('氏名', ''))]

                    if len(candidate_names_ordered) < 2:
                        insufficient_candidates = len(target_indices)
                    else:
                        # 3D025-G4960: スキル保有者全員で完全に均等に分散（既存割当を無視）
                        # 10ロット × 2人 = 20人の割当を、可能な限り均等に分散
                        usage_counts: Dict[str, int] = {name: 0 for name in candidate_names_ordered}
                        

                        def pick_min_usage(exclude: Set[str]) -> str:
                            """使用回数が最小の検査員を選択（同数の場合はランダム性を避けるためコード順、新規品チームは除外）"""
                            best_name = ''
                            best_count = None
                            best_code = None
                            for name in candidate_names_ordered:
                                if name in exclude:
                                    continue
                                # 候補者リストから候補を取得
                                candidate = next((c for c in candidates_sorted if _normalize_name(c.get('氏名', '')) == name), None)
                                if candidate is None:
                                    continue
                                
                                # 新規品チームは除外
                                if candidate.get('is_new_team', False):
                                    continue
                                
                                count = usage_counts.get(name, 0)
                                code = _get_code(candidate)
                                if best_count is None or count < best_count or (count == best_count and best_code and _code_key(code) < _code_key(best_code)):
                                    best_name = name
                                    best_count = count
                                    best_code = code
                            return best_name

                        # 既存の割当を無視して、完全に均等に再割当
                        for target_index in target_indices:
                            row = result_df.loc[target_index]
                            old_1 = _normalize_name(row.get('検査員1', ''))
                            old_2 = _normalize_name(row.get('検査員2', ''))

                            # 既存割当を無視して、使用回数が最小の検査員を選択
                            inspector1 = pick_min_usage(exclude=set())
                            if inspector1:
                                usage_counts[inspector1] = usage_counts.get(inspector1, 0) + 1
                            else:
                                insufficient_candidates += 1

                            inspector2 = pick_min_usage(exclude={inspector1} if inspector1 else set())
                            if inspector2:
                                usage_counts[inspector2] = usage_counts.get(inspector2, 0) + 1
                            else:
                                insufficient_candidates += 1

                            if (old_1 != inspector1) or (old_2 != inspector2):
                                changed_count += 1

                            # 結果反映（最大2人固定）
                            if '検査員1' in result_df.columns:
                                result_df.at[target_index, '検査員1'] = inspector1
                            if '検査員2' in result_df.columns:
                                result_df.at[target_index, '検査員2'] = inspector2
                            for i in range(3, 6):
                                col = f'検査員{i}'
                                if col in result_df.columns:
                                    result_df.at[target_index, col] = ''
                            result_df.at[target_index, '検査員人数'] = 2 if inspector1 and inspector2 else (1 if inspector1 else 0)

                            if '分割検査時間' in result_df.columns:
                                total_time = float(inspection_times.loc[target_index])
                                result_df.at[target_index, '分割検査時間'] = round(total_time / 2.0, 1)
                            if 'remaining_work_hours' in result_df.columns:
                                result_df.at[target_index, 'remaining_work_hours'] = 0.0 if (inspector1 and inspector2) else float(inspection_times.loc[target_index])
                            if 'assignability_status' in result_df.columns and (inspector1 and inspector2):
                                result_df.at[target_index, 'assignability_status'] = 'fully_assigned'

                            self.update_team_info(result_df, target_index, inspector_master_df, show_skill_values)
                            enforced_count += 1

                    if enforced_count:
                        msg = f"{target_product}: 2人割当を平準化しました (対象{enforced_count}件, 変更{changed_count}件, 候補{len(candidate_names_ordered)}名)"
                        if insufficient_candidates:
                            msg += f"（候補不足 {insufficient_candidates}件）"
                        self.log_message(msg)

                        # 平準化「後」の最終内訳をログに出す（途中の置き換えログとは別物）
                        try:
                            cols = [c for c in ['検査員1', '検査員2'] if c in result_df.columns]
                            if cols:
                                final_counts: Dict[str, int] = {}
                                pairs: List[str] = []
                                final_rows = result_df.loc[product_mask].copy()
                                # なるべく安定した順で（出荷予定日→生産ロットID）
                                for sort_col in ['出荷予定日', '生産ロットID']:
                                    if sort_col in final_rows.columns:
                                        final_rows[sort_col] = final_rows[sort_col].astype(str)
                                sort_keys = [c for c in ['出荷予定日', '生産ロットID'] if c in final_rows.columns]
                                if sort_keys:
                                    final_rows = final_rows.sort_values(sort_keys, na_position='last')

                                def _format_ship(v: Any) -> str:
                                    try:
                                        if pd.isna(v):
                                            return ''
                                    except Exception:
                                        pass
                                    s = str(v).strip()
                                    if not s or s.lower() == 'nat':
                                        return ''
                                    try:
                                        dt = pd.to_datetime(v, errors='coerce')
                                        if pd.notna(dt):
                                            return pd.Timestamp(dt).strftime('%Y/%m/%d')
                                    except Exception:
                                        pass
                                    return s

                                for idx_val, r in final_rows.iterrows():
                                    i1 = _normalize_name(r.get('検査員1', ''))
                                    i2 = _normalize_name(r.get('検査員2', ''))
                                    for name in (i1, i2):
                                        if name:
                                            final_counts[name] = final_counts.get(name, 0) + 1
                                    lot_id = str(r.get('生産ロットID', '') or '').strip()
                                    ship = _format_ship(r.get('出荷予定日', ''))
                                    if lot_id:
                                        key = lot_id
                                    else:
                                        key = f"idx{idx_val}"
                                        if ship:
                                            key = f"{key}({ship})"
                                    pairs.append(f"{key}:{i1}/{i2}")

                                used = sorted(final_counts.items(), key=lambda kv: (-kv[1], kv[0]))
                                if used:
                                    counts_compact = ", ".join([f"{n}={c}" for n, c in used])
                                    max_c = used[0][1]
                                    min_c = min(c for _, c in used)
                                    self.log_message(
                                        f"{target_product}: [FINAL] 割当内訳(検査員1+2) {counts_compact} (最大{max_c}, 最小{min_c}, 差{max_c - min_c})"
                                    )

                                # ロット別の最終割当（生産ロットIDが空でも idx/出荷予定日 をキーにして必ず全件出す）
                                if pairs and len(pairs) <= 30:
                                    self.log_message(
                                        f"{target_product}: [FINAL] ロット別({len(pairs)}件) { '; '.join(pairs) }"
                                    )
                        except Exception:
                            pass
            except Exception as e:
                self.log_message(f"3D025-G4960の2人割当保証中にエラーが発生しました: {e}", level='warning')

            self.log_message("チーム情報の再計算が完了しました")
            
            # 【改善】分割検査時間を最終的に再計算（チーム情報更新後、実際の検査員数に基づいて）
            # 列インデックスを事前に取得（itertuples()で高速化）
            inspection_time_col_idx_final = result_df.columns.get_loc('検査時間') if '検査時間' in result_df.columns else -1
            inspector_col_indices_final = {}
            for i in range(1, 6):
                col_name = f'検査員{i}'
                if col_name in result_df.columns:
                    inspector_col_indices_final[i] = result_df.columns.get_loc(col_name)
            
            for row_tuple in result_df.itertuples(index=True):
                index = row_tuple[0]  # インデックス
                inspection_time = row_tuple[inspection_time_col_idx_final + 1] if inspection_time_col_idx_final >= 0 and inspection_time_col_idx_final + 1 < len(row_tuple) else 0.0
                
                if inspection_time == 0 or pd.isna(inspection_time):
                    continue
                
                # 実際に割り当てられた検査員数をカウント
                actual_inspector_count = 0
                for i in range(1, 6):
                    if i not in inspector_col_indices_final:
                        continue
                    inspector_col_idx = inspector_col_indices_final[i]
                    inspector_value = row_tuple[inspector_col_idx + 1] if inspector_col_idx + 1 < len(row_tuple) else None
                    
                    if pd.notna(inspector_value) and str(inspector_value).strip() != '':
                        actual_inspector_count += 1
                
                # 分割検査時間を再計算: 検査時間 ÷ 実際の分割した検査人数
                if actual_inspector_count > 0:
                    actual_divided_time = inspection_time / actual_inspector_count
                    result_df.at[index, '分割検査時間'] = round(actual_divided_time, 1)
                else:
                    # 検査員が割り当てられていない場合は0
                    result_df.at[index, '分割検査時間'] = 0.0

            self.log_message("全体最適化が完了しました")
            # 列インデックスを事前に取得（itertuples()で高速化）
            inspector_count_col_idx_f = result_df.columns.get_loc('検査員人数') if '検査員人数' in result_df.columns else -1
            remaining_col_idx = result_df.columns.get_loc('remaining_work_hours') if 'remaining_work_hours' in result_df.columns else -1
            status_col_idx_f = result_df.columns.get_loc('assignability_status') if 'assignability_status' in result_df.columns else -1
            
            for row_tuple in result_df.itertuples(index=True):
                idx = row_tuple[0]  # インデックス
                inspector_count = row_tuple[inspector_count_col_idx_f + 1] if inspector_count_col_idx_f >= 0 and inspector_count_col_idx_f + 1 < len(row_tuple) else 0
                
                if inspector_count > 0:
                    remaining = row_tuple[remaining_col_idx + 1] if remaining_col_idx >= 0 and remaining_col_idx + 1 < len(row_tuple) else 0.0
                    if remaining is None or pd.isna(remaining):
                        remaining = 0.0
                    status = row_tuple[status_col_idx_f + 1] if status_col_idx_f >= 0 and status_col_idx_f + 1 < len(row_tuple) else ''
                    
                    if remaining <= 0.05 and status in {'capacity_shortage', 'capacity_shortage_partial', 'partial_assigned'}:
                        result_df.at[idx, 'assignability_status'] = 'fully_assigned'
            # 未割当カテゴリの可視化
            unresolved = result_df[(result_df['検査員人数'] == 0) | (result_df['remaining_work_hours'] > 0.05)]
            if not unresolved.empty:
                self.log_message("=== 未割当ロット内訳（最終） ===")
                status_counts = unresolved['assignability_status'].value_counts()
                for status, count in status_counts.items():
                    self.log_message(f"  - {status}: {count}件")
                    subset = unresolved[unresolved['assignability_status'] == status]
                    detail_rows = subset[['品番', '出荷予定日', 'remaining_work_hours', 'チーム情報']].head(5)
                    for idx, detail in detail_rows.iterrows():
                        remaining = detail.get('remaining_work_hours', 0.0) or 0.0
                        shipping_date = detail.get('出荷予定日', 'N/A')
                        info = detail.get('チーム情報', '')
                        self.log_message(
                            f"     ・{detail['品番']} ({shipping_date}) - 残時間 {remaining:.1f}h / {info}"
                        )
                self.log_message("==============================")
            else:
                self.log_message("未割当ロットはありません")
            self.relaxed_product_limit_assignments.clear()

            perf_logger.debug(
                "PERF {}: {:.1f} ms",
                "inspector_assignment.optimize.phase4.total",
                (perf_counter() - _t_perf_phase4_total) * 1000.0,
            )
             
            # 【高速化】ログバッファをフラッシュ
            if self.log_batch_enabled:
                self._flush_log_buffer()
            
            return result_df
            
        except Exception as e:
            # 【高速化】ログバッファをフラッシュ（エラー時も）
            if self.log_batch_enabled:
                self._flush_log_buffer()
            error_msg = f"全体最適化中にエラーが発生しました: {str(e)}"
            self.log_message(error_msg)
            logger.error(error_msg, exc_info=True)
            return result_df
    
    def fix_single_violation(
        self,
        index: int,
        inspector_code: str,
        inspector_name: str,
        divided_time: float,
        product_number: str,
        inspection_time: float,
        inspector_col_num: int,
        result_df: pd.DataFrame,
        inspector_master_df: pd.DataFrame,
        skill_master_df: pd.DataFrame,
        inspector_max_hours: Dict[str, float],
        current_date: date,
        show_skill_values: bool
    ) -> bool:
        """
        単一の違反（勤務時間超過または同一品番4時間超過）を是正
        
        Args:
            index: 行インデックス
            inspector_code: 検査員コード
            inspector_name: 検査員名
            divided_time: 分割検査時間
            product_number: 品番
            inspection_time: 検査時間
            inspector_col_num: 検査員列番号
            result_df: 結果DataFrame
            inspector_master_df: 検査員マスタ
            skill_master_df: スキルマスタ
            inspector_max_hours: 検査員ごとの最大勤務時間辞書
            current_date: 現在日付
            show_skill_values: スキル値を表示するか
        
        Returns:
            是正成功時はTrue、失敗時はFalse
        """
        try:
            # 違反件数をカウント
            self.violation_count += 1
            
            row = result_df.iloc[index]
            
            # 新規品かどうかを判定（スキルマスタに登録がない場合）
            skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
            is_new_product = skill_rows.empty
            
            # 出荷予定日を取得
            shipping_date = None
            is_within_two_weeks = False
            if '出荷予定日' in row.index:
                shipping_date = row.get('出荷予定日', None)
                if pd.notna(shipping_date):
                    shipping_date = pd.to_datetime(shipping_date, errors='coerce')
                    if pd.notna(shipping_date):
                        shipping_date_date = shipping_date.date()
                        # 本日から2週間以内の出荷予定日かどうかを判定
                        two_weeks_later = current_date + timedelta(days=14)
                        is_within_two_weeks = shipping_date_date <= two_weeks_later
            
            # 新規品で出荷予定日が2週間以内の場合は、代替が見つからない場合でも保護する
            protect_new_product = is_new_product and is_within_two_weeks
            
            # このロットの他の検査員を確認
            current_inspectors = []
            for i in range(1, 6):
                inspector_col = f'検査員{i}'
                if pd.notna(row.get(inspector_col)) and str(row[inspector_col]).strip() != '':
                    inspector_name_check = str(row[inspector_col]).strip()
                    if '(' in inspector_name_check:
                        inspector_name_check = inspector_name_check.split('(')[0].strip()
                    current_inspectors.append((i, inspector_name_check))
            
            # 超過している検査員を一時的に外す
            if len(current_inspectors) > 1:
                # 複数人で分担している場合は、超過している検査員を外して別の人に置き換え
                removed_inspector = None
                for i, name in current_inspectors:
                    if name == inspector_name:
                        removed_inspector = (i, name)
                        break
                
                if removed_inspector:
                    # 代替検査員を探す
                    process_number = row.get('現在工程番号', '')
                    lot_process_name = str(row.get('現在工程名', '') or '').strip()
                    # スキルマスタに登録があるか確認
                    skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                    is_new_product = skill_rows.empty
                    shipping_date = row.get('出荷予定日', None)
                    
                    # 【改善】3営業日以内のロットについては、スキルマッチングを緩和（新製品チームを追加）
                    is_within_three_business_days = False
                    if shipping_date and pd.notna(shipping_date):
                        try:
                            shipping_date_parsed = pd.to_datetime(shipping_date, errors='coerce')
                            if pd.notna(shipping_date_parsed):
                                shipping_date_date = shipping_date_parsed.date()
                                today = pd.Timestamp.now().date()
                                def add_business_days(start: date, business_days: int) -> date:
                                    result_date = start
                                    added = 0
                                    while added < business_days:
                                        result_date += timedelta(days=1)
                                        if result_date.weekday() < 5:
                                            added += 1
                                    return result_date
                                three_business_days_ahead = add_business_days(today, 3)
                                is_within_three_business_days = shipping_date_date <= three_business_days_ahead
                        except Exception:
                            pass
                    
                    # 3営業日以内のロットについては、新製品チームを追加候補にする
                    allow_new_team_fallback = is_new_product or is_within_three_business_days
                    
                    # デバッグログ: 3営業日以内の判定結果を出力
                    if shipping_date and pd.notna(shipping_date):
                        self.log_message(f"fix_single_violation: 品番 {product_number}, 出荷予定日 {shipping_date}, 3営業日以内: {is_within_three_business_days}, 新規品: {is_new_product}", debug=True)
                    
                    available_inspectors = self.get_available_inspectors(
                        product_number, process_number, skill_master_df, inspector_master_df,
                        shipping_date=shipping_date, allow_new_team_fallback=allow_new_team_fallback,
                        process_name_context=lot_process_name
                    )
                    # 新規品または3営業日以内のロットで利用可能な検査員が見つからない場合は新製品チームも取得
                    if not available_inspectors and (is_new_product or is_within_three_business_days):
                        self.log_message(f"品番 {product_number} (新規品: {is_new_product}, 3営業日以内: {is_within_three_business_days}): 新製品チームを取得します")
                        available_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                        if available_inspectors:
                            self.log_message(f"品番 {product_number}: 新製品チームから {len(available_inspectors)}人の検査員を取得しました")
                        else:
                            self.log_message(f"品番 {product_number}: 新製品チームからも検査員を取得できませんでした", level='warning')
                    
                    # 既に割り当てられている検査員を除外
                    current_codes = []
                    for _, name in current_inspectors:
                        if name != inspector_name:
                            info = inspector_master_df[inspector_master_df['#氏名'] == name]
                            if not info.empty:
                                current_codes.append(info.iloc[0]['#ID'])
                    
                    # 既に割り当てられている人以外で、勤務時間に余裕がある人を探す
                    replacement_candidates = []
                    excluded_reasons = {}  # 除外理由を記録
                    
                    # 当日洗浄上がり品かどうかを判定
                    shipping_date_raw = row.get('出荷予定日', None)
                    shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                    is_same_day_cleaning_lot = (
                        shipping_date_str == "当日洗浄上がり品" or
                        shipping_date_str == "当日洗浄品" or
                        "当日洗浄" in shipping_date_str
                    )
                    
                    # 当日洗浄上がり品の場合、既にこの品番に割り当てられた検査員を取得
                    force_assign_same_day = self._should_force_assign_same_day(shipping_date_raw)
                    already_assigned_to_this_product = set()
                    if is_same_day_cleaning_lot and (not force_assign_same_day):
                        already_assigned_to_this_product = self.same_day_cleaning_inspectors.get(product_number, set())
                    
                    # 【改善】段階的な制約緩和レベルを定義
                    relaxation_levels = [
                        {
                            'name': 'strict',
                            'same_day_cleaning_product_constraint': True,
                            'work_hours_buffer': 0.0,
                            'product_limit': self.product_limit_hard_threshold,
                        },
                        {
                            'name': 'relaxed_product_limit',
                            'same_day_cleaning_product_constraint': True,
                            'work_hours_buffer': 0.0,
                            'product_limit': PRODUCT_LIMIT_FINAL_TOLERANCE,
                        },
                        {
                            'name': 'relaxed_work_hours',
                            'same_day_cleaning_product_constraint': True,
                            'work_hours_buffer': 0.5,
                            'product_limit': PRODUCT_LIMIT_FINAL_TOLERANCE,
                        },
                        {
                            'name': 'relaxed_same_day_constraint',
                            'same_day_cleaning_product_constraint': False,
                            'work_hours_buffer': 0.5,
                            'product_limit': PRODUCT_LIMIT_FINAL_TOLERANCE,
                        },
                    ]
                    
                    # 各緩和レベルで候補を検索
                    selected_relaxation_level = None
                    for level in relaxation_levels:
                        level_candidates = []
                        level_excluded = {}
                        
                        for insp in available_inspectors:
                            if insp['コード'] not in current_codes:
                                code = insp['コード']
                                insp_name = insp['氏名']
                                
                                # 当日洗浄上がり品の品番単位制約チェック（緩和レベルに応じて）
                                if (level['same_day_cleaning_product_constraint'] and 
                                    is_same_day_cleaning_lot and 
                                    code in already_assigned_to_this_product):
                                    level_excluded[insp_name] = f"既にこの品番に割り当て済み（品番単位の制約）"
                                    continue
                                
                                max_hours = inspector_max_hours.get(code, 8.0)
                                daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                                
                                # 勤務時間チェック（緩和レベルに応じて）
                                if daily_hours + divided_time > max_hours - level['work_hours_buffer']:
                                    level_excluded[insp_name] = f"勤務時間超過 ({daily_hours:.1f}h + {divided_time:.1f}h > {max_hours:.1f}h - {level['work_hours_buffer']:.1f}h緩和)"
                                    continue
                                
                                # 同一品番4時間上限チェック（緩和レベルに応じて）
                                current_product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                                if current_product_hours + divided_time > level['product_limit']:
                                    level_excluded[insp_name] = f"同一品番超過 ({current_product_hours:.1f}h + {divided_time:.1f}h > {level['product_limit']:.1f}h)"
                                    continue
                                
                                total_hours = self.inspector_work_hours.get(code, 0.0)
                                level_candidates.append((total_hours, insp))
                        
                        if level_candidates:
                            replacement_candidates = level_candidates
                            excluded_reasons = level_excluded
                            selected_relaxation_level = level['name']
                            if level['name'] != 'strict':
                                self.log_message(f"制約緩和レベル '{level['name']}' で {len(level_candidates)}人の候補が見つかりました (品番: {product_number})", level='info')
                            break
                        else:
                            excluded_reasons.update(level_excluded)
                    
                    # 候補が見つからない場合、制約違反の程度で優先順位付けを試行
                    if not replacement_candidates:
                        scored_candidates = []
                        
                        for insp in available_inspectors:
                            if insp['コード'] not in current_codes:
                                code = insp['コード']
                                insp_name = insp['氏名']
                                
                                score = 0
                                violations = []
                                
                                # 各制約違反にペナルティを付与
                                if is_same_day_cleaning_lot and code in already_assigned_to_this_product:
                                    score += 100  # 高ペナルティ
                                    violations.append("品番単位制約")
                                
                                max_hours = inspector_max_hours.get(code, 8.0)
                                daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                                excess_work_hours = max(0, (daily_hours + divided_time) - max_hours)
                                if excess_work_hours > 0:
                                    score += excess_work_hours * 10
                                    violations.append(f"勤務時間超過{excess_work_hours:.1f}h")
                                
                                current_product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                                excess_product_hours = max(0, (current_product_hours + divided_time) - self.product_limit_hard_threshold)
                                if excess_product_hours > 0:
                                    score += excess_product_hours * 5
                                    violations.append(f"同一品番超過{excess_product_hours:.1f}h")
                                
                                scored_candidates.append((score, insp, violations))
                        
                        # スコアが低い順（違反が少ない順）にソート
                        scored_candidates.sort(key=self._priority_sort_key)
                        
                        # 最も違反が少ない候補を選択（許容可能な範囲内）
                        if scored_candidates:
                            best_score = scored_candidates[0][0]
                            # 当日洗浄上がり品の場合は閾値を緩和（100まで許容）
                            threshold = 100 if is_same_day_cleaning_lot else 50
                            if best_score < threshold:
                                best_candidate = scored_candidates[0]
                                self.log_message(
                                    f"制約を一部緩和して割り当て: '{best_candidate[1]['氏名']}' "
                                    f"(違反: {', '.join(best_candidate[2])}, 品番: {product_number}, スコア: {best_score:.1f})",
                                    level='warning',
                                )
                                replacement_candidates = [(0, best_candidate[1])]
                                selected_relaxation_level = 'scored_relaxation'
                    
                    if not replacement_candidates:
                        # 代替検査員が見つからない理由を詳細にログ出力（統計情報を追加）
                        total_candidates = len(available_inspectors)
                        excluded_by_reason = {}
                        for name, reason in excluded_reasons.items():
                            reason_type = reason.split(':')[0] if ':' in reason else reason.split('(')[0].strip()
                            if reason_type not in excluded_by_reason:
                                excluded_by_reason[reason_type] = 0
                            excluded_by_reason[reason_type] += 1
                        
                        self.log_message(f"代替検査員が見つかりません (品番: {product_number}, ロットインデックス: {index})", level='warning')
                        self.log_message(f"   候補検査員総数: {total_candidates}人, 除外された検査員: {len(excluded_reasons)}人", level='warning')
                        if excluded_by_reason:
                            self.log_message(f"   除外理由別統計: {excluded_by_reason}", level='warning')
                        if excluded_reasons:
                            for name, reason in list(excluded_reasons.items())[:5]:
                                self.log_message(f"     - {name}: {reason}", level='warning')
                            if len(excluded_reasons) > 5:
                                self.log_message(f"     ... 他{len(excluded_reasons) - 5}人", level='warning')
                        else:
                            self.log_message(f"   理由: 候補検査員が全て既に割り当て済み", level='warning')
                    
                    if replacement_candidates:
                        # 最も総勤務時間が少ない人を選択
                        replacement_candidates.sort(key=self._priority_sort_key)
                        replacement_inspector = replacement_candidates[0][1]
                        
                        # 置き換えを実行
                        i_col, _ = removed_inspector
                        if show_skill_values:
                            if replacement_inspector.get('is_new_team', False):
                                replacement_name = f"{replacement_inspector['氏名']}(新)"
                            else:
                                replacement_name = f"{replacement_inspector['氏名']}({replacement_inspector['スキル']})"
                        else:
                            if replacement_inspector.get('is_new_team', False):
                                replacement_name = f"{replacement_inspector['氏名']}(新)"
                            else:
                                replacement_name = replacement_inspector['氏名']
                        
                        result_df.at[index, f'検査員{i_col}'] = replacement_name
                        
                        # 履歴を更新（元の検査員から時間を引く）
                        old_code = inspector_code
                        old_daily = self.inspector_daily_assignments.get(old_code, {}).get(current_date, 0.0)
                        old_total = self.inspector_work_hours.get(old_code, 0.0)
                        self.inspector_daily_assignments[old_code][current_date] = max(0.0, old_daily - divided_time)
                        self.inspector_work_hours[old_code] = max(0.0, old_total - divided_time)
                        
                        # 品番別累計時間も更新
                        if old_code in self.inspector_product_hours:
                            if product_number in self.inspector_product_hours[old_code]:
                                self.inspector_product_hours[old_code][product_number] = max(0.0, self.inspector_product_hours[old_code][product_number] - divided_time)
                        self.relaxed_product_limit_assignments.discard((old_code, product_number))
                        
                        # swap成功時にカウント
                        self.swap_count += 1
                        
                        # 新しい検査員に時間を追加
                        new_code = replacement_inspector['コード']
                        if new_code not in self.inspector_daily_assignments:
                            self.inspector_daily_assignments[new_code] = {}
                        if current_date not in self.inspector_daily_assignments[new_code]:
                            self.inspector_daily_assignments[new_code][current_date] = 0.0
                        
                        self.inspector_daily_assignments[new_code][current_date] += divided_time
                        if new_code not in self.inspector_work_hours:
                            self.inspector_work_hours[new_code] = 0.0
                        self.inspector_work_hours[new_code] += divided_time
                        
                        # 品番別累計時間も更新
                        if new_code not in self.inspector_product_hours:
                            self.inspector_product_hours[new_code] = {}
                        self.inspector_product_hours[new_code][product_number] = (
                            self.inspector_product_hours[new_code].get(product_number, 0.0) + divided_time
                        )
                        # 改善ポイント: 設定時間超過の場合はrelaxed_product_limit_assignmentsに追加
                        if self.inspector_product_hours[new_code][product_number] > self.product_limit_hard_threshold:
                            self.relaxed_product_limit_assignments.add((new_code, product_number))
                        
                        # 【追加】制約緩和レベルが'scored_relaxation'の場合は、relaxed_product_limit_assignmentsに追加
                        if selected_relaxation_level == 'scored_relaxation':
                            # 同一品番4時間超過または勤務時間超過の場合はrelaxed_product_limit_assignmentsに追加
                            current_product_hours_after = self.inspector_product_hours[new_code][product_number]
                            if current_product_hours_after > self.product_limit_hard_threshold:
                                self.relaxed_product_limit_assignments.add((new_code, product_number))
                                self.log_message(
                                    f"制約緩和割り当てをrelaxed_product_limit_assignmentsに追加: {new_code}, {product_number} (同一品番累計: {current_product_hours_after:.1f}h)",
                                    level='debug',
                                )
                        
                        # 当日洗浄上がり品のロットの場合、same_day_cleaning_inspectorsを更新（品番単位）
                        if is_same_day_cleaning_lot:
                            # 元の検査員がこの品番に割り当てられていた場合、削除
                            if product_number in self.same_day_cleaning_inspectors:
                                self.same_day_cleaning_inspectors[product_number].discard(old_code)
                            # 新しい検査員をこの品番に割り当てられた検査員として記録
                            self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(new_code)
                        
                        # チーム情報を更新
                        self.update_team_info(result_df, index, inspector_master_df, show_skill_values)
                        
                        self.log_message(
                            f"置き換え: '{inspector_name}' → '{replacement_inspector['氏名']}' (品番: {product_number}, 出荷予定日: {row['出荷予定日']})",
                            level='debug',
                        )
                        return True
            
            elif len(current_inspectors) == 1:
                # 1人だけの場合、増員するか他の人に置き換え
                # 検査時間が3時間未満の場合は増員を行わず、置き換えのみを行う
                # 検査時間が3時間以上の場合は、検査員5名まで増員可能
                # ただし、既に5名の場合は置き換えのみ
                
                # 検査時間が設定時間未満の場合は増員をスキップして置き換えのみ
                if inspection_time < self.required_inspectors_threshold:
                    # 置き換え処理（増員ではなく）
                    process_number = row.get('現在工程番号', '')
                    lot_process_name = str(row.get('現在工程名', '') or '').strip()
                    # スキルマスタに登録があるか確認
                    skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                    is_new_product = skill_rows.empty
                    shipping_date = row.get('出荷予定日', None)
                    available_inspectors = self.get_available_inspectors(
                        product_number, process_number, skill_master_df, inspector_master_df,
                        shipping_date=shipping_date, allow_new_team_fallback=is_new_product,
                        process_name_context=lot_process_name
                    )
                    # 新規品の場合は新製品チームも取得
                    if not available_inspectors and is_new_product:
                        self.log_message(f"新規品 {product_number}: 置き換え用に新製品チームを取得します（検査時間3時間未満のため増員なし）")
                        available_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                    
                    # 既に割り当てられている検査員を除外
                    current_codes = [inspector_code]
                    
                    # 置き換え候補を探す（同一品番の4時間上限チェックも含む）
                    replacement_candidates = []
                    excluded_reasons = {}  # 除外理由を記録
                    
                    # 当日洗浄上がり品かどうかを判定
                    shipping_date_raw = row.get('出荷予定日', None)
                    shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                    is_same_day_cleaning_lot = (
                        shipping_date_str == "当日洗浄上がり品" or
                        shipping_date_str == "当日洗浄品" or
                        "当日洗浄" in shipping_date_str
                    )
                    
                    # 当日洗浄上がり品の場合、既にこの品番に割り当てられた検査員を取得
                    already_assigned_to_this_product = set()
                    if is_same_day_cleaning_lot:
                        already_assigned_to_this_product = self.same_day_cleaning_inspectors.get(product_number, set())
                    
                    # 【改善】段階的な制約緩和レベルを定義
                    relaxation_levels = [
                        {
                            'name': 'strict',
                            'same_day_cleaning_product_constraint': True,
                            'work_hours_buffer': 0.0,
                            'product_limit': self.product_limit_hard_threshold,
                        },
                        {
                            'name': 'relaxed_product_limit',
                            'same_day_cleaning_product_constraint': True,
                            'work_hours_buffer': 0.0,
                            'product_limit': PRODUCT_LIMIT_FINAL_TOLERANCE,
                        },
                        {
                            'name': 'relaxed_work_hours',
                            'same_day_cleaning_product_constraint': True,
                            'work_hours_buffer': 0.5,
                            'product_limit': PRODUCT_LIMIT_FINAL_TOLERANCE,
                        },
                        {
                            'name': 'relaxed_same_day_constraint',
                            'same_day_cleaning_product_constraint': False,
                            'work_hours_buffer': 0.5,
                            'product_limit': PRODUCT_LIMIT_FINAL_TOLERANCE,
                        },
                    ]
                    
                    # 各緩和レベルで候補を検索
                    selected_relaxation_level = None
                    for level in relaxation_levels:
                        level_candidates = []
                        level_excluded = {}
                        
                        for insp in available_inspectors:
                            if insp['コード'] not in current_codes:
                                code = insp['コード']
                                insp_name = insp['氏名']
                                
                                # 当日洗浄上がり品の品番単位制約チェック（緩和レベルに応じて）
                                if (level['same_day_cleaning_product_constraint'] and 
                                    is_same_day_cleaning_lot and 
                                    code in already_assigned_to_this_product):
                                    level_excluded[insp_name] = f"既にこの品番に割り当て済み（品番単位の制約）"
                                    continue
                                
                                max_hours = inspector_max_hours.get(code, 8.0)
                                daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                                
                                # 勤務時間チェック（緩和レベルに応じて）
                                if daily_hours + inspection_time > max_hours - level['work_hours_buffer']:
                                    level_excluded[insp_name] = f"勤務時間超過 ({daily_hours:.1f}h + {inspection_time:.1f}h > {max_hours:.1f}h - {level['work_hours_buffer']:.1f}h緩和)"
                                    continue
                                
                                # 同一品番4時間上限チェック（緩和レベルに応じて）
                                current_product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                                if current_product_hours + inspection_time > level['product_limit']:
                                    level_excluded[insp_name] = f"同一品番超過 ({current_product_hours:.1f}h + {inspection_time:.1f}h > {level['product_limit']:.1f}h)"
                                    continue
                                
                                total_hours = self.inspector_work_hours.get(code, 0.0)
                                level_candidates.append((total_hours, insp))
                        
                        if level_candidates:
                            replacement_candidates = level_candidates
                            excluded_reasons = level_excluded
                            selected_relaxation_level = level['name']
                            if level['name'] != 'strict':
                                self.log_message(
                                    f"制約緩和レベル '{level['name']}' で {len(level_candidates)}人の候補が見つかりました (品番: {product_number}, 置き換え処理)",
                                    level='debug',
                                )
                            break
                        else:
                            excluded_reasons.update(level_excluded)
                    
                    # 候補が見つからない場合、制約違反の程度で優先順位付けを試行
                    if not replacement_candidates:
                        scored_candidates = []
                        
                        for insp in available_inspectors:
                            if insp['コード'] not in current_codes:
                                code = insp['コード']
                                insp_name = insp['氏名']
                                
                                score = 0
                                violations = []
                                
                                # 各制約違反にペナルティを付与
                                if is_same_day_cleaning_lot and code in already_assigned_to_this_product:
                                    score += 100  # 高ペナルティ
                                    violations.append("品番単位制約")
                                
                                max_hours = inspector_max_hours.get(code, 8.0)
                                daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                                excess_work_hours = max(0, (daily_hours + inspection_time) - max_hours)
                                if excess_work_hours > 0:
                                    score += excess_work_hours * 10
                                    violations.append(f"勤務時間超過{excess_work_hours:.1f}h")
                                
                                current_product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                                excess_product_hours = max(0, (current_product_hours + inspection_time) - PRODUCT_LIMIT_HARD_THRESHOLD)
                                if excess_product_hours > 0:
                                    score += excess_product_hours * 5
                                    violations.append(f"同一品番超過{excess_product_hours:.1f}h")
                                
                                scored_candidates.append((score, insp, violations))
                        
                        # スコアが低い順（違反が少ない順）にソート
                        scored_candidates.sort(key=self._priority_sort_key)
                        
                        # 最も違反が少ない候補を選択（許容可能な範囲内）
                        if scored_candidates:
                            best_score = scored_candidates[0][0]
                            # 当日洗浄上がり品の場合は閾値を緩和（100まで許容）
                            threshold = 100 if is_same_day_cleaning_lot else 50
                            if best_score < threshold:
                                best_candidate = scored_candidates[0]
                                self.log_message(
                                    f"制約を一部緩和して割り当て: '{best_candidate[1]['氏名']}' "
                                    f"(違反: {', '.join(best_candidate[2])}, 品番: {product_number}, スコア: {best_score:.1f}, 置き換え処理)",
                                    level='warning',
                                )
                                replacement_candidates = [(0, best_candidate[1])]
                                selected_relaxation_level = 'scored_relaxation'
                    
                    if not replacement_candidates:
                        # 代替検査員が見つからない理由を詳細にログ出力（統計情報を追加）
                        total_candidates = len(available_inspectors)
                        excluded_by_reason = {}
                        for name, reason in excluded_reasons.items():
                            reason_type = reason.split(':')[0] if ':' in reason else reason.split('(')[0].strip()
                            if reason_type not in excluded_by_reason:
                                excluded_by_reason[reason_type] = 0
                            excluded_by_reason[reason_type] += 1
                        
                        self.log_message(
                            f"代替検査員が見つかりません (品番: {product_number}, ロットインデックス: {index}, 置き換え処理、検査時間3時間未満のため増員なし)",
                            level='warning',
                        )
                        self.log_message(f"   候補検査員総数: {total_candidates}人, 除外された検査員: {len(excluded_reasons)}人", level='warning')
                        if excluded_by_reason:
                            self.log_message(f"   除外理由別統計: {excluded_by_reason}", level='warning')
                        if excluded_reasons:
                            for name, reason in list(excluded_reasons.items())[:5]:
                                self.log_message(f"     - {name}: {reason}", level='warning')
                            if len(excluded_reasons) > 5:
                                self.log_message(f"     ... 他{len(excluded_reasons) - 5}人", level='warning')
                        else:
                            self.log_message(f"   理由: 候補検査員が全て既に割り当て済み", level='warning')
                        return False
                    
                    if replacement_candidates:
                        # 最も総勤務時間が少ない人に置き換え
                        replacement_candidates.sort(key=self._priority_sort_key)
                        replacement_inspector = replacement_candidates[0][1]
                        
                        if show_skill_values:
                            if replacement_inspector.get('is_new_team', False):
                                replacement_name = f"{replacement_inspector['氏名']}(新)"
                            else:
                                replacement_name = f"{replacement_inspector['氏名']}({replacement_inspector['スキル']})"
                        else:
                            if replacement_inspector.get('is_new_team', False):
                                replacement_name = f"{replacement_inspector['氏名']}(新)"
                            else:
                                replacement_name = replacement_inspector['氏名']
                        
                        result_df.at[index, '検査員1'] = replacement_name
                        
                        # 履歴を更新（元の検査員から時間を引く）
                        old_code = inspector_code
                        old_daily = self.inspector_daily_assignments.get(old_code, {}).get(current_date, 0.0)
                        old_total = self.inspector_work_hours.get(old_code, 0.0)
                        self.inspector_daily_assignments[old_code][current_date] = max(0.0, old_daily - inspection_time)
                        self.inspector_work_hours[old_code] = max(0.0, old_total - inspection_time)
                        
                        # 品番別累計時間も更新
                        if old_code in self.inspector_product_hours:
                            if product_number in self.inspector_product_hours[old_code]:
                                self.inspector_product_hours[old_code][product_number] = max(0.0, self.inspector_product_hours[old_code][product_number] - inspection_time)
                        
                        # 新しい検査員に時間を追加
                        new_code = replacement_inspector['コード']
                        if new_code not in self.inspector_daily_assignments:
                            self.inspector_daily_assignments[new_code] = {}
                        if current_date not in self.inspector_daily_assignments[new_code]:
                            self.inspector_daily_assignments[new_code][current_date] = 0.0
                        
                        self.inspector_daily_assignments[new_code][current_date] += inspection_time
                        if new_code not in self.inspector_work_hours:
                            self.inspector_work_hours[new_code] = 0.0
                        self.inspector_work_hours[new_code] += inspection_time
                        
                        # 品番別累計時間も更新
                        if new_code not in self.inspector_product_hours:
                            self.inspector_product_hours[new_code] = {}
                        self.inspector_product_hours[new_code][product_number] = (
                            self.inspector_product_hours[new_code].get(product_number, 0.0) + inspection_time
                        )
                        
                        # 【追加】制約緩和レベルが'scored_relaxation'の場合は、relaxed_product_limit_assignmentsに追加
                        if selected_relaxation_level == 'scored_relaxation':
                            # 同一品番4時間超過の場合はrelaxed_product_limit_assignmentsに追加
                            current_product_hours_after = self.inspector_product_hours[new_code][product_number]
                            if current_product_hours_after > self.product_limit_hard_threshold:
                                self.relaxed_product_limit_assignments.add((new_code, product_number))
                                self.log_message(
                                    f"制約緩和割り当てをrelaxed_product_limit_assignmentsに追加: {new_code}, {product_number} (同一品番累計: {current_product_hours_after:.1f}h, 置き換え処理)",
                                    level='debug',
                                )
                        
                        # 当日洗浄上がり品のロットの場合、same_day_cleaning_inspectorsを更新（品番単位）
                        if is_same_day_cleaning_lot:
                            # 元の検査員がこの品番に割り当てられていた場合、削除
                            if product_number in self.same_day_cleaning_inspectors:
                                self.same_day_cleaning_inspectors[product_number].discard(old_code)
                            # 新しい検査員をこの品番に割り当てられた検査員として記録
                            self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(new_code)
                        
                        result_df.at[index, '分割検査時間'] = round(inspection_time, 1)
                        
                        # チーム情報を更新
                        self.update_team_info(result_df, index, inspector_master_df, show_skill_values)
                        
                        self.log_message(
                            f"置き換え: '{inspector_name}' → '{replacement_inspector['氏名']}' (品番: {product_number}, 検査時間: {inspection_time:.1f}h < 3.0hのため増員なし、出荷予定日: {row['出荷予定日']})",
                            level='debug',
                        )
                        return True
                
                elif len(current_inspectors) >= 5:
                    # 置き換え処理（増員ではなく）
                    process_number = row.get('現在工程番号', '')
                    lot_process_name = str(row.get('現在工程名', '') or '').strip()
                    # スキルマスタに登録があるか確認
                    skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                    is_new_product = skill_rows.empty
                    shipping_date = row.get('出荷予定日', None)
                    available_inspectors = self.get_available_inspectors(
                        product_number, process_number, skill_master_df, inspector_master_df,
                        shipping_date=shipping_date, allow_new_team_fallback=is_new_product,
                        process_name_context=lot_process_name
                    )
                    # 新規品の場合は新製品チームも取得
                    if not available_inspectors and is_new_product:
                        self.log_message(
                            f"新規品 {product_number}: 置き換え用に新製品チームを取得します",
                            level='debug',
                        )
                        available_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                    
                    # 既に割り当てられている検査員を除外
                    current_codes = [inspector_code]
                    
                    # 置き換え候補を探す（同一品番の4時間上限チェックも含む）
                    replacement_candidates = []
                    excluded_reasons = {}  # 除外理由を記録
                    
                    # 当日洗浄上がり品かどうかを判定
                    shipping_date_raw = row.get('出荷予定日', None)
                    shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                    is_same_day_cleaning_lot = (
                        shipping_date_str == "当日洗浄上がり品" or
                        shipping_date_str == "当日洗浄品" or
                        "当日洗浄" in shipping_date_str
                    )
                    
                    # force_assign_same_dayを初期化（未初期化参照を防止）
                    force_assign_same_day = self._should_force_assign_same_day(shipping_date_raw)
                    
                    # 当日洗浄上がり品の場合、既にこの品番に割り当てられた検査員を取得
                    already_assigned_to_this_product = set()
                    if is_same_day_cleaning_lot:
                        already_assigned_to_this_product = set() if force_assign_same_day else self.same_day_cleaning_inspectors.get(product_number, set())
                    
                    for insp in available_inspectors:
                        if insp['コード'] not in current_codes:
                            code = insp['コード']
                            insp_name = insp['氏名']
                            
                            # 当日洗浄上がり品の場合、既にこの品番に割り当てられた検査員を除外（品番単位の制約）
                            if is_same_day_cleaning_lot and (not force_assign_same_day) and code in already_assigned_to_this_product:
                                excluded_reasons[insp_name] = f"既にこの品番に割り当て済み（品番単位の制約）"
                                continue
                            
                            max_hours = inspector_max_hours.get(code, 8.0)
                            # 勤務時間チェック
                            if not self.check_work_hours_capacity(code, inspection_time, max_hours, current_date):
                                daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                                excluded_reasons[insp_name] = f"勤務時間超過 ({daily_hours:.1f}h + {inspection_time:.1f}h > {max_hours:.1f}h)"
                                continue
                            # 改善ポイント: 最適化フェーズでの4時間上限チェック（厳格）
                            current_product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                            if current_product_hours + inspection_time > self.product_limit_hard_threshold:
                                excluded_reasons[insp_name] = f"同一品番{self.product_limit_hard_threshold:.1f}時間超過 ({current_product_hours:.1f}h + {inspection_time:.1f}h = {current_product_hours + inspection_time:.1f}h > {self.product_limit_hard_threshold:.1f}h)"
                                continue
                            total_hours = self.inspector_work_hours.get(code, 0.0)
                            replacement_candidates.append((total_hours, insp))
                    
                    if not replacement_candidates:
                        # 代替検査員が見つからない理由を詳細にログ出力
                        self.log_message(
                            f"代替検査員が見つかりません (品番: {product_number}, ロットインデックス: {index}, 置き換え処理)",
                            level='warning',
                        )
                        if excluded_reasons:
                            self.log_message(f"   除外された検査員: {len(excluded_reasons)}人", level='warning')
                            for name, reason in list(excluded_reasons.items())[:5]:  # 最大5人まで表示
                                self.log_message(f"     - {name}: {reason}", level='warning')
                            if len(excluded_reasons) > 5:
                                self.log_message(f"     ... 他{len(excluded_reasons) - 5}人", level='warning')
                        else:
                            self.log_message(f"   理由: 候補検査員が全て既に割り当て済み", level='warning')
                    
                    if replacement_candidates:
                        # 最も総勤務時間が少ない人に置き換え
                        replacement_candidates.sort(key=self._priority_sort_key)
                        replacement_inspector = replacement_candidates[0][1]
                        
                        if show_skill_values:
                            if replacement_inspector.get('is_new_team', False):
                                replacement_name = f"{replacement_inspector['氏名']}(新)"
                            else:
                                replacement_name = f"{replacement_inspector['氏名']}({replacement_inspector['スキル']})"
                        else:
                            if replacement_inspector.get('is_new_team', False):
                                replacement_name = f"{replacement_inspector['氏名']}(新)"
                            else:
                                replacement_name = replacement_inspector['氏名']
                        
                        result_df.at[index, '検査員1'] = replacement_name
                        
                        # 履歴を更新（元の検査員から時間を引く）
                        old_code = inspector_code
                        old_daily = self.inspector_daily_assignments.get(old_code, {}).get(current_date, 0.0)
                        old_total = self.inspector_work_hours.get(old_code, 0.0)
                        self.inspector_daily_assignments[old_code][current_date] = max(0.0, old_daily - inspection_time)
                        self.inspector_work_hours[old_code] = max(0.0, old_total - inspection_time)
                        
                        # 品番別累計時間も更新
                        if old_code in self.inspector_product_hours:
                            if product_number in self.inspector_product_hours[old_code]:
                                self.inspector_product_hours[old_code][product_number] = max(0.0, self.inspector_product_hours[old_code][product_number] - inspection_time)
                        
                        # 新しい検査員に時間を追加
                        new_code = replacement_inspector['コード']
                        if new_code not in self.inspector_daily_assignments:
                            self.inspector_daily_assignments[new_code] = {}
                        if current_date not in self.inspector_daily_assignments[new_code]:
                            self.inspector_daily_assignments[new_code][current_date] = 0.0
                        
                        self.inspector_daily_assignments[new_code][current_date] += inspection_time
                        if new_code not in self.inspector_work_hours:
                            self.inspector_work_hours[new_code] = 0.0
                        self.inspector_work_hours[new_code] += inspection_time
                        
                        # 品番別累計時間も更新
                        if new_code not in self.inspector_product_hours:
                            self.inspector_product_hours[new_code] = {}
                        self.inspector_product_hours[new_code][product_number] = (
                            self.inspector_product_hours[new_code].get(product_number, 0.0) + inspection_time
                        )
                        
                        # 当日洗浄上がり品のロットの場合、same_day_cleaning_inspectorsを更新（品番単位）
                        if is_same_day_cleaning_lot:
                            # 元の検査員がこの品番に割り当てられていた場合、削除
                            if product_number in self.same_day_cleaning_inspectors:
                                self.same_day_cleaning_inspectors[product_number].discard(old_code)
                            # 新しい検査員をこの品番に割り当てられた検査員として記録
                            self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(new_code)
                        
                        result_df.at[index, '分割検査時間'] = round(inspection_time, 1)
                        
                        # チーム情報を更新
                        self.update_team_info(result_df, index, inspector_master_df, show_skill_values)
                        
                        self.log_message(
                            f"置き換え: '{inspector_name}' → '{replacement_inspector['氏名']}' (品番: {product_number}, 検査時間: {inspection_time:.1f}h, 出荷予定日: {row['出荷予定日']})",
                            level='debug',
                        )
                        return True
                else:
                    # 検査員5名まで増員を試みる（3時間の条件は無視）
                    # 現在の検査員が5名未満の場合のみ増員を試みる
                    if len(current_inspectors) < 5:
                        process_number = row.get('現在工程番号', '')
                        lot_process_name = str(row.get('現在工程名', '') or '').strip()
                        # スキルマスタに登録があるか確認
                        skill_rows = skill_master_df[skill_master_df.iloc[:, 0] == product_number]
                        is_new_product = skill_rows.empty
                        shipping_date = row.get('出荷予定日', None)
                        available_inspectors = self.get_available_inspectors(
                            product_number, process_number, skill_master_df, inspector_master_df,
                            shipping_date=shipping_date, allow_new_team_fallback=is_new_product,
                            process_name_context=lot_process_name
                        )
                        # 新規品の場合は新製品チームも取得
                        if not available_inspectors and is_new_product:
                            self.log_message(
                                f"新規品 {product_number}: 増員用に新製品チームを取得します",
                                level='debug',
                            )
                            available_inspectors = self.get_new_product_team_inspectors(inspector_master_df)
                    
                    # 既に割り当てられている検査員を除外
                    current_codes = [inspector_code]
                    
                    # 追加できる検査員を探す
                    addition_candidates = []
                    excluded_reasons = {}  # 除外理由を記録
                    
                    # 当日洗浄上がり品かどうかを判定
                    shipping_date_raw = row.get('出荷予定日', None)
                    shipping_date_str = str(shipping_date_raw).strip() if pd.notna(shipping_date_raw) else ''
                    is_same_day_cleaning_lot = (
                        shipping_date_str == "当日洗浄上がり品" or
                        shipping_date_str == "当日洗浄品" or
                        "当日洗浄" in shipping_date_str
                    )
                    
                    # force_assign_same_dayを初期化（未初期化参照を防止）
                    force_assign_same_day = self._should_force_assign_same_day(shipping_date_raw)
                    
                    # 当日洗浄上がり品の場合、既にこの品番に割り当てられた検査員を取得
                    already_assigned_to_this_product = set()
                    if is_same_day_cleaning_lot:
                        already_assigned_to_this_product = set() if force_assign_same_day else self.same_day_cleaning_inspectors.get(product_number, set())
                    
                    for insp in available_inspectors:
                        if insp['コード'] not in current_codes:
                            code = insp['コード']
                            insp_name = insp['氏名']
                            
                            # 当日洗浄上がり品の場合、既にこの品番に割り当てられた検査員を除外（品番単位の制約）
                            if is_same_day_cleaning_lot and (not force_assign_same_day) and code in already_assigned_to_this_product:
                                excluded_reasons[insp_name] = f"既にこの品番に割り当て済み（品番単位の制約）"
                                continue
                            
                            max_hours = inspector_max_hours.get(code, 8.0)
                            if not self.check_work_hours_capacity(code, divided_time, max_hours, current_date):
                                daily_hours = self.inspector_daily_assignments.get(code, {}).get(current_date, 0.0)
                                excluded_reasons[insp_name] = f"勤務時間超過 ({daily_hours:.1f}h + {divided_time:.1f}h > {max_hours:.1f}h)"
                                continue
                            # 改善ポイント: 最適化フェーズでの4時間上限チェック（厳格）
                            current_product_hours = self.inspector_product_hours.get(code, {}).get(product_number, 0.0)
                            # force_assign_same_dayを初期化（未初期化参照を防止）
                            if 'force_assign_same_day' not in locals():
                                force_assign_same_day = self._should_force_assign_same_day(shipping_date_raw)
                            if (not force_assign_same_day) and current_product_hours + divided_time > self.product_limit_hard_threshold:
                                excluded_reasons[insp_name] = f"同一品番{self.product_limit_hard_threshold:.1f}時間超過 ({current_product_hours:.1f}h + {divided_time:.1f}h = {current_product_hours + divided_time:.1f}h > {self.product_limit_hard_threshold:.1f}h)"
                                continue
                            total_hours = self.inspector_work_hours.get(code, 0.0)
                            addition_candidates.append((total_hours, insp))
                    
                    if not addition_candidates and len(current_inspectors) < 5:
                        # 追加検査員が見つからない理由を詳細にログ出力
                        self.log_message(
                            f"追加検査員が見つかりません (品番: {product_number}, ロットインデックス: {index}, 増員処理)",
                            level='warning',
                        )
                        if excluded_reasons:
                            self.log_message(f"   除外された検査員: {len(excluded_reasons)}人", level='warning')
                            for name, reason in list(excluded_reasons.items())[:5]:  # 最大5人まで表示
                                self.log_message(f"     - {name}: {reason}", level='warning')
                            if len(excluded_reasons) > 5:
                                self.log_message(f"     ... 他{len(excluded_reasons) - 5}人", level='warning')
                        else:
                            self.log_message(f"   理由: 候補検査員が全て既に割り当て済み", level='warning')
                    
                    if addition_candidates and len(current_inspectors) < 5:
                        # 最も総勤務時間が少ない人を追加
                        addition_candidates.sort(key=self._priority_sort_key)
                        addition_inspector = addition_candidates[0][1]
                        new_count = len(current_inspectors) + 1
                        
                        if show_skill_values:
                            if addition_inspector.get('is_new_team', False):
                                addition_name = f"{addition_inspector['氏名']}(新)"
                            else:
                                addition_name = f"{addition_inspector['氏名']}({addition_inspector['スキル']})"
                        else:
                            if addition_inspector.get('is_new_team', False):
                                addition_name = f"{addition_inspector['氏名']}(新)"
                            else:
                                addition_name = addition_inspector['氏名']
                        
                        result_df.at[index, f'検査員{new_count}'] = addition_name
                        result_df.at[index, '検査員人数'] = new_count
                        
                        # 分割検査時間の計算: 検査時間 ÷ 実際の分割した検査人数
                        actual_divided_time = inspection_time / new_count
                        
                        # 新しい検査員に時間を追加
                        new_code = addition_inspector['コード']
                        if new_code not in self.inspector_daily_assignments:
                            self.inspector_daily_assignments[new_code] = {}
                        if current_date not in self.inspector_daily_assignments[new_code]:
                            self.inspector_daily_assignments[new_code][current_date] = 0.0
                        
                        # 元の検査員の時間を新しい分割時間に調整
                        old_daily = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
                        old_divided_time = inspection_time / (new_count - 1)  # 増員前の人数
                        new_divided_time = actual_divided_time  # 増員後の分割時間
                        
                        # 元の検査員の時間を調整（新しい分割時間に合わせる）
                        adjustment = old_divided_time - new_divided_time
                        self.inspector_daily_assignments[inspector_code][current_date] = old_daily - adjustment
                        self.inspector_work_hours[inspector_code] = max(0.0, self.inspector_work_hours.get(inspector_code, 0.0) - adjustment)
                        
                        # 品番別累計時間も更新
                        if inspector_code in self.inspector_product_hours:
                            if product_number in self.inspector_product_hours[inspector_code]:
                                self.inspector_product_hours[inspector_code][product_number] = max(0.0, self.inspector_product_hours[inspector_code][product_number] - adjustment)
                        
                        # 新しい検査員に時間を追加
                        self.inspector_daily_assignments[new_code][current_date] += new_divided_time
                        if new_code not in self.inspector_work_hours:
                            self.inspector_work_hours[new_code] = 0.0
                        self.inspector_work_hours[new_code] += new_divided_time
                        
                        # 品番別累計時間も更新
                        if new_code not in self.inspector_product_hours:
                            self.inspector_product_hours[new_code] = {}
                        self.inspector_product_hours[new_code][product_number] = (
                            self.inspector_product_hours[new_code].get(product_number, 0.0) + new_divided_time
                        )
                        
                        # 当日洗浄上がり品のロットの場合、same_day_cleaning_inspectorsを更新（品番単位）
                        if is_same_day_cleaning_lot:
                            # 新しい検査員をこの品番に割り当てられた検査員として記録
                            self.same_day_cleaning_inspectors.setdefault(product_number, set()).add(new_code)
                        
                        result_df.at[index, '分割検査時間'] = round(actual_divided_time, 1)
                        
                        # チーム情報を更新
                        self.update_team_info(result_df, index, inspector_master_df, show_skill_values)
                        
                        self.log_message(f"増員: '{inspector_name}' に '{addition_inspector['氏名']}' を追加 (品番: {product_number}, 出荷予定日: {row['出荷予定日']})")
                        return True
            
            # 是正できなかった場合の処理
            force_assign_same_day = self._should_force_assign_same_day(row.get('出荷予定日', None))
            if force_assign_same_day:
                shipping_date_str = str(shipping_date) if shipping_date is not None else 'N/A'
                self.log_message(
                    f"当日洗浄 {product_number} (出荷予定日: {shipping_date_str}) のルール違反を是正できませんでしたが、最優先のため割り当てを維持します",
                    level='warning',
                )
                current_product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                if current_product_hours + divided_time > self.product_limit_hard_threshold:
                    self.relaxed_product_limit_assignments.add((inspector_code, product_number))
                return True

            if protect_new_product:
                # 新規品で出荷予定日が2週間以内の場合は、代替が見つからなくても保護する
                # 勤務時間超過を一時的に許容する（relaxed_product_limit_assignmentsに追加）
                shipping_date_str = str(shipping_date) if shipping_date is not None else 'N/A'
                self.log_message(
                    f"新規品 {product_number} (出荷予定日: {shipping_date_str}) のルール違反を是正できませんでしたが、出荷予定日が2週間以内のため保護します"
                )
                # 現在の品番別累計時間を取得
                current_product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                # 改善ポイント: 設定時間超過の場合は、relaxed_product_limit_assignmentsに追加
                if current_product_hours + divided_time > self.product_limit_hard_threshold:
                    self.relaxed_product_limit_assignments.add((inspector_code, product_number))
                return True  # 保護したのでTrueを返す
            else:
                # 通常の場合は未割当にする
                # 詳細な理由をログ出力
                violation_type = "同一品番4時間超過" if inspector_code in self.inspector_product_hours and product_number in self.inspector_product_hours.get(inspector_code, {}) else "勤務時間超過"
                current_product_hours = self.inspector_product_hours.get(inspector_code, {}).get(product_number, 0.0)
                daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
                max_hours = inspector_max_hours.get(inspector_code, 8.0)
                
                self.log_message(f"ルール違反を是正できませんでした。品番 {product_number} のロットを未割当にします", level='warning')
                self.log_message(f"   違反検査員: {inspector_name} (コード: {inspector_code})", level='warning')
                self.log_message(f"   違反内容: {violation_type}", level='warning')
                if violation_type == "同一品番4時間超過":
                    self.log_message(f"   同一品番累計: {current_product_hours:.1f}h (追加予定: {divided_time:.1f}h)", level='warning')
                else:
                    self.log_message(f"   勤務時間: {daily_hours:.1f}h / 最大: {max_hours:.1f}h (追加予定: {divided_time:.1f}h)", level='warning')
                self.clear_assignment(result_df, index)
                return False
            
        except Exception as e:
            self.log_message(f"違反是正中にエラーが発生しました: {str(e)}")
            # エラー時も未割当にする
            try:
                self.clear_assignment(result_df, index)
            except Exception as e:
                logger.debug(f"割り当てクリア処理でエラーが発生しました（無視）: {e}")
                pass
            return False
    
    def clear_assignment(self, result_df: pd.DataFrame, index: int) -> None:
        """
        ロットの割り当てをクリア（未割当にする）
        
        Args:
            result_df: 結果DataFrame
            index: 行インデックス
        """
        try:
            current_date = pd.Timestamp.now().date()
            row = result_df.iloc[index]
            shipping_date = row.get('出荷予定日', None)
            product_number = row.get('品番', '')
            divided_time = row.get('分割検査時間', 0.0)
            total_required = row.get('検査時間', divided_time)

            # 要日出荷（当日/過去）・当日洗浄上がり品は、最終検証で違反が出ても未割当に落とさず割当維持を優先する
            try:
                has_any_assignment = False
                for i in range(1, 6):
                    inspector_col = f'検査員{i}'
                    inspector_name = row.get(inspector_col, '')
                    if pd.notna(inspector_name) and str(inspector_name).strip() != '':
                        has_any_assignment = True
                        break
                if has_any_assignment and self._should_force_keep_assignment(shipping_date):
                    current_team_info = result_df.at[index, 'チーム情報'] if 'チーム情報' in result_df.columns else ''
                    if not (pd.notna(current_team_info) and "割当維持" in str(current_team_info)):
                        if 'チーム情報' in result_df.columns:
                            result_df.at[index, 'チーム情報'] = '要日優先のため割当維持(制約違反あり)'
                    self.log_message(
                        f"要日優先: ロットの割当解除をスキップしました: 品番 {product_number}, 出荷予定日 {shipping_date}",
                        level='warning',
                    )
                    return
            except Exception:
                pass
            
            # 履歴からこのロットの時間を引く（割り当てされている検査員の時間を戻す）
            for i in range(1, 6):
                inspector_col = f'検査員{i}'
                inspector_name = row.get(inspector_col, '')
                if pd.notna(inspector_name) and str(inspector_name).strip() != '':
                    inspector_name_str = str(inspector_name).strip()
                    if '(' in inspector_name_str:
                        inspector_name_str = inspector_name_str.split('(')[0].strip()
                    
                    # 検査員コードを取得して履歴から時間を引く
                    # 注意: この時点で履歴が既に再計算されている場合、時間を引く処理は不要
                    # 履歴の再計算は呼び出し元で行われる前提
            
            # 検査員1～5をクリア
            for i in range(1, 6):
                result_df.at[index, f'検査員{i}'] = ''
            result_df.at[index, '検査員人数'] = 0
            result_df.at[index, '分割検査時間'] = 0.0
            
            # チーム情報を保持（既に詳細な理由が設定されている場合は保持、そうでない場合は「未割当」）
            current_team_info = result_df.at[index, 'チーム情報']
            if pd.notna(current_team_info) and str(current_team_info).strip().startswith('未割当'):
                # 既に詳細な理由が設定されている場合は保持
                pass
            else:
                result_df.at[index, 'チーム情報'] = '未割当(最終検証で違反検出)'
            
            result_df.at[index, 'remaining_work_hours'] = round(total_required if pd.notna(total_required) else 0.0, 2)
            result_df.at[index, 'assignability_status'] = 'final_product_limit_violation'
            result_df.at[index, 'over_product_limit_flag'] = False
            
            self.log_message(f"ロットを未割当にしました: 品番 {product_number}, インデックス {index}, チーム情報: '{result_df.at[index, 'チーム情報']}'")
            
        except Exception as e:
            self.log_message(f"未割当処理中にエラーが発生しました: {str(e)}")
    
    def check_work_hours_capacity(
        self,
        inspector_code: str,
        additional_hours: float,
        max_hours: float,
        current_date: date
    ) -> bool:
        """
        検査員の勤務時間に余裕があるかチェック（10%超過まで許容）
        
        Args:
            inspector_code: 検査員コード
            additional_hours: 追加する時間
            max_hours: 最大勤務時間
            current_date: 現在日付
        
        Returns:
            余裕がある場合はTrue、それ以外はFalse（10%超過まで許容）
        """
        try:
            daily_hours = self.inspector_daily_assignments.get(inspector_code, {}).get(current_date, 0.0)
            # 10%超過まで許容
            allowed_max_hours = max_hours * (1.0 + WORK_HOURS_OVERRUN_RATE)
            # 0.05時間（3分）の余裕を持たせる
            return daily_hours + additional_hours <= allowed_max_hours - WORK_HOURS_BUFFER
        except Exception as e:
            logger.debug(f"勤務時間チェック処理でエラーが発生しました（デフォルト: False）: {e}")
            return False
    
    def update_team_info(
        self,
        result_df: pd.DataFrame,
        index: int,
        inspector_master_df: pd.DataFrame,
        show_skill_values: bool = False
    ) -> None:
        """
        チーム情報を更新（最適化後に呼び出す）
        
        Args:
            result_df: 結果DataFrame
            index: 行インデックス
            inspector_master_df: 検査員マスタ
            show_skill_values: スキル値を表示するか
        """
        try:
            # 検査員人数が0の場合、既存のチーム情報が「未割当」で始まる場合は保持
            inspector_count = result_df.at[index, '検査員人数']
            if inspector_count == 0 or pd.isna(inspector_count) or inspector_count == 0:
                current_team_info = result_df.at[index, 'チーム情報']
                if pd.notna(current_team_info) and str(current_team_info).strip().startswith('未割当'):
                    # 既に未割当の理由が設定されている場合は保持
                    self.log_message(f"チーム情報更新: インデックス{index}の未割当理由を保持 '{current_team_info}'")
                    return str(current_team_info)
                elif pd.notna(current_team_info) and str(current_team_info).strip() != '':
                    # 既に何か設定されている場合は保持
                    self.log_message(f"チーム情報更新: インデックス{index}の既存情報を保持 '{current_team_info}'")
                    return str(current_team_info)
            
            team_members = []
            
            # 現在の検査員を取得（検査員1～5）
            for i in range(1, 6):
                inspector_col = f'検査員{i}'
                inspector_name = result_df.at[index, inspector_col]
                
                if pd.notna(inspector_name) and str(inspector_name).strip() != '':
                    # スキル値や(新)を除去して実名を取得
                    actual_name = str(inspector_name)
                    if '(' in actual_name:
                        actual_name = actual_name.split('(')[0].strip()
                    
                    team_members.append(actual_name)
            
            # チーム情報を設定
            if len(team_members) > 1:
                team_info = f"チーム: {', '.join(team_members)}"
            elif len(team_members) == 1:
                team_info = f"個人: {team_members[0]}"
            else:
                # 検査員人数が0の場合、既存のチーム情報を確認
                current_team_info = result_df.at[index, 'チーム情報']
                if pd.notna(current_team_info) and str(current_team_info).strip().startswith('未割当'):
                    team_info = str(current_team_info)  # 既存の未割当理由を保持
                else:
                    team_info = "未割当"  # 理由が不明な場合はデフォルト
            
            result_df.at[index, 'チーム情報'] = team_info
            return team_info
            
        except Exception as e:
            self.log_message(f"チーム情報更新中にエラーが発生しました: {str(e)}")
            return ""
    
    def add_products_to_master(
        self,
        new_products: List[Dict[str, Any]],
        product_master_path: str
    ) -> None:
        """
        製品マスタ.xlsxに新しい品番を追加
        
        Args:
            new_products: 追加する品番のリスト（各要素は辞書形式）
            product_master_path: 製品マスタファイルのパス
        """
        try:
            if not new_products:
                return
            
            file_path = Path(product_master_path)
            if not file_path.exists():
                self.log_message(f"製品マスタファイルが見つかりません: {file_path}")
                return
            
            # Excelファイルを開く
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # B列（品番列）の最終行を取得
            last_row = ws.max_row
            b_column_values = [cell.value for cell in ws['B'] if cell.value is not None]
            
            # ヘッダー行を除いた最終データ行を取得
            if len(b_column_values) > 1:  # ヘッダー行を除く
                last_data_row = len(b_column_values)
            else:
                last_data_row = 1  # ヘッダーのみの場合
            
            # 新しい行を追加（最終データ行の次の行から）
            start_row = last_data_row + 1
            
            # 列のマッピング（列名から列番号に変換）
            header_row = 1
            column_map = {}
            for col_idx, cell in enumerate(ws[header_row], start=1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                column_map[cell.value] = col_letter
            
            # 品番列（B列）の検索
            product_column = 'B'  # 品番はB列
            product_name_column = 'C'  # 品名はC列
            process_column = 'D'  # 工程番号はD列
            inspection_time_column = 'E'  # 検査時間はE列
            auto_add_column = 'H'  # 自動追加はH列
            
            # 既存の品番を取得（重複チェック用）
            existing_products = set()
            for row in range(2, last_data_row + 1):
                product_cell = ws[f'{product_column}{row}']
                if product_cell.value:
                    existing_products.add(str(product_cell.value))
            
            # 新しい品番を追加
            added_count = 0
            current_row = start_row
            
            for product in new_products:
                product_number = str(product['品番'])
                
                # 重複チェック
                if product_number in existing_products:
                    self.log_message(f"ℹ️ 品番 '{product_number}' は既に製品マスタに存在するため、スキップします")
                    continue
                
                # 行を追加
                ws[f'{product_column}{current_row}'] = product_number
                ws[f'{product_name_column}{current_row}'] = product.get('品名', '')
                # D列（工程番号）には何も出力しない
                # ws[f'{process_column}{current_row}'] = product.get('工程番号', '')
                ws[f'{inspection_time_column}{current_row}'] = product.get('検査時間', 15.0)
                ws[f'{auto_add_column}{current_row}'] = True
                
                existing_products.add(product_number)
                added_count += 1
                current_row += 1
                
                self.log_message(f"製品マスタに追加: 品番={product_number}, 品名={product.get('品名', '')}, 工程番号={product.get('工程番号', '')}, 検査時間={product.get('検査時間', 15.0)}秒/個")
            
            # ファイルを保存
            if added_count > 0:
                wb.save(file_path)
                self.log_message(f"製品マスタファイルを保存しました: {added_count}件の品番を追加")
            else:
                wb.close()
                
        except Exception as e:
            error_msg = f"製品マスタへの追加処理中にエラーが発生しました: {str(e)}"
            self.log_message(error_msg)
            logger.error(error_msg, exc_info=True)
            raise
    
    def reset_assignment_history(self) -> None:
        """割り当て履歴をリセット"""
        self.inspector_assignment_count = {}
        self.inspector_last_assignment = {}
        self.inspector_work_hours = {}
        self.inspector_daily_assignments = {}
        self.inspector_product_hours = {}
        self.relaxed_product_limit_assignments = set()
        self.log_message("検査員割り当て履歴と勤務時間をリセットしました")
