from typing import Optional, Any
import pandas as pd
import os
from datetime import datetime
from tkinter import filedialog, messagebox
from loguru import logger
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelExporter:
    """Excelエクスポート機能を提供するクラス"""
    
    def __init__(self) -> None:
        """初期化"""
        pass
    
    def _is_file_open(self, file_path: str) -> bool:
        """
        ファイルが開かれているかチェック
        
        Args:
            file_path: ファイルパス
        
        Returns:
            ファイルが開かれている場合はTrue、それ以外はFalse
        """
        try:
            if os.path.exists(file_path):
                with open(file_path, 'a'):
                    pass
                return False
            return False
        except IOError:
            return True
    
    def _apply_header_style(
        self,
        writer: Any,
        sheet_name: str,
        df: pd.DataFrame
    ) -> None:
        """
        ヘッダー行にスタイルを適用
        
        Args:
            writer: ExcelWriterオブジェクト
            sheet_name: シート名
            df: DataFrame
        """
        try:
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # ヘッダー行のスタイル
            header_font = Font(name='Yu Gothic', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ヘッダー行（1行目）にスタイルを適用
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                
        except Exception as e:
            logger.error(f"ヘッダースタイル適用エラー: {str(e)}")
    
    def _adjust_column_widths(
        self,
        writer: Any,
        sheet_name: str,
        df: pd.DataFrame
    ) -> None:
        """
        列幅を自動調整（ベクトル化処理で高速化）
        
        Args:
            writer: ExcelWriterオブジェクト
            sheet_name: シート名
            df: DataFrame
        """
        try:
            worksheet = writer.sheets[sheet_name]
            
            for col_num, column_title in enumerate(df.columns, 1):
                # ヘッダーの長さを考慮
                header_length = len(str(column_title))
                
                # データの最大長を計算（ベクトル化処理で高速化）
                if df[column_title].notna().any():
                    # 文字列に変換してベクトル処理
                    column_str = df[column_title].astype(str)
                    # 文字列長を計算
                    str_lengths = column_str.str.len()
                    # 日本語文字（ASCII以外）の数を計算
                    japanese_chars = column_str.str.count(r'[^\x00-\x7F]')
                    # 日本語文字は2倍の幅として計算
                    total_lengths = str_lengths + japanese_chars
                    # 最大長を取得
                    max_data_length = int(total_lengths.max()) if len(total_lengths) > 0 else 0
                else:
                    max_data_length = 0
                
                # 列幅を計算（ヘッダーとデータの最大長を考慮）
                column_width = max(header_length, max_data_length) + 2
                
                # 最小幅と最大幅を設定
                column_width = max(8, min(column_width, 50))
                
                # 列幅を設定
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = column_width
                
        except Exception as e:
            logger.error(f"列幅調整エラー: {str(e)}")
    
    def export_lot_assignment_to_excel(self, assignment_df: pd.DataFrame) -> bool:
        """
        ロット割り当て結果をExcelファイルにエクスポート
        
        Args:
            assignment_df: ロット割り当て結果のDataFrame
        
        Returns:
            エクスポート成功時はTrue、失敗時はFalse
        """
        try:
            if assignment_df is None or assignment_df.empty:
                messagebox.showwarning("警告", "エクスポートするロット割り当てデータがありません。")
                return False
            
            # デフォルトファイル名を生成
            current_date = datetime.now().strftime("%Y%m%d")
            default_filename = f"出荷不足ロット_{current_date}.xlsx"
            
            # ファイル保存ダイアログを表示
            file_path = filedialog.asksaveasfilename(
                title="ロット割り当て結果をExcelファイルに保存",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename,
                initialdir=os.path.expanduser("~/Desktop")
            )
            
            if not file_path:
                return False
            
            # ファイルが既に開かれているかチェック
            if self._is_file_open(file_path):
                messagebox.showerror("エクスポートエラー", 
                    f"ファイルが既に開かれています: {file_path}\n"
                    "ファイルを閉じてから再度実行してください。")
                return False
            
            # 日付列をフォーマット
            formatted_df = assignment_df.copy()
            date_columns = ['出荷予定日', '指示日']
            for col in date_columns:
                if col in formatted_df.columns:
                    formatted_df[col] = pd.to_datetime(formatted_df[col]).dt.strftime('%Y/%m/%d')
            
            # Excelファイルに書き込み
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # ロット割り当て結果シート
                formatted_df.to_excel(
                    writer, 
                    sheet_name='ロット割り当て結果', 
                    index=False, 
                    startrow=0
                )
                
                # サマリーシート
                summary_data = {
                    '項目': ['エクスポート日時', '総件数', '総不足数'],
                    '値': [
                        datetime.now().strftime("%Y/%m/%d"),
                        len(formatted_df),
                        f"{formatted_df['不足数'].sum():.0f}" if '不足数' in formatted_df.columns else "0"
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(
                    writer, 
                    sheet_name='サマリー', 
                    index=False, 
                    startrow=0
                )
                
                # スタイルを適用
                self._apply_header_style(writer, 'ロット割り当て結果', formatted_df)
                self._apply_header_style(writer, 'サマリー', summary_df)
                self._adjust_column_widths(writer, 'ロット割り当て結果', formatted_df)
                self._adjust_column_widths(writer, 'サマリー', summary_df)
            
            messagebox.showinfo("エクスポート完了", f"ロット割り当て結果をエクスポートしました:\n{file_path}")
            logger.info(f"ロット割り当て結果Excelエクスポート完了: {file_path}")
            return True
            
        except PermissionError as e:
            error_msg = (
                f"ファイルの書き込み権限がありません: {file_path}\n\n"
                "考えられる原因:\n"
                "1. ファイルが他のアプリケーションで開かれている\n"
                "2. ファイルが読み取り専用になっている\n"
                "3. ディレクトリへの書き込み権限がない\n\n"
                "解決方法:\n"
                "1. ファイルを閉じてから再度実行してください\n"
                "2. 別のファイル名または場所を選択してください"
            )
            logger.error(f"ロット割り当て結果Excelエクスポート権限エラー: {str(e)}")
            messagebox.showerror("エクスポートエラー", error_msg)
            return False
        except Exception as e:
            error_msg = f"Excelエクスポート中にエラーが発生しました: {str(e)}"
            logger.error(error_msg)
            messagebox.showerror("エクスポートエラー", error_msg)
            return False
    
    def export_main_data_to_excel(self, main_df):
        """抽出データをExcelファイルにエクスポート"""
        try:
            if main_df is None or main_df.empty:
                messagebox.showwarning("警告", "エクスポートする抽出データがありません。")
                return False
            
            # デフォルトファイル名を生成
            current_date = datetime.now().strftime("%Y%m%d")
            default_filename = f"抽出データ_{current_date}.xlsx"
            
            # ファイル保存ダイアログを表示
            file_path = filedialog.asksaveasfilename(
                title="抽出データをExcelファイルに保存",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename,
                initialdir=os.path.expanduser("~/Desktop")
            )
            
            if not file_path:
                return False
            
            # ファイルが既に開かれているかチェック
            if self._is_file_open(file_path):
                messagebox.showerror("エクスポートエラー", 
                    f"ファイルが既に開かれています: {file_path}\n"
                    "ファイルを閉じてから再度実行してください。")
                return False
            
            # 日付列をフォーマット
            formatted_df = main_df.copy()
            date_columns = ['出荷予定日']
            for col in date_columns:
                if col in formatted_df.columns:
                    formatted_df[col] = pd.to_datetime(formatted_df[col]).dt.strftime('%Y/%m/%d')
            
            # Excelファイルに書き込み
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 抽出データシート
                formatted_df.to_excel(
                    writer, 
                    sheet_name='抽出データ', 
                    index=False, 
                    startrow=0
                )
                
                # サマリーシート
                summary_data = {
                    '項目': ['エクスポート日時', '総件数', '不足数マイナス件数'],
                    '値': [
                        datetime.now().strftime("%Y/%m/%d"),
                        len(formatted_df),
                        len(formatted_df[formatted_df['不足数'] < 0]) if '不足数' in formatted_df.columns else "0"
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(
                    writer, 
                    sheet_name='サマリー', 
                    index=False, 
                    startrow=0
                )
                
                # スタイルを適用
                self._apply_header_style(writer, '抽出データ', formatted_df)
                self._apply_header_style(writer, 'サマリー', summary_df)
                self._adjust_column_widths(writer, '抽出データ', formatted_df)
                self._adjust_column_widths(writer, 'サマリー', summary_df)
            
            messagebox.showinfo("エクスポート完了", f"抽出データをエクスポートしました:\n{file_path}")
            logger.info(f"抽出データExcelエクスポート完了: {file_path}")
            return True
            
        except PermissionError as e:
            error_msg = (
                f"ファイルの書き込み権限がありません: {file_path}\n\n"
                "考えられる原因:\n"
                "1. ファイルが他のアプリケーションで開かれている\n"
                "2. ファイルが読み取り専用になっている\n"
                "3. ディレクトリへの書き込み権限がない\n\n"
                "解決方法:\n"
                "1. ファイルを閉じてから再度実行してください\n"
                "2. 別のファイル名または場所を選択してください"
            )
            logger.error(f"メインデータExcelエクスポート権限エラー: {str(e)}")
            messagebox.showerror("エクスポートエラー", error_msg)
            return False
        except Exception as e:
            error_msg = f"Excelエクスポート中にエラーが発生しました: {str(e)}"
            logger.error(error_msg)
            messagebox.showerror("エクスポートエラー", error_msg)
            return False
    
    def export_inspector_assignment_to_excel(self, inspector_df):
        """検査員割振り結果をExcelファイルにエクスポート"""
        try:
            if inspector_df is None or inspector_df.empty:
                messagebox.showwarning("警告", "エクスポートする検査員割振りデータがありません。")
                return False
            
            # デフォルトファイル名を生成
            current_date = datetime.now().strftime("%Y%m%d")
            default_filename = f"検査員割振り結果_{current_date}.xlsx"
            
            # ファイル保存ダイアログを表示
            file_path = filedialog.asksaveasfilename(
                title="検査員割振り結果をExcelファイルに保存",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename,
                initialdir=os.path.expanduser("~/Desktop")
            )
            
            if not file_path:
                return False
            
            # ファイルが既に開かれているかチェック
            if self._is_file_open(file_path):
                messagebox.showerror("エクスポートエラー", 
                    f"ファイルが既に開かれています: {file_path}\n"
                    "ファイルを閉じてから再度実行してください。")
                return False
            
            # 日付列をフォーマット
            formatted_df = inspector_df.copy()
            date_columns = ['出荷予定日', 'ロット日']
            for col in date_columns:
                if col in formatted_df.columns:
                    formatted_df[col] = pd.to_datetime(formatted_df[col]).dt.strftime('%Y/%m/%d')
            
            # Excelファイルに書き込み
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 検査員割振り結果シート
                formatted_df.to_excel(
                    writer, 
                    sheet_name='検査員割振り結果', 
                    index=False, 
                    startrow=0
                )
                
                # サマリーシート
                summary_data = {
                    '項目': ['エクスポート日時', '総件数', '総検査時間(時間)'],
                    '値': [
                        datetime.now().strftime("%Y/%m/%d"),
                        len(formatted_df),
                        f"{formatted_df['検査時間'].sum():.1f}" if '検査時間' in formatted_df.columns else "0.0"
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(
                    writer, 
                    sheet_name='サマリー', 
                    index=False, 
                    startrow=0
                )
                
                # スタイルを適用
                self._apply_header_style(writer, '検査員割振り結果', formatted_df)
                self._apply_header_style(writer, 'サマリー', summary_df)
                self._adjust_column_widths(writer, '検査員割振り結果', formatted_df)
                self._adjust_column_widths(writer, 'サマリー', summary_df)
            
            messagebox.showinfo("エクスポート完了", f"検査員割振り結果をエクスポートしました:\n{file_path}")
            logger.info(f"検査員割振り結果Excelエクスポート完了: {file_path}")
            return True
            
        except PermissionError as e:
            error_msg = (
                f"ファイルの書き込み権限がありません: {file_path}\n\n"
                "考えられる原因:\n"
                "1. ファイルが他のアプリケーションで開かれている\n"
                "2. ファイルが読み取り専用になっている\n"
                "3. ディレクトリへの書き込み権限がない\n\n"
                "解決方法:\n"
                "1. ファイルを閉じてから再度実行してください\n"
                "2. 別のファイル名または場所を選択してください"
            )
            logger.error(f"検査員割振り結果Excelエクスポート権限エラー: {str(e)}")
            messagebox.showerror("エクスポートエラー", error_msg)
            return False
        except Exception as e:
            error_msg = f"Excelエクスポート中にエラーが発生しました: {str(e)}"
            logger.error(error_msg)
            messagebox.showerror("エクスポートエラー", error_msg)
            return False# �g�p��:
# exporter = ExcelExporter()
# exporter.export_main_data_to_excel(...)
